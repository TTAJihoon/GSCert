import os
import re
import sys
import time
import logging
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook, Workbook

# playwright / pywinauto 는 레거시 'playwright' 다운로드 경로에서만 필요하므로
# top-level 에서 import 하지 않는다(HTTP 직접연동 경로는 requests 만 있으면 동작).
# 각각 해당 분기/함수 안에서 lazy import 한다.

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    # `python main/utils/weekly.py` 로 실행하면 script 디렉터리만 path 에 들어가므로,
    # main.views.review.ecm_http_client 를 import 할 수 있도록 저장소 루트를 추가한다.
    sys.path.insert(0, str(PROJECT_ROOT))
DATA_DIR = PROJECT_ROOT / "main" / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = DATA_DIR / "weekly_gs_sync.log"
REFERENCE_SHEET_NAME = "인증획득제품리스트"

MASTER_BASE_COLUMN_COUNT = 14  # reference.xlsx A~N
MASTER_RECERT_TYPE_COLUMN = 15  # reference.xlsx O
MASTER_PREV_CERT_INFO_COLUMN = 16  # reference.xlsx P
MASTER_KOLAS_COLUMN = 17  # reference.xlsx Q
MASTER_TOTAL_COLUMN_COUNT = MASTER_KOLAS_COLUMN
ECM_RECERT_TYPE_COLUMN = 25  # ECM 인증획득제품리스트 Y
ECM_PREV_CERT_INFO_COLUMN = 26  # ECM 인증획득제품리스트 Z
RECERT_TYPE_HEADER = "재인증구분"
PREV_CERT_INFO_HEADER = "기인증번호제품정보버전"
KOLAS_HEADER = "KOLAS"
# '기타정보' 시트: I열(성적서구분)이 KOLAS인 행의 B열(인증번호)을 모아
# '인증획득제품리스트' B열과 매칭시켜 reference.xlsx Q열에 표시한다.
OTHER_INFO_SHEET_NAME = "기타정보"
OTHER_INFO_DATA_START_ROW = 4
OTHER_INFO_CERT_NUMBER_COLUMN = 2  # 기타정보 B: 인증번호
OTHER_INFO_REPORT_TYPE_COLUMN = 9  # 기타정보 I: 성적서구분
KOLAS_REPORT_TYPE_VALUE = "KOLAS"
MASTER_HEADERS = [
    "일련번호",
    "인증번호",
    "인증일자",
    "회사명",
    "제품",
    "등급",
    "시험번호",
    "SW분류",
    "제품설명",
    "총WD",
    "재계약",
    "특이사항",
    "시작날짜종료날짜",
    "시험원",
    RECERT_TYPE_HEADER,
    PREV_CERT_INFO_HEADER,
    KOLAS_HEADER,
]


def _env_path(name: str, default: Path) -> Path:
    value = os.environ.get(name)
    return Path(value) if value else default


def _env_optional_path(name: str) -> Path | None:
    value = os.environ.get(name)
    return Path(value) if value else None


def _default_python_executable() -> Path:
    candidates = [
        PROJECT_ROOT / ".venv" / "Scripts" / "python.exe",
        PROJECT_ROOT / "venv" / "Scripts" / "python.exe",
        PROJECT_ROOT.parent / ".venv" / "Scripts" / "python.exe",
        PROJECT_ROOT.parent / "venv" / "Scripts" / "python.exe",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return Path(sys.executable)


# =========================
# 설정
# =========================
@dataclass
class Config:
    # ===== 테스트용 (끝나면 False로 바꾸거나 지우면 됨) =====
    test_click_doc_enabled = False

    project_root: Path = PROJECT_ROOT

    # 기준 파일(이제 xlsx로 관리)
    master_xlsx: Path = _env_path("GSCERT_REFERENCE_XLSX", DATA_DIR / "reference.xlsx")

    # 다운로드 폴더
    download_folder: Path = _env_path("GSCERT_WEEKLY_DOWNLOAD_DIR", Path(r"C:\Users\Administrator"))

    # 특정 주차를 수동 갱신할 때 사용한다. 예: 20260511
    target_monday: str = os.environ.get("GSCERT_WEEKLY_TARGET_DATE", "").strip()

    # 이미 받은 xlsx를 직접 기준 파일에 반영할 때 사용한다.
    source_xlsx: Path | None = _env_optional_path("GSCERT_WEEKLY_SOURCE_XLSX")

    # 시작 URL (= 분당 Destiny ECM 서버)
    start_url: str = os.environ.get("GSCERT_WEEKLY_START_URL", "http://210.104.181.10")

    # 다운로드 방식: "http"(서버 HTTP 직접연동, 기본) / "playwright"(레거시 브라우저+팝업).
    # 문제 시 GSCERT_WEEKLY_SOURCE=playwright 로 즉시 롤백한다.
    download_source: str = os.environ.get("GSCERT_WEEKLY_SOURCE", "http").strip().lower()

    # HTTP 직접연동 root OID(분당 = C_ROOT). 자격증명은 ECM_USERNAME_BUNDANG/ECM_PASSWORD_BUNDANG.
    ecm_root_oid: str = os.environ.get("ECM_ROOT_OID_BUNDANG", "") or "C_ROOT"

    # 로그인 세션 저장(레거시 playwright 경로에서만 사용: 최초 1회 로그인 후 재사용)
    storage_state: Path = _env_path("GSCERT_EDM_STORAGE_STATE", DATA_DIR / "edm_storage_state.json")

    # reference.xlsx -> PostgreSQL 적재
    python_executable: Path = _env_path("GSCERT_PYTHON", _default_python_executable())
    manage_py: Path = _env_path("GSCERT_MANAGE_PY", PROJECT_ROOT / "manage.py")
    django_settings: str = os.environ.get("GSCERT_DJANGO_SETTINGS", "myproject.settings")

    # 타임아웃/대기
    pw_timeout_ms: int = 30_000
    dialog_wait_sec: int = 15
    download_wait_sec: int = 180

    # 트리/문서 규칙
    year_folder_suffix: str = "시험서비스"
    zero_folder_prefix_re: re.Pattern = re.compile(r"^00\s")
    doc_prefix: str = "인증획득제품"

CFG = Config()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.StreamHandler(sys.stdout), logging.FileHandler(LOG_FILE, encoding="utf-8")],
)

class _StreamToLogger:
    def __init__(self, logger, level):
        self.logger = logger
        self.level = level
        self._buf = ""

    def write(self, message):
        if not message:
            return
        self._buf += message
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            line = line.rstrip()
            if line:
                self.logger.log(self.level, line)

    def flush(self):
        if self._buf.strip():
            self.logger.log(self.level, self._buf.strip())
        self._buf = ""

# stdout/stderr를 logging으로 흡수 (print/traceback 포함)
_root = logging.getLogger()
sys.stdout = _StreamToLogger(_root, logging.INFO)
sys.stderr = _StreamToLogger(_root, logging.ERROR)
    

# =========================
# 공통 유틸
# =========================
def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _reference_sheet(wb):
    if REFERENCE_SHEET_NAME in wb.sheetnames:
        return wb[REFERENCE_SHEET_NAME]
    return wb.active


def _serial_value(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            if float(value).is_integer():
                return int(value)
        except (OverflowError, ValueError):
            return None
    s = str(value).strip().strip('"').strip("'")
    m = re.fullmatch(r"(\d+)(?:\.0+)?", s)
    if m:
        return int(m.group(1))
    return None


# =========================
# 날짜(전 주 월요일)
# =========================
def this_week_monday_yyyymmdd(tz: str = "Asia/Seoul") -> str:
    target_monday = resolve_target_monday_arg()
    if target_monday:
        try:
            datetime.strptime(target_monday, "%Y%m%d")
        except ValueError as exc:
            raise ValueError("대상 날짜는 YYYYMMDD 형식이어야 합니다.") from exc
        return target_monday

    now = datetime.now(ZoneInfo(tz))
    monday = now - timedelta(days=now.weekday()) - timedelta(days=7)
    return monday.strftime("%Y%m%d")


def resolve_target_monday_arg() -> str:
    if len(sys.argv) > 1 and sys.argv[1].strip():
        return sys.argv[1].strip()
    return CFG.target_monday


# =========================
# 기준 파일(xlsx) 처리
# =========================
def read_last_serial_from_master_xlsx(xlsx_path: Path) -> int:
    """
    reference.xlsx에서 A열(일련번호) 마지막 숫자를 robust하게 찾는다.
    - A열을 아래에서 위로 훑으며 가장 마지막 숫자(정수)를 반환
    - 숫자가 하나도 없으면 최초 적재로 보고 0을 반환
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"master file not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    ws = _reference_sheet(wb)

    for r in range(ws.max_row, 0, -1):
        serial = _serial_value(ws.cell(row=r, column=1).value)
        if serial is not None:
            return serial

    logging.info("master A열에서 일련번호를 찾지 못했습니다. 최초 적재로 간주하고 0부터 시작합니다.")
    return 0


def append_rows_to_master_xlsx(master_xlsx: Path, rows: list[list], ensure_cols: bool = True) -> None:
    """
    master.xlsx 마지막 행 다음에 다운로드 A~N + ECM Y/Z를 append.
    ECM Y/Z는 reference.xlsx O/P(15/16열)에 입력한다.
    줄바꿈(\n) 포함 문자열은 그대로 셀에 들어감.
    """
    master_xlsx.parent.mkdir(parents=True, exist_ok=True)

    if master_xlsx.exists():
        wb = load_workbook(master_xlsx)
    else:
        wb = Workbook()

    ws = _reference_sheet(wb)

    # 헤더행이 없으면 최초 적재 파일로 보고 A~P 헤더를 만든다.
    # 이미 헤더가 있어도 O/P는 적재용 표준 헤더로 맞춘다. O열이 기존 K열 "재계약"과
    # 중복되면 pandas 적재 단계에서 컬럼이 모호해지므로 여기서 바로잡는다.
    if _is_blank(ws.cell(row=1, column=1).value):
        for c_idx, header in enumerate(MASTER_HEADERS, start=1):
            ws.cell(row=1, column=c_idx, value=header)
    else:
        ws.cell(row=1, column=MASTER_RECERT_TYPE_COLUMN, value=RECERT_TYPE_HEADER)
        ws.cell(row=1, column=MASTER_PREV_CERT_INFO_COLUMN, value=PREV_CERT_INFO_HEADER)
        ws.cell(row=1, column=MASTER_KOLAS_COLUMN, value=KOLAS_HEADER)

    # 마지막 "의미 있는" 행 찾기: A열이 비어있지 않은 마지막 행 기준
    last = ws.max_row
    while last > 1 and _is_blank(ws.cell(row=last, column=1).value):
        last -= 1
    write_row = last + 1

    for row in rows:
        if ensure_cols:
            row = (row + [None] * MASTER_TOTAL_COLUMN_COUNT)[:MASTER_TOTAL_COLUMN_COUNT]

        for c_idx in range(1, MASTER_TOTAL_COLUMN_COUNT + 1):  # A..P
            v = row[c_idx - 1]
            ws.cell(row=write_row, column=c_idx, value=v)
        write_row += 1

    wb.save(master_xlsx)


# =========================
# 다운로드 xlsx에서 범위 추출
# =========================
def extract_a_to_n_rows_after_serial(xlsx_path: Path, start_serial: int, sheet_name: str | None = None) -> list[list]:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")

    target_sheet = sheet_name or REFERENCE_SHEET_NAME
    wb = load_workbook(xlsx_path, data_only=True)
    if target_sheet not in wb.sheetnames:
        raise ValueError(f"시트 '{target_sheet}' 를 찾지 못했습니다. 현재 시트: {wb.sheetnames}")
    ws = wb[target_sheet]

    found_row = None
    if int(start_serial) <= 0:
        for r in range(1, ws.max_row + 1):
            serial = _serial_value(ws.cell(row=r, column=1).value)
            if serial is not None and serial >= 1:
                found_row = r
                break

        if found_row is None:
            raise ValueError("다운로드 엑셀 A열에서 가져올 일련번호(1 이상)를 찾지 못했습니다.")

        start_row = found_row
    else:
        for r in range(1, ws.max_row + 1):
            serial = _serial_value(ws.cell(row=r, column=1).value)
            if serial == int(start_serial):
                found_row = r
                break

        if found_row is None:
            raise ValueError(f"다운로드 엑셀 A열에서 일련번호 {start_serial} 를 찾지 못했습니다.")

        start_row = found_row + 1

    last_data_row = 0
    for r in range(1, ws.max_row + 1):
        any_val = False
        for c in range(1, MASTER_BASE_COLUMN_COUNT + 1):  # A..N
            if ws.cell(row=r, column=c).value not in (None, ""):
                any_val = True
                break
        if any_val:
            last_data_row = r

    if last_data_row < start_row:
        return []

    out = []
    for r in range(start_row, last_data_row + 1):
        an = [ws.cell(row=r, column=c).value for c in range(1, MASTER_BASE_COLUMN_COUNT + 1)]  # A..N
        if all(v in (None, "") for v in an):
            continue
        # ECM 시트 Y/Z 값을 reference.xlsx O/P로 쓰기 위해 A~N 뒤에 붙인다.
        yz = [
            ws.cell(row=r, column=ECM_RECERT_TYPE_COLUMN).value,
            ws.cell(row=r, column=ECM_PREV_CERT_INFO_COLUMN).value,
        ]
        out.append(an + yz)
    return out


# =========================
# '기타정보' 시트 KOLAS 매칭
# =========================
def read_kolas_cert_numbers(xlsx_path: Path) -> set[str]:
    """'기타정보' 시트에서 I열(성적서구분)이 KOLAS인 행의 B열(인증번호)을 모아 반환한다."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if OTHER_INFO_SHEET_NAME not in wb.sheetnames:
        logging.warning("'%s' 시트를 찾지 못해 KOLAS 매칭을 건너뜁니다.", OTHER_INFO_SHEET_NAME)
        return set()

    ws = wb[OTHER_INFO_SHEET_NAME]
    result: set[str] = set()
    for row in ws.iter_rows(
        min_row=OTHER_INFO_DATA_START_ROW,
        max_col=OTHER_INFO_REPORT_TYPE_COLUMN,
        values_only=True,
    ):
        report_type = row[OTHER_INFO_REPORT_TYPE_COLUMN - 1]
        if report_type is None or str(report_type).strip() != KOLAS_REPORT_TYPE_VALUE:
            continue
        cert_number = row[OTHER_INFO_CERT_NUMBER_COLUMN - 1]
        if cert_number is not None:
            result.add(str(cert_number).strip())
    return result


def append_kolas_column(rows: list[list], kolas_cert_numbers: set[str]) -> list[list]:
    """A~N + Y/Z(16개) 행 각각에 KOLAS 여부(17번째 값)를 덧붙인다.

    B열(인증번호, row[1])이 kolas_cert_numbers 에 있으면 'KOLAS', 없으면 빈 문자열.
    """
    out = []
    for row in rows:
        cert_number = row[1]
        is_kolas = cert_number is not None and str(cert_number).strip() in kolas_cert_numbers
        out.append(list(row) + [KOLAS_HEADER if is_kolas else ""])
    return out


# =========================
# 행 정규화(요청사항 반영)
# =========================
def normalize_rows(rows: list[list]) -> list[list]:
    """
    요청 반영 규칙:

    (1) 추가 기업명 행 합치기:
        - 조건: A 비어있고 AND C 비어있고 AND D 존재
        - 동작: 바로 위 행 D에 '\\n' + D를 붙이고 현재 행은 삭제

    (2) A만 비어있는 경우 삭제:
        - 조건: A 비어있고 AND (B 또는 C 값이 존재)
        - 동작: 현재 행은 삭제

    주의: (1)이 (2)보다 먼저 평가되어야 함.
    """
    out: list[list] = []

    for row in rows:
        row = (row + [None] * MASTER_TOTAL_COLUMN_COUNT)[:MASTER_TOTAL_COLUMN_COUNT]  # A..N + O/P/Q 고정

        # 완전 빈 행 제거
        if all(_is_blank(v) for v in row):
            continue

        a = row[0]  # A: 일련번호
        b = row[1]  # B: 인증번호
        c = row[2]  # C: 인증일자
        d = row[3]  # D: 회사명

        # (1) 추가 기업명 행 합치기: A blank & C blank & D has value
        if _is_blank(a) and _is_blank(c) and not _is_blank(d) and out:
            prev = out[-1]
            prev_d = "" if _is_blank(prev[3]) else str(prev[3]).strip()
            cur_d = str(d).strip()
            # 줄바꿈 유지
            out[-1][3] = (prev_d + "\n" + cur_d).strip()
            continue  # 현재 행 삭제

        # (2) A blank & (B or C has value) => 삭제
        if _is_blank(a) and (not _is_blank(b) or not _is_blank(c)):
            continue

        out.append(row)

    return out


# =========================
# source_xlsx 경로 해석
# =========================
def _xlsx_date_key(p: Path) -> str:
    m = re.search(r"(\d{8})", p.name)
    if m:
        return m.group(1)
    return datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y%m%d")


def _resolve_source_xlsx(path: Path) -> Path:
    """파일이면 그대로, 디렉토리면 '인증획득제품' 포함 xlsx 중 가장 최신 파일 반환."""
    if path.is_file():
        return path
    if path.is_dir():
        candidates = sorted(path.glob("*인증획득제품*.xlsx"), key=_xlsx_date_key, reverse=True)
        if not candidates:
            raise FileNotFoundError(f"디렉토리에서 '인증획득제품' xlsx를 찾을 수 없습니다: {path}")
        logging.info("'인증획득제품' 최신 파일 자동 선택: %s", candidates[0].name)
        return candidates[0]
    raise FileNotFoundError(f"경로가 존재하지 않습니다: {path}")


# =========================
# UIA: "폴더 찾아보기" 대화상자 처리 (Enter만)
# =========================
def confirm_browse_dialog_by_enter(wait_popup_sec: int = 15, after_popup_sec: float = 3.0):
    """
    '폴더 찾아보기' 팝업이 뜬 뒤 after_popup_sec(기본 3초) 기다렸다가 Enter만 눌러 진행.
    - 기본 폴더/최근 폴더가 이미 원하는 경로로 선택되어 있다는 전제
    - 레거시 playwright 경로 전용(HTTP 직접연동에서는 이 팝업 자체가 없다).
    """
    from pywinauto import Desktop
    from pywinauto.keyboard import send_keys

    # 1) 팝업이 뜰 때까지 최대 wait_popup_sec 동안 기다림 (Win32로 가볍게 체크)
    dlg = None
    end = time.time() + max(1, int(wait_popup_sec))
    while time.time() < end:
        try:
            cand = Desktop(backend="win32").window(title="폴더 찾아보기", class_name="#32770")
            if cand.exists(timeout=0.2):
                dlg = cand
                break
        except Exception:
            pass
        time.sleep(0.2)

    # 2) 팝업이 확인되면 포커스 주고(가능하면), 3초 기다렸다가 Enter
    if dlg is not None:
        try:
            dlg.set_focus()
        except Exception:
            pass

    logging.info("폴더 선택 팝업 대기 후 Enter 입력: %.1fs", after_popup_sec)
    time.sleep(max(0.0, float(after_popup_sec)))

    send_keys("{ENTER}")


# =========================
# 다운로드 완료 대기(파일 생성 + 크기 안정화)
# =========================
def wait_for_file_complete(folder: Path, expected_name: str, timeout_sec: int) -> Path:
    folder = folder.resolve()
    end = time.time() + timeout_sec
    target = folder / expected_name

    last_size = -1
    stable_count = 0

    while time.time() < end:
        if target.exists():
            try:
                size = target.stat().st_size
            except OSError:
                time.sleep(0.3)
                continue

            if size == last_size and size > 0:
                stable_count += 1
            else:
                stable_count = 0
            last_size = size

            if stable_count >= 3:
                return target

        time.sleep(0.5)

    raise TimeoutError(f"다운로드 파일이 완료되지 않았습니다: {target}")


# =========================
# Playwright: 웹 탐색/저장 트리거
# =========================
def ensure_page(p):
    # ECM 서버는 사설망(로컬 네트워크)이라 Chrome의 Local Network Access 검사에 막힐 수 있다.
    # chrome://flags 수동 변경은 Playwright가 띄우는 별도 Chromium에는 적용되지 않으므로,
    # 다운로드 워커(launch_browser)와 동일하게 실행 인자로 해당 검사를 비활성화한다.
    browser = p.chromium.launch(
        headless=False,  # UI 팝업 때문에 headful 필수
        args=[
            "--disable-features=LocalNetworkAccessCheck,LocalNetworkAccessChecks",
            "--disable-local-network-access-check",
        ],
    )
    context_kwargs = {
        "accept_downloads": False,  # 브라우저 다운로드가 아니라 OS 팝업 방식이라 의미 없음
        "viewport": {"width": 1400, "height": 900},
    }
    if CFG.storage_state.exists():
        context_kwargs["storage_state"] = str(CFG.storage_state)

    ctx = browser.new_context(**context_kwargs)
    page = ctx.new_page()
    page.set_default_timeout(CFG.pw_timeout_ms)

    page.goto(CFG.start_url, wait_until="domcontentloaded")

    if not CFG.storage_state.exists():
        logging.warning("처음 실행: 브라우저에서 로그인/SSO 완료 후 Enter를 누르세요.")
        input("로그인 완료 후 Enter: ")
        ctx.storage_state(path=str(CFG.storage_state))
        logging.info("storage_state 저장 완료: %s", CFG.storage_state)

    return browser, ctx, page


def click_year_and_first_00_folder(page, year: int):
    from playwright.sync_api import TimeoutError as PWTimeout

    year_text = f"{year} {CFG.year_folder_suffix}"
    logging.info("연도 폴더 클릭: %s", year_text)

    page.locator('a[menuname="edm-folder-context-tree"]').first.wait_for(state="visible", timeout=60_000)

    year_a = page.locator(
        'a[menuname="edm-folder-context-tree"]',
        has_text=re.compile(rf"^\s*{year}\s*{re.escape(CFG.year_folder_suffix)}\s*$")
    ).first

    year_a.scroll_into_view_if_needed()
    year_a.click()
    page.wait_for_timeout(300)

    year_li = year_a.locator("xpath=ancestor::li[1]")

    cls = (year_li.get_attribute("class") or "")
    if "jstree-closed" in cls:
        expander = year_li.locator("ins.jstree-icon").first
        expander.click()
        page.wait_for_function(
            """(el) => (el.className || '').includes('jstree-open')""",
            arg=year_li,
            timeout=60_000
        )
    else:
        try:
            page.wait_for_function(
                """(el) => (el.querySelectorAll('ul li a[menuname="edm-folder-context-tree"]').length > 0)""",
                arg=year_li,
                timeout=10_000
            )
        except Exception:
            pass

    zero_a = year_li.locator('a[menuname="edm-folder-context-tree"]').filter(
        has_text=re.compile(r"^\s*00\s")
    ).first

    try:
        zero_a.wait_for(state="visible", timeout=60_000)
    except PWTimeout:
        zero_a = page.locator('a[menuname="edm-folder-context-tree"]').filter(
            has_text=re.compile(rf"^\s*00\s+{year}년")
        ).first
        zero_a.wait_for(state="visible", timeout=60_000)

    zero_name = zero_a.inner_text()
    logging.info("00 폴더 클릭: %s", zero_name)

    zero_a.scroll_into_view_if_needed()
    zero_a.click()
    page.wait_for_timeout(800)


def open_doc_properties_by_monday(page, monday: str):
    doc_title = f"{CFG.doc_prefix}({monday})"
    logging.info("문서 클릭: %s", doc_title)
    page.locator("span.document-list-item-name-text-span", has_text=doc_title).first.click()
    page.wait_for_timeout(900)


def trigger_save_icon_for_attachment(page, monday: str):
    filename = f"{CFG.doc_prefix}({monday}).xlsx"
    logging.info("첨부파일 저장 아이콘 클릭(폴더 선택 팝업 유발): %s", filename)

    row = page.locator('tr.prop-view-file-list-item', has_text=filename).first
    save_btn = row.locator('div[events="document-fileSave-click"]').first
    save_btn.click()


def sync_reference_db():
    import subprocess

    if not CFG.manage_py.exists():
        raise FileNotFoundError(f"manage.py 파일을 찾을 수 없습니다: {CFG.manage_py}")

    cmd = [
        str(CFG.python_executable),
        str(CFG.manage_py),
        "import_reference_db",
        "--source-xlsx",
        str(CFG.master_xlsx),
        "--settings",
        CFG.django_settings,
    ]

    env = os.environ.copy()
    env.setdefault("PYTHONIOENCODING", "utf-8")
    env.setdefault("PYTHONUTF8", "1")

    logging.info("reference PostgreSQL 적재 실행: %s", " ".join(cmd))
    completed = subprocess.run(
        cmd,
        cwd=str(CFG.project_root),
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        check=False,
        env=env,
    )
    if completed.stdout.strip():
        logging.info("reference DB 적재 stdout: %s", completed.stdout.strip())
    if completed.stderr.strip():
        logging.error("reference DB 적재 stderr: %s", completed.stderr.strip())
    if completed.returncode != 0:
        raise RuntimeError(f"reference DB 적재 실패: exit_code={completed.returncode}")

    logging.info("reference PostgreSQL DB 적재 완료")


def sync_new_reference_projects(since_serial: int):
    """SwData에 이번 실행으로 새로 들어온 건을 ReferenceProject(점검대상 프로젝트 목록)에 반영한다.

    예전에는 launcher.ps1의 'G'(Google Sheets 동기화)가 구글시트 전체를 기준으로
    ReferenceProject를 채웠지만, 이제는 SwData(인증획득목록 엑셀)가 기준이라
    weekly 동기화가 방금 추가한 신규 건만 골라 구글시트(신청일/계약일 보완용)와
    매칭해 반영한다. 'G'는 더 이상 필요 없어 launcher 메뉴에서 제거했다(관리
    명령 자체는 전체 재동기화/복구용으로 남겨둠).
    """
    import subprocess

    if not CFG.manage_py.exists():
        raise FileNotFoundError(f"manage.py 파일을 찾을 수 없습니다: {CFG.manage_py}")

    cmd = [
        str(CFG.python_executable),
        str(CFG.manage_py),
        "sync_new_certified_projects",
        "--since-serial",
        str(since_serial),
        "--settings",
        CFG.django_settings,
    ]

    env = os.environ.copy()
    env.setdefault("PYTHONIOENCODING", "utf-8")
    env.setdefault("PYTHONUTF8", "1")

    logging.info("신규 점검대상 프로젝트 반영 실행: %s", " ".join(cmd))
    completed = subprocess.run(
        cmd,
        cwd=str(CFG.project_root),
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        check=False,
        env=env,
    )
    if completed.stdout.strip():
        logging.info("신규 프로젝트 반영 stdout: %s", completed.stdout.strip())
    if completed.stderr.strip():
        logging.error("신규 프로젝트 반영 stderr: %s", completed.stderr.strip())
    if completed.returncode != 0:
        raise RuntimeError(f"신규 프로젝트 반영 실패: exit_code={completed.returncode}")

    logging.info("신규 점검대상 프로젝트 반영 완료")


# =========================
# HTTP 직접연동 다운로드 (서버 → ECM, requests)
# =========================
_LIST_DATE_RE = re.compile(r"(\d{8})")


def _weekly_http_client():
    """분당 ECM(HTTP 직접연동) 클라이언트 생성 + 로그인.

    XOR 로그인이 브라우저 SSO(storage_state)를 대체한다. 자격증명은 워커와 동일하게
    ECM_USERNAME_BUNDANG/ECM_PASSWORD_BUNDANG 환경변수에서만 읽는다.
    """
    from main.views.review.ecm_http_client import DestinyECM

    user = os.environ.get("ECM_USERNAME_BUNDANG", "") or os.environ.get("ECM_USERNAME", "")
    pw = os.environ.get("ECM_PASSWORD_BUNDANG", "") or os.environ.get("ECM_PASSWORD", "")
    if not user or not pw:
        raise RuntimeError(
            "HTTP 다운로드에는 ECM_USERNAME_BUNDANG/ECM_PASSWORD_BUNDANG 환경변수가 필요합니다."
        )
    client = DestinyECM(CFG.start_url, CFG.ecm_root_oid, user, pw)
    client.login()
    logging.info("ECM HTTP 로그인 완료: %s (root=%s)", CFG.start_url, CFG.ecm_root_oid)
    return client


def _find_zero_folder_oid(client, year: int):
    """{year} 시험서비스 아래의 '00 …' 폴더 OID 를 찾는다."""
    service_oid = client.find_year_folder(str(year))
    if not service_oid:
        return None
    for child in client.children(service_oid):
        name = str(child.get("name", "")).strip()
        if CFG.zero_folder_prefix_re.match(name):
            return client.oid(child)
    return None


def select_latest_list_file(client, years):
    """00 폴더의 파일 중 '인증획득제품…(YYYYMMDD).xlsx' 이름의 날짜가 가장 최근인 파일 메타 반환.

    years 를 순서대로 훑어(예: 올해→전년) 후보가 나오면 그 연도에서 최댓값을 고른다.
    """
    for year in years:
        oid = _find_zero_folder_oid(client, year)
        if not oid:
            continue
        candidates = []
        for f in client.files(oid):
            name = f.get("fileName") or ""
            if CFG.doc_prefix in name and name.lower().endswith(".xlsx"):
                m = _LIST_DATE_RE.search(name)
                if m:
                    candidates.append((m.group(1), f))
        if candidates:
            candidates.sort(key=lambda x: x[0], reverse=True)
            latest_date, meta = candidates[0]
            logging.info(
                "최근 목록 파일 선택: %s (날짜 %s, 후보 %d개, %d년 폴더)",
                meta.get("fileName"), latest_date, len(candidates), year,
            )
            return meta
    return None


def download_latest_list_via_http() -> Path:
    """HTTP 직접연동으로 가장 최근 날짜의 인증획득제품 목록 xlsx 를 내려받는다."""
    client = _weekly_http_client()
    this_year = int(this_week_monday_yyyymmdd()[:4])
    meta = select_latest_list_file(client, [this_year, this_year - 1])
    if not meta:
        raise RuntimeError(
            f"ECM 에서 '{CFG.doc_prefix} …(YYYYMMDD).xlsx' 목록 파일을 찾지 못했습니다."
        )
    data = client.download_bytes(meta)
    if not data.startswith(b"PK"):
        raise RuntimeError(
            f"다운로드 결과가 xlsx(PK) 형식이 아닙니다: {meta.get('fileName')} ({data[:4]!r})"
        )
    CFG.download_folder.mkdir(parents=True, exist_ok=True)
    out = CFG.download_folder / meta["fileName"]
    tmp = out.with_name(out.name + ".part")
    tmp.write_bytes(data)
    tmp.replace(out)
    logging.info("HTTP 다운로드 완료: %s (%d bytes)", out, len(data))
    return out


# =========================
# main
# =========================
def main():
    CFG.download_folder.mkdir(parents=True, exist_ok=True)

    # 1) 기준 파일 A열 마지막 숫자
    last_serial = read_last_serial_from_master_xlsx(CFG.master_xlsx)
    logging.info("master 마지막 일련번호(A열): %s", last_serial)

    # 2) 이번주 월요일 파일명 결정
    monday = this_week_monday_yyyymmdd()
    if getattr(CFG, "test_click_doc_enabled", False):
        monday = "20260105"  # 테스트 끝나면 이 블록 삭제하거나 False로
        logging.info("[TEST] monday 강제 설정: %s", monday)
    if resolve_target_monday_arg():
        logging.info("대상 주차 수동 지정: %s", monday)

    xlsx_name = f"{CFG.doc_prefix}({monday}).xlsx"
    expected_path = CFG.download_folder / xlsx_name

    if CFG.source_xlsx:
        downloaded = _resolve_source_xlsx(CFG.source_xlsx)
        logging.info("다운로드 단계 생략, 지정된 xlsx 사용: %s", downloaded)
    elif CFG.download_source == "http":
        # 3-a) HTTP 직접연동: 브라우저/팝업 없이 가장 최근 날짜의 목록 파일을 내려받는다.
        logging.info("다운로드 방식: HTTP 직접연동(requests)")
        downloaded = download_latest_list_via_http()
    else:
        # 3-b) 레거시 playwright: 브라우저 탐색 + 폴더 선택 팝업 처리 + 파일 생성 대기.
        from playwright.sync_api import sync_playwright

        logging.info("다운로드 방식: 레거시 playwright(브라우저+팝업)")
        if expected_path.exists():
            try:
                expected_path.unlink()
            except Exception:
                pass

        year = int(monday[:4])

        with sync_playwright() as p:
            browser, ctx, page = ensure_page(p)
            try:
                click_year_and_first_00_folder(page, year)
                open_doc_properties_by_monday(page, monday)

                trigger_save_icon_for_attachment(page, monday)

                # 폴더 찾아보기 팝업: 2~3초 후 Enter면 충분하다고 했으니 그대로 반영
                confirm_browse_dialog_by_enter(wait_popup_sec=CFG.dialog_wait_sec, after_popup_sec=3.0)

                downloaded = wait_for_file_complete(CFG.download_folder, xlsx_name, timeout_sec=CFG.download_wait_sec)
                logging.info("다운로드 완료 확인: %s", downloaded)

            finally:
                ctx.close()
                browser.close()

    # 4) xlsx에서 last_serial 아래부터 A~N + Y/Z 추출 -> 정규화 -> master.xlsx A~N + O/P append
    logging.info("추가분 추출 대상 xlsx: %s", downloaded)
    rows = extract_a_to_n_rows_after_serial(downloaded, start_serial=last_serial, sheet_name="인증획득제품리스트")
    logging.info("추출된 행 수(A~N + Y/Z, 정규화 전): %d", len(rows))

    normalized_row_count = 0
    if rows:
        kolas_cert_numbers = read_kolas_cert_numbers(downloaded)
        logging.info("'%s' 시트에서 KOLAS 인증번호 %d건 확인", OTHER_INFO_SHEET_NAME, len(kolas_cert_numbers))
        rows = append_kolas_column(rows, kolas_cert_numbers)

        rows2 = normalize_rows(rows)
        normalized_row_count = len(rows2)
        logging.info("정규화 후 행 수(A~N + O/P/Q): %d", normalized_row_count)

        if rows2:
            append_rows_to_master_xlsx(CFG.master_xlsx, rows2, ensure_cols=True)
            logging.info("master append 완료(xlsx): %s", CFG.master_xlsx)
        else:
            logging.info("정규화 결과 추가할 데이터가 없습니다. master 변경 없음.")
    else:
        logging.info("추가할 데이터가 없습니다. master 변경 없음.")

    # 5) 저장소 기준 DB 적재
    sync_reference_db()

    # 6) 방금 SwData에 새로 들어온 건을 점검대상 프로젝트(ReferenceProject)에 반영
    #    (last_serial보다 큰 SwData 행 = 이번에 새로 추가된 건)
    try:
        sync_new_reference_projects(last_serial)
    except Exception:
        logging.exception("신규 점검대상 프로젝트 반영 실패(SwData 적재 자체는 완료됨)")

    logging.info("DONE")
    return normalized_row_count


if __name__ == "__main__":
    # 종료 코드로 후속 자동화(스케줄러)에 결과를 전달한다.
    #   0 = 정상 종료, 정규화 후 신규 행 없음
    #   2 = 정상 종료, 정규화 후 신규 행 있음(A~N + O/P/Q > 0)
    #   1 = 처리 중 예외 발생
    try:
        normalized_row_count = main()
    except Exception:
        logging.exception("UNHANDLED ERROR")
        sys.exit(1)
    else:
        sys.exit(2 if normalized_row_count else 0)
        raise
