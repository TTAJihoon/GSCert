import json
import sqlite3
import tempfile
import zipfile
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch
from xml.sax.saxutils import escape

from django.core.management import call_command
from django.db.utils import DatabaseError, OperationalError, ProgrammingError
from django.test import RequestFactory, SimpleTestCase, TestCase, override_settings
from django.utils import timezone

from main.models import (
    DownloadReviewJob,
    DownloadReviewJobStatus,
    DownloadReviewLog,
    DownloadReviewLogLevel,
    DownloadReviewManualOverride,
    DownloadReviewProject,
    DownloadReviewProjectReviewStatus,
    DownloadReviewProjectStatus,
    DownloadReviewRule,
    DownloadReviewRuleResult,
    DownloadReviewRuleStatus,
    ReferenceCenterPl,
    ReferenceProject,
    SwData,
)
from main.views.review.ecm_reference_db import (
    ARTIFACT_REVIEW_COLUMNS,
    ReferenceDbError,
    ReferenceDbMissing,
    ReferenceQueryError,
    write_project_review_result,
)
from main.views.review.ecm_download_review_worker import run_worker_once
from main.views.review.ecm_manual_override import apply_manual_override_to_evaluation, manual_overrides_for_project
from main.views.review.ecm_download_review_inspection import (
    ExcelSheet,
    ExcelWorkbook,
    _check_defect_report_environment,
    _list_mismatches,
    cleanup_download_dir,
    cleanup_stale_project_history,
    get_rule_output_variables,
    run_download_inspection,
)
from main.views.review.ecm_llm_review import (
    LlmReviewFileContext,
    LlmReviewRuleContext,
    build_llm_review_payload,
    parse_llm_review_response,
)
from main.views.review.ecm_download_verify import verify_downloaded_files
from main.views.review.ecm_download_review_api import (
    active_job,
    job_cancel,
    job_detail,
    job_project_change_note,
    job_project_results_excel,
    job_project_results,
    job_results_excel,
    job_projects,
    jobs,
    latest_project_results,
    local_review_rules_bundle,
    local_review_rules_manifest,
    pl_assignments,
    pl_assignments_apply,
    project_full_documents_download,
    projects,
    rule_result_manual_pass,
    rule_result_artifact,
)
from main.views.csrf import csrf_failure
from main.utils.ecm_reference_sheet import parse_sheet_projects, read_csv_rows, split_company_product
from gscert_review_core import engine
from gscert_review_core.result_display import friendly_message


def _docx_bytes(*, paragraphs=None, tables=None, blocks=None, header=None, footer=None):
    paragraphs = paragraphs or []
    tables = tables or []
    body_parts = []
    for block in blocks or []:
        if block["type"] == "paragraph":
            body_parts.append(_docx_paragraph_xml(block["text"]))
        elif block["type"] == "table":
            body_parts.append(_docx_table_xml(block["rows"]))
    for paragraph in paragraphs:
        body_parts.append(_docx_paragraph_xml(paragraph))
    for table in tables:
        body_parts.append(_docx_table_xml(table))
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        + "".join(body_parts)
        + "</w:body></w:document>"
    )
    bytes_buffer = tempfile.SpooledTemporaryFile()
    with zipfile.ZipFile(bytes_buffer, "w") as archive:
        archive.writestr("word/document.xml", document_xml.encode("utf-8"))
        if header is not None:
            header_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<w:hdr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                + _docx_paragraph_xml(header)
                + "</w:hdr>"
            )
            archive.writestr("word/header1.xml", header_xml.encode("utf-8"))
        if footer is not None:
            footer_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<w:ftr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                + _docx_paragraph_xml(footer)
                + "</w:ftr>"
            )
            archive.writestr("word/footer1.xml", footer_xml.encode("utf-8"))
    bytes_buffer.seek(0)
    data = bytes_buffer.read()
    bytes_buffer.close()
    return data


def _docx_paragraph_xml(paragraph):
    return f"<w:p><w:r><w:t>{escape(paragraph)}</w:t></w:r></w:p>"


def _docx_table_xml(table):
    rows_xml = []
    for row in table:
        cells_xml = []
        for cell in row:
            cells_xml.append(
                "<w:tc><w:p><w:r><w:t>"
                + escape(cell)
                + "</w:t></w:r></w:p></w:tc>"
            )
        rows_xml.append("<w:tr>" + "".join(cells_xml) + "</w:tr>")
    return "<w:tbl>" + "".join(rows_xml) + "</w:tbl>"


def _pdf_bytes(lines):
    import fitz

    document = fitz.open()
    page = document.new_page()
    y = 72
    for line in lines:
        page.insert_text((72, y), line, fontsize=11)
        y += 18
    data = document.tobytes()
    document.close()
    return data


def _xlsx_bytes(*, rows=None, sheet_name="Sheet1", header=None, footer=None):
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    if header is not None:
        worksheet.oddHeader.center.text = header
    if footer is not None:
        worksheet.oddFooter.center.text = footer
    for row in rows or []:
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _xlsx_workbook_bytes(sheets, *, header=None, footer=None):
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets:
        worksheet = workbook.create_sheet(sheet_name)
        if header is not None:
            worksheet.oddHeader.center.text = header
        if footer is not None:
            worksheet.oddFooter.center.text = footer
        for row in rows:
            worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _defect_report_xlsx(project_number, sheets, *, header=None, footer=None):
    rows_by_sheet = []
    for sheet_name in sheets:
        if sheet_name in ("최종결함리포트", "시험분석자료"):
            date_text = "2026년 05월 31일"
        elif sheet_name.startswith("1차"):
            date_text = "2026년 05월 10일"
        else:
            date_text = "2026년 05월 20일"

        rows = [
            [f"{project_number} {sheet_name} 보고일자: {date_text}", "시험환경 Windows 11"],
            [],
            [],
        ]
        if sheet_name == "최종결함리포트":
            rows.extend([
                [],
                ["", "", "", "", "", "", "품질특성"],
                ["", "", "", "", "", "", "잔여-1"],
                ["", "", "", "", "", "", "잔여-2"],
            ])
        elif sheet_name == "시험분석자료":
            rows.extend([
                ["", "", "High", "3"],
                ["", "", "수정전"],
                [],
                [],
                [],
                [],
                ["", "", "7"],
                ["", "", "결함-1", "", "H"],
                ["", "", "-", "", "H"],
                ["", "", "", "", "H"],
                ["", "", "결함-2", "", "H"],
                ["", "", "결함-3", "", "H"],
            ])
        rows_by_sheet.append((sheet_name, rows))
    return _xlsx_workbook_bytes(rows_by_sheet, header=header, footer=footer)


def _defect_report_xlsx_split_title(project_number, sheets, *, header=None, footer=None):
    rows_by_sheet = []
    for sheet_name in sheets:
        if sheet_name in ("최종결함리포트", "시험분석자료"):
            date_text = "2026.05.31."
        elif sheet_name.startswith("1차"):
            date_text = "2026.04.14."
        else:
            date_text = "2026.04.20."

        rows = [
            [project_number, "시험환경 Windows 11"],
            [sheet_name],
            [f"보고일자: {date_text}"],
            [],
            [],
        ]
        if sheet_name == "최종결함리포트":
            rows.extend([
                [],
                ["", "", "", "", "", "", "잔여특성"],
                ["", "", "", "", "", "", "잔여-1"],
                ["", "", "", "", "", "", "잔여-2"],
            ])
        elif sheet_name == "시험분석자료":
            rows.extend([
                ["", "", "High", "3"],
                ["", "", "수정전"],
                [],
                [],
                [],
                [],
                ["", "", "7"],
                ["", "", "결함-1", "", "H"],
                ["", "", "-", "", "H"],
                ["", "", "", "", "H"],
                ["", "", "결함-2", "", "H"],
                ["", "", "결함-3", "", "H"],
            ])
        rows_by_sheet.append((sheet_name, rows))
    return _xlsx_workbook_bytes(rows_by_sheet, header=header, footer=footer)


def _defect_report_zero_residual_xlsx(project_number):
    return _xlsx_workbook_bytes([
        (
            "최종결함리포트",
            [
                [f"{project_number} 최종결함리포트\n보고일자: 2026년 05월 31일", "", "", "", "[시험환경 : Windows 11]"],
                [],
                [],
                ["차시", "순번", "시험환경\nOS", "결함요약", "결함정도", "발생빈도", "품질특성", "결함 설명"],
            ],
        ),
        (
            "시험분석자료",
            [
                [f"{project_number} 시험분석자료\n보고일자: 2026년 05월 31일", "", "", "", "[시험환경 : Windows 11]"],
                [],
                [],
                ["", "", "High", "3"],
                ["", "", "수정전"],
                [],
                [],
                [],
                [],
                ["", "", "7"],
                ["", "", "결함-1", "", "H"],
                ["", "", "-", "", "H"],
                ["", "", "", "", "H"],
                ["", "", "결함-2", "", "H"],
                ["", "", "결함-3", "", "H"],
            ],
        ),
    ])


def _test_case_xlsx(
    project_number,
    *,
    pl="김준호",
    start_date="2026.05.01.",
    end_date="2026.05.31.",
    residual_count=2,
    footer=None,
):
    rows = [
        [f"{project_number} 테스트케이스"],
        [f"작성자: {pl}"],
        ["검토자: 김진영"],
        [f"작성일: {start_date} ~ {end_date}"],
        [],
        ["TC ID", "상세 테스트 결과"],
    ]
    for index in range(residual_count):
        rows.append([f"TC-F-{index + 1}", "F"])
    rows.append(["TC-P-1", "P"])
    return _xlsx_bytes(rows=rows, footer=footer)


def _test_plan_docx(project_number, *, product="테스트제품", version="v1.0", pl="김준호", wd="10"):
    first_table = [
        ["시험시작일", "2026.05.01."],
        ["비고", ""],
        ["담당자", "김진영"],
        ["시험PL", pl],
    ]
    second_table = [
        ["소프트웨어 명", product],
        ["버전", version],
        ["시험신청번호", project_number],
    ]
    configuration_table = [
        ["구분", "형상항목 ID"],
        ["소스", f"{project_number}-SRC"],
        ["문서", f"{project_number}-DOC"],
    ]
    schedule_table = [
        ["구분", "WD"],
        ["준비", "1"],
        ["분석", "1"],
        ["시험", str(int(wd) - 3)],
        ["종료", "1"],
    ]
    spec_table = [["항목", "값"], ["OS", "Windows"]]
    return _docx_bytes(
        blocks=[
            {"type": "table", "rows": first_table},
            {"type": "table", "rows": second_table},
            {"type": "paragraph", "text": "5.1 형상항목 식별 규칙"},
            {"type": "table", "rows": configuration_table},
            {"type": "paragraph", "text": "2.2 시험일정"},
            {"type": "table", "rows": schedule_table},
            {"type": "paragraph", "text": "<세부사양>"},
            {"type": "table", "rows": spec_table},
        ],
        footer="Copyright 2026 TTA",
    )


def _inspection_checklist_xlsx(
    project_number,
    *,
    pl="김준호",
    wd="10",
    high="3",
    before="7",
    start_date="2026.05.01.",
    end_date="2026.05.31.",
):
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)

    def sheet(name):
        worksheet = workbook.create_sheet(name)
        worksheet.oddHeader.center.text = f"프로젝트번호: {project_number}"
        worksheet.oddFooter.center.text = "한국정보통신기술협회"
        return worksheet

    cover = sheet("표지")
    cover["A1"] = f"{project_number} 점검표"
    cover["A2"] = f"{start_date} ~ {end_date}"
    cover["A3"] = f"김 진 영 / {pl}"

    feature = sheet("기능별 점검표")
    feature["A8"] = 1
    feature["B8"] = "대분류1"
    feature["C8"] = "중분류1"
    feature["D8"] = "기능1"
    feature["A9"] = 2
    feature["B9"] = "대분류2"
    feature["C9"] = "중분류2"
    feature["D9"] = "기능2"
    for row in range(8, 10):
        for column in range(5, 35):
            feature.cell(row=row, column=column, value="O")

    suitability = sheet("2. 기능적합성")
    suitability["A16"] = "대분류1"
    suitability["B16"] = "중분류1"
    suitability["C16"] = "기능1"
    suitability["D16"] = "O"
    suitability["A17"] = "대분류2"
    suitability["B17"] = "중분류2"
    suitability["C17"] = "기능2"
    suitability["D17"] = "O"

    reliability = sheet("6. 신뢰성")
    reliability["C5"] = wd
    reliability["C11"] = high
    reliability["E11"] = before

    scores = sheet("측정항목별 점수표")
    for index, row in enumerate(range(7, 91), start=1):
        scores.cell(row=row, column=4, value=f"score-{index}")

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _quality_inspection_table_xlsx(project_number, *, score_overrides=None, footer=None):
    from openpyxl import Workbook

    score_overrides = score_overrides or {}
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{project_number} 품질검사표"
    if footer is not None:
        worksheet.oddFooter.center.text = footer
    for index, row in enumerate(range(4, 88), start=1):
        worksheet.cell(row=row, column=4, value=score_overrides.get(index, f"score-{index}"))
    for index, row in enumerate(range(4, 37), start=1):
        worksheet.cell(row=row, column=5, value=f"quality-{index}")

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _quality_evaluation_report_docx(project_number, *, company="에이치소프트"):
    quality_values = [
        *[f"quality-{index}" for index in range(4, 27)],
        *[f"quality-{index}" for index in range(28, 34)],
        "quality-1",
        "quality-2",
        "quality-3",
    ]
    quality_table = [["품질특성", "구분", "평가결과", "비고"]]
    for index, value in enumerate(quality_values, start=1):
        quality_table.append([f"항목{index}", "", value, ""])

    paragraphs = [
        f"{project_number} 품질평가보고서",
        project_number,
        project_number,
        project_number,
        project_number,
        project_number,
        "성  명 : 김  성  희",
        "정  성  룡     (서명)",
        "신청일자 : 2026년 5월 2일",
        "계약일자 : 2026년 5월 3일",
        "제품시험평가 : 2026년 5월 1일 ~ 2026년 5월 31일",
        "품질인증심의위원회 : 2026년 6월 1일",
        "<품질특성별 세부 평가결과>",
    ]
    blocks = [
        *({"type": "paragraph", "text": paragraph} for paragraph in paragraphs),
        {"type": "table", "rows": [["목차", "페이지"], ["품질특성별 세부 평가결과", "12"]]},
        {"type": "table", "rows": quality_table},
        {"type": "table", "rows": [["회사(기관)명", company]]},
    ]
    return _docx_bytes(
        blocks=blocks,
    )


class EcmReferenceSheetParserTests(SimpleTestCase):
    def test_split_company_product_removes_parentheses_and_splits_once(self):
        company, product = split_company_product("주식회사 테스트(상암)-제품명-v1.0")

        self.assertEqual(company, "주식회사 테스트")
        self.assertEqual(product, "제품명-v1.0")

    def test_parse_sheet_projects_uses_date_block_and_pl_center(self):
        csv_text = "\n".join([
            ",2026년 6월 22일(월),,,,,,,",
            ",,,,,,,,",
            ",,,,,,,,",
            ",회사A(비고)-제품A,18,2026.06.01,2026.06.02,2026.06.03,2026.06.30,박지훈, TTA-26-00001",
            ",회사B(비고)-제품B,12,2026.06.04,2026.06.05,2026.06.06,2026.07.01,임우섭, TTA-26-00002",
            ",,,,,,,,",
        ])

        rows = parse_sheet_projects(read_csv_rows(csv_text))

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].project_number, "TTA-26-00001")
        self.assertEqual(rows[0].cert_date, "6/22")
        self.assertEqual(rows[0].company, "회사A")
        self.assertEqual(rows[0].product, "제품A")
        self.assertEqual(rows[0].center_code, "sangam")
        self.assertEqual(rows[1].center_code, "bundang")


class SyncReferenceProjectsFromSheetCommandTests(SimpleTestCase):
    def test_assign_unknown_pl_reclassifies_projects_in_current_run(self):
        csv_path = self._write_source_csv([
            ",회사C(비고)-제품C,20,2026.06.07,2026.06.08,2026.06.09,2026.07.02,신규PL, TTA-26-00003",
        ])
        out = StringIO()

        with patch("builtins.input", side_effect=["1", "상암"]):
            call_command(
                "sync_reference_projects_from_sheet",
                "--source-csv",
                str(csv_path),
                "--dry-run",
                "--assign-unknown-pl",
                stdout=out,
            )

        output = out.getvalue()
        self.assertIn("상암 1건", output)
        self.assertNotIn("미분류 1건", output)

    def test_assign_unknown_pl_can_skip_unassigned_names(self):
        csv_path = self._write_source_csv([
            ",회사D(비고)-제품D,21,2026.06.10,2026.06.11,2026.06.12,2026.07.03,보류PL, TTA-26-00004",
        ])
        out = StringIO()

        with patch("builtins.input", side_effect=[""]):
            call_command(
                "sync_reference_projects_from_sheet",
                "--source-csv",
                str(csv_path),
                "--dry-run",
                "--assign-unknown-pl",
                stdout=out,
            )

        self.assertIn("센터 미분류 PL: 보류PL", out.getvalue())

    def _write_source_csv(self, project_rows):
        csv_text = "\n".join([
            ",2026년 6월 22일(월),,,,,,,",
            ",,,,,,,,",
            ",,,,,,,,",
            *project_rows,
            ",,,,,,,,",
        ])
        temp = tempfile.NamedTemporaryFile("w", encoding="utf-8-sig", newline="", suffix=".csv", delete=False)
        with temp:
            temp.write(csv_text)
        return Path(temp.name)


class SyncNewCertifiedProjectsCommandTests(TestCase):
    """weekly(W) 동기화가 SwData 신규 건을 ReferenceProject에 반영하는 신규 명령어 테스트.

    'G'(Google Sheets 동기화) 메뉴가 하던 일을, 이제는 SwData(인증획득목록 엑셀)를
    기준으로 새로 추가된 건만 구글시트(신청일/계약일 보완용)와 매칭해 반영한다.
    """

    databases = {"reference"}

    def setUp(self):
        ReferenceCenterPl.objects.using("reference").create(
            center_code="sangam", center_label="상암", name="박지훈", display_order=1,
        )
        # since_serial=10 기준으로 11번(신규)만 대상이 되도록 함.
        SwData.objects.using("reference").create(
            serial_number=10, test_number="TTA-26-00099", company="이전회사", product="이전제품",
            test_lab="박지훈", cert_date="6/1", start_date="2026.06.01", end_date="2026.06.02",
            total_wd="10",
        )
        SwData.objects.using("reference").create(
            serial_number=11, test_number="TTA-26-00001", company="회사A", product="제품A",
            test_lab="박지훈", cert_date="6/22", start_date="2026.06.03", end_date="2026.06.30",
            total_wd="18",
        )

    def test_new_swdata_row_matched_to_sheet_creates_reference_project(self):
        csv_path = self._write_source_csv([
            ",회사A(비고)-제품A,18,2026.06.01,2026.06.02,2026.06.03,2026.06.30,박지훈, TTA-26-00001",
        ])
        out = StringIO()

        call_command(
            "sync_new_certified_projects",
            "--since-serial", "10",
            "--source-csv", str(csv_path),
            stdout=out,
        )

        project = ReferenceProject.objects.using("reference").get(project_number="TTA-26-00001")
        # 회사명/제품명/PL/WD/시험기간은 SwData(기준 원본) 값을 그대로 씀.
        self.assertEqual(project.company, "회사A")
        self.assertEqual(project.product, "제품A")
        self.assertEqual(project.primary_tester, "박지훈")
        self.assertEqual(project.wd, "18")
        self.assertEqual(project.start_date, "2026.06.03")
        self.assertEqual(project.expected_end_date, "2026.06.30")
        # 신청일/계약일은 SwData에 없으므로 구글시트에서 보완.
        self.assertEqual(project.request_date, "2026.06.01")
        self.assertEqual(project.contract_date, "2026.06.02")
        # 센터는 ReferenceCenterPl 매핑을 통해 유도.
        self.assertEqual(project.center_code, "sangam")
        # 이전(10번) 건은 대상이 아니므로 생성되지 않음.
        self.assertFalse(
            ReferenceProject.objects.using("reference").filter(project_number="TTA-26-00099").exists()
        )

    def test_new_swdata_row_not_in_sheet_is_skipped(self):
        csv_path = self._write_source_csv([])  # 구글시트에 해당 프로젝트번호가 없음
        out = StringIO()

        call_command(
            "sync_new_certified_projects",
            "--since-serial", "10",
            "--source-csv", str(csv_path),
            stdout=out,
        )

        self.assertFalse(
            ReferenceProject.objects.using("reference").filter(project_number="TTA-26-00001").exists()
        )
        self.assertIn("건너뜀", out.getvalue())

    def test_existing_reference_project_keeps_review_result(self):
        ReferenceProject.objects.using("reference").create(
            project_number="TTA-26-00001",
            center_code="unknown",
            company="옛회사명",
            review_result="X",
            inspection_date="2026.06.20 10:00",
        )
        csv_path = self._write_source_csv([
            ",회사A(비고)-제품A,18,2026.06.01,2026.06.02,2026.06.03,2026.06.30,박지훈, TTA-26-00001",
        ])

        call_command(
            "sync_new_certified_projects",
            "--since-serial", "10",
            "--source-csv", str(csv_path),
            stdout=StringIO(),
        )

        project = ReferenceProject.objects.using("reference").get(project_number="TTA-26-00001")
        self.assertEqual(project.company, "회사A")  # SwData 값으로 최신화됨
        self.assertEqual(project.review_result, "X")  # 점검 결과는 보존됨
        self.assertEqual(project.inspection_date, "2026.06.20 10:00")

    def _write_source_csv(self, project_rows):
        csv_text = "\n".join([
            ",2026년 6월 22일(월),,,,,,,",
            ",,,,,,,,",
            ",,,,,,,,",
            *project_rows,
            ",,,,,,,,",
        ])
        temp = tempfile.NamedTemporaryFile("w", encoding="utf-8-sig", newline="", suffix=".csv", delete=False)
        with temp:
            temp.write(csv_text)
        return Path(temp.name)


class DownloadVerifyTests(SimpleTestCase):
    def test_zero_byte_file_is_warning_not_failure(self):
        # 0바이트 파일은 다운로드 확인에서 실패시키지 않고 경고로만 남긴다.
        # (파싱이 필요한 점검 대상이면 점검규칙이 부적합/오류로 잡는다.)
        with tempfile.TemporaryDirectory() as temp_dir:
            (Path(temp_dir) / "TTA-26-00010_안내.txt").write_bytes(b"")
            (Path(temp_dir) / "TTA-26-00010_보고서.pdf").write_bytes(b"content")

            result = verify_downloaded_files(temp_dir, "TTA-26-00010")

        self.assertTrue(result.success)
        self.assertTrue(any("0 byte" in w for w in result.warnings))

    def test_missing_project_number_is_warning_not_failure(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "downloaded-report.pdf"
            file_path.write_bytes(b"content")

            result = verify_downloaded_files(temp_dir, "TTA-26-00010")

        self.assertTrue(result.success)
        self.assertFalse(result.has_project_number_files)
        self.assertTrue(result.warnings)


class DownloadReviewInspectionCompareTests(SimpleTestCase):
    def test_result_message_omits_expected_and_actual_details(self):
        pass_message = friendly_message(DownloadReviewRuleStatus.PASS, "", "A", "A")
        fail_message = friendly_message(
            DownloadReviewRuleStatus.FAIL,
            "차이: 보고일자가 다릅니다.\n기대값: 2026.05.01\n실제값: 2026.05.02",
            "보고일자 2026.05.01",
            "2026.05.02",
        )
        unsupported_message = friendly_message(
            "unsupported",
            "",
            "문서 본문 직접 확인",
            "-",
        )

        self.assertEqual(pass_message, "기준을 충족했습니다.")
        self.assertEqual(fail_message, "차이: 문서의 날짜/기간 값이 기준정보와 다릅니다.")
        self.assertNotIn("기대값", fail_message)
        self.assertNotIn("실제값", fail_message)
        self.assertNotIn("2026.05.01", fail_message)
        self.assertNotIn("2026.05.02", fail_message)
        self.assertNotIn("문서 본문 직접 확인", unsupported_message)

    def test_product_version_split_accepts_numeric_and_word_suffixes(self):
        cases = {
            "자료분석 플랫폼 v1": ("자료분석 플랫폼", "v1"),
            "자료분석 플랫폼 v1.0": ("자료분석 플랫폼", "v1.0"),
            "자료분석 플랫폼 1": ("자료분석 플랫폼", "1"),
            "자료분석 플랫폼 3.0": ("자료분석 플랫폼", "3.0"),
            "EBS ISM3.0": ("EBS ISM", "3.0"),
            "자료분석 플랫폼 Enterprise": ("자료분석 플랫폼", "Enterprise"),
        }

        for raw, expected in cases.items():
            with self.subTest(raw=raw):
                self.assertEqual(engine._split_product_and_version(raw), expected)

    def test_version_matches_accepts_word_versions_case_insensitively(self):
        self.assertTrue(engine._version_matches("Enterprise", "enterprise"))
        self.assertTrue(engine._version_matches("v1", "1"))
        self.assertFalse(engine._version_matches("Enterprise", "Standard"))

    def test_version_matches_ignores_trailing_zero_precision_differences(self):
        # 실제 사례(TTA-26-01501): 등록된 제품명은 'v1.0'인데 시험계획서 문서에는
        # 'v1'로만 적혀 있어, 예전에는 문자열 비교('1.0' != '1')로 오탐 부적합
        # 처리됐다. 숫자로는 같은 버전이므로 일치로 봐야 한다.
        self.assertTrue(engine._version_matches("v1.0", "v1"))
        self.assertTrue(engine._version_matches("v1.0", "1"))
        self.assertTrue(engine._version_matches("2.10.0", "2.10"))
        # 끝자리 0 무시일 뿐 실제로 다른 버전까지 같다고 보면 안 된다.
        self.assertFalse(engine._version_matches("2.10", "2.1"))
        self.assertFalse(engine._version_matches("v1.0", "v2"))

    def test_defect_report_sheet_names_ignore_whitespace(self):
        workbook = engine.ExcelWorkbook([
            engine.ExcelSheet("1차결함리포트", []),
        ])
        versioned_files = {
            1: engine.FileInfo(name="TTA-26-00010 결함리포트 v1.0.xlsx", path=""),
        }

        result = engine._check_defect_report_sheets({1: workbook}, versioned_files, defect_round_count=1)

        self.assertTrue(result["passed"], result)
        self.assertEqual(result["details"][0]["missing_sheets"], [])
        self.assertEqual(result["details"][0]["extra_sheets"], [])

    def test_test_case_missing_file_emits_all_expected_sub_checks(self):
        rule = SimpleNamespace(
            config_json={
                "filename_keywords": ["testcase", "{project_number}"],
                "extensions": [".xlsx"],
                "exact_count": 1,
                "forbidden_footer_terms": [{"text": "TPG"}],
                "forbidden_header_terms": [{"text": "TTA"}],
                "required_footer_terms": [{"text": "TTA"}],
                "title_text": "{project_number} testcase",
                "author_label": "author",
                "reviewer_label": "reviewer:",
                "reviewer_expected": "reviewer",
                "date_label": "date",
                "result_header": "result",
            },
            target_file_type="any",
            target_file_pattern="",
        )
        context = engine.RuleContext(
            project_number="TTA-26-00010",
            product_raw="",
            product="",
            version="",
            company="",
            pl="PL",
            wd="",
            start_date="2026.05.01",
            end_date="2026.05.02",
            year="2026",
            request_date="",
            contract_date="",
            certification_committee_date="",
            derived_variables={"잔여결함수": 2},
            center="bundang",
        )
        verify_result = SimpleNamespace(files=[])

        evaluation = engine._evaluate_test_case_check(
            rule,
            1,
            SimpleNamespace(project_number="TTA-26-00010"),
            context,
            verify_result,
        )

        sub_checks = evaluation.raw_detail["sub_checks"]
        self.assertEqual(evaluation.status, DownloadReviewRuleStatus.FAIL)
        self.assertEqual(len(sub_checks), 8)
        self.assertTrue(all(item.get("passed") is False for item in sub_checks))
        self.assertEqual(sub_checks[0]["sub_check_key"], "sub-1")
        self.assertEqual(sub_checks[-1]["sub_check_key"], "sub-8")
        self.assertTrue(sub_checks[-1]["blocked_by_prerequisite"])
        self.assertNotIn("_expected_sub_check_templates", evaluation.raw_detail)

    def test_artifact_revision_selection_uses_latest_minor_across_folders(self):
        rule = SimpleNamespace(
            config_json={"folder_keyword_chain": ["시험", "계획"]},
            target_file_pattern="",
            target_file_type="any",
        )
        verify_result = SimpleNamespace(files=[
            engine.FileInfo(
                name="TTA-26-00010 기능리스트 v1.0.xlsx",
                path="4.시험/가.계획/TTA-26-00010 기능리스트 v1.0.xlsx",
                extension=".xlsx",
            ),
            engine.FileInfo(
                name="TTA-26-00010 기능리스트 v1.1.xlsx",
                path="추가제출/TTA-26-00010 기능리스트 v1.1.xlsx",
                extension=".xlsx",
            ),
        ])

        selected, selected_folder = engine._files_in_configured_folder(rule, verify_result)

        self.assertEqual(selected_folder, "4.시험/가.계획")
        self.assertEqual([file_info.name for file_info in selected], ["TTA-26-00010 기능리스트 v1.1.xlsx"])

    def test_artifact_revision_selection_prefers_versioned_file_over_unversioned(self):
        rule = SimpleNamespace(
            config_json={"folder_keyword_chain": ["시험", "계획"]},
            target_file_pattern="",
            target_file_type="any",
        )
        verify_result = SimpleNamespace(files=[
            engine.FileInfo(
                name="TTA-26-00010 기능리스트.xlsx",
                path="4.시험/가.계획/TTA-26-00010 기능리스트.xlsx",
                extension=".xlsx",
            ),
            engine.FileInfo(
                name="TTA-26-00010 기능리스트 v1.1.xlsx",
                path="추가제출/TTA-26-00010 기능리스트 v1.1.xlsx",
                extension=".xlsx",
            ),
        ])

        selected, _selected_folder = engine._files_in_configured_folder(rule, verify_result)

        self.assertEqual([file_info.name for file_info in selected], ["TTA-26-00010 기능리스트 v1.1.xlsx"])

    def test_artifact_revision_selection_prefers_latest_modified_for_same_version(self):
        older = engine.FileInfo(
            name="TTA-26-00010 기능리스트 v1.1.xlsx",
            path="old/TTA-26-00010 기능리스트 v1.1.xlsx",
            extension=".xlsx",
            modified_at=datetime(2026, 5, 1, 9, 0, 0),
        )
        newer = engine.FileInfo(
            name="TTA-26-00010 기능리스트 v1.1.xlsx",
            path="new/TTA-26-00010 기능리스트 v1.1.xlsx",
            extension=".xlsx",
            modified_at=datetime(2026, 5, 2, 9, 0, 0),
        )

        selected = engine._latest_revision_files([older, newer])

        self.assertEqual([file_info.path for file_info in selected], ["new/TTA-26-00010 기능리스트 v1.1.xlsx"])

    def test_artifact_revision_parser_does_not_treat_trailing_date_as_revision(self):
        file_info = engine.FileInfo(
            name="TTA-26-00010 시험계획서 2026.05.10.xlsx",
            path="TTA-26-00010 시험계획서 2026.05.10.xlsx",
            extension=".xlsx",
        )

        self.assertIsNone(engine._artifact_revision_info(file_info))

    def test_defect_report_versions_keep_latest_minor_per_major(self):
        files = [
            engine.FileInfo(name="TTA-26-00010 결함리포트 v1.0.xlsx", path="v1.0", extension=".xlsx"),
            engine.FileInfo(name="TTA-26-00010 결함리포트 v1.1.xlsx", path="v1.1", extension=".xlsx"),
            engine.FileInfo(name="TTA-26-00010 결함리포트 v2.0.xlsx", path="v2.0", extension=".xlsx"),
        ]

        versioned = engine._defect_report_versioned_files(engine._latest_revision_files(files), {})

        self.assertEqual(set(versioned), {1, 2})
        self.assertEqual(versioned[1].name, "TTA-26-00010 결함리포트 v1.1.xlsx")

    def test_list_mismatches_compares_numeric_text_by_value(self):
        mismatches = _list_mismatches(
            ["1", "1.0", "0.125", "1,000", "NA", "30분"],
            ["1.00", "1", "0.1250", "1000.00", "NA", "30"],
            start_index=2,
        )

        self.assertEqual(mismatches, [{"row": 7, "expected": "30분", "actual": "30"}])

    def test_spec_table_comparison_ignores_spaces_and_newlines(self):
        context = engine.build_context(project_number="TTA-26-00010")
        context.derived_variables["시험성적서_세부사양표"] = [
            ["No", "설치 SW"],
            ["1", "- 시험대상\n제품 (PC 프로그램)"],
        ]

        check = engine._test_plan_spec_table_check(
            [
                ["No", "설치 SW"],
                ["1", "- 시험 대상 제품 (PC 프로그램)"],
            ],
            {},
            context,
        )

        self.assertTrue(check["passed"])
        self.assertEqual(check["actual"], "일치")
        self.assertEqual(check["comparison_mode"], "ignore_whitespace")

    def test_spec_table_comparison_still_reports_content_mismatch(self):
        context = engine.build_context(project_number="TTA-26-00010")
        context.derived_variables["시험성적서_세부사양표"] = [
            ["No", "설치 SW"],
            ["1", "- 시험대상 제품 (서버 프로그램)"],
        ]

        check = engine._test_plan_spec_table_check(
            [
                ["No", "설치 SW"],
                ["1", "- 시험 대상 제품 (PC 프로그램)"],
            ],
            {},
            context,
        )

        self.assertFalse(check["passed"])
        self.assertIn("B2 계획서", check["actual"])


class DownloadReviewApiPureTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_project_full_documents_download_get_streams_zip(self):
        request = self.factory.get(
            "/api/projects/TTA-26-00010/full-documents-download/",
            {"cert_date": "2026-05-13"},
        )

        with patch(
            "main.views.testing.history_download.iter_full_project_documents_zip",
            return_value=iter([b"PK", b"zip-data"]),
        ) as mocked_stream:
            response = project_full_documents_download(request, "TTA-26-00010")
            content = b"".join(response.streaming_content)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        self.assertEqual(response["Content-Disposition"], 'attachment; filename="TTA-26-00010.zip"')
        self.assertEqual(response["X-Accel-Buffering"], "no")
        self.assertEqual(content, b"PKzip-data")
        mocked_stream.assert_called_once_with("TTA-26-00010", "2026-05-13")

    def test_project_full_documents_download_reuses_history_full_download(self):
        request = self.factory.post(
            "/api/projects/TTA-26-00010/full-documents-download/",
            data=json.dumps({"cert_date": "2026-05-13"}),
            content_type="application/json",
        )

        with patch(
            "main.views.testing.history_download.download_full_project_documents",
            return_value={"doc_count": 2, "center": "sangam"},
        ) as mocked_download:
            response = project_full_documents_download(request, "TTA-26-00010")
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(data["success"])
        self.assertEqual(data["download_url"], "/history/report/TTA-26-00010/download/")
        self.assertEqual(data["doc_count"], 2)
        self.assertEqual(data["center"], "sangam")
        mocked_download.assert_called_once_with("TTA-26-00010", "2026-05-13")


class LlmReviewInterfaceTests(SimpleTestCase):
    databases = {"reference"}

    def test_payload_builder_creates_provider_neutral_messages(self):
        payload = build_llm_review_payload(
            project={"project_number": "TTA-26-00010", "company": "Example"},
            rule=LlmReviewRuleContext(
                code="manual_llm_rule",
                name="Manual LLM rule",
                prompt="Check whether the report is acceptable.",
                artifact_column="Report",
            ),
            files=[
                LlmReviewFileContext(
                    name="TTA-26-00010_report.pdf",
                    path="TTA-26-00010/TTA-26-00010_report.pdf",
                    size=123,
                    extension=".pdf",
                )
            ],
        )

        data = payload.to_dict()

        self.assertEqual(data["schema_version"], "download-review-llm-v1")
        self.assertEqual(data["provider_hint"], "manual-claude")
        self.assertEqual(data["messages"][0]["role"], "system")
        self.assertIn("response_schema", data)
        self.assertEqual(data["files"][0]["path"], "TTA-26-00010/TTA-26-00010_report.pdf")

    def test_parser_accepts_fenced_json_response(self):
        parsed = parse_llm_review_response(
            """
            ```json
            {
              "status": "fail",
              "expected": "Company name matches",
              "actual": "Different company name",
              "message": "The supplied evidence does not match.",
              "evidence": [{"file_name": "report.pdf", "reason": "Mismatch"}],
              "confidence": 0.82
            }
            ```
            """
        )

        self.assertEqual(parsed.status, DownloadReviewRuleStatus.FAIL)
        self.assertEqual(parsed.confidence, 0.82)
        self.assertEqual(parsed.evidence[0]["file_name"], "report.pdf")

    def test_build_llm_review_prompt_command_outputs_json(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            report = Path(temp_dir) / "TTA-26-00010_report.txt"
            report.write_text("sample document text", encoding="utf-8")
            out = StringIO()

            call_command(
                "build_llm_review_prompt",
                "--project-number",
                "TTA-26-00010",
                "--download-dir",
                temp_dir,
                "--rule-name",
                "Manual rule",
                "--rule-prompt",
                "Check the supplied document.",
                stdout=out,
            )

        data = json.loads(out.getvalue())

        self.assertEqual(data["schema_version"], "download-review-llm-v1")
        self.assertEqual(data["project"]["project_number"], "TTA-26-00010")
        self.assertEqual(data["files"][0]["name"], "TTA-26-00010_report.txt")
        self.assertEqual(data["rule"]["name"], "Manual rule")


class DownloadReviewRuleSeedCommandTests(TestCase):
    databases = {"default", "workflow", "reference"}

    def test_seed_creates_disabled_rules_by_default(self):
        out = StringIO()

        call_command("seed_download_review_rules", stdout=out)

        self.assertEqual(DownloadReviewRule.objects.count(), len(ARTIFACT_REVIEW_COLUMNS))
        first_rule = DownloadReviewRule.objects.order_by("sort_order").first()
        self.assertEqual(first_rule.code, "artifact_01")
        self.assertEqual(first_rule.config_json["artifact_column"], ARTIFACT_REVIEW_COLUMNS[0])
        self.assertFalse(first_rule.enabled)
        report_rule = DownloadReviewRule.objects.get(name="시험성적서(PDF)")
        defect_rule = DownloadReviewRule.objects.get(name="결함리포트")
        self.assertLess(report_rule.sort_order, defect_rule.sort_order)
        self.assertIn("created=", out.getvalue())

    def test_seed_can_enable_and_update_existing_rules(self):
        DownloadReviewRule.objects.create(code="artifact_01", name="old", enabled=False)
        out = StringIO()

        call_command("seed_download_review_rules", "--enable", "--update-existing", stdout=out)

        rule = DownloadReviewRule.objects.get(code="artifact_01")
        self.assertEqual(rule.name, ARTIFACT_REVIEW_COLUMNS[0])
        self.assertTrue(rule.enabled)
        self.assertIn("updated=", out.getvalue())

    def test_seed_only_real_creates_implemented_rules(self):
        out = StringIO()

        call_command("seed_download_review_rules", "--only-real", "--enable", stdout=out)

        self.assertEqual(DownloadReviewRule.objects.count(), 18)
        self.assertEqual(
            set(DownloadReviewRule.objects.values_list("name", flat=True)),
            {
                "계약서",
                "합의서(PDF)",
                "수수료산정표",
                "시험환경구성도",
                "품질특성별제품정보기재사항",
                "기능리스트",
                "시험계획서(PDF)",
                "최초/최종형상RawData",
                "테스트케이스",
                "결함리포트",
                "점검표(PDF)",
                "1차/2차/성능/보안RawData",
                "시험성적서(PDF)",
                "시험기록서",
                "품질평가보고서",
                "품질검사표",
                "SW저작권확인서",
                "홍보이미지",
            },
        )
        rule = DownloadReviewRule.objects.get(name="계약서")
        self.assertEqual(rule.rule_type, "required_artifact_file")
        self.assertTrue(rule.enabled)
        rawdata_rule = DownloadReviewRule.objects.get(name="1차/2차/성능/보안RawData")
        security_check = next(
            check
            for check in rawdata_rule.config_json["folder_checks"]
            if check["keyword"] == "보안"
        )
        self.assertTrue(security_check["txt_only_pass"])


class DownloadReviewProjectsApiTests(TestCase):
    databases = {"default", "workflow", "reference"}

    def setUp(self):
        self.factory = RequestFactory()
        self._seed_reference_projects(
            "sangam",
            [
                dict(
                    project_number="TTA-26-00009",
                    cert_date="05/12",
                    cert_committee_date=date(2026, 5, 12),
                    company="우리데이터 주식회사",
                    product="우리데이터클리닝 V1.0",
                    pl="박지훈",
                    review_result="O",
                    inspection_date="2026.05.12 20:30",
                ),
                dict(
                    project_number="TTA-26-00010",
                    cert_date="05/13",
                    cert_committee_date=date(2026, 5, 13),
                    company="에이치소프트",
                    product="SecureFlow 2.1",
                    pl="김준호",
                    review_result="",
                    inspection_date="",
                ),
                dict(
                    project_number="TTA-26-00008",
                    cert_date="05/11",
                    cert_committee_date=date(2026, 5, 11),
                    company="넥스트랩",
                    product="NextLab QA Suite",
                    pl="최유진",
                    review_result="X",
                    inspection_date="2026.05.11 21:00",
                ),
            ],
        )

    def test_projects_are_sorted_by_cert_date_desc_by_default(self):
        data = self._get_projects()

        self.assertEqual(data["pagination"]["total"], 3)
        self.assertEqual(
            [item["project_number"] for item in data["items"]],
            ["TTA-26-00010", "TTA-26-00009", "TTA-26-00008"],
        )
        self.assertTrue(data["items"][0]["selectable"])
        self.assertFalse(data["items"][1]["selectable"])
        self.assertEqual(data["items"][1]["active_state_label"], "완료")

    def test_projects_filter_uses_allowlisted_query_params(self):
        data = self._get_projects({"company": "우리", "limit": "1"})

        self.assertEqual(data["pagination"]["total"], 1)
        self.assertEqual(data["items"][0]["company"], "우리데이터 주식회사")

    def test_projects_reject_unknown_query_params(self):
        response = self._request({"raw_sql": "SELECT * FROM ecm_list"})
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 400)
        self.assertFalse(data["success"])
        self.assertEqual(data["error_code"], "invalid_query")

    def test_projects_still_return_when_workflow_state_db_is_unavailable(self):
        class BrokenProjectManager:
            def select_related(self, *args):
                return self

            def filter(self, *args, **kwargs):
                return self

            def order_by(self, *args):
                return self

            def __iter__(self):
                raise DatabaseError("workflow database is unavailable")

        class BrokenDownloadReviewProject:
            objects = BrokenProjectManager()

        request = self.factory.get("/api/projects/", {"limit": "2", "center": "sangam"})
        with patch(
            "main.views.review.ecm_download_review_jobs.DownloadReviewProject",
            BrokenDownloadReviewProject,
        ):
            response = projects(request)
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(data["items"]), 2)
        self.assertIsNone(data["items"][0]["active_job_id"])
        self.assertEqual(data["items"][0]["active_state_label"], "")
        self.assertEqual(data["items"][1]["active_state_label"], "완료")

    def test_projects_show_latest_failed_held_result_as_failed(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            failed_project_count=1,
        )
        DownloadReviewProject.objects.create(
            job=job,
            center_code="sangam",
            project_number="TTA-26-00010",
            ecm_row_json={"project_number": "TTA-26-00010", "company": "에이치소프트"},
            status=DownloadReviewProjectStatus.FAILED,
            review_status=DownloadReviewProjectReviewStatus.HELD,
            current_step="다운로드 파일 확인",
            error_message="다운로드 파일을 찾을 수 없습니다.",
            completed_at=timezone.now(),
        )

        data = self._get_projects({"project_number": "TTA-26-00010"})

        self.assertEqual(data["items"][0]["review"], "실패")
        self.assertEqual(data["items"][0]["review_raw"], "실패")

    def test_projects_can_read_yeongnam_center_db(self):
        self._seed_reference_projects(
            "yeongnam",
            [
                dict(
                    project_number="TTA-26-09999",
                    cert_date="05/14",
                    cert_committee_date=date(2026, 5, 14),
                    company="영남테스트",
                    product="Yeongnam Suite",
                    pl="김영남",
                    review_result="",
                    inspection_date="",
                ),
            ],
        )

        request = self.factory.get("/api/projects/", {"center": "yeongnam"})
        response = projects(request)
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(data["items"][0]["center_code"], "yeongnam")
        self.assertEqual(data["items"][0]["center_label"], "영남")
        self.assertEqual(data["items"][0]["project_number"], "TTA-26-09999")

    def test_projects_default_to_bundang_on_bundang_server_host(self):
        self._seed_reference_projects(
            "bundang",
            [
                dict(
                    project_number="TTA-26-00099",
                    cert_date="05/15",
                    cert_committee_date=date(2026, 5, 15),
                    company="분당기업",
                    product="Bundang Suite",
                    pl="이분당",
                    review_result="",
                    inspection_date="",
                ),
            ],
        )

        request = self.factory.get("/api/projects/", {"limit": "1"}, HTTP_HOST="210.96.71.194")
        response = projects(request)
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(data["items"][0]["center_code"], "bundang")
        self.assertEqual(data["items"][0]["center_label"], "분당")

    def test_projects_reject_other_center_on_bundang_server_host(self):
        request = self.factory.get("/api/projects/", {"center": "sangam"}, HTTP_HOST="210.96.71.194")
        # 센터 제한 로직 자체를 검증하기 위해 이 호스트의 허용 센터를 명시적으로 좁힌다
        # (운영 settings.py의 현재 허용 목록 값에 테스트가 의존하지 않도록).
        with self.settings(DOWNLOAD_REVIEW_ALLOWED_CENTERS_BY_HOST={"210.96.71.194": {"bundang"}}):
            response = projects(request)
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 400)
        self.assertEqual(data["error_code"], "invalid_query")
        self.assertIn("이 서버에서 처리하지 않는 센터", data["message"])

    def _get_projects(self, params=None):
        response = self._request(params or {})
        self.assertEqual(response.status_code, 200)
        return json.loads(response.content.decode("utf-8"))

    def _request(self, params):
        # 이 클래스의 기본 픽스처는 sangam 센터에 있으므로, 호출부가 다른 센터/호스트를
        # 명시하지 않는 한 sangam으로 고정한다(전역 DOWNLOAD_REVIEW_DEFAULT_CENTER 값과 무관하게).
        params = {**{"center": "sangam"}, **params}
        request = self.factory.get("/api/projects/", params)
        return projects(request)

    def _seed_reference_projects(self, center_code, rows):
        ReferenceProject.objects.using("reference").bulk_create(
            [ReferenceProject(center_code=center_code, **row) for row in rows]
        )


class PlAssignmentApiTests(TestCase):
    """'PL 배정 목록' 기능(구 'DB 새로고침' 버튼 자리) API 회귀 테스트."""

    databases = {"default", "workflow", "reference"}

    def setUp(self):
        self.factory = RequestFactory()
        ReferenceCenterPl.objects.using("reference").bulk_create([
            ReferenceCenterPl(center_code="sangam", center_label="상암", name="김진영", display_order=1),
            ReferenceCenterPl(center_code="bundang", center_label="분당", name="임우섭", display_order=1),
        ])
        ReferenceProject.objects.using("reference").bulk_create([
            ReferenceProject(
                project_number="TTA-26-90001", center_code="sangam", center_label="상암",
                primary_tester="김진영",
            ),
            ReferenceProject(
                project_number="TTA-26-90002", center_code="unknown", center_label="미분류",
                primary_tester="신규PL",
            ),
            ReferenceProject(
                project_number="TTA-26-90003", center_code="unknown", center_label="미분류",
                primary_tester="신규PL",
            ),
        ])

    def _get_assignments(self):
        request = self.factory.get("/api/pl-assignments/")
        response = pl_assignments(request)
        return json.loads(response.content)

    def _apply(self, changes):
        request = self.factory.post(
            "/api/pl-assignments/apply/",
            data=json.dumps({"changes": changes}),
            content_type="application/json",
        )
        return pl_assignments_apply(request)

    def test_get_lists_centers_and_unassigned_with_project_counts(self):
        data = self._get_assignments()

        self.assertTrue(data["success"])
        center_codes = {center["code"] for center in data["centers"]}
        self.assertEqual(center_codes, {"sangam", "bundang", "yeongnam", "unknown"})
        self.assertEqual(
            data["assignments"]["sangam"],
            [{"name": "김진영", "project_count": 1}],
        )
        self.assertEqual(
            data["assignments"]["unknown"],
            [{"name": "신규PL", "project_count": 2}],
        )

    def test_unassigned_to_center_creates_mapping_and_moves_existing_projects(self):
        response = self._apply([{"name": "신규PL", "from_center": "unknown", "to_center": "bundang"}])
        payload = json.loads(response.content)

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["success"])
        self.assertEqual(payload["moved_project_count"], 2)
        self.assertEqual(payload["updated_pl_count"], 1)

        pl_row = ReferenceCenterPl.objects.using("reference").get(name="신규PL")
        self.assertEqual(pl_row.center_code, "bundang")

        moved_numbers = set(
            ReferenceProject.objects.using("reference")
            .filter(primary_tester="신규PL")
            .values_list("project_number", "center_code")
        )
        self.assertEqual(
            moved_numbers,
            {("TTA-26-90002", "bundang"), ("TTA-26-90003", "bundang")},
        )

    def test_unassigned_to_center_leaves_already_inspected_projects_behind(self):
        # 실제로 발생 가능한 시나리오: A센터(상암)에서 점검까지 끝난 프로젝트가 있는
        # PL을 미배정으로 옮기면, 다음 시트 동기화가 center_code를 'unknown'으로
        # 덮어써도 review_result는 그대로 남는다(sync는 review_result를 안 건드림).
        # 이 상태에서 그 PL을 다른 센터로 재배정해도, 이미 점검된 프로젝트까지
        # 새 센터로 휩쓸려 가면 안 된다 - 미배정 그대로 남아야 한다.
        ReferenceProject.objects.using("reference").create(
            project_number="TTA-26-90004",
            center_code="unknown",
            center_label="미분류",
            primary_tester="신규PL",
            review_result="X",
        )

        data = self._get_assignments()
        self.assertEqual(
            data["assignments"]["unknown"],
            [{"name": "신규PL", "project_count": 2}],
        )

        response = self._apply([{"name": "신규PL", "from_center": "unknown", "to_center": "bundang"}])
        payload = json.loads(response.content)

        self.assertTrue(payload["success"])
        self.assertEqual(payload["moved_project_count"], 2)

        already_inspected = ReferenceProject.objects.using("reference").get(project_number="TTA-26-90004")
        self.assertEqual(already_inspected.center_code, "unknown")

    def test_center_to_center_move_updates_mapping_but_keeps_existing_projects(self):
        response = self._apply([{"name": "김진영", "from_center": "sangam", "to_center": "yeongnam"}])
        payload = json.loads(response.content)

        self.assertTrue(payload["success"])
        self.assertEqual(payload["moved_project_count"], 0)
        self.assertEqual(payload["updated_pl_count"], 1)

        pl_row = ReferenceCenterPl.objects.using("reference").get(name="김진영")
        self.assertEqual(pl_row.center_code, "yeongnam")

        # 기존 프로젝트는 그대로 상암에 남아있어야 한다(다음 시트 동기화부터 반영).
        existing_project = ReferenceProject.objects.using("reference").get(project_number="TTA-26-90001")
        self.assertEqual(existing_project.center_code, "sangam")

    def test_center_to_unassigned_deletes_mapping_but_keeps_existing_projects(self):
        response = self._apply([{"name": "김진영", "from_center": "sangam", "to_center": "unknown"}])
        payload = json.loads(response.content)

        self.assertTrue(payload["success"])
        self.assertFalse(
            ReferenceCenterPl.objects.using("reference").filter(name="김진영").exists()
        )
        existing_project = ReferenceProject.objects.using("reference").get(project_number="TTA-26-90001")
        self.assertEqual(existing_project.center_code, "sangam")

    def test_apply_rejects_empty_changes(self):
        response = self._apply([])
        payload = json.loads(response.content)

        self.assertEqual(response.status_code, 400)
        self.assertFalse(payload["success"])


class DownloadReviewChangeNoteApiTests(TestCase):
    databases = {"default", "workflow"}

    def setUp(self):
        self.factory = RequestFactory()
        self.temp_dir = tempfile.TemporaryDirectory()

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_job_project_change_note_endpoint_returns_txt_content(self):
        job, project = self._make_project()
        project_dir = Path(self.temp_dir.name) / "downloads" / "TTA-26-00010"
        project_dir.mkdir(parents=True)
        (project_dir / "수정 내용.txt").write_text("기능리스트 v1.1 추가", encoding="cp949")
        project.download_dir = str(project_dir)
        project.save(update_fields=["download_dir", "updated_at"])
        self._add_rule_result(project)

        results_response = job_project_results(
            self.factory.get(f"/api/job-projects/{project.id}/results/"),
            project.id,
        )
        results_data = json.loads(results_response.content.decode("utf-8"))
        note_response = job_project_change_note(
            self.factory.get(f"/api/job-projects/{project.id}/change-note/"),
            project.id,
        )
        note_data = json.loads(note_response.content.decode("utf-8"))

        self.assertEqual(results_response.status_code, 200)
        self.assertTrue(results_data["project"]["change_note"]["available"])
        self.assertEqual(results_data["project"]["change_note"]["file_name"], "수정 내용.txt")
        self.assertEqual(note_response.status_code, 200)
        self.assertEqual(note_data["job"]["id"], str(job.id))
        self.assertIn("기능리스트 v1.1 추가", note_data["change_note"]["content"])

    def test_job_project_change_note_endpoint_uses_log_fallback_after_cleanup(self):
        job, project = self._make_project()
        DownloadReviewLog.objects.create(
            job=job,
            job_project=project,
            level=DownloadReviewLogLevel.INFO,
            event_code="change_note_detected",
            message="수정 내용 파일 확인",
            detail_json={
                "available": True,
                "file_name": "수정 내용.txt",
                "file_path": "TTA-26-00010/수정 내용.txt",
                "content": "정리 후에도 보이는 수정 내용",
                "source": "file",
            },
        )

        results_response = job_project_results(
            self.factory.get(f"/api/job-projects/{project.id}/results/"),
            project.id,
        )
        results_data = json.loads(results_response.content.decode("utf-8"))
        note_response = job_project_change_note(
            self.factory.get(f"/api/job-projects/{project.id}/change-note/"),
            project.id,
        )
        note_data = json.loads(note_response.content.decode("utf-8"))

        self.assertEqual(results_response.status_code, 200)
        self.assertTrue(results_data["project"]["change_note"]["available"])
        self.assertEqual(note_response.status_code, 200)
        self.assertEqual(note_data["job"]["id"], str(job.id))
        self.assertEqual(note_data["change_note"]["content"], "정리 후에도 보이는 수정 내용")

    def _make_project(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            center_code="bundang",
            project_number="TTA-26-00010",
            ecm_row_json={"project_number": "TTA-26-00010", "company": "에이치소프트"},
        )
        return job, project

    def _add_rule_result(self, project):
        DownloadReviewRuleResult.objects.create(
            job_project=project,
            rule_code="required-report",
            rule_name="시험성적서 PDF 존재",
            sequence=1,
            status=DownloadReviewRuleStatus.PASS,
            expected="파일 존재",
            actual="파일 존재",
            message="정상 확인",
        )


class CsrfFailureViewTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    def test_api_csrf_failure_returns_json(self):
        request = self.factory.post(
            "/api/rule-results/00000000-0000-0000-0000-000000000000/manual-pass/",
            HTTP_ACCEPT="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        response = csrf_failure(request, reason="Referer checking failed")
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 403)
        self.assertFalse(data["success"])
        self.assertEqual(data["error_code"], "csrf_failed")
        self.assertIn("새로고침", data["message"])
        self.assertEqual(response["Cache-Control"], "no-store")

    def test_manual_pass_unexpected_error_returns_json(self):
        request = self.factory.post(
            "/api/rule-results/00000000-0000-0000-0000-000000000000/manual-pass/",
            data=json.dumps({"memo": "원본 파일 확인"}),
            content_type="application/json",
        )

        with self.assertLogs("main.views.review.ecm_download_review_api", level="ERROR"):
            with patch(
                "main.views.review.ecm_download_review_api.mark_rule_result_manual_pass",
                side_effect=RuntimeError("boom"),
            ):
                response = rule_result_manual_pass(
                    request,
                    "00000000-0000-0000-0000-000000000000",
                )
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 500)
        self.assertFalse(data["success"])
        self.assertIn("수동 적합 처리", data["message"])
        self.assertEqual(response["Cache-Control"], "no-store")


class DownloadReviewManualOverrideRobustnessTests(TestCase):
    databases = {"default", "workflow", "reference"}

    def setUp(self):
        self.factory = RequestFactory()

    def test_manual_pass_succeeds_when_project_recalculation_fails(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00030"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            center_code="bundang",
            project_number="TTA-26-00030",
            review_status=DownloadReviewProjectReviewStatus.NEEDS_FIX,
        )
        result = DownloadReviewRuleResult.objects.create(
            job_project=project,
            rule_code="artifact_14",
            rule_name="시험기록서",
            sequence=14,
            status=DownloadReviewRuleStatus.FAIL,
            expected="파일명에 기록서 포함",
            actual="조건에 맞는 파일을 찾지 못했습니다.",
            message="시험기록서 없음",
        )

        with self.assertLogs("main.views.review.ecm_download_review_jobs", level="ERROR"):
            with patch(
                "main.views.review.ecm_download_review_jobs._recalculate_project_review_after_manual_override",
                side_effect=RuntimeError("recalculate failed"),
            ):
                response = rule_result_manual_pass(
                    self.factory.post(
                        f"/api/rule-results/{result.id}/manual-pass/",
                        data=json.dumps({"memo": "외부 자료로 확인"}),
                        content_type="application/json",
                    ),
                    result.id,
                )
        data = json.loads(response.content.decode("utf-8"))
        result.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(data["success"])
        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS)
        self.assertTrue(
            DownloadReviewManualOverride.objects.filter(
                project_number="TTA-26-00030",
                rule_code="artifact_14",
            ).exists()
        )

    def test_manual_pass_succeeds_when_manual_override_table_is_unavailable(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00031"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            center_code="bundang",
            project_number="TTA-26-00031",
            review_status=DownloadReviewProjectReviewStatus.NEEDS_FIX,
        )
        result = DownloadReviewRuleResult.objects.create(
            job_project=project,
            rule_code="artifact_17",
            rule_name="SW저작권확인서",
            sequence=17,
            status=DownloadReviewRuleStatus.FAIL,
            expected="파일명에 확인서 포함",
            actual="조건에 맞는 파일을 찾지 못했습니다.",
            message="확인서 없음",
        )

        with self.assertLogs("main.views.review.ecm_download_review_jobs", level="ERROR"):
            with patch(
                "main.views.review.ecm_download_review_jobs.DownloadReviewManualOverride.objects.update_or_create",
                side_effect=OperationalError('relation "inspection_manual_override" does not exist'),
            ):
                response = rule_result_manual_pass(
                    self.factory.post(
                        f"/api/rule-results/{result.id}/manual-pass/",
                        data=json.dumps({"memo": "제출자 확인"}),
                        content_type="application/json",
                    ),
                    result.id,
                )
        data = json.loads(response.content.decode("utf-8"))
        result.refresh_from_db()
        overrides = manual_overrides_for_project(project, ["artifact_17"])

        self.assertEqual(response.status_code, 200)
        self.assertTrue(data["success"])
        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS)
        self.assertEqual(result.raw_detail_json["manual_override"]["memo"], "제출자 확인")
        self.assertFalse(DownloadReviewManualOverride.objects.exists())
        self.assertEqual(overrides["artifact_17"].memo, "제출자 확인")

    def test_manual_pass_succeeds_when_postgres_reports_missing_override_table_in_korean(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00032"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            center_code="sangam",
            project_number="TTA-26-00032",
            review_status=DownloadReviewProjectReviewStatus.NEEDS_FIX,
        )
        result = DownloadReviewRuleResult.objects.create(
            job_project=project,
            rule_code="artifact_09",
            rule_name="테스트케이스",
            sequence=9,
            status=DownloadReviewRuleStatus.FAIL,
            expected="작성일 일치",
            actual="작성일 불일치",
            message="테스트케이스 작성일 불일치",
        )

        with self.assertLogs("main.views.review.ecm_download_review_jobs", level="ERROR"):
            with patch(
                "main.views.review.ecm_download_review_jobs.DownloadReviewManualOverride.objects.update_or_create",
                side_effect=ProgrammingError('"inspection_manual_override" 이름의 릴레이션(relation)이 없습니다'),
            ):
                response = rule_result_manual_pass(
                    self.factory.post(
                        f"/api/rule-results/{result.id}/manual-pass/",
                        data=json.dumps({"memo": "인증위 제출 엑셀을 잘못 기입함. 실제 날짜는 맞음."}),
                        content_type="application/json",
                    ),
                    result.id,
                )

        data = json.loads(response.content.decode("utf-8"))
        result.refresh_from_db()
        overrides = manual_overrides_for_project(project, ["artifact_09"])

        self.assertEqual(response.status_code, 200)
        self.assertTrue(data["success"])
        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS)
        self.assertEqual(
            result.raw_detail_json["manual_override"]["memo"],
            "인증위 제출 엑셀을 잘못 기입함. 실제 날짜는 맞음.",
        )
        self.assertFalse(DownloadReviewManualOverride.objects.exists())
        self.assertEqual(overrides["artifact_09"].memo, "인증위 제출 엑셀을 잘못 기입함. 실제 날짜는 맞음.")


class DownloadReviewManualOverrideTests(TestCase):
    databases = {"default", "workflow", "reference"}

    def setUp(self):
        self.factory = RequestFactory()

    def test_manual_pass_requires_memo(self):
        _job, project = self._make_project()
        result = self._add_rule_result(project, status=DownloadReviewRuleStatus.FAIL)

        response = rule_result_manual_pass(
            self.factory.post(
                f"/api/rule-results/{result.id}/manual-pass/",
                data=json.dumps({"memo": "   "}),
                content_type="application/json",
            ),
            result.id,
        )
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 400)
        self.assertFalse(data["success"])
        self.assertIn("사유", data["message"])
        self.assertFalse(DownloadReviewManualOverride.objects.exists())

    def test_manual_pass_updates_current_result_and_project_payload(self):
        self._seed_rule()
        _job, project = self._make_project()
        result = self._add_rule_result(project, status=DownloadReviewRuleStatus.FAIL)

        response = rule_result_manual_pass(
            self.factory.post(
                f"/api/rule-results/{result.id}/manual-pass/",
                data=json.dumps({"memo": "원본 파일을 별도 확인해 정상으로 판단"}),
                content_type="application/json",
                REMOTE_ADDR="127.0.0.1",
            ),
            result.id,
        )
        data = json.loads(response.content.decode("utf-8"))
        result.refresh_from_db()
        project.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS)
        self.assertEqual(project.review_status, DownloadReviewProjectReviewStatus.COMPLETED)
        self.assertEqual(data["items"][0]["manual_override"]["memo"], "원본 파일을 별도 확인해 정상으로 판단")
        self.assertEqual(data["display_items"][0]["status"], DownloadReviewRuleStatus.PASS)
        self.assertEqual(data["display_items"][0]["manual_override"]["memo"], "원본 파일을 별도 확인해 정상으로 판단")

    def test_manual_pass_updates_only_selected_sub_check(self):
        self._seed_rule()
        _job, project = self._make_project()
        result = DownloadReviewRuleResult.objects.create(
            job_project=project,
            rule_code="required-contract",
            rule_name="Contract",
            sequence=1,
            status=DownloadReviewRuleStatus.FAIL,
            expected="A / B",
            actual="bad A / bad B",
            message="A failed",
            raw_detail_json={
                "sub_checks": [
                    {"expected": "[A] A", "actual": "bad A", "passed": False, "message": "A failed"},
                    {"expected": "[B] B", "actual": "bad B", "passed": False, "message": "B failed"},
                ],
            },
        )

        response = rule_result_manual_pass(
            self.factory.post(
                f"/api/rule-results/{result.id}/manual-pass/",
                data=json.dumps({"memo": "A was verified manually", "sub_check_key": "sub-1"}),
                content_type="application/json",
            ),
            result.id,
        )
        data = json.loads(response.content.decode("utf-8"))
        result.refresh_from_db()
        project.refresh_from_db()

        sub_checks = result.raw_detail_json["sub_checks"]
        self.assertEqual(response.status_code, 200)
        self.assertEqual(result.status, DownloadReviewRuleStatus.FAIL)
        self.assertEqual(project.review_status, DownloadReviewProjectReviewStatus.NEEDS_FIX)
        self.assertTrue(sub_checks[0]["passed"])
        self.assertEqual(sub_checks[0]["manual_override"]["sub_check_key"], "sub-1")
        self.assertFalse(sub_checks[1]["passed"])
        self.assertNotIn("manual_override", sub_checks[1])
        self.assertTrue(
            DownloadReviewManualOverride.objects.filter(
                project_number=project.project_number,
                rule_code="required-contract",
                sub_check_key="sub-1",
            ).exists()
        )
        self.assertEqual(data["display_items"][0]["status"], DownloadReviewRuleStatus.PASS)
        self.assertEqual(data["display_items"][0]["manual_override"]["sub_check_key"], "sub-1")
        self.assertEqual(data["display_items"][1]["status"], DownloadReviewRuleStatus.FAIL)
        self.assertIsNone(data["display_items"][1]["manual_override"])

    def test_sub_check_manual_pass_persists_only_for_matching_sub_check(self):
        _job, project = self._make_project()
        DownloadReviewManualOverride.objects.create(
            center_code="bundang",
            project_number=project.project_number,
            rule_code="required-contract",
            sub_check_key="sub-1",
            rule_name="Contract - A",
            memo="A was verified manually",
        )
        evaluation = engine.RuleEvaluation(
            rule=SimpleNamespace(code="required-contract", name="Contract"),
            sequence=1,
            status=DownloadReviewRuleStatus.FAIL,
            expected="A / B",
            actual="bad A / bad B",
            message="A failed",
            raw_detail={
                "sub_checks": [
                    {"expected": "[A] A", "actual": "bad A", "passed": False, "message": "A failed"},
                    {"expected": "[B] B", "actual": "bad B", "passed": False, "message": "B failed"},
                ],
            },
        )

        overrides = manual_overrides_for_project(project, ["required-contract"])
        updated = apply_manual_override_to_evaluation(evaluation, overrides["required-contract"])

        self.assertEqual(updated.status, DownloadReviewRuleStatus.FAIL)
        self.assertTrue(updated.raw_detail["sub_checks"][0]["passed"])
        self.assertEqual(updated.raw_detail["sub_checks"][0]["manual_override"]["memo"], "A was verified manually")
        self.assertFalse(updated.raw_detail["sub_checks"][1]["passed"])
        self.assertNotIn("manual_override", updated.raw_detail["sub_checks"][1])

    def test_sub_check_manual_pass_is_hidden_when_auto_check_passes(self):
        _job, project = self._make_project()
        DownloadReviewManualOverride.objects.create(
            center_code="bundang",
            project_number=project.project_number,
            rule_code="required-contract",
            sub_check_key="sub-1",
            rule_name="Contract - A",
            memo="A was verified manually",
        )
        evaluation = engine.RuleEvaluation(
            rule=SimpleNamespace(code="required-contract", name="Contract"),
            sequence=1,
            status=DownloadReviewRuleStatus.FAIL,
            expected="A / B",
            actual="A ok / bad B",
            message="B failed",
            raw_detail={
                "sub_checks": [
                    {"expected": "[A] A", "actual": "A ok", "passed": True, "message": "A passed"},
                    {"expected": "[B] B", "actual": "bad B", "passed": False, "message": "B failed"},
                ],
            },
        )

        overrides = manual_overrides_for_project(project, ["required-contract"])
        updated = apply_manual_override_to_evaluation(evaluation, overrides["required-contract"])

        self.assertEqual(updated.status, DownloadReviewRuleStatus.FAIL)
        self.assertTrue(updated.raw_detail["sub_checks"][0]["passed"])
        self.assertNotIn("manual_override", updated.raw_detail["sub_checks"][0])
        self.assertFalse(updated.raw_detail["sub_checks"][1]["passed"])

    def test_manual_pass_persists_across_next_inspection(self):
        self._seed_rule()
        DownloadReviewManualOverride.objects.create(
            center_code="bundang",
            project_number="TTA-26-00020",
            rule_code="required-contract",
            rule_name="계약서",
            memo="ECM 외부 대조로 계약서를 확인함",
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            download_root = Path(tmpdir)
            project_dir = download_root / "TTA-26-00020"
            project_dir.mkdir(parents=True)
            _job, project = self._make_project(download_dir=str(project_dir))
            verify_result = verify_downloaded_files(str(project_dir), project.project_number)

            with self.settings(AGENT_DOWNLOAD_BASE_DIR=download_root):
                outcome = run_download_inspection(project, verify_result, {})

        result = DownloadReviewRuleResult.objects.get(job_project=project, rule_code="required-contract")
        override = DownloadReviewManualOverride.objects.get(project_number=project.project_number, rule_code="required-contract")

        self.assertEqual(outcome.failed_count, 0)
        self.assertEqual(outcome.artifact_results["계약서"], "O")
        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS)
        self.assertEqual(result.raw_detail_json["manual_override"]["memo"], "ECM 외부 대조로 계약서를 확인함")
        self.assertIsNotNone(override.last_applied_at)

    def test_manual_pass_is_hidden_when_next_inspection_passes_automatically(self):
        self._seed_rule()
        DownloadReviewManualOverride.objects.create(
            center_code="bundang",
            project_number="TTA-26-00020",
            rule_code="required-contract",
            rule_name="계약서",
            memo="ECM 외부 대조로 계약서를 확인함",
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            download_root = Path(tmpdir)
            project_dir = download_root / "TTA-26-00020"
            project_dir.mkdir(parents=True)
            (project_dir / "TTA-26-00020 계약서.pdf").write_bytes(b"contract")
            _job, project = self._make_project(download_dir=str(project_dir))
            verify_result = verify_downloaded_files(str(project_dir), project.project_number)

            with self.settings(AGENT_DOWNLOAD_BASE_DIR=download_root):
                outcome = run_download_inspection(project, verify_result, {})

        result = DownloadReviewRuleResult.objects.get(job_project=project, rule_code="required-contract")
        override = DownloadReviewManualOverride.objects.get(project_number=project.project_number, rule_code="required-contract")

        self.assertEqual(outcome.failed_count, 0)
        self.assertEqual(outcome.artifact_results["계약서"], "O")
        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS)
        self.assertNotIn("manual_override", result.raw_detail_json)
        self.assertIsNone(override.last_applied_at)

    def _seed_rule(self):
        DownloadReviewRule.objects.create(
            code="required-contract",
            name="계약서",
            rule_type="required_file_name_contains",
            config_json={"contains": "계약서", "artifact_column": "계약서"},
            sort_order=1,
        )

    def _make_project(self, *, download_dir=""):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00020"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            center_code="bundang",
            project_number="TTA-26-00020",
            download_dir=download_dir,
            ecm_row_json={
                "project_number": "TTA-26-00020",
                "company": "에이치소프트",
                "product": "SecureFlow V1.0",
                "pl": "김준호",
            },
        )
        return job, project

    def _add_rule_result(self, project, *, status):
        return DownloadReviewRuleResult.objects.create(
            job_project=project,
            rule_code="required-contract",
            rule_name="계약서",
            sequence=1,
            status=status,
            expected="계약서 파일 1개",
            actual="조건에 맞는 파일 없음",
            message="계약서 파일을 찾지 못했습니다.",
            raw_detail_json={"artifact_column": "계약서"},
        )


@override_settings(DOWNLOAD_REVIEW_DEFAULT_CENTER="bundang")
class DownloadReviewJobsApiTests(TestCase):
    databases = {"default", "workflow", "reference"}

    def setUp(self):
        self.factory = RequestFactory()
        self.temp_dir = tempfile.TemporaryDirectory()
        # 이 클래스의 요청들은 center 를 지정하지 않으므로 fixture 센터인 bundang을
        # 기본 센터로 고정한다.
        self._seed_reference_projects(
            "bundang",
            [
                dict(
                    project_number="TTA-26-00009",
                    cert_date="05/12",
                    cert_committee_date=date(2026, 5, 12),
                    company="우리데이터 주식회사",
                    product="우리데이터클리닝 V1.0",
                    pl="박지훈",
                    review_result="O",
                    inspection_date="2026.05.12 20:30",
                ),
                dict(
                    project_number="TTA-26-00010",
                    cert_date="05/13",
                    cert_committee_date=date(2026, 5, 13),
                    company="에이치소프트",
                    product="SecureFlow 2.1",
                    pl="김준호",
                    review_result="",
                    inspection_date="",
                ),
                dict(
                    project_number="TTA-26-00011",
                    cert_date="05/14",
                    cert_committee_date=date(2026, 5, 14),
                    company="브릿지웨어",
                    product="BridgeHub",
                    pl="박지훈",
                    review_result="",
                    inspection_date="",
                ),
                dict(
                    project_number="TTA-26-00012",
                    cert_date="05/15",
                    cert_committee_date=date(2026, 5, 15),
                    company="넥스트랩",
                    product="NextLab QA Suite",
                    pl="최유진",
                    review_result="",
                    inspection_date="",
                ),
            ],
        )

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_job_request_creates_job_and_projects(self):
        response = self._post_job(["TTA-26-00010"])
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 201)
        self.assertTrue(data["success"])
        self.assertEqual(data["requested_project_count"], 1)
        self.assertIn(data["status"], {"scheduled", "queued"})
        self.assertEqual(DownloadReviewJob.objects.count(), 1)
        self.assertEqual(DownloadReviewProject.objects.count(), 1)

    def test_job_request_uses_selected_center_without_cross_center_conflict(self):
        # reference_project.project_number 는 센터 전역에서 유일하므로(더 이상 센터별
        # 파일로 분리되지 않음), 같은 프로젝트번호를 두 센터에 동시에 둘 수 없다.
        # 센터마다 별도 프로젝트번호를 써서 "센터 선택이 서로 간섭하지 않는지"를 검증한다.
        self._seed_reference_projects(
            "yeongnam",
            [
                dict(
                    project_number="TTA-26-00020",
                    cert_date="05/16",
                    cert_committee_date=date(2026, 5, 16),
                    company="영남기업",
                    product="Yeongnam Suite",
                    pl="김영남",
                    review_result="",
                    inspection_date="",
                ),
            ],
        )
        self._seed_reference_projects(
            "sangam",
            [
                dict(
                    project_number="TTA-26-00021",
                    cert_date="05/17",
                    cert_committee_date=date(2026, 5, 17),
                    company="상암기업",
                    product="Sangam Suite",
                    pl="박상암",
                    review_result="",
                    inspection_date="",
                ),
            ],
        )

        yeongnam_response = self._post_job(["TTA-26-00020"], center="yeongnam")
        yeongnam_data = json.loads(yeongnam_response.content.decode("utf-8"))

        sangam_response = self._post_job(["TTA-26-00021"], center="sangam")
        sangam_data = json.loads(sangam_response.content.decode("utf-8"))

        self.assertEqual(yeongnam_response.status_code, 201)
        self.assertEqual(yeongnam_data["center_code"], "yeongnam")
        self.assertEqual(sangam_response.status_code, 201)
        self.assertEqual(sangam_data["center_code"], "sangam")
        self.assertEqual(DownloadReviewJob.objects.count(), 2)
        self.assertEqual(
            set(DownloadReviewProject.objects.values_list("center_code", flat=True)),
            {"sangam", "yeongnam"},
        )

    def test_job_request_rejects_completed_projects(self):
        response = self._post_job(["TTA-26-00009"])
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 400)
        self.assertFalse(data["success"])
        self.assertEqual(data["error_code"], "completed_project_not_allowed")
        self.assertEqual(data["details"]["completed_project_numbers"], ["TTA-26-00009"])
        self.assertEqual(DownloadReviewJob.objects.count(), 0)

    def test_job_request_rejects_active_duplicate_projects(self):
        first_response = self._post_job(["TTA-26-00010"])
        self.assertEqual(first_response.status_code, 201)

        response = self._post_job(["TTA-26-00010"])
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 409)
        self.assertFalse(data["success"])
        self.assertEqual(data["error_code"], "active_project_conflict")
        self.assertEqual(data["details"]["conflicts"][0]["project_number"], "TTA-26-00010")
        self.assertEqual(DownloadReviewJob.objects.count(), 1)

    def test_active_job_returns_no_polling_when_empty(self):
        response = active_job(self.factory.get("/api/jobs/active/"))
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertIsNone(data["active_job"])
        self.assertFalse(data["polling"]["should_poll"])
        self.assertIsNone(data["polling"]["recommended_interval_ms"])
        self.assertIsNone(data["polling"]["wake_at"])

    def test_active_job_and_detail_endpoints_return_created_job(self):
        created = json.loads(self._post_job(["TTA-26-00010"]).content.decode("utf-8"))

        active_response = active_job(self.factory.get("/api/jobs/active/"))
        active_data = json.loads(active_response.content.decode("utf-8"))
        detail_response = job_detail(
            self.factory.get(f"/api/jobs/{created['job_id']}/"),
            created["job_id"],
        )
        detail_data = json.loads(detail_response.content.decode("utf-8"))

        self.assertEqual(active_response.status_code, 200)
        self.assertEqual(active_data["active_job"]["id"], created["job_id"])
        if active_data["active_job"]["status"] == "scheduled":
            self.assertFalse(active_data["polling"]["should_poll"])
            self.assertIsNone(active_data["polling"]["recommended_interval_ms"])
            self.assertIsNotNone(active_data["polling"]["wake_at"])
        else:
            self.assertTrue(active_data["polling"]["should_poll"])
            self.assertEqual(active_data["polling"]["recommended_interval_ms"], 3000)
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_data["job"]["selected_project_numbers"], ["TTA-26-00010"])

    def test_active_job_endpoint_uses_global_worker_queue_even_with_center_query(self):
        sangam = DownloadReviewJob.objects.create(
            center_code="sangam",
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        yeongnam = DownloadReviewJob.objects.create(
            center_code="yeongnam",
            status=DownloadReviewJobStatus.SCHEDULED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00011"],
        )

        response = active_job(self.factory.get("/api/jobs/active/", {"center": "yeongnam"}))
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(data["active_job"]["id"], str(sangam.id))
        self.assertNotEqual(data["active_job"]["id"], str(yeongnam.id))
        self.assertEqual(data["active_job_count"], 2)
        self.assertNotIn("center_active_job_count", data)

    def test_projects_api_marks_active_project_as_not_selectable(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.SCHEDULED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        DownloadReviewProject.objects.create(
            job=job,
            center_code="bundang",
            project_number="TTA-26-00010",
            ecm_row_json={"project_number": "TTA-26-00010", "company": "에이치소프트"},
        )

        response = projects(
            self.factory.get("/api/projects/", {"project_number": "TTA-26-00010"}),
        )
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(data["items"][0]["active_job_id"], str(job.id))
        self.assertEqual(data["items"][0]["active_state_label"], "예약중")
        self.assertFalse(data["items"][0]["selectable"])

    def test_jobs_list_endpoint_returns_recent_jobs_and_filters_status(self):
        completed = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=2,
            completed_project_count=1,
            failed_project_count=1,
            selected_projects_json=["TTA-26-00010", "TTA-26-00011"],
        )
        DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.SCHEDULED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00012"],
        )

        all_response = jobs(self.factory.get("/api/jobs/", {"status": "all", "limit": "10"}))
        all_data = json.loads(all_response.content.decode("utf-8"))
        finished_response = jobs(self.factory.get("/api/jobs/", {"status": "finished"}))
        finished_data = json.loads(finished_response.content.decode("utf-8"))
        completed_response = jobs(self.factory.get("/api/jobs/", {"status": "completed"}))
        completed_data = json.loads(completed_response.content.decode("utf-8"))

        self.assertEqual(all_response.status_code, 200)
        self.assertEqual(all_data["pagination"]["total"], 2)
        self.assertEqual(all_data["items"][0]["status"], DownloadReviewJobStatus.SCHEDULED)
        self.assertEqual(finished_response.status_code, 200)
        self.assertEqual(finished_data["pagination"]["total"], 1)
        self.assertEqual(finished_data["items"][0]["id"], str(completed.id))
        self.assertEqual(completed_response.status_code, 200)
        self.assertEqual(completed_data["pagination"]["total"], 1)
        self.assertEqual(completed_data["items"][0]["id"], str(completed.id))
        self.assertEqual(completed_data["items"][0]["completed_project_count"], 1)
        self.assertEqual(completed_data["items"][0]["failed_project_count"], 1)

    def test_jobs_list_endpoint_filters_explicit_center_query(self):
        DownloadReviewJob.objects.create(
            center_code="sangam",
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        yeongnam = DownloadReviewJob.objects.create(
            center_code="yeongnam",
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00011"],
        )

        response = jobs(self.factory.get("/api/jobs/", {"status": "all", "center": "yeongnam"}))
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(data["pagination"]["total"], 1)
        self.assertEqual(data["items"][0]["id"], str(yeongnam.id))
        self.assertEqual(data["items"][0]["center_code"], "yeongnam")

    def test_jobs_list_endpoint_without_center_accumulates_all_centers(self):
        DownloadReviewJob.objects.create(
            center_code="sangam",
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        DownloadReviewJob.objects.create(
            center_code="yeongnam",
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00011"],
        )

        response = jobs(self.factory.get("/api/jobs/", {"status": "all", "limit": "10"}))
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(data["pagination"]["total"], 2)
        self.assertEqual({item["center_code"] for item in data["items"]}, {"sangam", "yeongnam"})
        self.assertIsNone(data["center"])

    def test_cancel_scheduled_job_marks_projects_skipped(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.SCHEDULED,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            ecm_row_json={"project_number": "TTA-26-00010", "company": "에이치소프트"},
        )

        response = job_cancel(self.factory.post(f"/api/jobs/{job.id}/cancel/"), job.id)
        data = json.loads(response.content.decode("utf-8"))
        job.refresh_from_db()
        project.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(data["success"])
        self.assertEqual(job.status, DownloadReviewJobStatus.CANCELED)
        self.assertIsNotNone(job.canceled_at)
        self.assertEqual(project.status, DownloadReviewProjectStatus.SKIPPED)
        self.assertEqual(project.current_step, "사용자 취소")

    def test_cancel_running_job_is_rejected(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )

        response = job_cancel(self.factory.post(f"/api/jobs/{job.id}/cancel/"), job.id)
        data = json.loads(response.content.decode("utf-8"))
        job.refresh_from_db()

        self.assertEqual(response.status_code, 409)
        self.assertFalse(data["success"])
        self.assertEqual(data["error_code"], "job_cancel_not_allowed")
        self.assertEqual(job.status, DownloadReviewJobStatus.RUNNING)

    def test_jobs_list_endpoint_rejects_unknown_filters(self):
        cases = [
            {"raw_sql": "SELECT * FROM automation_job"},
            {"status": "unknown"},
            {"limit": "101"},
            {"offset": "-1"},
        ]

        for params in cases:
            with self.subTest(params=params):
                response = jobs(self.factory.get("/api/jobs/", params))
                data = json.loads(response.content.decode("utf-8"))

                self.assertEqual(response.status_code, 400)
                self.assertFalse(data["success"])
                self.assertEqual(data["error_code"], "invalid_job_request")

    def test_job_projects_and_results_endpoints_return_project_data(self):
        created = json.loads(self._post_job(["TTA-26-00010"]).content.decode("utf-8"))
        project = DownloadReviewProject.objects.get(project_number="TTA-26-00010")
        project.status = DownloadReviewProjectStatus.INSPECTING
        project.current_step = "zip 검사 중"
        project.save(update_fields=["status", "current_step"])
        DownloadReviewRuleResult.objects.create(
            job_project=project,
            rule_code="required-report",
            rule_name="시험성적서 PDF 존재",
            sequence=1,
            file_path="TTA-26-00010/시험성적서.pdf",
            file_name="시험성적서.pdf",
            status=DownloadReviewRuleStatus.PASS,
            expected="파일 존재",
            actual="파일 존재",
            message="정상 확인",
        )

        projects_response = job_projects(
            self.factory.get(f"/api/jobs/{created['job_id']}/projects/"),
            created["job_id"],
        )
        projects_data = json.loads(projects_response.content.decode("utf-8"))
        results_response = job_project_results(
            self.factory.get(f"/api/job-projects/{project.id}/results/"),
            project.id,
        )
        results_data = json.loads(results_response.content.decode("utf-8"))

        self.assertEqual(projects_response.status_code, 200)
        self.assertEqual(projects_data["items"][0]["project_number"], "TTA-26-00010")
        self.assertEqual(projects_data["items"][0]["status_label"], "검사중")
        self.assertEqual(results_response.status_code, 200)
        self.assertEqual(results_data["items"][0]["status_label"], "정상")

    def test_result_excel_endpoints_return_workbooks(self):
        from openpyxl import load_workbook

        created = json.loads(self._post_job(["TTA-26-00010"]).content.decode("utf-8"))
        project = DownloadReviewProject.objects.get(project_number="TTA-26-00010")
        DownloadReviewRuleResult.objects.create(
            job_project=project,
            rule_code="required-report",
            rule_name="시험성적서 PDF 존재",
            sequence=1,
            file_path="TTA-26-00010/report.pdf",
            file_name="report.pdf",
            status=DownloadReviewRuleStatus.PASS,
            expected="파일 존재",
            actual="파일 존재",
            message="정상 확인",
        )

        project_response = job_project_results_excel(
            self.factory.get(f"/api/job-projects/{project.id}/results.xlsx"),
            project.id,
        )
        job_response = job_results_excel(
            self.factory.get(f"/api/jobs/{created['job_id']}/results.xlsx"),
            created["job_id"],
        )

        self.assertEqual(project_response.status_code, 200)
        self.assertIn("spreadsheetml.sheet", project_response["Content-Type"])
        self.assertIn("filename*=UTF-8''", project_response["Content-Disposition"])
        project_workbook = load_workbook(BytesIO(project_response.content))
        project_values = [
            cell
            for row in project_workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]
        self.assertIn("TTA-26-00010", project_values)
        self.assertIn("시험성적서 PDF 존재", project_values)

        self.assertEqual(job_response.status_code, 200)
        job_workbook = load_workbook(BytesIO(job_response.content))
        job_values = [
            cell
            for row in job_workbook.active.iter_rows(values_only=True)
            for cell in row
            if cell is not None
        ]
        self.assertIn("TTA-26-00010", job_values)
        self.assertIn("시험성적서 PDF 존재", job_values)

    def test_latest_project_results_endpoint_returns_most_recent_finished_project(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.COMPLETED,
            requested_project_count=1,
            completed_project_count=1,
            selected_projects_json=["TTA-26-00009"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            center_code="bundang",
            project_number="TTA-26-00009",
            status=DownloadReviewProjectStatus.COMPLETED,
            review_status=DownloadReviewProjectReviewStatus.COMPLETED,
            ecm_row_json={
                "project_number": "TTA-26-00009",
                "company": "우리데이터 주식회사",
                "product": "우리데이터클리닝 V1.0",
            },
        )
        DownloadReviewRuleResult.objects.create(
            job_project=project,
            rule_code="required-report",
            rule_name="시험성적서 PDF 존재",
            sequence=1,
            file_path="TTA-26-00009/시험성적서.pdf",
            file_name="시험성적서.pdf",
            status=DownloadReviewRuleStatus.PASS,
            expected="파일 존재",
            actual="파일 존재",
            message="정상 확인",
        )

        response = latest_project_results(
            self.factory.get("/api/projects/TTA-26-00009/latest-results/"),
            "TTA-26-00009",
        )
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(data["project"]["project_number"], "TTA-26-00009")
        self.assertEqual(data["items"][0]["rule_name"], "시험성적서 PDF 존재")
        self.assertEqual(data["items"][0]["status_label"], "정상")

    def test_download_inspection_records_results_and_cleanup_deletes_download_dir(self):
        download_root = Path(self.temp_dir.name) / "downloads"
        project_dir = download_root / "TTA-26-00010_1"
        project_dir.mkdir(parents=True)
        (project_dir / "계약서_TTA-26-00010.pdf").write_bytes(b"contract")
        (project_dir / "readme.txt").write_bytes(b"readme")
        DownloadReviewRule.objects.create(
            code="계약서",
            name="계약서",
            rule_type="required_file_name_contains",
            config_json={"contains": "계약서", "artifact_column": "계약서"},
            sort_order=1,
        )
        DownloadReviewRule.objects.create(
            code="시험성적서(PDF)",
            name="시험성적서(PDF)",
            rule_type="required_file_name_contains",
            config_json={"contains": "시험성적서", "artifact_column": "시험성적서(PDF)"},
            sort_order=2,
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            center_code="bundang",
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={
                "project_number": "TTA-26-00010",
                "company": "에이치소프트",
                "pl": "김준호",
            },
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")
        file_summary = {
            "file_count": verify_result.file_count,
            "file_names": [file_info.name for file_info in verify_result.files],
        }

        with self.settings(AGENT_DOWNLOAD_BASE_DIR=download_root):
            outcome = run_download_inspection(project, verify_result, file_summary)
            write_project_review_result(
                project.project_number,
                outcome.reference_review,
                artifact_results=outcome.artifact_results,
            )
            cleanup = cleanup_download_dir(project)

        rows = self._reference_rows(
            ["TTA-26-00010"],
            ["점검결과", "계약서", "시험성적서(PDF)"],
        )
        project.refresh_from_db()

        self.assertEqual(outcome.reference_review, "X")
        self.assertEqual(outcome.failed_count, 1)
        self.assertGreaterEqual(DownloadReviewRuleResult.objects.filter(job_project=project).count(), 2)
        self.assertEqual(rows["TTA-26-00010"]["점검결과"], "X")
        self.assertEqual(rows["TTA-26-00010"]["계약서"], "O")
        self.assertEqual(rows["TTA-26-00010"]["시험성적서(PDF)"], "X")
        self.assertTrue(cleanup.deleted)
        self.assertFalse(project_dir.exists())
        self.assertIsNotNone(project.zip_deleted_at)

    def test_actual_artifact_rules_inspect_zip_entries(self):
        project_dir = Path(self.temp_dir.name) / "downloads"
        project_dir.mkdir(parents=True)
        master_db_path = Path(self.temp_dir.name) / "reference.db"
        self._create_master_reference_db(master_db_path)
        zip_path = project_dir / "TTA-26-00010.zip"
        rawdata_zip_path = project_dir / "TTA-26-00010 rawdata.zip"
        with zipfile.ZipFile(rawdata_zip_path, "w") as rawdata_archive:
            rawdata_archive.writestr("결함/defect.png", b"image")
            rawdata_archive.writestr("보안/1차/raw.txt", b"security")
            rawdata_archive.writestr("보안/2차/raw.txt", b"security")
            rawdata_archive.writestr("성능/1차/raw.txt", b"performance")
            rawdata_archive.writestr("성능/2차/raw.txt", b"performance")
            for folder_name in ("최초정상", "최종정상"):
                for index in range(5):
                    info = zipfile.ZipInfo(
                        f"3.설계/제품스크린샷/{folder_name}/image-{index}.png",
                        date_time=(2026, 5, 20, 12, 0, 0),
                    )
                    rawdata_archive.writestr(info, b"image")
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.writestr("2.계약/TTA-26-00010 계약서.pdf", b"contract")
            archive.writestr(
                "2.계약/TTA-26-00010 시험합의서.docx",
                _docx_bytes(
                    tables=[[["시험신청번호", "TTA-26-00010"]]],
                    header="TTA-26-00010",
                    footer="TIS-0101-3 (00)",
                ),
            )
            archive.writestr(
                "2.계약/TTA-26-00010 시험합의서.pdf",
                _pdf_bytes(["ApplicationNo", "TTA-26-00010"]),
            )
            archive.writestr("2.계약/TTA-26-00010 수수료산정표.xlsx", b"fee")
            archive.writestr("4.시험/가.계획/TTA-26-00010 시험환경구성도.pptx", b"diagram")
            archive.writestr(
                "4.시험/가.계획/TTA-26-00010 기능리스트.xlsx",
                _xlsx_bytes(
                    rows=[
                        ["TTA-26-00010 기능리스트"],
                        ["작성자: 김준호"],
                        ["대분류", "중분류", "기능"],
                        ["대분류1", "중분류1", "기능1"],
                    ],
                ),
            )
            archive.writestr(
                "4.시험/가.계획/TTA-26-00010 시험계획서.docx",
                _test_plan_docx("TTA-26-00010", product="테스트제품", version="v1.0", pl="김준호", wd="10"),
            )
            archive.writestr(
                "4.시험/가.계획/TTA-26-00010 시험계획서.pdf",
                _pdf_bytes(["TTA-26-00010 시험계획서"]),
            )
            archive.writestr(
                "4.시험/가.계획/TTA-26-00010 품질특성별 제품 정보 기재사항.docx",
                _docx_bytes(
                    paragraphs=[
                        "(TTA-26-00010) 품질특성별",
                        "(2026.5.28)",
                    ],
                ),
            )
            for folder_name in ("최초형상", "최종형상"):
                for index in range(5):
                    info = zipfile.ZipInfo(
                        f"3.설계/제품스크린샷/{folder_name}/image-{index}.png",
                        date_time=(2026, 5, 20, 12, 0, 0),
                    )
                    archive.writestr(info, b"image")
            archive.writestr(
                "3.설계/TTA-26-00010 테스트케이스.xlsx",
                _test_case_xlsx("TTA-26-00010", pl="김준호", residual_count=2, footer="TTA"),
            )
            archive.writestr(
                "3.설계/TTA-26-00010 점검표.xlsx",
                _inspection_checklist_xlsx("TTA-26-00010", pl="김준호", wd="10", high="3", before="7"),
            )
            archive.writestr(
                "3.설계/TTA-26-00010 점검표.pdf",
                _pdf_bytes(["TTA-26-00010 점검표"]),
            )
            archive.writestr("5.수행/결함리포트/raw.txt", b"defect")
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v1.0.xlsx",
                _defect_report_xlsx("TTA-26-00010", ["1차 결함리포트"]),
            )
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v2.0.xlsx",
                _defect_report_xlsx("TTA-26-00010", ["1차 결함리포트", "2차 결함리포트"]),
            )
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v3.0.xlsx",
                _defect_report_xlsx(
                    "TTA-26-00010",
                    ["1차 결함리포트", "2차 결함리포트", "최종결함리포트", "시험분석자료"],
                ),
            )
            archive.writestr("5.수행/보안/1차/raw.txt", b"security")
            archive.writestr("5.수행/보안/2차/raw.txt", b"security")
            archive.writestr("5.수행/성능/1차/raw.txt", b"performance")
            archive.writestr("5.수행/성능/2차/raw.txt", b"performance")
            archive.writestr(
                "6.시험/나.종료/TTA-26-00010 시험성적서.docx",
                _docx_bytes(
                    header="TTA-26-00010",
                    footer="TPG-1016-5(02)",
                    paragraphs=["<세부사양>"],
                    tables=[
                        [["항목", "값"], ["OS", "Windows"]],
                        [["결함리포트 송부 1차: 2026.05.10 2차: 2026.05.20"]],
                    ],
                ),
            )
            archive.writestr(
                "6.시험/나.종료/TTA-26-00010 시험성적서.pdf",
                _pdf_bytes(["TTA-26-00010 시험성적서"]),
            )
            archive.writestr(
                "6.시험/나.종료/TTA-26-00010 시험기록서.pdf",
                _pdf_bytes(["TTA-26-00010 시험기록서"]),
            )
            archive.writestr(
                "6.시험/나.종료/v2.0 2026.05.10. 변수확인.txt",
                b"variable probe",
            )
            archive.writestr(
                "6.시험/인증관련/TTA-26-00010 품질검사표.xlsx",
                _quality_inspection_table_xlsx("TTA-26-00010", footer="한국정보통신기술협회"),
            )
            archive.writestr(
                "6.시험/인증관련/TTA-26-00010 품질평가보고서.docx",
                _quality_evaluation_report_docx("TTA-26-00010"),
            )
            archive.writestr("6.시험/인증관련/SW저작권확인서.pdf", b"copyright")
            archive.writestr("7.홍보자료/promo.jpg", b"promo")

        call_command("seed_download_review_rules", "--only-real", "--enable", stdout=StringIO())
        agreement_rule = DownloadReviewRule.objects.get(name="합의서(PDF)")
        agreement_config = agreement_rule.config_json
        agreement_config["content_checks"][1]["label"] = "ApplicationNo"
        agreement_rule.config_json = agreement_config
        agreement_rule.save(update_fields=["config_json", "updated_at"])
        DownloadReviewRule.objects.create(
            code="variable_probe",
            name="변수 전달 테스트",
            rule_type="required_artifact_file",
            target_file_type="any",
            enabled=True,
            sort_order=900,
            config_json={
                "folder_keyword_chain": ["시험", "종료"],
                "filename_keywords": ["v{결함차수}.0", "{1차}"],
                "extensions": [".txt"],
                "min_count": 1,
                "pass_message": "이전 규칙 산출 변수를 확인했습니다.",
            },
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={
                "project_number": "TTA-26-00010",
                "company": "에이치소프트",
                "product": "테스트제품 v1.0",
                "pl": "김준호",
                "wd": "10",
                "신청일": "2026.05.02.",
                "계약일": "2026.05.03.",
                "인증일자": "2026.06.01.",
            },
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")

        artifact_dir = Path(self.temp_dir.name) / "artifacts"
        with self.settings(
            DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH=master_db_path,
            DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir,
        ):
            outcome = run_download_inspection(project, verify_result, {})
        results = {
            result.rule_name: result
            for result in DownloadReviewRuleResult.objects.filter(job_project=project)
        }

        failed_rules = [
            (result.rule_name, result.message)
            for result in DownloadReviewRuleResult.objects.filter(job_project=project, status=DownloadReviewRuleStatus.FAIL)
        ]
        self.assertEqual(outcome.reference_review, "O", failed_rules)
        self.assertEqual(outcome.failed_count, 0)
        self.assertEqual(outcome.artifact_results["계약서"], "O")
        self.assertEqual(outcome.artifact_results["합의서(PDF)"], "O")
        self.assertEqual(outcome.artifact_results["수수료산정표"], "O")
        self.assertEqual(outcome.artifact_results["시험환경구성도"], "O")
        self.assertEqual(outcome.artifact_results["품질특성별제품정보기재사항"], "O")
        self.assertEqual(outcome.artifact_results["기능리스트"], "O")
        self.assertEqual(outcome.artifact_results["시험계획서(PDF)"], "O")
        self.assertEqual(outcome.artifact_results["최초/최종형상RawData"], "O")
        self.assertEqual(outcome.artifact_results["테스트케이스"], "O")
        self.assertEqual(outcome.artifact_results["결함리포트"], "O")
        self.assertEqual(outcome.artifact_results["점검표(PDF)"], "O")
        self.assertEqual(outcome.artifact_results["1차/2차/성능/보안RawData"], "O")
        self.assertEqual(outcome.artifact_results["시험성적서(PDF)"], "O")
        self.assertEqual(outcome.artifact_results["시험기록서"], "O")
        self.assertEqual(outcome.artifact_results["품질평가보고서"], "O")
        self.assertEqual(outcome.artifact_results["품질검사표"], "O")
        self.assertEqual(outcome.artifact_results["SW저작권확인서"], "O")
        self.assertEqual(outcome.artifact_results["홍보이미지"], "O")
        self.assertIn("2.계약", results["계약서"].file_path)
        self.assertIn("4.시험/가.계획", results["시험환경구성도"].file_path)
        agreement_result = results["합의서(PDF)"]
        feature_result = results["기능리스트"]
        plan_result = results["시험계획서(PDF)"]
        self.assertTrue(
            any(
                check.get("part") == "header" and check["passed"]
                for check in agreement_result.raw_detail_json["content_checks"]
            )
        )
        self.assertTrue(
            any(
                check.get("part") == "footer" and check["passed"]
                for check in agreement_result.raw_detail_json["content_checks"]
            )
        )
        self.assertEqual(results["변수 전달 테스트"].status, DownloadReviewRuleStatus.PASS)
        report_result = results["시험성적서(PDF)"]
        self.assertEqual(report_result.raw_detail_json["variables"]["결함차수"], 2)
        self.assertEqual(report_result.raw_detail_json["variables"]["1차"], "2026.05.10.")
        self.assertEqual(report_result.raw_detail_json["variables"]["2차"], "2026.05.20.")
        self.assertEqual(report_result.raw_detail_json["variables"]["시험성적서_세부사양표"], [["항목", "값"], ["OS", "Windows"]])
        self.assertEqual(get_rule_output_variables(project)["결함차수"], 2)
        self.assertTrue(
            any(
                check.get("details")
                and check["details"][0]["term"] == "TIS-"
                and check["details"][0]["passed"]
                for check in plan_result.raw_detail_json["checks"]
            )
        )
        self.assertEqual(plan_result.raw_detail_json["checks"][-1]["name"], "spec_table")
        self.assertTrue(plan_result.raw_detail_json["checks"][-1]["passed"])
        defect_result = results["결함리포트"]
        self.assertEqual(defect_result.raw_detail_json["variables"]["잔여결함수"], 2)
        self.assertEqual(defect_result.raw_detail_json["variables"]["H"], "3")
        self.assertEqual(defect_result.raw_detail_json["variables"]["R"], "7")
        self.assertTrue(defect_result.raw_detail_json["print_text_checks"]["forbidden_headers"]["passed"])
        self.assertTrue(defect_result.raw_detail_json["print_text_checks"]["forbidden_footers"]["passed"])
        test_case_result = results["테스트케이스"]
        self.assertTrue(test_case_result.raw_detail_json["footer_check"]["passed"])
        self.assertEqual(test_case_result.raw_detail_json["residual_defect_check"]["expected_count"], 2)
        self.assertEqual(test_case_result.raw_detail_json["residual_defect_check"]["actual_count"], 2)
        self.assertEqual(test_case_result.raw_detail_json["residual_defect_check"]["failed_rows"], [7, 8])
        checklist_result = results["점검표(PDF)"]
        self.assertTrue(checklist_result.raw_detail_json["footer_checks"]["forbidden"]["passed"])
        self.assertTrue(checklist_result.raw_detail_json["footer_checks"]["required"]["passed"])
        self.assertEqual(len(checklist_result.raw_detail_json["variables"]["측정항목별점수표"]), 84)
        self.assertEqual(checklist_result.raw_detail_json["variables"]["측정항목별점수표"][0], "score-1")
        quality_result = results["품질검사표"]
        quality_values = quality_result.raw_detail_json["variables"]["품질부특성측정값"]
        self.assertEqual(len(quality_values), 32)
        self.assertEqual(quality_values[0], "quality-4")
        self.assertNotIn("quality-27", quality_values)
        self.assertEqual(quality_values[-1], "quality-3")
        report_quality_result = results["품질평가보고서"]
        self.assertEqual(report_quality_result.raw_detail_json["project_number_count"], 6)
        self.assertEqual(len(report_quality_result.raw_detail_json["quality_value_check"]["actual_values"]), 32)
        for result in (agreement_result, feature_result, plan_result, report_result):
            self.assertEqual(len(result.raw_detail_json["artifacts"]), 1)
            self.assertTrue(
                (artifact_dir / result.raw_detail_json["artifacts"][0]["relative_path"]).is_file()
            )
        self.assertEqual(len(checklist_result.raw_detail_json["artifacts"]), 1)
        self.assertTrue(
            (artifact_dir / checklist_result.raw_detail_json["artifacts"][0]["relative_path"]).is_file()
        )
        test_record_result = results["시험기록서"]
        test_record_artifact = test_record_result.raw_detail_json["artifacts"][0]
        self.assertFalse(test_record_artifact["download"])
        self.assertEqual(test_record_artifact["content_type"], "image/png")
        self.assertTrue(
            (artifact_dir / test_record_artifact["relative_path"]).is_file()
        )
        with self.settings(DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir):
            artifact_response = rule_result_artifact(
                self.factory.get(f"/api/rule-results/{report_result.id}/artifacts/pdf_first_page/"),
                report_result.id,
                "pdf_first_page",
            )
        artifact_bytes = b"".join(artifact_response.streaming_content)
        self.assertEqual(artifact_response.status_code, 200)
        self.assertEqual(artifact_response["Content-Type"], "image/png")
        self.assertTrue(artifact_bytes.startswith(b"\x89PNG"))
        with self.settings(DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir):
            agreement_artifact_response = rule_result_artifact(
                self.factory.get(f"/api/rule-results/{agreement_result.id}/artifacts/pdf_first_page/"),
                agreement_result.id,
                "pdf_first_page",
            )
        agreement_artifact_bytes = b"".join(agreement_artifact_response.streaming_content)
        self.assertEqual(agreement_artifact_response.status_code, 200)
        self.assertEqual(agreement_artifact_response["Content-Type"], "image/png")
        self.assertTrue(agreement_artifact_bytes.startswith(b"\x89PNG"))
        with self.settings(DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir):
            feature_artifact_response = rule_result_artifact(
                self.factory.get(f"/api/rule-results/{feature_result.id}/artifacts/feature_list_area/"),
                feature_result.id,
                "feature_list_area",
            )
        feature_artifact_bytes = b"".join(feature_artifact_response.streaming_content)
        self.assertEqual(feature_artifact_response.status_code, 200)
        self.assertEqual(feature_artifact_response["Content-Type"], "image/png")
        self.assertTrue(feature_artifact_bytes.startswith(b"\x89PNG"))
        with self.settings(DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir):
            checklist_artifact_response = rule_result_artifact(
                self.factory.get(f"/api/rule-results/{checklist_result.id}/artifacts/pdf_first_page/"),
                checklist_result.id,
                "pdf_first_page",
            )
        checklist_artifact_bytes = b"".join(checklist_artifact_response.streaming_content)
        self.assertEqual(checklist_artifact_response.status_code, 200)
        self.assertEqual(checklist_artifact_response["Content-Type"], "image/png")
        self.assertTrue(checklist_artifact_bytes.startswith(b"\x89PNG"))
        results_response = job_project_results(
            self.factory.get(f"/api/job-projects/{project.id}/results/"),
            project.id,
        )
        results_payload = json.loads(results_response.content.decode("utf-8"))
        report_payload = next(
            item for item in results_payload["items"]
            if item["rule_name"] == "시험성적서(PDF)"
        )
        self.assertEqual(report_payload["artifacts"][0]["id"], "pdf_first_page")
        self.assertNotIn("relative_path", report_payload["artifacts"][0])
        self.assertNotIn("relative_path", report_payload["raw_detail"]["artifacts"][0])
        agreement_payload = next(
            item for item in results_payload["items"]
            if item["rule_name"] == "합의서(PDF)"
        )
        feature_payload = next(
            item for item in results_payload["items"]
            if item["rule_name"] == "기능리스트"
        )
        self.assertEqual(agreement_payload["artifacts"][0]["label"], "합의서 1페이지")
        self.assertEqual(feature_payload["artifacts"][0]["id"], "feature_list_area")

    def test_cleanup_stale_project_history_keeps_only_latest_per_project(self):
        """같은 프로젝트번호를 재점검하면 이전 job_project 의 산출물 폴더·DB 행(결과/로그)만
        지워지고, 최신 것과 다른 프로젝트번호/센터의 것은 그대로 남아야 한다."""
        artifact_dir = Path(self.temp_dir.name) / "artifacts"
        artifact_dir.mkdir()

        def _make_job():
            return DownloadReviewJob.objects.create(
                status=DownloadReviewJobStatus.SCHEDULED,
                requested_project_count=1,
                selected_projects_json=["TTA-26-00010"],
            )

        # 재점검은 매번 새 Job 을 만든다(같은 Job 안에서는 project_number 가 유일해야 함).
        old_project = DownloadReviewProject.objects.create(
            job=_make_job(), project_number="TTA-26-00010", center_code="sangam",
        )
        new_project = DownloadReviewProject.objects.create(
            job=_make_job(), project_number="TTA-26-00010", center_code="sangam",
        )
        other_center_project = DownloadReviewProject.objects.create(
            job=_make_job(), project_number="TTA-26-00010", center_code="yeongnam",
        )
        other_project_number = DownloadReviewProject.objects.create(
            job=_make_job(), project_number="TTA-26-99999", center_code="sangam",
        )

        for project in (old_project, new_project, other_center_project, other_project_number):
            folder = artifact_dir / str(project.id)
            folder.mkdir()
            (folder / "artifact_09_pdf_first_page.png").write_bytes(b"fake-png")
            DownloadReviewRuleResult.objects.create(
                job_project=project, rule_name="테스트케이스",
                status=DownloadReviewRuleStatus.PASS,
            )
            DownloadReviewLog.objects.create(
                job=project.job, job_project=project,
                level=DownloadReviewLogLevel.INFO, message="점검 완료",
            )

        with self.settings(DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir):
            summary = cleanup_stale_project_history(new_project)

        self.assertEqual(summary["artifact_dirs_removed"], 1)
        self.assertEqual(summary["project_rows_removed"], 1)
        self.assertEqual(summary["cascaded_rows_removed"], 2)  # RuleResult 1 + Log 1

        self.assertFalse((artifact_dir / str(old_project.id)).exists())
        self.assertTrue((artifact_dir / str(new_project.id)).exists())
        # 다른 센터(같은 프로젝트번호)·다른 프로젝트번호의 산출물은 건드리지 않는다.
        self.assertTrue((artifact_dir / str(other_center_project.id)).exists())
        self.assertTrue((artifact_dir / str(other_project_number.id)).exists())

        self.assertFalse(DownloadReviewProject.objects.filter(id=old_project.id).exists())
        self.assertFalse(DownloadReviewRuleResult.objects.filter(job_project_id=old_project.id).exists())
        self.assertFalse(DownloadReviewLog.objects.filter(job_project_id=old_project.id).exists())
        # 다른 행들의 결과/로그는 그대로 남아야 한다.
        self.assertTrue(DownloadReviewProject.objects.filter(id=new_project.id).exists())
        self.assertTrue(DownloadReviewRuleResult.objects.filter(job_project_id=new_project.id).exists())
        self.assertTrue(DownloadReviewProject.objects.filter(id=other_center_project.id).exists())
        self.assertTrue(DownloadReviewProject.objects.filter(id=other_project_number.id).exists())

    def _write_valid_rawdata_zip(self, rawdata_zip_path):
        with zipfile.ZipFile(rawdata_zip_path, "w") as archive:
            archive.writestr("결함/defect.png", b"image")
            archive.writestr("보안/1차/raw.txt", b"security")
            archive.writestr("보안/2차/raw.txt", b"security")
            archive.writestr("성능/1차/raw.txt", b"performance")
            archive.writestr("성능/2차/raw.txt", b"performance")
            for folder_name in ("최초정상", "최종정상"):
                for index in range(5):
                    info = zipfile.ZipInfo(
                        f"3.설계/제품스크린샷/{folder_name}/image-{index}.png",
                        date_time=(2026, 5, 20, 12, 0, 0),
                    )
                    archive.writestr(info, b"image")

    def _rawdata_only_project(self, project_dir):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        return DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={
                "project_number": "TTA-26-00010",
                "company": "에이치소프트",
                "product": "테스트제품 v1.0",
                "pl": "김준수",
                "wd": "10",
                "신청일": "2026.05.02.",
                "계약일": "2026.05.03.",
                "인증일자": "2026.06.01.",
            },
        )

    def test_rawdata_rules_run_when_only_rawdata_zip_exists(self):
        project_dir = Path(self.temp_dir.name) / "rawdata_only"
        project_dir.mkdir(parents=True)
        master_db_path = Path(self.temp_dir.name) / "reference_rawdata_only.db"
        self._create_master_reference_db(master_db_path)
        self._write_valid_rawdata_zip(project_dir / "raw_data.zip")

        call_command("seed_download_review_rules", "--only-real", "--enable", stdout=StringIO())
        project = self._rawdata_only_project(project_dir)
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")

        artifact_dir = Path(self.temp_dir.name) / "artifacts_rawdata_only"
        with self.settings(
            DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH=master_db_path,
            DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir,
        ):
            outcome = run_download_inspection(project, verify_result, {})

        results = {
            result.rule_name: result
            for result in DownloadReviewRuleResult.objects.filter(job_project=project)
        }
        self.assertTrue(verify_result.success)
        self.assertFalse(verify_result.has_project_number_files)
        self.assertEqual(outcome.artifact_results["최초/최종형상RawData"], "O")
        self.assertEqual(outcome.artifact_results["1차/2차/성능/보안RawData"], "O")
        self.assertEqual(outcome.artifact_results["계약서"], "X")
        self.assertEqual(results["최초/최종형상RawData"].status, DownloadReviewRuleStatus.PASS)
        self.assertEqual(results["1차/2차/성능/보안RawData"].status, DownloadReviewRuleStatus.PASS)

    def test_rawdata_rules_continue_when_submission_zip_is_unreadable(self):
        project_dir = Path(self.temp_dir.name) / "corrupt_submission_with_rawdata"
        project_dir.mkdir(parents=True)
        master_db_path = Path(self.temp_dir.name) / "reference_corrupt_submission.db"
        self._create_master_reference_db(master_db_path)
        (project_dir / "TTA-26-00010.zip").write_bytes(b"not a zip")
        self._write_valid_rawdata_zip(project_dir / "TTA-26-00010 raw-data.zip")

        call_command("seed_download_review_rules", "--only-real", "--enable", stdout=StringIO())
        project = self._rawdata_only_project(project_dir)
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")

        artifact_dir = Path(self.temp_dir.name) / "artifacts_corrupt_submission"
        with self.settings(
            DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH=master_db_path,
            DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir,
        ):
            outcome = run_download_inspection(project, verify_result, {})

        results = {
            result.rule_name: result
            for result in DownloadReviewRuleResult.objects.filter(job_project=project)
        }
        self.assertTrue(verify_result.success)
        self.assertEqual(outcome.artifact_results["최초/최종형상RawData"], "O")
        self.assertEqual(outcome.artifact_results["1차/2차/성능/보안RawData"], "O")
        self.assertEqual(results["최초/최종형상RawData"].status, DownloadReviewRuleStatus.PASS)
        self.assertEqual(results["1차/2차/성능/보안RawData"].status, DownloadReviewRuleStatus.PASS)
        self.assertTrue(getattr(verify_result, "_inspection_zip_errors"))

    def test_environment_diagram_accepts_png_and_pptx_together(self):
        project_dir = Path(self.temp_dir.name) / "environment_diagram"
        project_dir.mkdir(parents=True)
        zip_path = project_dir / "TTA-26-00010.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.writestr("4.시험/가.계획/TTA-26-00010 환경구성도.png", b"png")
            archive.writestr("4.시험/가.계획/TTA-26-00010 환경구성도.pptx", b"pptx")

        DownloadReviewRule.objects.create(
            code="artifact_04",
            name="시험환경구성도",
            rule_type="required_artifact_file",
            target_file_type="any",
            enabled=True,
            sort_order=40,
            config_json={
                "artifact_column": "시험환경구성도",
                "folder_keyword_chain": ["시험", "계획"],
                "filename_keywords": ["구성도", "{project_number}"],
                "extensions": [".png", ".pptx"],
                "min_count": 1,
            },
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={"project_number": "TTA-26-00010"},
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")

        outcome = run_download_inspection(project, verify_result, {})
        result = DownloadReviewRuleResult.objects.get(job_project=project, rule_name="시험환경구성도")

        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS)
        self.assertEqual(outcome.artifact_results["시험환경구성도"], "O")
        self.assertEqual(result.raw_detail_json["matched_file_count"], 2)

    def test_performance_rawdata_passes_when_folder_has_any_entry(self):
        project_dir = Path(self.temp_dir.name) / "rawdata_performance"
        project_dir.mkdir(parents=True)
        rawdata_zip_path = project_dir / "TTA-26-00010 rawdata.zip"
        with zipfile.ZipFile(rawdata_zip_path, "w") as archive:
            archive.writestr("결함/raw.txt", b"defect")
            archive.writestr("보안/1차/raw.txt", b"security")
            archive.writestr("보안/2차/raw.txt", b"security")
            archive.writestr("성능시험/측정자료/raw.txt", b"performance")

        DownloadReviewRule.objects.create(
            code="artifact_12",
            name="1차/2차/성능/보안RawData",
            rule_type="rawdata_folder_structure_check",
            target_file_type="any",
            enabled=True,
            sort_order=120,
            config_json={
                "artifact_column": "1차/2차/성능/보안RawData",
                "folder_checks": [
                    {"keyword": "결함", "failure_message": "결함리포트 rawdata 확인 불가"},
                    {
                        "keyword": "보안",
                        "exact_child_folders": 2,
                        "each_child_has_entry": True,
                        "txt_only_pass": True,
                        "unwrap_single_folder": True,
                        "failure_message": "보안성 rawdata 확인 불가",
                    },
                    {
                        "keyword": "성능",
                        "min_entries": 1,
                        "failure_message": "성능 rawdata 확인 불가",
                    },
                ],
                "pass_message": "rawdata 폴더 구조를 확인했습니다.",
            },
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={"project_number": "TTA-26-00010"},
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")

        outcome = run_download_inspection(project, verify_result, {})
        result = DownloadReviewRuleResult.objects.get(job_project=project, rule_name="1차/2차/성능/보안RawData")

        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS, result.raw_detail_json)
        self.assertEqual(outcome.artifact_results["1차/2차/성능/보안RawData"], "O")

    def test_security_rawdata_passes_on_exception_marker_file_even_without_subfolders(self):
        # 실제 사례(TTA-26-01501): 보안시험 폴더 안에 하위 폴더 2개 구조 없이
        # "인빅티 대상아님"이라는 0바이트 안내 파일만 있어도, pass_if_file_name_contains
        # 목록의 단어 중 하나(예: '대상')만 포함되면 폴더 구조와 무관하게 적합 처리한다.
        project_dir = Path(self.temp_dir.name) / "rawdata_security_exception"
        project_dir.mkdir(parents=True)
        rawdata_zip_path = project_dir / "TTA-26-00010 rawdata.zip"
        with zipfile.ZipFile(rawdata_zip_path, "w") as archive:
            archive.writestr("결함/raw.txt", b"defect")
            archive.writestr("보안시험/인빅티 대상아님", b"")  # 0바이트, 하위 폴더 없음
            archive.writestr("성능시험/측정자료/raw.txt", b"performance")

        DownloadReviewRule.objects.create(
            code="artifact_12",
            name="1차/2차/성능/보안RawData",
            rule_type="rawdata_folder_structure_check",
            target_file_type="any",
            enabled=True,
            sort_order=120,
            config_json={
                "artifact_column": "1차/2차/성능/보안RawData",
                "folder_checks": [
                    {"keyword": "결함", "failure_message": "결함리포트 rawdata 확인 불가"},
                    {
                        "keyword": "보안",
                        "exact_child_folders": 2,
                        "each_child_has_entry": True,
                        "txt_only_pass": True,
                        "unwrap_single_folder": True,
                        "pass_if_file_name_contains": ["인빅티", "invicti", "수행", "대상", "시험", "면제"],
                        "failure_message": "보안성 rawdata 확인 불가",
                    },
                    {
                        "keyword": "성능",
                        "min_entries": 1,
                        "failure_message": "성능 rawdata 확인 불가",
                    },
                ],
                "pass_message": "rawdata 폴더 구조를 확인했습니다.",
            },
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={"project_number": "TTA-26-00010"},
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")

        outcome = run_download_inspection(project, verify_result, {})
        result = DownloadReviewRuleResult.objects.get(job_project=project, rule_name="1차/2차/성능/보안RawData")

        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS, result.raw_detail_json)
        self.assertEqual(outcome.artifact_results["1차/2차/성능/보안RawData"], "O")

    def test_image_screenshot_date_failure_lists_period_dates_and_count(self):
        project_dir = Path(self.temp_dir.name) / "downloads"
        project_dir.mkdir(parents=True)
        master_db_path = Path(self.temp_dir.name) / "reference.db"
        self._create_master_reference_db(master_db_path)
        zip_path = project_dir / "TTA-26-00010.zip"
        with zipfile.ZipFile(zip_path, "w"):
            pass
        rawdata_zip_path = project_dir / "TTA-26-00010 rawdata.zip"
        with zipfile.ZipFile(rawdata_zip_path, "w") as archive:
            for folder_name, date_time in (
                ("최초형상", (2026, 4, 30, 12, 0, 0)),
                ("최종형상", (2026, 6, 1, 12, 0, 0)),
            ):
                for index in range(5):
                    info = zipfile.ZipInfo(
                        f"3.설계/제품스크린샷/{folder_name}/image-{index}.png",
                        date_time=date_time,
                    )
                    archive.writestr(info, b"image")

        DownloadReviewRule.objects.create(
            code="artifact_8",
            name="최초/최종형상RawData",
            rule_type="image_screenshot_folder_date_check",
            target_file_type="any",
            enabled=True,
            sort_order=80,
            config_json={
                "artifact_column": "최초/최종형상RawData",
                "folder_keyword_chain": ["설계"],
                "min_images_per_folder": 5,
                "required_candidate_folder_count": 2,
            },
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={"project_number": "TTA-26-00010"},
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")

        with self.settings(DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH=master_db_path):
            outcome = run_download_inspection(project, verify_result, {})

        result = DownloadReviewRuleResult.objects.get(
            job_project=project,
            rule_name="최초/최종형상RawData",
        )
        message = "시험기간은 2026.05.01.~2026.05.31.인데 수정일자가 2026.04.30. 또는 2026.06.01.인 이미지가 10개 존재함"
        self.assertEqual(outcome.failed_count, 1)
        self.assertEqual(result.status, DownloadReviewRuleStatus.FAIL)
        self.assertEqual(result.message, message)
        self.assertEqual(result.actual, message)
        self.assertEqual(result.raw_detail_json["out_of_range_date_counts"], {
            "2026.04.30.": 5,
            "2026.06.01.": 5,
        })

    def test_defect_report_count_mismatch_uses_test_report_message(self):
        project_dir = Path(self.temp_dir.name) / "downloads"
        project_dir.mkdir(parents=True)
        zip_path = project_dir / "TTA-26-00010.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.writestr(
                "6.시험/나.종료/TTA-26-00010 시험성적서.docx",
                _docx_bytes(
                    header="TTA-26-00010",
                    tables=[
                        [["결함리포트 송부 1차: 2026.05.10 2차: 2026.05.20"]],
                    ],
                ),
            )
            archive.writestr(
                "6.시험/나.종료/TTA-26-00010 시험성적서.pdf",
                _pdf_bytes(["TTA-26-00010 시험성적서"]),
            )
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v1.0.xlsx",
                _defect_report_xlsx("TTA-26-00010", ["1차 결함리포트"]),
            )
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v2.0.xlsx",
                _defect_report_xlsx("TTA-26-00010", ["1차 결함리포트", "2차 결함리포트"]),
            )

        DownloadReviewRule.objects.create(
            code="artifact_13",
            name="시험성적서(PDF)",
            rule_type="test_report_document_check",
            target_file_type="any",
            enabled=True,
            sort_order=95,
            config_json={
                "artifact_column": "시험성적서(PDF)",
                "folder_keyword_chain": ["시험", "종료"],
                "filename_keywords": ["시험성적서", "{project_number}"],
                "required_files": [
                    {"extensions": [".docx"], "exact_count": 1},
                    {"extensions": [".pdf"], "exact_count": 1},
                ],
                "spec_marker": "<세부사양>",
                "pdf_artifact_label": "시험성적서 1페이지",
            },
        )
        DownloadReviewRule.objects.create(
            code="artifact_10",
            name="결함리포트",
            rule_type="defect_report_check",
            target_file_type="any",
            enabled=True,
            sort_order=100,
            config_json={
                "artifact_column": "결함리포트",
                "folder_keyword_chain": ["수행"],
                "filename_keywords": ["결함리포트", "{project_number}"],
                "extensions": [".xlsx", ".xls"],
                "count_mismatch_message": "시험성적서의 결함 차수와 결함리포트 개수가 다름",
            },
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={"project_number": "TTA-26-00010"},
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")
        artifact_dir = Path(self.temp_dir.name) / "artifacts"

        with self.settings(DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir):
            outcome = run_download_inspection(project, verify_result, {})

        result = DownloadReviewRuleResult.objects.get(job_project=project, rule_name="결함리포트")
        self.assertEqual(outcome.failed_count, 1)
        self.assertEqual(result.status, DownloadReviewRuleStatus.FAIL)
        self.assertEqual(result.message, "시험성적서의 결함 차수와 결함리포트 개수가 다름")
        self.assertEqual(result.actual, "결함리포트 Excel 파일 2개")

    def test_defect_report_dates_accept_split_title_and_report_date_cells(self):
        project_dir = Path(self.temp_dir.name) / "downloads"
        project_dir.mkdir(parents=True)
        master_db_path = Path(self.temp_dir.name) / "reference.db"
        self._create_master_reference_db(master_db_path)
        zip_path = project_dir / "TTA-26-00010.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v1.0.xlsx",
                _defect_report_xlsx_split_title("TTA-26-00010", ["1차 결함리포트"]),
            )
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v2.0.xlsx",
                _defect_report_xlsx_split_title("TTA-26-00010", ["1차 결함리포트", "2차 결함리포트"]),
            )
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v3.0.xlsx",
                _defect_report_xlsx_split_title(
                    "TTA-26-00010",
                    ["1차 결함리포트", "2차 결함리포트", "최종결함리포트", "시험분석자료"],
                ),
            )
            archive.writestr(
                "6.시험/다.종료/TTA-26-00010 시험성적서.docx",
                _docx_bytes(
                    tables=[
                        [["결함리포트 송부 1차: 2026.04.14 2차: 2026.04.20"]],
                    ],
                ),
            )
            archive.writestr(
                "6.시험/다.종료/TTA-26-00010 시험성적서.pdf",
                _pdf_bytes(["TTA-26-00010 시험성적서"]),
            )

        DownloadReviewRule.objects.create(
            code="artifact_13",
            name="시험성적서(PDF)",
            rule_type="test_report_document_check",
            target_file_type="any",
            enabled=True,
            sort_order=95,
            config_json={
                "artifact_column": "시험성적서(PDF)",
                "folder_keyword_chain": ["시험", "종료"],
                "filename_keywords": ["시험성적서", "{project_number}"],
                "required_files": [
                    {"extensions": [".docx"], "exact_count": 1},
                    {"extensions": [".pdf"], "exact_count": 1},
                ],
                "spec_marker": "<일반사항>",
                "pdf_artifact_label": "시험성적서 1페이지",
            },
        )
        DownloadReviewRule.objects.create(
            code="artifact_10",
            name="결함리포트",
            rule_type="defect_report_check",
            target_file_type="any",
            enabled=True,
            sort_order=100,
            config_json={
                "artifact_column": "결함리포트",
                "folder_keyword_chain": ["수행"],
                "filename_keywords": ["결함리포트", "{project_number}"],
                "extensions": [".xlsx", ".xls"],
                "count_mismatch_message": "시험성적서의 결함 차수와 결함리포트 개수가 다름",
                "report_date_message": "프로젝트 번호, 결함 차시, 보고일자 중 잘못된 값이 작성됨",
            },
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={"project_number": "TTA-26-00010"},
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")
        artifact_dir = Path(self.temp_dir.name) / "artifacts"

        with self.settings(
            DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH=master_db_path,
            DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir,
        ):
            run_download_inspection(project, verify_result, {})

        result = DownloadReviewRuleResult.objects.get(job_project=project, rule_name="결함리포트")
        self.assertEqual(result.status, DownloadReviewRuleStatus.PASS, result.raw_detail_json["report_date_checks"])
        self.assertEqual(result.raw_detail_json["report_date_checks"][0]["sheet_text"], "1차 결함리포트")
        self.assertEqual(result.raw_detail_json["report_date_checks"][0]["actual_date"], "2026.04.14.")

    def test_zero_residual_defect_variable_is_available_when_defect_count_missing(self):
        project_dir = Path(self.temp_dir.name) / "downloads"
        project_dir.mkdir(parents=True)
        master_db_path = Path(self.temp_dir.name) / "reference.db"
        self._create_master_reference_db(master_db_path)
        zip_path = project_dir / "TTA-26-00010.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v3.0.xlsx",
                _defect_report_zero_residual_xlsx("TTA-26-00010"),
            )
            archive.writestr(
                "3.설계/TTA-26-00010 테스트케이스.xlsx",
                _test_case_xlsx("TTA-26-00010", pl="김준호", residual_count=0),
            )

        DownloadReviewRule.objects.create(
            code="artifact_10",
            name="결함리포트",
            rule_type="defect_report_check",
            target_file_type="any",
            enabled=True,
            sort_order=100,
            config_json={
                "artifact_column": "결함리포트",
                "folder_keyword_chain": ["수행"],
                "filename_keywords": ["결함리포트", "{project_number}"],
                "extensions": [".xlsx", ".xls"],
                "count_mismatch_message": "시험성적서의 결함 차수와 결함리포트 개수가 다름",
            },
        )
        DownloadReviewRule.objects.create(
            code="artifact_09",
            name="테스트케이스",
            rule_type="test_case_check",
            target_file_type="any",
            enabled=True,
            sort_order=105,
            config_json={
                "artifact_column": "테스트케이스",
                "folder_keyword_chain": ["설계"],
                "filename_keywords": ["테스트케이스", "{project_number}"],
                "extensions": [".xlsx", ".xls"],
                "exact_count": 1,
                "title_text": "{project_number} 테스트케이스",
                "author_label": "작성자",
                "reviewer_label": "검토자:",
                "reviewer_expected": "김진영",
                "date_label": "작성일",
                "result_header": "상세 테스트 결과",
                "residual_message": "잔여 결함이 작성되지 않음",
            },
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={
                "project_number": "TTA-26-00010",
                "pl": "김준호",
            },
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")

        with self.settings(DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH=master_db_path):
            run_download_inspection(project, verify_result, {})

        defect_result = DownloadReviewRuleResult.objects.get(job_project=project, rule_name="결함리포트")
        test_case_result = DownloadReviewRuleResult.objects.get(job_project=project, rule_name="테스트케이스")
        self.assertEqual(defect_result.raw_detail_json["variables"]["잔여결함수"], 0)
        self.assertEqual(test_case_result.raw_detail_json["residual_defect_check"]["expected_count"], 0)
        self.assertEqual(test_case_result.raw_detail_json["residual_defect_check"]["actual_count"], 0)
        self.assertTrue(
            next(
                item for item in test_case_result.raw_detail_json["sub_checks"]
                if item["expected"].startswith("[잔여결함]")
            )["passed"]
        )

    def test_defect_report_environment_uses_value_in_right_cell(self):
        workbook_by_version = {
            1: ExcelWorkbook(sheets=[
                ExcelSheet("1차 결함리포트", [["", "", "", "[시험환경 :", "관리자 PC : Windows 11]"]]),
            ]),
            2: ExcelWorkbook(sheets=[
                ExcelSheet("2차 결함리포트", [["", "", "", "[시험환경 :", "관리자 PC : Windows 11]"]]),
            ]),
            3: ExcelWorkbook(sheets=[
                ExcelSheet("최종결함리포트", [["", "", "", "[시험환경 :", "관리자 PC : Windows 11]"]]),
                ExcelSheet("시험분석자료", [["", "", "", "", "[ 시험환경: 관리자 PC : Windows 11]"]]),
            ]),
        }

        result = _check_defect_report_environment(workbook_by_version)

        self.assertTrue(result["passed"], result)
        self.assertIn("관리자 PC : Windows 11", result["actual"])

    def test_defect_report_variables_are_kept_when_print_text_check_fails(self):
        project_dir = Path(self.temp_dir.name) / "downloads"
        project_dir.mkdir(parents=True)
        master_db_path = Path(self.temp_dir.name) / "reference.db"
        self._create_master_reference_db(master_db_path)
        zip_path = project_dir / "TTA-26-00010.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.writestr(
                "4.시험/나.설계/TTA-26-00010 테스트케이스.xlsx",
                _test_case_xlsx("TTA-26-00010", pl="김준호"),
            )
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v1.0.xlsx",
                _defect_report_xlsx("TTA-26-00010", ["1차 결함리포트"]),
            )
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v2.0.xlsx",
                _defect_report_xlsx("TTA-26-00010", ["1차 결함리포트", "2차 결함리포트"]),
            )
            archive.writestr(
                "5.수행/TTA-26-00010 결함리포트 v3.0.xlsx",
                _defect_report_xlsx(
                    "TTA-26-00010",
                    ["1차 결함리포트", "2차 결함리포트", "최종결함리포트", "시험분석자료"],
                    footer="소프트웨어시험인증연구소",
                ),
            )
            archive.writestr(
                "6.시험/나.종료/TTA-26-00010 시험성적서.docx",
                _docx_bytes(
                    header="TTA-26-00010",
                    tables=[
                        [["결함리포트 송부 1차: 2026.05.10 2차: 2026.05.20"]],
                    ],
                ),
            )
            archive.writestr(
                "6.시험/나.종료/TTA-26-00010 시험성적서.pdf",
                _pdf_bytes(["TTA-26-00010 시험성적서"]),
            )

        DownloadReviewRule.objects.create(
            code="artifact_13",
            name="시험성적서(PDF)",
            rule_type="test_report_document_check",
            target_file_type="any",
            enabled=True,
            sort_order=95,
            config_json={
                "artifact_column": "시험성적서(PDF)",
                "folder_keyword_chain": ["시험", "종료"],
                "filename_keywords": ["시험성적서", "{project_number}"],
                "required_files": [
                    {"extensions": [".docx"], "exact_count": 1},
                    {"extensions": [".pdf"], "exact_count": 1},
                ],
                "spec_marker": "<세부사양>",
                "pdf_artifact_label": "시험성적서 1페이지",
            },
        )
        DownloadReviewRule.objects.create(
            code="artifact_10",
            name="결함리포트",
            rule_type="defect_report_check",
            target_file_type="any",
            enabled=True,
            sort_order=100,
            config_json={
                "artifact_column": "결함리포트",
                "folder_keyword_chain": ["수행"],
                "filename_keywords": ["결함리포트", "{project_number}"],
                "extensions": [".xlsx", ".xls"],
                "version_pattern": r"(?i)v(\d+)\.0",
                "forbidden_footer_terms": [
                    {
                        "text": "소프트웨어시험인증연구소",
                        "message": "결함리포트 바닥글에 '소프트웨어시험인증연구소'라는 단어가 잘못 작성됨",
                    },
                ],
            },
        )
        DownloadReviewRule.objects.create(
            code="artifact_9",
            name="테스트케이스",
            rule_type="test_case_check",
            target_file_type="any",
            enabled=True,
            sort_order=110,
            config_json={
                "artifact_column": "테스트케이스",
                "folder_keyword_chain": ["설계"],
                "filename_keywords": ["테스트케이스", "{project_number}"],
                "extensions": [".xlsx", ".xls"],
                "exact_count": 1,
                "title_text": "{project_number} 테스트케이스",
                "author_label": "작성자:",
                "reviewer_label": "검토자:",
                "reviewer_expected": "김진영",
                "date_text": "작성일: {시작일} ~ {종료일}",
                "result_header": "상세 테스트 결과",
                "residual_message": "잔여 결함이 작성되지 않음",
            },
        )
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={
                "project_number": "TTA-26-00010",
                "product": "테스트제품 v1.0",
                "pl": "김준호",
            },
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")
        artifact_dir = Path(self.temp_dir.name) / "artifacts"

        with self.settings(
            DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH=master_db_path,
            DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir,
        ):
            outcome = run_download_inspection(project, verify_result, {})

        results = {
            result.rule_name: result
            for result in DownloadReviewRuleResult.objects.filter(job_project=project)
        }
        defect_result = results["결함리포트"]
        test_case_result = results["테스트케이스"]

        self.assertEqual(outcome.failed_count, 1)
        self.assertEqual(defect_result.status, DownloadReviewRuleStatus.FAIL)
        self.assertEqual(defect_result.raw_detail_json["variables"]["잔여결함수"], 2)
        self.assertEqual(defect_result.raw_detail_json["variables"]["H"], "3")
        self.assertEqual(defect_result.raw_detail_json["variables"]["R"], "7")
        self.assertEqual(test_case_result.status, DownloadReviewRuleStatus.PASS)
        self.assertEqual(test_case_result.raw_detail_json["residual_defect_check"]["expected_count"], 2)
        self.assertEqual(get_rule_output_variables(project)["잔여결함수"], 2)

    def test_quality_report_still_uses_quality_values_when_quality_table_score_compare_fails(self):
        project_dir = Path(self.temp_dir.name) / "downloads"
        project_dir.mkdir(parents=True)
        master_db_path = Path(self.temp_dir.name) / "reference.db"
        self._create_master_reference_db(master_db_path)
        zip_path = project_dir / "TTA-26-00010.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.writestr(
                "4.시험/나.설계/TTA-26-00010 점검표.xlsx",
                _inspection_checklist_xlsx("TTA-26-00010", pl="김준호", wd="10", high="", before=""),
            )
            archive.writestr(
                "4.시험/나.설계/TTA-26-00010 점검표.pdf",
                _pdf_bytes(["TTA-26-00010 점검표"]),
            )
            archive.writestr(
                "6.시험/인증관련/TTA-26-00010 품질검사표.xlsx",
                _quality_inspection_table_xlsx("TTA-26-00010", score_overrides={21: "NA"}),
            )
            archive.writestr(
                "6.시험/인증관련/TTA-26-00010 품질평가보고서.docx",
                _quality_evaluation_report_docx("TTA-26-00010"),
            )

        call_command("seed_download_review_rules", "--only-real", "--enable", stdout=StringIO())
        DownloadReviewRule.objects.exclude(
            name__in=["점검표(PDF)", "품질검사표", "품질평가보고서"]
        ).update(enabled=False)
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
            selected_projects_json=["TTA-26-00010"],
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            project_number="TTA-26-00010",
            download_dir=str(project_dir),
            ecm_row_json={
                "project_number": "TTA-26-00010",
                "company": "에이치소프트",
                "product": "테스트제품 v1.0",
                "pl": "김준호",
                "wd": "10",
                "request_date": "2026.05.02.",
                "contract_date": "2026.05.03.",
                "cert_date": "2026.06.01.",
            },
        )
        verify_result = verify_downloaded_files(str(project_dir), "TTA-26-00010")
        artifact_dir = Path(self.temp_dir.name) / "artifacts"

        with self.settings(
            DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH=master_db_path,
            DOWNLOAD_REVIEW_ARTIFACT_DIR=artifact_dir,
        ):
            outcome = run_download_inspection(project, verify_result, {})

        results = {
            result.rule_name: result
            for result in DownloadReviewRuleResult.objects.filter(job_project=project)
        }
        quality_table_result = results["품질검사표"]
        quality_report_result = results["품질평가보고서"]

        self.assertEqual(outcome.failed_count, 1)
        self.assertEqual(quality_table_result.status, DownloadReviewRuleStatus.FAIL)
        self.assertEqual(
            quality_table_result.message,
            "점검표와 품질검사표의 품질부특성 값이 총 84개의 값 중에 1개의 값이 다름"
            "(현재 값: 총 84개 중 1개 값이 다름)",
        )
        self.assertEqual(quality_table_result.raw_detail_json["score_compare"]["total_count"], 84)
        self.assertEqual(quality_table_result.raw_detail_json["score_compare"]["mismatch_count"], 1)
        self.assertEqual(len(quality_table_result.raw_detail_json["variables"]["품질부특성측정값"]), 32)
        self.assertNotIn("quality-27", quality_table_result.raw_detail_json["variables"]["품질부특성측정값"])
        self.assertEqual(quality_report_result.status, DownloadReviewRuleStatus.PASS)
        self.assertTrue(quality_report_result.raw_detail_json["quality_value_check"]["passed"])
        self.assertEqual(len(quality_report_result.raw_detail_json["quality_value_check"]["actual_values"]), 32)
        self.assertIn("품질부특성측정값", get_rule_output_variables(project))

    def test_dry_run_worker_completes_job_with_mixed_project_results(self):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.QUEUED,
            requested_project_count=3,
            selected_projects_json=["TTA-26-00010", "TTA-26-00011", "TTA-26-00012"],
            progress_message="대기열 등록 완료",
        )
        # 워커의 참조 DB 반영(_write_reference_result_safely)은 project.center_code 를 그대로
        # 넘기므로, setUp에서 심어둔 bundang 픽스처와 맞추기 위해 명시적으로 지정한다
        # (DownloadReviewProject.center_code 모델 기본값은 'sangam'이라 지정하지 않으면 어긋난다).
        DownloadReviewProject.objects.bulk_create(
            [
                DownloadReviewProject(
                    job=job,
                    center_code="bundang",
                    project_number="TTA-26-00010",
                    ecm_row_json={"project_number": "TTA-26-00010", "company": "에이치소프트"},
                ),
                DownloadReviewProject(
                    job=job,
                    center_code="bundang",
                    project_number="TTA-26-00011",
                    ecm_row_json={"project_number": "TTA-26-00011", "company": "브릿지웨어"},
                ),
                DownloadReviewProject(
                    job=job,
                    center_code="bundang",
                    project_number="TTA-26-00012",
                    ecm_row_json={"project_number": "TTA-26-00012", "company": "넥스트랩"},
                ),
            ]
        )

        result = run_worker_once(dry_run=True)
        job.refresh_from_db()
        job_projects = list(job.projects.order_by("project_number"))
        reference_rows = self._reference_rows(
            ["TTA-26-00010", "TTA-26-00011", "TTA-26-00012"],
            ["점검결과", "점검날짜", "회사명", "계약서", "시험성적서(PDF)"],
        )
        projects_response = projects(
            self.factory.get("/api/projects/", {"project_number": "TTA-26-00010"}),
        )
        projects_data = json.loads(projects_response.content.decode("utf-8"))

        self.assertTrue(result.processed)
        self.assertEqual(result.status, "completed")
        self.assertEqual(job.status, DownloadReviewJobStatus.COMPLETED)
        self.assertEqual(job.completed_project_count, 2)
        self.assertEqual(job.failed_project_count, 1)
        self.assertEqual(job_projects[0].review_status, DownloadReviewProjectReviewStatus.COMPLETED)
        self.assertEqual(job_projects[1].review_status, DownloadReviewProjectReviewStatus.NEEDS_FIX)
        self.assertEqual(job_projects[2].review_status, DownloadReviewProjectReviewStatus.HELD)
        self.assertEqual(DownloadReviewRuleResult.objects.filter(job_project=job_projects[0]).count(), 30)
        self.assertEqual(
            DownloadReviewRuleResult.objects.filter(
                job_project=job_projects[1],
                status=DownloadReviewRuleStatus.FAIL,
            ).count(),
            1,
        )
        self.assertEqual(DownloadReviewRuleResult.objects.filter(job_project=job_projects[2]).count(), 0)
        self.assertEqual(reference_rows["TTA-26-00010"]["점검결과"], "O")
        self.assertEqual(reference_rows["TTA-26-00010"]["회사명"], "에이치소프트")
        self.assertEqual(reference_rows["TTA-26-00010"]["계약서"], "O")
        self.assertEqual(reference_rows["TTA-26-00010"]["시험성적서(PDF)"], "O")
        self.assertNotEqual(reference_rows["TTA-26-00010"]["점검날짜"], "")
        self.assertFalse(projects_data["items"][0]["selectable"])
        self.assertEqual(reference_rows["TTA-26-00011"]["점검결과"], "X")
        self.assertEqual(reference_rows["TTA-26-00011"]["시험성적서(PDF)"], "X")
        self.assertEqual(reference_rows["TTA-26-00012"]["점검결과"], "")
        self.assertEqual(reference_rows["TTA-26-00012"]["계약서"], "")

    def test_write_back_rejects_non_review_columns(self):
        with self.assertRaises(ReferenceQueryError):
            write_project_review_result(
                "TTA-26-00010",
                "완료",
                artifact_results={"회사명": "변조된 회사명"},
            )

        row = self._reference_rows(["TTA-26-00010"], ["점검결과", "회사명"])["TTA-26-00010"]
        self.assertEqual(row["점검결과"], "")
        self.assertEqual(row["회사명"], "에이치소프트")

    def test_write_back_rejects_unknown_project_number(self):
        # PostgreSQL 전환 후에는 "DB 파일이 없음" 시나리오가 성립하지 않는다(테이블은 항상
        # 존재). 대신 reference_project 에 없는 프로젝트번호에 대한 오류 처리를 검증한다.
        with self.assertRaises(ReferenceDbError):
            write_project_review_result("TTA-26-99999", "완료")

    def test_write_back_accepts_failed_review_result(self):
        result = write_project_review_result("TTA-26-00010", "실패")
        response = projects(self.factory.get("/api/projects/", {"project_number": "TTA-26-00010"}))

        row = self._reference_rows(["TTA-26-00010"], ["점검결과"])["TTA-26-00010"]
        data = json.loads(response.content.decode("utf-8"))

        self.assertEqual(result["updated_columns"], ["점검결과"])
        self.assertEqual(row["점검결과"], "실패")
        self.assertEqual(data["items"][0]["review"], "실패")

    def test_write_back_sets_review_result_and_artifact_column(self):
        # 기존 SQLite 시절 "점검날짜 컬럼이 없는 스키마" 시나리오를 대체한다. reference_project
        # 는 항상 고정된 컬럼(inspection_date 등)을 가지므로 그 시나리오는 더 이상 성립하지
        # 않고, 대신 점검결과+산출물 컬럼+점검일자가 함께 기록되는 정상 경로를 검증한다.
        self._seed_reference_projects(
            "bundang",
            [
                dict(
                    project_number="TTA-26-00099",
                    cert_date="05/12",
                    cert_committee_date=date(2026, 5, 12),
                    company="옵션테스트",
                    product="NoDate",
                    pl="박지훈",
                    review_result="",
                    inspection_date="",
                ),
            ],
        )

        result = write_project_review_result(
            "TTA-26-00099",
            "완료",
            artifact_results={"계약서": "정상"},
            inspected_at="2026.05.12 20:00",
        )

        row = self._reference_rows(["TTA-26-00099"], ["점검결과", "계약서"])["TTA-26-00099"]

        self.assertEqual(result["updated_columns"], ["점검결과", "계약서"])
        self.assertEqual(row["점검결과"], "O")
        self.assertEqual(row["계약서"], "O")

    def _post_job(self, project_numbers, *, center=None):
        payload = {"project_numbers": project_numbers}
        if center:
            payload["center"] = center
        request = self.factory.post(
            "/api/jobs/",
            data=json.dumps(payload),
            content_type="application/json",
            REMOTE_ADDR="127.0.0.1",
        )
        return jobs(request)

    def _seed_reference_projects(self, center_code, rows):
        ReferenceProject.objects.using("reference").bulk_create(
            [ReferenceProject(center_code=center_code, **row) for row in rows]
        )

    def _create_master_reference_db(self, db_path):
        conn = sqlite3.connect(db_path)
        try:
            conn.execute(
                """
                CREATE TABLE sw_data (
                    "시험번호" TEXT,
                    "시작일자" TEXT,
                    "종료일자" TEXT
                )
                """
            )
            conn.execute(
                """
                INSERT INTO sw_data ("시험번호", "시작일자", "종료일자")
                VALUES (?, ?, ?)
                """,
                ("TTA-26-00010", "2026-05-01", "2026-05-31"),
            )
            conn.commit()
        finally:
            conn.close()

    def _reference_project_field(self, project, column):
        mapping = {"점검결과": "review_result", "점검날짜": "inspection_date", "회사명": "company"}
        if column in mapping:
            return getattr(project, mapping[column])
        return (project.artifact_results_json or {}).get(column, "")

    def _reference_rows(self, project_numbers, columns, center_code="bundang"):
        projects_by_number = {
            project.project_number: project
            for project in ReferenceProject.objects.using("reference").filter(
                center_code=center_code, project_number__in=project_numbers
            )
        }
        return {
            number: {
                column: self._reference_project_field(project, column)
                for column in columns
            }
            for number, project in projects_by_number.items()
        }


class LocalReviewRulebaseApiTests(TestCase):
    databases = {"default", "workflow", "reference"}

    def test_rulebase_manifest_and_bundle_return_enabled_rules(self):
        DownloadReviewRule.objects.create(
            code="artifact_01",
            name="Contract",
            rule_type="required_artifact_file",
            config_json={"artifact_column": "Contract"},
            enabled=True,
            sort_order=1,
        )
        DownloadReviewRule.objects.create(
            code="artifact_disabled",
            name="Disabled",
            rule_type="required_artifact_file",
            enabled=False,
            sort_order=2,
        )

        manifest_response = local_review_rules_manifest(
            RequestFactory().get("/api/local-review/rules/manifest/")
        )
        manifest = json.loads(manifest_response.content.decode("utf-8"))

        self.assertEqual(manifest_response.status_code, 200)
        self.assertTrue(manifest["success"])
        self.assertEqual(manifest["rule_count"], 1)
        self.assertTrue(manifest["checksum"].startswith("sha256:"))

        bundle_response = local_review_rules_bundle(
            RequestFactory().get("/api/local-review/rules/bundle/")
        )
        bundle = json.loads(bundle_response.content.decode("utf-8"))

        self.assertEqual(bundle_response.status_code, 200)
        self.assertTrue(bundle["success"])
        self.assertEqual(bundle["rule_count"], 1)
        self.assertEqual(bundle["rules"][0]["code"], "artifact_01")
        self.assertEqual(bundle["rules"][0]["config_json"]["artifact_column"], "Contract")


class RuleConfigValidationTests(SimpleTestCase):
    """config_json 검증기(main.rule_config_validation) 단위 테스트."""

    def test_all_seeded_specs_pass_validation(self):
        # 현재 시드되는 18개 실제 규칙은 모두 검증을 통과해야 한다.
        from main.management.commands.seed_download_review_rules import _rule_specs
        from main.rule_config_validation import validate_rule_spec

        specs = _rule_specs(only_real=False)
        self.assertEqual(len(specs), 18)
        for spec in specs:
            errors, warnings = validate_rule_spec(spec)
            self.assertEqual(errors, [], f"{spec.get('code')} 검증 실패: {errors}")
            self.assertEqual(warnings, [], f"{spec.get('code')} 경고: {warnings}")

    def test_unknown_rule_type_is_error(self):
        from main.rule_config_validation import validate_rule_config

        errors, _ = validate_rule_config("made_up_check", {"filename_keywords": []})
        self.assertTrue(any("rule_type" in e for e in errors))

    def test_config_must_be_object(self):
        from main.rule_config_validation import validate_rule_config

        errors, _ = validate_rule_config("required_artifact_file", ["oops"])
        self.assertTrue(any("config_json" in e for e in errors))

    def test_missing_required_key_is_error(self):
        from main.rule_config_validation import validate_rule_config

        errors, _ = validate_rule_config(
            "required_artifact_file", {"folder_keyword_chain": ["계약"]}
        )
        self.assertTrue(any("filename_keywords" in e for e in errors))

    def test_string_where_list_expected_is_error(self):
        from main.rule_config_validation import validate_rule_config

        errors, _ = validate_rule_config(
            "required_artifact_file",
            {"filename_keywords": "계약서", "extensions": ".pdf"},
        )
        self.assertTrue(any("filename_keywords" in e for e in errors))
        self.assertTrue(any("extensions" in e for e in errors))

    def test_non_integer_count_is_error(self):
        from main.rule_config_validation import validate_rule_config

        errors, _ = validate_rule_config(
            "required_artifact_file", {"filename_keywords": [], "exact_count": "1"}
        )
        self.assertTrue(any("exact_count" in e for e in errors))

    def test_extension_without_dot_is_error(self):
        from main.rule_config_validation import validate_rule_config

        errors, _ = validate_rule_config(
            "required_artifact_file", {"filename_keywords": [], "extensions": ["pdf"]}
        )
        self.assertTrue(any("extensions" in e for e in errors))

    def test_invalid_version_pattern_regex_is_error(self):
        from main.rule_config_validation import validate_rule_config

        errors, _ = validate_rule_config(
            "defect_report_check",
            {"filename_keywords": [], "version_pattern": "v(\\d+"},
        )
        self.assertTrue(any("version_pattern" in e for e in errors))

    def test_unknown_content_check_type_is_error(self):
        from main.rule_config_validation import validate_rule_config

        errors, _ = validate_rule_config(
            "document_artifact_check",
            {
                "required_files": [{"extensions": [".pdf"]}],
                "content_checks": [{"type": "docx_magic", "text": "x"}],
            },
        )
        self.assertTrue(any("content_check" in e for e in errors))

    def test_content_check_missing_required_key_is_error(self):
        from main.rule_config_validation import validate_rule_config

        errors, _ = validate_rule_config(
            "document_artifact_check",
            {
                "required_files": [{"extensions": [".pdf"]}],
                "content_checks": [
                    {"type": "docx_table_next_cell_equals", "label": "시험신청번호"}
                ],
            },
        )
        self.assertTrue(any("expected" in e for e in errors))

    def test_missing_artifact_column_is_warning_not_error(self):
        from main.rule_config_validation import validate_rule_config

        errors, warnings = validate_rule_config(
            "required_artifact_file", {"filename_keywords": ["계약서"]}
        )
        self.assertEqual(errors, [])
        self.assertTrue(any("artifact_column" in w for w in warnings))


class RuleGraphValidationTests(SimpleTestCase):
    """requires/produces 의존 그래프 검증 단위 테스트."""

    def test_seeded_specs_graph_is_valid(self):
        # 현재 시드 규칙의 requires/produces 그래프는 sort_order 와 정합해야 한다.
        from main.management.commands.seed_download_review_rules import _rule_specs
        from main.rule_config_validation import validate_rule_graph_from_specs

        errors, _warnings = validate_rule_graph_from_specs(_rule_specs(only_real=False))
        self.assertEqual(errors, [], f"그래프 검증 실패: {errors}")

    def test_missing_producer_is_error(self):
        from main.rule_config_validation import _graph_entry, validate_rule_graph

        entries = [_graph_entry("consumer", "", 10, {"requires": ["잔여결함수"]})]
        errors, _ = validate_rule_graph(entries)
        self.assertTrue(any("잔여결함수" in e for e in errors))

    def test_producer_after_consumer_is_error(self):
        from main.rule_config_validation import _graph_entry, validate_rule_graph

        # producer 가 consumer 보다 늦게(더 큰 sort_order) 실행되면 오류.
        entries = [
            _graph_entry("consumer", "", 10, {"requires": ["X"]}),
            _graph_entry("producer", "", 20, {"produces": ["X"]}),
        ]
        errors, _ = validate_rule_graph(entries)
        self.assertTrue(any("sort_order" in e for e in errors))

    def test_producer_before_consumer_is_valid(self):
        from main.rule_config_validation import _graph_entry, validate_rule_graph

        entries = [
            _graph_entry("producer", "", 10, {"produces": ["X"]}),
            _graph_entry("consumer", "", 20, {"requires": ["X"]}),
        ]
        errors, _ = validate_rule_graph(entries)
        self.assertEqual(errors, [])

    def test_disabled_producer_simulated_as_missing_is_error(self):
        from main.rule_config_validation import _graph_entry, validate_rule_graph

        # 비활성 producer 는 호출자가 entries 에서 제외 → consumer 만 남으면 오류로 잡힌다.
        entries = [_graph_entry("consumer", "", 20, {"requires": ["측정항목별점수표"]})]
        errors, _ = validate_rule_graph(entries)
        self.assertTrue(any("측정항목별점수표" in e for e in errors))

    def test_duplicate_producer_is_warning(self):
        from main.rule_config_validation import _graph_entry, validate_rule_graph

        entries = [
            _graph_entry("p1", "", 10, {"produces": ["X"]}),
            _graph_entry("p2", "", 11, {"produces": ["X"]}),
            _graph_entry("c", "", 20, {"requires": ["X"]}),
        ]
        errors, warnings = validate_rule_graph(entries)
        self.assertEqual(errors, [])
        self.assertTrue(any("여러 규칙이 생성" in w for w in warnings))


class WorkerDownloadDirCleanupTests(TestCase):
    """워커가 다운로드 시작 전 프로젝트 폴더를 비우는 동작(팝업 잔여물 방지)의 회귀 테스트.

    실제 ECM/Agent 없이 합성 다운로드 폴더로 '이전 산출물이 남은' 상태를 재생한다.
    """

    databases = {"default", "workflow"}

    def _make_project(self, base, project_number="TTA-26-00010"):
        job = DownloadReviewJob.objects.create(
            status=DownloadReviewJobStatus.RUNNING,
            requested_project_count=1,
        )
        project = DownloadReviewProject.objects.create(
            job=job,
            center_code="sangam",
            project_number=project_number,
            ecm_row_json={"project_number": project_number},
            status=DownloadReviewProjectStatus.RUNNING,
        )
        return job, project

    def test_clears_leftover_artifacts_before_download(self):
        from main.views.review.ecm_download_review_worker import _clear_project_download_dir

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                job, project = self._make_project(base)
                proj_dir = Path(base) / project.project_number
                (proj_dir / "sub").mkdir(parents=True)
                (proj_dir / "old.txt").write_text("leftover", encoding="utf-8")
                (proj_dir / "sub" / "nested.pdf").write_bytes(b"%PDF-1.4 leftover")

                _clear_project_download_dir(job, project)

                # 폴더가 통째로 비워졌고(삭제), 사전 정리 로그가 남아야 한다.
                self.assertFalse(proj_dir.exists())
                self.assertTrue(
                    DownloadReviewLog.objects.filter(
                        job_project=project, event_code="download_preclean"
                    ).exists()
                )

    def test_noop_when_folder_absent(self):
        from main.views.review.ecm_download_review_worker import _clear_project_download_dir

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                job, project = self._make_project(base)
                # 폴더가 없으면 아무 일도 일어나지 않고 로그도 남지 않는다.
                _clear_project_download_dir(job, project)
                self.assertFalse((Path(base) / project.project_number).exists())
                self.assertFalse(
                    DownloadReviewLog.objects.filter(
                        job_project=project, event_code="download_preclean"
                    ).exists()
                )

    def test_only_target_project_folder_is_removed(self):
        from main.views.review.ecm_download_review_worker import _clear_project_download_dir

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                job, project = self._make_project(base, "TTA-26-00010")
                target = Path(base) / "TTA-26-00010"
                other = Path(base) / "TTA-26-99999"
                target.mkdir()
                (target / "a.txt").write_text("x", encoding="utf-8")
                other.mkdir()
                (other / "b.txt").write_text("y", encoding="utf-8")

                _clear_project_download_dir(job, project)

                # 대상 프로젝트 폴더만 지우고, 다른 프로젝트 폴더는 보존(동시 작업 안전).
                self.assertFalse(target.exists())
                self.assertTrue(other.exists())


class ArtifactSourceSeamTests(SimpleTestCase):
    """산출물 source 추상화: ECM 없이도 다른 source 를 끼워 동작함을 검증.

    LocalFolderArtifactSource 는 ECM 을 떼고 다른 저장소를 붙일 때의 첫 구현이자,
    ECM 없이 워커 흐름을 돌리는 fake-live 테스트 더블이다.
    """

    def _run(self, coro):
        import asyncio

        return asyncio.run(coro)

    def test_local_source_copies_project_folder_into_download_dir(self):
        from types import SimpleNamespace
        from main.views.review.artifact_source import LocalFolderArtifactSource

        async def _noop(*_args):
            return None

        async def _not_canceled():
            return False

        with tempfile.TemporaryDirectory() as root, tempfile.TemporaryDirectory() as base:
            src = Path(root) / "TTA-26-00010"
            (src / "계약").mkdir(parents=True)
            (src / "계약" / "계약서.pdf").write_bytes(b"%PDF-1.4")
            (src / "note.txt").write_text("x", encoding="utf-8")

            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                source = LocalFolderArtifactSource(source_root=root)
                result = self._run(
                    source.fetch(
                        SimpleNamespace(project_number="TTA-26-00010"),
                        on_progress=_noop,
                        is_canceled=_not_canceled,
                    )
                )

            self.assertTrue(result.success)
            dst = Path(base) / "TTA-26-00010"
            self.assertTrue((dst / "계약" / "계약서.pdf").exists())
            self.assertTrue((dst / "note.txt").exists())
            self.assertEqual(result.download_dir, str(dst))

    def test_local_source_reports_missing_source_folder(self):
        from types import SimpleNamespace
        from main.views.review.artifact_source import LocalFolderArtifactSource

        async def _noop(*_args):
            return None

        async def _not_canceled():
            return False

        with tempfile.TemporaryDirectory() as root, tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                source = LocalFolderArtifactSource(source_root=root)
                result = self._run(
                    source.fetch(
                        SimpleNamespace(project_number="TTA-26-99999"),
                        on_progress=_noop,
                        is_canceled=_not_canceled,
                    )
                )

            self.assertFalse(result.success)
            self.assertEqual(result.error_step, "로컬 폴더 확인")

    def test_fetch_raises_when_canceled(self):
        from types import SimpleNamespace
        from main.views.review.artifact_source import (
            JobCanceledError,
            LocalFolderArtifactSource,
        )

        async def _noop(*_args):
            return None

        async def _canceled():
            return True

        source = LocalFolderArtifactSource(source_root="/nonexistent")
        with self.assertRaises(JobCanceledError):
            self._run(
                source.fetch(
                    SimpleNamespace(project_number="TTA-26-00010"),
                    on_progress=_noop,
                    is_canceled=_canceled,
                )
            )

    def test_factory_builds_known_sources_and_rejects_unknown(self):
        from main.views.review.artifact_source import (
            HttpEcmArtifactSource,
            LocalFolderArtifactSource,
            build_artifact_source,
        )

        # 레거시 Playwright source('ecm')는 제거됨 — 호환을 위해 'ecm-http' 로 별칭 처리된다.
        self.assertIsInstance(build_artifact_source("ecm"), HttpEcmArtifactSource)
        self.assertIsInstance(build_artifact_source("ecm-http"), HttpEcmArtifactSource)
        self.assertIsInstance(build_artifact_source("local"), LocalFolderArtifactSource)
        with self.assertRaises(ValueError):
            build_artifact_source("dropbox")


class _FakeEcmClient:
    """HttpEcmArtifactSource 계약 테스트용 mock ECM 클라이언트(네트워크 없음).

    트리는 {oid: {"contents": {"folders":[{name,oid}],"files":[meta...]}}} 로 표현.
    """

    def __init__(self, *, project_oid, tree, blobs, project_name="GS-A-23-336(완료)"):
        self._project_oid = project_oid
        self._tree = tree
        self._blobs = blobs
        self._project_name = project_name
        self.login_calls = 0
        self.download_calls = []

    def login(self):
        self.login_calls += 1

    def find_project_folder(self, test_no, cert_date="", grade=""):
        if self._project_oid is None:
            return None
        return {"oid": self._project_oid, "name": self._project_name}

    def folder_contents(self, oid):
        return self._tree.get(oid, {"folders": [], "files": []})

    def download_bytes(self, meta):
        self.download_calls.append(meta.get("storageFileID"))
        return self._blobs[meta["storageFileID"]]


class HttpEcmArtifactSourceTests(SimpleTestCase):
    """HTTP 직접연동 source 계약: 레이아웃/진행/무결성/취소/미탐색(네트워크 없음)."""

    def _run(self, coro):
        import asyncio

        return asyncio.run(coro)

    def _project(self, number="GS-A-23-0336", center="sangam"):
        from types import SimpleNamespace

        return SimpleNamespace(project_number=number, center_code=center, ecm_row_json={})

    def _fetch(self, source, project, *, canceled=False):
        progressed = []

        async def _on_progress(rel, count):
            progressed.append((list(rel), count))

        async def _is_canceled():
            return canceled

        result = self._run(
            source.fetch(project, on_progress=_on_progress, is_canceled=_is_canceled)
        )
        return result, progressed

    def test_downloads_reproduce_project_relative_layout_with_nfc(self):
        from main.views.review.artifact_source import HttpEcmArtifactSource

        tree = {
            "P": {
                "folders": [{"name": "계약", "oid": "C"}],
                "files": [{"fileName": "표지.pdf", "storageFileID": "f0", "fileSize": 5}],
            },
            "C": {
                "folders": [],
                "files": [{"fileName": "계약서.pdf", "storageFileID": "f1", "fileSize": 8}],
            },
        }
        blobs = {"f0": b"%PDF-", "f1": b"%PDF-1.4"}
        client = _FakeEcmClient(project_oid="P", tree=tree, blobs=blobs)
        source = HttpEcmArtifactSource(client_factory=lambda center: client)

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                result, progressed = self._fetch(source, self._project())

            self.assertTrue(result.success, result.error_message)
            dst = Path(base) / "GS-A-23-0336"
            self.assertTrue((dst / "표지.pdf").exists())
            self.assertTrue((dst / "계약" / "계약서.pdf").exists())
            self.assertEqual(result.download_dir, str(dst))
            self.assertEqual(result.downloaded_folder_count, 2)
            # on_progress 는 파일이 있는 폴더마다 상대경로+건수로 호출된다.
            self.assertIn(([], 1), progressed)
            self.assertIn((["계약"], 1), progressed)

    def test_integrity_failure_fails_project_after_retry(self):
        from main.views.review.artifact_source import HttpEcmArtifactSource

        tree = {"P": {"folders": [], "files": [
            {"fileName": "계약서.pdf", "storageFileID": "f1", "fileSize": 10},
        ]}}
        # 매직바이트가 %PDF 가 아니고 크기도 다름 → 검증 실패.
        blobs = {"f1": b"garbage"}
        client = _FakeEcmClient(project_oid="P", tree=tree, blobs=blobs)
        source = HttpEcmArtifactSource(client_factory=lambda center: client)

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                result, _ = self._fetch(source, self._project())

            self.assertFalse(result.success)
            self.assertEqual(result.error_step, "무결성 검증")
            # 1회 재다운로드까지 시도한다.
            self.assertEqual(len(client.download_calls), 2)

    def test_missing_project_folder_reports_error(self):
        from main.views.review.artifact_source import HttpEcmArtifactSource

        client = _FakeEcmClient(project_oid=None, tree={}, blobs={})
        source = HttpEcmArtifactSource(client_factory=lambda center: client)

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                result, _ = self._fetch(source, self._project())

            self.assertFalse(result.success)
            self.assertEqual(result.error_step, "프로젝트 폴더 탐색")

    def test_fetch_raises_when_canceled(self):
        from main.views.review.artifact_source import (
            HttpEcmArtifactSource,
            JobCanceledError,
        )

        client = _FakeEcmClient(project_oid="P", tree={}, blobs={})
        source = HttpEcmArtifactSource(client_factory=lambda center: client)

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                with self.assertRaises(JobCanceledError):
                    self._fetch(source, self._project(), canceled=True)

    def test_verify_downloaded_bytes_helper(self):
        from main.views.review.artifact_source import verify_downloaded_bytes

        self.assertEqual(verify_downloaded_bytes(b"%PDF-1.4", "a.pdf", 8), "")
        self.assertEqual(verify_downloaded_bytes(b"PK\x03\x04", "a.xlsx", 4), "")
        # 원본이 0바이트인 파일(마커 등)은 빈 데이터도 정상.
        self.assertEqual(verify_downloaded_bytes(b"", "홍보를 원치않음.txt", 0), "")
        # expected>0 인데 비어 있으면 잘림/실패.
        self.assertIn("빈 응답", verify_downloaded_bytes(b"", "a.pdf", 100))
        self.assertIn("크기", verify_downloaded_bytes(b"%PDF", "a.pdf", 999))
        self.assertIn("매직바이트", verify_downloaded_bytes(b"nope", "a.pdf", 4))

    def test_legacy_office_extension_helper(self):
        from main.views.review.artifact_source import legacy_office_extension

        ole = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 8
        self.assertEqual(legacy_office_extension(ole, "결함리포트.xlsx"), "xls")
        self.assertEqual(legacy_office_extension(ole, "합의서.docx"), "doc")
        self.assertEqual(legacy_office_extension(ole, "슬라이드.pptx"), "ppt")
        # 정상 zip 기반 파일은 대상이 아니다.
        self.assertIsNone(legacy_office_extension(b"PK\x03\x04", "a.xlsx"), None)
        # OLE 매직바이트가 아니면 대상이 아니다.
        self.assertIsNone(legacy_office_extension(b"garbage", "a.xlsx"))
        # 신형↔구형 매핑에 없는 확장자는 대상이 아니다.
        self.assertIsNone(legacy_office_extension(ole, "a.pdf"))

    def test_legacy_ole_file_mislabeled_as_xlsx_is_recovered(self):
        from main.views.review.artifact_source import HttpEcmArtifactSource

        ole_bytes = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 8
        tree = {"P": {"folders": [], "files": [
            {"fileName": "결함리포트 v2.0.xlsx", "storageFileID": "f1", "fileSize": len(ole_bytes)},
        ]}}
        client = _FakeEcmClient(project_oid="P", tree=tree, blobs={"f1": ole_bytes})
        source = HttpEcmArtifactSource(client_factory=lambda center: client)

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_DOWNLOAD_BASE_DIR=base):
                result, _ = self._fetch(source, self._project())

            self.assertTrue(result.success, result.error_message)
            dst = Path(base) / "GS-A-23-0336"
            # 신형 확장자(.xlsx) 대신 실제 포맷(.xls)으로 저장되어야 한다.
            self.assertTrue((dst / "결함리포트 v2.0.xls").exists())
            self.assertFalse((dst / "결함리포트 v2.0.xlsx").exists())
            # 매직바이트만으로 즉시 복구되므로 재다운로드는 필요 없다.
            self.assertEqual(len(client.download_calls), 1)


class EcmHttpClientPureFunctionTests(SimpleTestCase):
    """ecm_http_client 의 순수 함수(네트워크 없음): 매칭/연도/점수/파일수집."""

    def test_xor_encrypt_roundtrip_is_deterministic(self):
        from main.views.review.ecm_http_client import xor_encrypt

        self.assertEqual(xor_encrypt("abc"), xor_encrypt("abc"))
        self.assertTrue(xor_encrypt("secret"))

    def test_test_no_patterns_match_zero_padding_both_ways(self):
        from main.views.review.ecm_http_client import DestinyECM

        patterns = DestinyECM.test_no_patterns("GS-A-23-0336")
        self.assertTrue(any(p.search("GS-A-23-336(완료)") for p in patterns))
        self.assertTrue(any(p.search("GS-A-23-0336 계약") for p in patterns))
        # 뒤 숫자 경계: 336 이 3360 에 매칭되면 안 된다.
        self.assertFalse(any(p.search("GS-A-23-3360") for p in patterns))

    def test_year_candidates_prefers_cert_date_then_test_no(self):
        from main.views.review.ecm_http_client import DestinyECM

        self.assertEqual(
            DestinyECM.year_candidates("GS-A-23-0336", "2024-01-01"),
            ["2024", "2023"],
        )
        self.assertEqual(DestinyECM.year_candidates("GS-A-23-0336"), ["2023"])

    def test_project_match_score_prioritizes_completed(self):
        from main.views.review.ecm_http_client import DestinyECM

        self.assertGreater(
            DestinyECM.project_match_score("GS-A-23-336(완료)"),
            DestinyECM.project_match_score("GS-A-23-336(신청)"),
        )
        self.assertLess(DestinyECM.project_match_score("GS-A-23-336(취소)"), 0)

    def test_find_project_folder_stops_after_contiguous_match_block(self):
        from main.views.review.ecm_http_client import DestinyECM

        # 같은 프로젝트 3개(신청/계약/완료) 뒤에 다른 프로젝트 폴더 1000개가 이어진 목록.
        children = (
            [{"name": f"GS-A-23-336({s})", "OID": s} for s in ("신청", "계약", "완료")]
            + [{"name": f"GS-A-23-{n}", "OID": str(n)} for n in range(400, 1400)]
        )

        class _CountingClient(DestinyECM):
            def __init__(self):
                self.root_oid = "R"
                self.scanned = 0

            def year_candidates(self, test_no, cert_date=""):
                return ["2023"]

            def find_year_folder(self, year):
                return "SVC"

            def gs_candidate_roots(self, service_oid, grade=""):
                return [("GROOT", "GS")]

            def children(self, oid):
                for c in children:
                    self.scanned += 1
                    yield c

        client = _CountingClient()
        best = client.find_project_folder("GS-A-23-336")
        # 완료 폴더가 점수로 선택된다.
        self.assertIn("완료", best["name"])
        # 블록(3) + 종료를 유발한 다른 프로젝트 1개 = 4개만 훑고 멈춘다(1003개 아님).
        self.assertEqual(client.scanned, 4)

    def test_find_year_folder_handles_direct_and_nested_center_layout(self):
        from main.views.review.ecm_http_client import DestinyECM

        class _Client(DestinyECM):
            def __init__(self, tree):
                self.root_oid = "ROOT"
                self._tree = tree

            def children(self, oid):
                return self._tree.get(oid, [])

        # 분당식: root 직속에 년도 폴더.
        direct = _Client({"ROOT": [{"name": "2023 시험서비스", "OID": "Y"}]})
        self.assertEqual(direct.find_year_folder("2023"), "Y")

        # 상암/영남식: root 아래 {센터명} 폴더가 한 단계 더 있는 경우.
        nested = _Client({
            "ROOT": [{"name": "상암AX센터", "OID": "C"}],
            "C": [{"name": "2023 시험서비스", "OID": "Y2"}],
        })
        self.assertEqual(nested.find_year_folder("2023"), "Y2")

    def test_collect_files_recurses_and_needs_filename_and_storageid(self):
        from main.views.review.ecm_http_client import DestinyECM

        payload = {
            "params": {
                "rows": [
                    {"fileName": "a.pdf", "storageFileID": "s1", "OID": "o1", "fileSize": "10"},
                    {"fileName": "no_storage.pdf"},  # storageFileID 없음 → 제외
                    {"nested": {"fileName": "b.xlsx", "storageFileID": "s2", "fileSize": 20}},
                ]
            }
        }
        out = []
        DestinyECM.collect_files(payload, out)
        ids = {f["storageFileID"] for f in out}
        self.assertEqual(ids, {"s1", "s2"})


class HistoryDocumentHttpTests(SimpleTestCase):
    """시험 이력 '문서 다운로드' HTTP 전환: 인증위원회 트리 탐색 + 시험성적서 Word 필터."""

    def test_parse_cert_date_parts(self):
        from main.views.review.ecm_http_client import DestinyECM

        self.assertEqual(DestinyECM.parse_cert_date_parts("2022-08-15"), ("2022", "20220815"))
        self.assertEqual(DestinyECM.parse_cert_date_parts("2022.8.5"), ("2022", "20220805"))
        self.assertEqual(DestinyECM.parse_cert_date_parts("bad"), ("", ""))

    def test_select_report_documents_keeps_only_report_word_by_default(self):
        from main.views.review.ecm_http_client import DestinyECM

        files = [
            {"fileName": "GS-B-22-355 시험성적서 v1.0.docx", "storageFileID": "a"},
            {"fileName": "GS-B-22-355 시험성적서 v1.0.pdf", "storageFileID": "b"},  # PDF 제외
            {"fileName": "GS-B-22-355 시험계획서 v1.0.docx", "storageFileID": "c"},  # 성적서 아님
            {"fileName": "GS-B-22-355 시험성적서.doc", "storageFileID": "d"},
        ]
        report = DestinyECM.select_report_documents(files, report_only=True)
        self.assertEqual({f["storageFileID"] for f in report}, {"a", "d"})
        # report_only=False 면 전체 유지.
        self.assertEqual(len(DestinyECM.select_report_documents(files, report_only=False)), 4)

    def test_find_full_project_folder_bundang_requires_gs_and_grade1(self):
        from main.views.review.ecm_http_client import DestinyECM

        tree = {
            "ROOT": [{"name": "2023 시험서비스", "OID": "SVC"}],
            "SVC": [
                {"name": "02 GS시험인증(2등급)", "OID": "GS2"},   # 1등급 아님 → 제외
                {"name": "03 GS시험인증(1등급)", "OID": "GS1"},
            ],
            "GS2": [{"name": "GS-B-23-067(완료)", "OID": "WRONG"}],
            "GS1": [{"name": "바. GS-B-23-067(완료)", "OID": "RIGHT"}],
        }

        class _C(DestinyECM):
            def __init__(self):
                self.root_oid = "ROOT"

            def children(self, oid):
                return tree.get(oid, [])

        found = _C().find_full_project_folder("GS-B-23-067", "2023-07-24", "bundang")
        self.assertEqual(found["oid"], "RIGHT")

    def test_find_full_project_folder_sangam_enters_center_folder_first(self):
        from main.views.review.ecm_http_client import DestinyECM

        tree = {
            "ROOT": [{"name": "상암AX센터", "OID": "SANGAM"}, {"name": "영남AX센터", "OID": "YN"}],
            "SANGAM": [{"name": "2026 시험서비스", "OID": "SVC"}],
            "YN": [{"name": "2026 시험서비스", "OID": "SVC_YN"}],
            "SVC": [{"name": "01 GS인증시험(1등급)", "OID": "GS1"}],
            "SVC_YN": [{"name": "01 GS인증시험(1등급)", "OID": "GS_YN"}],
            "GS1": [{"name": "00266 TTA-26-00266(완료)", "OID": "RIGHT"}],
            "GS_YN": [{"name": "00266 TTA-26-00266(완료)", "OID": "WRONG"}],
        }

        class _C(DestinyECM):
            def __init__(self):
                self.root_oid = "ROOT"

            def children(self, oid):
                return tree.get(oid, [])

        found = _C().find_full_project_folder("TTA-26-00266", "2026-01-01", "sangam")
        self.assertEqual(found["oid"], "RIGHT")  # 영남 폴더로 새지 않는다

    def test_full_project_download_falls_back_across_centers(self):
        from unittest.mock import patch
        from main.views.testing import history_download

        class _Client:
            def __init__(self, center):
                self.center = center

            def login(self):
                pass

            def find_full_project_folder(self, test_no, cert_date, center_code):
                # 분당/상암엔 없고 영남에만 있는 프로젝트.
                if self.center == "yeongnam":
                    return {"oid": "P", "name": "GS-C-24-0003(완료)"}
                return None

            def walk_files(self, oid):
                yield [], {"fileName": "성적서.docx", "storageFileID": "1", "fileSize": 4}

            def download_bytes(self, meta):
                return b"PK\x03\x04"

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_REPORT_BASE_DIR=base):
                # reference_project 미해석(None) → 분당→상암→영남 순으로 폴백.
                with patch.object(history_download, "_resolve_project_center", return_value=(None, None)), \
                     patch("main.views.review.ecm_http_client.build_client", side_effect=lambda c: _Client(c)):
                    result = history_download.download_full_project_documents("GS-C-24-0003", "2024-01-01")
                self.assertEqual(result["center"], "yeongnam")
                self.assertEqual(result["doc_count"], 1)

    def test_download_report_streams_zip_then_deletes_files(self):
        from main.views.testing import history_download, history_report

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_REPORT_BASE_DIR=base):
                folder = history_download.all_dir("GS-B-23-067")
                folder.mkdir(parents=True)
                (folder / "a.docx").write_bytes(b"PK\x03\x04a")
                zp = history_download.zip_path("GS-B-23-067")
                zp.parent.mkdir(parents=True, exist_ok=True)
                zp.write_bytes(b"PKZIPDATA")

                req = RequestFactory().get("/history/report/GS-B-23-067/download/")
                resp = history_report.download_report(req, "GS-B-23-067")
                # 스트리밍 본문을 모두 소비해야 finally 정리가 실행된다.
                body = b"".join(resp.streaming_content)
                self.assertEqual(body, b"PKZIPDATA")
                # 전송 후 원본 폴더와 zip 이 삭제된다.
                self.assertFalse(folder.exists())
                self.assertFalse(zp.exists())

    def test_download_report_document_deletes_folder_after_serving(self):
        from main.views.testing import history_download, history_report

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_REPORT_BASE_DIR=base):
                folder = history_download.doc_dir("GS-B-23-067")
                folder.mkdir(parents=True)
                (folder / "성적서.docx").write_bytes(b"PK\x03\x04doc")

                req = RequestFactory().get("/history/report/GS-B-23-067/document/")
                resp = history_report.download_report_document(req, "GS-B-23-067")
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(resp.content, b"PK\x03\x04doc")
                # 서빙 직후 성적서 폴더가 삭제된다.
                self.assertFalse(folder.exists())

    def test_find_committee_test_folder_navigates_year_committee_date_test(self):
        from main.views.review.ecm_http_client import DestinyECM

        tree = {
            "ROOT": [{"name": "2022 시험서비스", "OID": "SVC"}],
            "SVC": [{"name": "00 2022년 GS인증심의위원회", "OID": "COM"}, {"name": "03 GS시험인증(1등급)", "OID": "GS"}],
            # 인증일자 폴더가 위원회 직속이 아니라 회차 폴더 밑에 있는 경우.
            "COM": [{"name": "01 1차 품질인증심의위원회", "OID": "ROUND"}],
            "ROUND": [{"name": "20220815 심의", "OID": "DATE"}],
            "DATE": [
                {"name": "GS-B-22-354 something", "OID": "X"},
                {"name": "GS-B-22-355 대상", "OID": "TARGET"},
            ],
        }

        class _TreeClient(DestinyECM):
            def __init__(self):
                self.root_oid = "ROOT"

            def children(self, oid):
                return tree.get(oid, [])

        client = _TreeClient()
        found = client.find_committee_test_folder("GS-B-22-355", "2022-08-15")
        self.assertIsNotNone(found)
        self.assertEqual(found["oid"], "TARGET")

    def test_doc_and_all_dirs_are_separate(self):
        from main.views.testing import history_download

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_REPORT_BASE_DIR=base):
                # 성적서(#1)는 __report/<번호>, 전체(#2)는 <번호> 로 완전히 분리된 폴더.
                doc = history_download.doc_dir("GS-B-22-355")
                full = history_download.all_dir("GS-B-22-355")
                self.assertNotEqual(doc, full)
                self.assertEqual(full.name, "GS-B-22-355")
                self.assertEqual(doc.parent.name, "__report")
                # 성적서 폴더에 파일을 둬도 전체 폴더 경로와 겹치지 않는다.
                doc.mkdir(parents=True)
                (doc / "시험성적서.docx").write_bytes(b"PK\x03\x04")
                self.assertFalse(str(doc).startswith(str(full) + "/") or str(doc).startswith(str(full) + "\\"))

    def test_walk_files_yields_relative_paths(self):
        from main.views.review.ecm_http_client import DestinyECM

        tree = {
            "P": {"folders": [{"name": "4.시험", "oid": "S"}],
                  "files": [{"fileName": "a.doc", "storageFileID": "1"}]},
            "S": {"folders": [], "files": [{"fileName": "b.xlsx", "storageFileID": "2"}]},
        }

        class _C(DestinyECM):
            def __init__(self):
                pass

            def folder_contents(self, oid):
                return tree.get(oid, {"folders": [], "files": []})

        items = list(_C().walk_files("P"))
        rels = {tuple(r): m["storageFileID"] for r, m in items}
        self.assertEqual(rels, {(): "1", ("4.시험",): "2"})

    def test_full_project_download_writes_tree_and_all_marker(self):
        from unittest.mock import patch
        from main.views.testing import history_download

        class _FakeClient:
            def login(self):
                pass

            def find_full_project_folder(self, test_no, cert_date, center_code):
                return {"oid": "P", "name": "GS-B-23-067(완료)"}

            def walk_files(self, oid):
                yield [], {"fileName": "성적서.docx", "storageFileID": "1", "fileSize": 4}
                yield ["4.시험"], {"fileName": "계획.docx", "storageFileID": "2", "fileSize": 4}

            def download_bytes(self, meta):
                return b"PK\x03\x04"

        with tempfile.TemporaryDirectory() as base:
            with override_settings(AGENT_REPORT_BASE_DIR=base):
                with patch.object(
                    history_download, "_resolve_project_center",
                    return_value=("bundang", {"cert_date": "2023.07.24"}),
                ), patch(
                    "main.views.review.ecm_http_client.build_client",
                    return_value=_FakeClient(),
                ):
                    result = history_download.download_full_project_documents(
                        "GS-B-23-067", "2023-07-24"
                    )

            self.assertEqual(result["doc_count"], 2)
            self.assertEqual(result["center"], "bundang")
            folder = Path(base) / "GS-B-23-067"
            self.assertTrue((folder / "성적서.docx").exists())
            self.assertTrue((folder / "4.시험" / "계획.docx").exists())
            # ZIP 이 다운로드 단계에서 미리 생성돼 있어야 GET 이 즉시 스트리밍한다.
            import zipfile as _zip
            zp = Path(result["zip_path"])
            self.assertTrue(zp.is_file())
            with _zip.ZipFile(zp) as zf:
                names = set(zf.namelist())
            self.assertEqual(names, {"성적서.docx", "4.시험/계획.docx"})


class WeeklyHttpDownloadTests(SimpleTestCase):
    """weekly.py HTTP 전환: 00 폴더에서 가장 최근 날짜 목록 파일 선택(네트워크 없음)."""

    class _FakeWeeklyClient:
        def __init__(self, files_by_oid):
            self._files_by_oid = files_by_oid

        def find_year_folder(self, year):
            return "S" + str(year)

        def children(self, oid):
            year = oid[1:]
            return [{"name": f"00 {year}년 시험서비스", "OID": "Z" + year}, {"name": "GS", "OID": "g"}]

        @staticmethod
        def oid(row):
            return row.get("OID") or row.get("oid") or ""

        def files(self, oid):
            return self._files_by_oid.get(oid, [])

    def test_extract_includes_y_z_columns(self):
        from openpyxl import Workbook
        from main.utils.weekly import extract_a_to_n_rows_after_serial, REFERENCE_SHEET_NAME

        wb = Workbook()
        ws = wb.active
        ws.title = REFERENCE_SHEET_NAME
        ws.cell(row=1, column=1, value="일련번호")   # 헤더
        ws.cell(row=2, column=1, value=100)          # 앵커(일련번호)
        ws.cell(row=3, column=1, value=101)          # A
        ws.cell(row=3, column=2, value="24-0001")    # B
        ws.cell(row=3, column=25, value="재인증")     # Y
        ws.cell(row=3, column=26, value="기인증 v1")  # Z

        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "t.xlsx"
            wb.save(path)
            rows = extract_a_to_n_rows_after_serial(path, 100, REFERENCE_SHEET_NAME)

        self.assertEqual(len(rows), 1)
        self.assertEqual(len(rows[0]), 16)          # A..N(14) + Y + Z
        self.assertEqual(rows[0][14], "재인증")       # Y(25열)
        self.assertEqual(rows[0][15], "기인증 v1")    # Z(26열)

    def test_selects_most_recent_dated_list_file(self):
        from main.utils import weekly

        client = self._FakeWeeklyClient({
            "Z2026": [
                {"fileName": "인증획득제품(20260629).xlsx", "storageFileID": "a", "fileSize": 9},
                {"fileName": "인증획득제품(20260706).xlsx", "storageFileID": "b", "fileSize": 9},
                {"fileName": "관련없음.xlsx", "storageFileID": "c", "fileSize": 1},
            ],
        })
        meta = weekly.select_latest_list_file(client, [2026, 2025])
        self.assertEqual(meta["storageFileID"], "b")

    def test_falls_back_to_previous_year_when_current_empty(self):
        from main.utils import weekly

        client = self._FakeWeeklyClient({
            "Z2025": [{"fileName": "인증획득제품(20251230).xlsx", "storageFileID": "p", "fileSize": 5}],
        })
        meta = weekly.select_latest_list_file(client, [2026, 2025])
        self.assertEqual(meta["storageFileID"], "p")

    def test_returns_none_when_no_dated_list_file(self):
        from main.utils import weekly

        client = self._FakeWeeklyClient({"Z2026": [{"fileName": "메모.txt", "storageFileID": "x"}]})
        self.assertIsNone(weekly.select_latest_list_file(client, [2026, 2025]))


class WorkerSourceSelectionTests(SimpleTestCase):
    """source 선택이 CLI(--source)에서 워커까지 전달되는지 검증."""

    def test_source_option_is_passed_to_worker(self):
        from main.views.review.ecm_download_review_worker import WorkerRunResult

        with patch(
            "main.management.commands.run_download_worker.run_worker_once"
        ) as mock_run:
            mock_run.return_value = WorkerRunResult(
                processed=False, status="idle", message="시작 가능한 작업이 없습니다."
            )
            call_command("run_download_worker", "--once", "--dry-run", "--source=local")

        self.assertEqual(mock_run.call_args.kwargs.get("source_name"), "local")


class ParseKoreanDateRangeTests(SimpleTestCase):
    """'시작날짜/종료날짜'(M열)의 다양한 표기에서 시작/종료일을 뽑아내는지 검증."""

    def _parse(self, value):
        from main.utils.xlsx_to_sqlite import parse_korean_date_range

        return parse_korean_date_range(value)

    def test_dot_separated_formats(self):
        self.assertEqual(self._parse("2003.03.17~2003.03.28"), ("2003-03-17", "2003-03-28"))
        self.assertEqual(self._parse("2007.06.04 ~ 2007.06.25"), ("2007-06-04", "2007-06-25"))
        # 0-패딩 유무가 섞여도 인식한다.
        self.assertEqual(self._parse("2013.3.3 ~ 2013.3.26"), ("2013-03-03", "2013-03-26"))
        # 구분자 앞뒤 공백('2015.12. 21', '2016. 1 .12')도 흡수한다.
        self.assertEqual(self._parse("2015.12. 21 ~ 2016. 1 .12"), ("2015-12-21", "2016-01-12"))
        # 끝에 마침표가 붙는 표기.
        self.assertEqual(self._parse("2019.1.21. ~ 2019.2.27."), ("2019-01-21", "2019-02-27"))

    def test_dash_and_slash_separators(self):
        self.assertEqual(self._parse("2020-04-29 ~ 2020-05-22"), ("2020-04-29", "2020-05-22"))
        self.assertEqual(self._parse("2014/10/13 ~ 2014/11/5"), ("2014-10-13", "2014-11-05"))
        # 한 값 안에서 구분자가 섞이는 경우.
        self.assertEqual(self._parse("2008-06.30~2008-07.18"), ("2008-06-30", "2008-07-18"))

    def test_korean_year_month_day(self):
        self.assertEqual(
            self._parse("2015년 8월 3일 ~ 2015년 9월 7일"), ("2015-08-03", "2015-09-07")
        )
        # 월/일 앞 공백이 두 칸인 표기.
        self.assertEqual(
            self._parse("2016년  2월 25일 ~ 2016년  3월 15일"), ("2016-02-25", "2016-03-15")
        )

    def test_two_digit_year(self):
        self.assertEqual(self._parse("14. 6. 25 ~ 14. 7. 22"), ("2014-06-25", "2014-07-22"))
        self.assertEqual(self._parse("25.02.06 ~ 25.02.21"), ("2025-02-06", "2025-02-21"))
        self.assertEqual(self._parse("25/11/05~25/11/21"), ("2025-11-05", "2025-11-21"))

    def test_excel_newline_escape_and_fullwidth_digits(self):
        # 엑셀 셀 줄바꿈 escape('_x000D_')가 날짜 사이에 끼어드는 경우.
        self.assertEqual(
            self._parse("2024년 08월 26일_x000D_\n2024년 09월 12일"),
            ("2024-08-26", "2024-09-12"),
        )
        # 전각 숫자.
        self.assertEqual(self._parse("2023.８.1. ~ 2023.9.６."), ("2023-08-01", "2023-09-06"))

    def test_multiple_ranges_returns_overall_span(self):
        # 재시험/재계약 등 여러 구간이 있으면 전체 최소~최대를 반환한다.
        self.assertEqual(
            self._parse(
                "2006.08.21 ~ 2006.10.20\n2007.07.12 ~ 2007.08.23\n(1차 재시험)"
            ),
            ("2006-08-21", "2007-08-23"),
        )
        self.assertEqual(
            self._parse(
                "(최초) 2020.05.06 ~ 2020.05.29\n(1차) 2020.08.20 ~ 2020.08.24"
            ),
            ("2020-05-06", "2020-08-24"),
        )

    def test_unparseable_values_return_none(self):
        for value in ["-", "2014.00.00 ~ 2014.00.00", "20XX.XX.XX ~ 20XX.XX.XX", "", None, "nan"]:
            self.assertEqual(self._parse(value), (None, None), msg=value)


class NormalizeCellTextTests(SimpleTestCase):
    """엑셀 셀의 '_x000D_'(CR) 표기를 줄바꿈으로 정규화하는지 검증."""

    def _norm(self, value):
        from main.utils.xlsx_to_sqlite import normalize_cell_text

        return normalize_cell_text(value)

    def test_x000d_becomes_newline(self):
        self.assertEqual(
            self._norm("㈜웨어비즈_x000D_ WAREBIZ Co., Ltd."),
            "㈜웨어비즈\n WAREBIZ Co., Ltd.",
        )
        self.assertEqual(
            self._norm("Chumdan Control_x000D_System Co., Ltd."),
            "Chumdan Control\nSystem Co., Ltd.",
        )

    def test_crlf_pair_collapses_to_single_newline(self):
        # 원래 CRLF('_x000D_\n')였던 자리가 '\n\n' 으로 겹치지 않아야 한다.
        self.assertEqual(self._norm("2024.06.21.~_x000D_\n2024.07.12."), "2024.06.21.~\n2024.07.12.")

    def test_existing_newline_preserved_and_none_passthrough(self):
        self.assertEqual(self._norm("정상 회사명\n영문명"), "정상 회사명\n영문명")
        self.assertIsNone(self._norm(None))


class SimilarSummarySelectionTests(SimpleTestCase):
    def setUp(self):
        self.factory = RequestFactory()

    @patch("main.views.testing.similar_summary.generate_recommended_summaries")
    def test_manual_prepare_returns_five_recommendations_and_checked_original(
        self,
        generate_recommendations,
    ):
        from main.views.testing.similar_summary import summarize_document

        generate_recommendations.return_value = [
            f"추천 제품 개요 {index}"
            for index in range(1, 6)
        ]
        request = self.factory.post(
            "/summarize_document/",
            {
                "action": "prepare",
                "fileType": "manual",
                "manualInput": "원본 제품 개요 문장",
            },
        )

        response = summarize_document(request)
        payload = json.loads(response.content)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["mode"], "manual")
        self.assertEqual(len(payload["options"]), 6)
        self.assertEqual(payload["options"][-1]["text"], "원본 제품 개요 문장")
        self.assertTrue(payload["options"][-1]["is_original"])
        self.assertEqual(payload["default_selected_ids"], ["original"])
        generate_recommendations.assert_called_once_with(
            "원본 제품 개요 문장",
            count=5,
            max_chars=60,
        )

    @patch("main.views.testing.similar_GPT.generate_gemma_text")
    def test_manual_recommendation_prompt_requests_semantic_synonyms(
        self,
        generate_text,
    ):
        from main.views.testing.similar_GPT import generate_recommended_summaries

        generate_text.return_value = json.dumps(
            {
                "recommendations": [
                    "기업 문서를 분류하고 찾아주는 자료 관리 솔루션",
                    "업무 자료를 체계화하고 조회하는 콘텐츠 관리 시스템",
                ]
            },
            ensure_ascii=False,
        )

        recommendations = generate_recommended_summaries(
            "기업 문서를 관리하고 검색하는 문서 관리 시스템",
            count=2,
            max_chars=60,
        )

        prompt = generate_text.call_args.args[0]
        self.assertEqual(len(recommendations), 2)
        self.assertIn("동의어 또는 문맥상 같은 뜻의 대체어", prompt)
        self.assertIn("단순한 어순 변경", prompt)
        self.assertIn("내용 단어를 2개 이상", prompt)
        self.assertIn("의미를 넓히거나 좁히", prompt)

    @patch("main.views.testing.similar_summary.generate_recommended_summaries")
    @patch("main.views.testing.similar_summary.run_gemini_gemma")
    @patch("main.views.testing.similar_summary.parse_file")
    def test_file_prepare_returns_four_recommendations_plus_original_summary(
        self,
        parse_file,
        run_summary,
        generate_recommendations,
    ):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from main.views.testing.similar_summary import summarize_document

        parse_file.return_value = "업로드 문서에서 추출한 충분한 제품 설명 내용입니다."
        run_summary.return_value = "원본 추출 요약 문장"
        generate_recommendations.return_value = [
            f"추천 제품 개요 {index}"
            for index in range(1, 5)
        ]
        request = self.factory.post(
            "/summarize_document/",
            {
                "action": "prepare",
                "fileType": "functionList",
                "file": SimpleUploadedFile("manual.pdf", b"fake pdf"),
            },
        )

        response = summarize_document(request)
        payload = json.loads(response.content)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["mode"], "file")
        self.assertEqual(len(payload["options"]), 5)
        self.assertEqual(payload["options"][-1]["text"], "원본 추출 요약 문장")
        self.assertEqual(payload["default_selected_ids"], ["recommendation-1"])
        generate_recommendations.assert_called_once_with(
            "업로드 문서에서 추출한 충분한 제품 설명 내용입니다.",
            count=4,
            max_chars=60,
        )

    @patch("main.views.testing.similar_summary.rerank_multiple_similar_candidates")
    @patch("main.views.testing.similar_summary.compare_multiple_from_index")
    def test_search_uses_all_selected_sentences_and_returns_average_ranking(
        self,
        compare_multiple,
        rerank_multiple,
    ):
        from main.views.testing.similar_summary import summarize_document

        faiss_rows = [
            {"일련번호": 1, "제품설명": "후보", "similarity": 0.75},
        ]
        reranked_rows = [
            {
                "일련번호": 1,
                "제품설명": "후보",
                "llm_score": 80.0,
                "similarity": 0.8,
            },
        ]
        compare_multiple.return_value = (faiss_rows, [0.75])
        rerank_multiple.return_value = reranked_rows
        selected = ["첫 번째 문장", "두 번째 문장"]
        request = self.factory.post(
            "/summarize_document/",
            {
                "action": "search",
                "inputMode": "manual",
                "selectedSummaries": json.dumps(selected, ensure_ascii=False),
            },
        )

        response = summarize_document(request)
        payload = json.loads(response.content)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["summary"], selected)
        self.assertEqual(payload["similarities"], [0.8])
        self.assertEqual(
            payload["search_period"],
            {"start": "2017-01-01", "end": date.today().isoformat()},
        )
        compare_multiple.assert_called_once_with(
            selected,
            k=30,
            cert_date_from=date(2017, 1, 1),
            cert_date_to=date.today(),
        )
        rerank_multiple.assert_called_once_with(selected, faiss_rows)

    def test_search_rejects_empty_selection(self):
        from main.views.testing.similar_summary import summarize_document

        request = self.factory.post(
            "/summarize_document/",
            {
                "action": "search",
                "selectedSummaries": "[]",
            },
        )

        response = summarize_document(request)

        self.assertEqual(response.status_code, 400)
        self.assertIn("1개 이상", json.loads(response.content)["response"])

    @patch("main.views.testing.similar_compare.SwData.objects.using")
    def test_similar_rows_include_same_notes_buttons_as_history(self, using):
        from main.views.testing.similar_compare import select_data_from_db

        class FakeQuerySet:
            def filter(self, **kwargs):
                self.filter_kwargs = kwargs
                return self

            def values(self):
                return [
                    {
                        "serial_number": 7,
                        "test_number": "TTA-26-00007",
                        "total_wd": "18",
                        "recert_type": "재인증",
                        "prev_cert_info": "GS-A-25-0001 제품 1.0",
                        "renewal": "재계약으로 시험 기간 조정",
                        "kolas": "KOLAS 인정",
                        "notes": "기능 변경으로 WD 2일 조정",
                    }
                ]

        queryset = FakeQuerySet()
        using.return_value = queryset

        rows = select_data_from_db([7])

        self.assertEqual(queryset.filter_kwargs, {"serial_number__in": [7]})
        self.assertEqual(
            [button["label"] for button in rows[0]["특이사항_버튼"]],
            ["재인증", "재계약", "KOLAS", "특이사항"],
        )
        self.assertEqual(
            rows[0]["특이사항_버튼"][-1]["tooltip"],
            "기능 변경으로 WD 2일 조정",
        )

    @patch("main.views.testing.similar_summary.rerank_multiple_similar_candidates")
    @patch("main.views.testing.similar_summary.compare_multiple_from_index")
    def test_search_accepts_default_and_custom_summary_sentences(
        self,
        compare_multiple,
        rerank_multiple,
    ):
        from main.views.testing.similar_summary import summarize_document

        selected = [f"검색 문장 {index}" for index in range(1, 9)]
        compare_multiple.return_value = ([], [])
        rerank_multiple.return_value = []
        request = self.factory.post(
            "/summarize_document/",
            {
                "action": "search",
                "inputMode": "file",
                "selectedSummaries": json.dumps(selected, ensure_ascii=False),
            },
        )

        response = summarize_document(request)

        self.assertEqual(response.status_code, 200)
        compare_multiple.assert_called_once()
        self.assertEqual(compare_multiple.call_args.args[0], selected)

    @patch("main.views.testing.similar_compare.select_data_from_db")
    @patch("main.views.testing.similar_compare._get_model")
    @patch("main.views.testing.similar_compare._get_index")
    def test_multiple_faiss_scores_are_averaged_per_product(
        self,
        get_index,
        get_model,
        select_data,
    ):
        import numpy as np
        from main.views.testing.similar_compare import compare_multiple_from_index

        class FakeIndex:
            ntotal = 2

            def search(self, query_vectors, count):
                self.query_vectors = query_vectors
                self.count = count
                return (
                    np.array([[0.9, 0.4], [0.8, 0.6]], dtype="float32"),
                    np.array([[1, 2], [2, 1]], dtype="int64"),
                )

        class FakeModel:
            def encode(self, texts, normalize_embeddings):
                self.texts = texts
                self.normalize_embeddings = normalize_embeddings
                return np.array([[1.0, 0.0], [0.0, 1.0]], dtype="float32")

        get_index.return_value = FakeIndex()
        get_model.return_value = FakeModel()
        select_data.return_value = [
            {"일련번호": 1, "제품설명": "첫 제품"},
            {"일련번호": 2, "제품설명": "둘째 제품"},
        ]

        rows, similarities = compare_multiple_from_index(["문장 1", "문장 2"], k=2)

        self.assertEqual([row["일련번호"] for row in rows], [1, 2])
        self.assertAlmostEqual(similarities[0], 0.75)
        self.assertAlmostEqual(similarities[1], 0.6)
        self.assertAlmostEqual(rows[0]["faiss_scores"][0], 0.9)
        self.assertAlmostEqual(rows[0]["faiss_scores"][1], 0.6)

    @patch("main.views.testing.similar_compare.select_data_from_db")
    @patch("main.views.testing.similar_compare._get_model")
    @patch("main.views.testing.similar_compare._get_index")
    def test_multiple_search_filters_candidates_by_certification_date(
        self,
        get_index,
        get_model,
        select_data,
    ):
        import numpy as np
        from main.views.testing.similar_compare import compare_multiple_from_index

        class FakeIndex:
            ntotal = 2

            def search(self, query_vectors, count):
                return (
                    np.array([[0.9, 0.8]], dtype="float32"),
                    np.array([[1, 2]], dtype="int64"),
                )

        class FakeModel:
            def encode(self, texts, normalize_embeddings):
                return np.array([[1.0, 0.0]], dtype="float32")

        get_index.return_value = FakeIndex()
        get_model.return_value = FakeModel()
        select_data.return_value = [
            {"일련번호": 1, "제품설명": "이전 제품", "인증일자": "2016.12.31"},
            {"일련번호": 2, "제품설명": "대상 제품", "인증일자": "2017-01-02"},
        ]

        rows, similarities = compare_multiple_from_index(
            ["문장"],
            k=2,
            cert_date_from=date(2017, 1, 1),
        )

        self.assertEqual([row["일련번호"] for row in rows], [2])
        self.assertAlmostEqual(similarities[0], 0.8)

    @patch("main.views.testing.similar_GPT.generate_gemma_text")
    def test_multiple_llm_scores_are_averaged_per_product(self, generate_text):
        from main.views.testing.similar_GPT import rerank_multiple_similar_candidates

        generate_text.return_value = json.dumps(
            {
                "results": [
                    {"id": "1", "scores": [90, 70]},
                    {"id": "2", "scores": [60, 70]},
                ]
            }
        )
        candidates = [
            {
                "일련번호": 1,
                "제품설명": "첫 제품",
                "faiss_similarity": 0.7,
                "faiss_scores": [0.8, 0.6],
            },
            {
                "일련번호": 2,
                "제품설명": "둘째 제품",
                "faiss_similarity": 0.65,
                "faiss_scores": [0.6, 0.7],
            },
        ]

        rows = rerank_multiple_similar_candidates(["문장 1", "문장 2"], candidates)

        self.assertEqual([row["일련번호"] for row in rows], [1, 2])
        self.assertEqual(rows[0]["llm_score"], 80)
        self.assertEqual(rows[0]["similarity"], 0.8)
        self.assertNotIn("faiss_scores", rows[0])


class SimilarDocumentParserTests(SimpleTestCase):
    def test_supported_extensions_match_product_requirement(self):
        from main.views.testing.similar_documents import SUPPORTED_EXTENSIONS

        self.assertEqual(
            SUPPORTED_EXTENSIONS,
            {
                ".pdf",
                ".doc",
                ".docx",
                ".xls",
                ".xlsx",
                ".hwp",
                ".hwpx",
                ".ppt",
                ".pptx",
                ".md",
            },
        )

    def test_markdown_parser_preserves_section_locator(self):
        from pathlib import Path
        from tempfile import TemporaryDirectory
        from main.views.testing.similar_documents import parse_document

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "product.md"
            path.write_text(
                "# 제품 목적\n문서를 관리합니다.\n## 핵심 기능\n검색을 제공합니다.",
                encoding="utf-8",
            )
            parsed = parse_document(path)

        self.assertEqual(len(parsed.units), 2)
        self.assertIn("SECTION:1|제품 목적", parsed.units[0].source_id)
        self.assertIn("문서를 관리합니다.", parsed.units[0].text)
        self.assertIn("검색을 제공합니다.", parsed.units[1].text)

    def test_extension_content_mismatch_is_rejected(self):
        from pathlib import Path
        from tempfile import TemporaryDirectory
        from main.views.testing.similar_documents import (
            DocumentParseError,
            parse_document,
        )

        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "fake.pdf"
            path.write_text("not a pdf", encoding="utf-8")
            with self.assertRaises(DocumentParseError):
                parse_document(path)

    @patch("main.views.testing.similar_analysis._final_options")
    @patch("main.views.testing.similar_analysis.count_gemma_tokens")
    def test_analysis_deduplicates_units_and_reports_coverage(
        self,
        count_tokens,
        final_options,
    ):
        from main.views.testing.similar_analysis import analyze_documents
        from main.views.testing.similar_documents import (
            DocumentUnit,
            ParsedDocument,
        )

        count_tokens.return_value = 20
        final_options.return_value = (
            "원본",
            ["추천1", "추천2", "추천3", "추천4"],
            ["문서 검색", "사용자 권한 관리"],
        )
        duplicate_text = "문서를 관리하고 검색하는 제품입니다."
        documents = [
            ParsedDocument(
                "a.md",
                ".md",
                [
                    DocumentUnit("A:1", "a.md", "section", "1", duplicate_text),
                    DocumentUnit("A:2", "a.md", "section", "2", "사용자 권한을 관리합니다."),
                ],
            ),
            ParsedDocument(
                "b.md",
                ".md",
                [DocumentUnit("B:1", "b.md", "section", "1", duplicate_text)],
            ),
        ]

        original, recommendations, key_features, coverage = analyze_documents(documents)

        self.assertEqual(original, "원본")
        self.assertEqual(len(recommendations), 4)
        self.assertEqual(key_features, ["문서 검색", "사용자 권한 관리"])
        self.assertEqual(coverage.extracted_units, 3)
        self.assertEqual(coverage.selected_units, 2)
        self.assertEqual(coverage.duplicate_units, 1)
        self.assertEqual(coverage.strategy, "direct")

    @patch("main.views.testing.similar_analysis.generate_gemma_text")
    @patch("main.views.testing.similar_analysis.count_gemma_tokens")
    def test_analysis_reports_actual_llm_input_and_output_tokens(
        self,
        count_tokens,
        generate_text,
    ):
        from main.views.testing.similar_analysis import analyze_documents
        from main.views.testing.similar_documents import (
            DocumentUnit,
            ParsedDocument,
        )

        count_tokens.return_value = 30

        def fake_generate(prompt, usage_callback=None, **kwargs):
            usage_callback(
                {
                    "input_tokens": 123,
                    "output_tokens": 45,
                    "total_tokens": 168,
                }
            )
            return json.dumps(
                {
                    "original_summary": "문서를 관리하고 검색하는 문서 관리 솔루션",
                    "recommendations": [
                        "문서 등록과 검색을 제공하는 문서 관리 시스템",
                        "권한별 문서 공유를 제공하는 문서 관리 프로그램",
                        "문서 버전 관리를 제공하는 기업용 관리 솔루션",
                        "감사 이력과 검색을 제공하는 문서 관리 소프트웨어",
                    ],
                    "key_features": [
                        "기업 문서 등록 및 검색",
                        "사용자별 공유 권한 관리",
                        "문서 변경 및 감사 이력 조회",
                    ],
                },
                ensure_ascii=False,
            )

        generate_text.side_effect = fake_generate
        source_text = "기업 문서의 등록, 검색, 공유와 권한 관리를 제공합니다."
        document = ParsedDocument(
            "product.md",
            ".md",
            [DocumentUnit("F:1", "product.md", "section", "1", source_text)],
        )

        _, _, key_features, coverage = analyze_documents([document])

        self.assertEqual(coverage.extracted_chars, len(source_text))
        self.assertEqual(coverage.llm_input_tokens, 123)
        self.assertEqual(coverage.llm_output_tokens, 45)
        self.assertEqual(coverage.llm_total_tokens, 168)
        self.assertEqual(coverage.llm_call_count, 1)
        self.assertEqual(len(key_features), 3)
        prompt = generate_text.call_args.args[0]
        self.assertIn("의미 보존형 추천 문장", prompt)
        self.assertIn("동의어 또는", prompt)
        self.assertIn("단순 어순 변경", prompt)
        self.assertIn("내용 단어를", prompt)

    @patch("main.views.testing.similar_analysis.generate_gemma_text")
    def test_file_summary_retries_instead_of_cutting_long_sentences(
        self,
        generate_text,
    ):
        from main.views.testing.similar_analysis import _final_options

        too_long = "가" * 61
        valid_payload = {
            "original_summary": "문서를 등록하고 검색하는 문서 관리 솔루션",
            "recommendations": [
                "업무 자료를 수집하고 조회하는 콘텐츠 관리 시스템",
                "기업 기록을 보관하고 탐색하는 자료 운영 프로그램",
                "사내 문서를 축적하고 찾아주는 정보 관리 소프트웨어",
                "조직 자료를 저장하고 검색하도록 지원하는 관리 솔루션",
            ],
            "key_features": [
                "문서 등록 및 전문 검색",
                "사용자별 접근 권한 관리",
            ],
        }
        generate_text.side_effect = [
            json.dumps(
                {
                    "original_summary": too_long,
                    "recommendations": valid_payload["recommendations"],
                    "key_features": valid_payload["key_features"],
                },
                ensure_ascii=False,
            ),
            json.dumps(valid_payload, ensure_ascii=False),
        ]

        original, recommendations, key_features = _final_options("문서 관리 제품 자료")

        self.assertEqual(generate_text.call_count, 2)
        self.assertLessEqual(len(original), 60)
        self.assertTrue(all(len(item) <= 60 for item in recommendations))
        self.assertEqual(len(key_features), 2)
        retry_prompt = generate_text.call_args.args[0]
        self.assertIn("직전 응답", retry_prompt)
        self.assertIn("문장 중간", generate_text.call_args_list[0].args[0])
