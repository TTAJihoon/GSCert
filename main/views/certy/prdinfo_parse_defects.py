from openpyxl import load_workbook
import re
from io import BytesIO

# --- 로컬 유틸(모듈 외부 의존 X) --------------------------------
def _normalize_spaces(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()

def _flat(s):
    # 공백/개행/탭/하이픈/콜론을 제거하고 소문자로 비교
    if s is None:
        return ""
    s = str(s).lower()
    s = re.sub(r"[\s\u00A0]+", "", s)
    s = s.replace(":", "").replace("-", "")
    return s

def _to_int(s):
    if s is None:
        return 0
    s2 = re.sub(r"[,\s]", "", str(s))
    m = re.search(r"(\d+)", s2)
    return int(m.group(1)) if m else 0

def _defect_round_from_filename(filename: str) -> int:
    # 예: "... v3.0.xlsx" -> 3
    m = re.search(r"v\s*(\d+)", filename or "", re.IGNORECASE)
    return int(m.group(1)) if m else 0

# --- 요청 규칙: ‘시험분석자료’ 시트 / D열 키워드 / E열 범위 ---------
QUAL_ORDER = [
    "적합성", "효율성", "호환성", "사용성", "신뢰성",
    "보안성", "유지보수성", "이식성", "요구사항",
]
DEG_ORDER = ["High", "Medium", "Low"]

def extract_process3_xlsx_defects(byts_or_io, filename):
    # bytes/BytesIO 모두 허용
    if hasattr(byts_or_io, "read"):
        data = byts_or_io.read()
    else:
        data = byts_or_io
    wb = load_workbook(BytesIO(data), data_only=True)

    out = {"결함차수": _defect_round_from_filename(filename)}
    for k in QUAL_ORDER:
        out[k] = {"수정전": 0, "최종": 0}
    for k in DEG_ORDER:
        out[k] = {"수정전": 0, "최종": 0}

    # 시트 찾기
    sheet = None
    for name in wb.sheetnames:
        if _flat(name) == _flat("시험분석자료"):
            sheet = wb[name]
            break
    if sheet is None:
        return out  # 시트 없으면 초기값 리턴

    # values[r][c] (0-based)
    values = [[(cell.value if cell.value is not None else "") for cell in row] for row in sheet.iter_rows()]
    H = len(values)

    def txt(r, c):
        try:
            return _normalize_spaces(values[r][c])
        except Exception:
            return ""

    def find_row_in_col_d(keyword: str):
        target = _flat(keyword)
        for r in range(H):
            if _flat(txt(r, 3)) == target:  # D열(3)에서 정확 매칭
                return r
        # 정확 매칭 실패 시 포함 매칭으로 한 번 더 탐색(안전망)
        for r in range(H):
            if target in _flat(txt(r, 3)):
                return r
        return None

    def pick_e_series(start_r: int, end_r: int):
        # E열(index 4) start_r..end_r inclusive
        res = []
        for r in range(max(0, start_r), min(H, end_r + 1)):
            res.append(_to_int(txt(r, 4)))
        return res

    # 1) 품질특성별 결함내역: D에서 찾고 E열 r+2..r+10 → 9개
    rA = find_row_in_col_d("품질특성별 결함내역")
    if rA is not None:
        vals = pick_e_series(rA + 2, rA + 10)
        vals = (vals + [0] * 9)[:9]
        for i, key in enumerate(QUAL_ORDER):
            out[key]["수정전"] = vals[i]

    # 2) 결함정도별 결함내역: D에서 찾고 E열 r+2..r+5 → 4개 (High, Medium, Low, 합계)
    rB = find_row_in_col_d("결함정도별 결함내역")
    if rB is not None:
        vals = pick_e_series(rB + 2, rB + 5)
        vals = (vals + [0] * 4)[:4]
        for i, key in enumerate(DEG_ORDER):
            out[key]["수정전"] = vals[i]

    return out
