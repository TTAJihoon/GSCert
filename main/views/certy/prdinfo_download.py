import os
import json
import threading
from io import BytesIO

from django.conf import settings
from django.http import HttpResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_protect
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import coordinate_to_tuple, get_column_letter

# 원본 템플릿 경로 (기존과 동일)
ORIGIN_XLSX_PATH = os.path.join(settings.BASE_DIR, "main/data/prdinfo.xlsx")

# ── 템플릿 바이트 캐시 (디스크 I/O 병목 제거)
_TEMPLATE_BYTES = None
_TEMPLATE_LOCK = threading.Lock()

def _get_template_bytes() -> bytes:
    global _TEMPLATE_BYTES
    if _TEMPLATE_BYTES is not None:
        return _TEMPLATE_BYTES
    with _TEMPLATE_LOCK:
        if _TEMPLATE_BYTES is None:
            with open(ORIGIN_XLSX_PATH, "rb") as f:
                _TEMPLATE_BYTES = f.read()
    return _TEMPLATE_BYTES

def _write_row(ws, start_cell: str, values):
    """
    start_cell부터 오른쪽으로 values를 순서대로 기록(B5 -> C5 -> ...).
    """
    row, col = coordinate_to_tuple(start_cell)
    for i, v in enumerate(values):
        ws[f"{get_column_letter(col + i)}{row}"] = "" if v is None else v

def _enable_wrap(ws, addr: str):
    """
    대상 셀에 wrap_text=True를 적용(기존 정렬은 유지, 수직은 기본 top로 지정).
    """
    cell = ws[addr]
    old = cell.alignment or Alignment()
    cell.alignment = Alignment(
        horizontal=old.horizontal,
        vertical=old.vertical or "top",
        wrap_text=True
    )

def _enable_wrap_row(ws, row: int, col_start_letter: str, col_end_letter: str):
    """
    같은 행의 연속 범위(B..N 같은) 모든 셀에 wrap 적용.
    """
    start_col = ord(col_start_letter.upper())
    end_col = ord(col_end_letter.upper())
    for c in range(start_col, end_col + 1):
        addr = f"{chr(c)}{row}"
        _enable_wrap(ws, addr)

@require_POST
@csrf_protect
def download_filled_prdinfo(request):
    """
    Luckysheet에서 지정된 범위의 값을 받아 템플릿 사본에 채워서 즉시 다운로드 응답.
    서버 파일에는 저장하지 않음.
    """
    # ── JSON 파싱
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    # 기대 페이로드:
    # {
    #   "prdinfo": {
    #       "row_B5_N5": [...13],
    #       "row_B7_N7": [...13],
    #       "B9": str, "D9": str, "F9": str, "G9": str, "H9": str, "J9": str, "L9": str
    #   },
    #   "defect": {
    #       "row_B4_O4": [...14]    # ← (신규) B..O
    #       # (하위호환) "row_B4_N4": [...13]
    #   }
    # }
    try:
        prdinfo = payload.get("prdinfo", {})
        defect  = payload.get("defect", {})

        vals_B5_N5 = prdinfo.get("row_B5_N5", [])
        vals_B7_N7 = prdinfo.get("row_B7_N7", [])
        B9 = prdinfo.get("B9", ""); D9 = prdinfo.get("D9", "")
        F9 = prdinfo.get("F9", ""); G9 = prdinfo.get("G9", "")
        H9 = prdinfo.get("H9", ""); J9 = prdinfo.get("J9", "")
        L9 = prdinfo.get("L9", "")

        # 결함정보: O열까지 우선 사용, 없으면 N열까지 하위호환
        vals_B4_O4 = defect.get("row_B4_O4")
        if vals_B4_O4 is None:
            vals_B4_O4 = defect.get("row_B4_N4", [])
    except Exception:
        return HttpResponseBadRequest("Missing fields")

    # ── 템플릿 메모리 사본 로드
    tpl_bytes = _get_template_bytes()
    wb = load_workbook(BytesIO(tpl_bytes))
    try:
        ws_prd = wb["제품 정보 요청"]
        ws_def = wb["결함정보"]
    except KeyError:
        return HttpResponseBadRequest("템플릿에 필요한 시트가 없습니다. (제품 정보 요청 / 결함정보)")

    # ── 값 채우기(멀티라인은 문자열에 '\n' 포함 그대로 기록)
    # 제품 정보 요청
    _write_row(ws_prd, "B5", vals_B5_N5)   # B5 ~ N5
    _write_row(ws_prd, "B7", vals_B7_N7)   # B7 ~ N7
    ws_prd["B9"] = B9; ws_prd["D9"] = D9; ws_prd["F9"] = F9
    ws_prd["G9"] = G9; ws_prd["H9"] = H9; ws_prd["J9"] = J9; ws_prd["L9"] = L9

    # 결함정보 (B4 ~ O4)
    if vals_B4_O4 is not None:
        _write_row(ws_def, "B4", vals_B4_O4)

    # ── 랩 텍스트(wrap) 적용: 엑셀에서 \n 줄바꿈이 보이도록
    _enable_wrap_row(ws_prd, 5, "B", "N")
    _enable_wrap_row(ws_prd, 7, "B", "N")
    for addr in ["B9", "D9", "F9", "G9", "H9", "J9", "L9"]:
        _enable_wrap(ws_prd, addr)

    # 결함정보 B..O (O까지)
    _enable_wrap_row(ws_def, 4, "B", "O")

    # ── 메모리에 저장 후 즉시 다운로드
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    resp = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="prdinfo_filled.xlsx"'
    resp["X-Content-Type-Options"] = "nosniff"
    return resp
