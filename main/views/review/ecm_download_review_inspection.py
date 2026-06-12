import fnmatch
import json
import re
import shutil
import sqlite3
import struct
from dataclasses import dataclass
from datetime import datetime, time
from io import BytesIO
from pathlib import Path, PurePosixPath
from zipfile import BadZipFile, ZipFile

from django.conf import settings
from django.utils import timezone
from lxml import etree

from main.models import (
    DownloadReviewProjectReviewStatus,
    DownloadReviewRule,
    DownloadReviewRuleResult,
    DownloadReviewRuleStatus,
)
from main.views.review.ecm_download_verify import FileInfo
from main.views.review.ecm_reference_db import ARTIFACT_REVIEW_COLUMNS


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".gif"}


class DownloadReviewInspectionError(RuntimeError):
    """검사 규칙을 실행할 수 없을 때 발생한다."""


class DownloadReviewCleanupSafetyError(RuntimeError):
    """다운로드 폴더 삭제 대상이 안전하지 않을 때 발생한다."""


@dataclass(frozen=True)
class RuleEvaluation:
    rule: DownloadReviewRule
    sequence: int
    status: str
    expected: str
    actual: str
    message: str
    file_path: str = ""
    file_name: str = ""
    raw_detail: dict | None = None


@dataclass(frozen=True)
class RuleContext:
    project_number: str
    product_raw: str
    product: str
    version: str
    company: str
    pl: str
    wd: str
    start_date: str
    end_date: str
    year: str
    request_date: str
    contract_date: str
    certification_committee_date: str
    derived_variables: dict[str, object]


@dataclass(frozen=True)
class InspectionOutcome:
    project_review_status: str
    reference_review: str
    artifact_results: dict
    passed_count: int
    failed_count: int
    result_count: int


@dataclass(frozen=True)
class CleanupOutcome:
    deleted: bool
    message: str
    file_count: int = 0


@dataclass(frozen=True)
class ExcelSheet:
    name: str
    rows: list[list[str]]
    header_text: str = ""


@dataclass(frozen=True)
class ExcelWorkbook:
    sheets: list[ExcelSheet]


def run_download_inspection(project, verify_result, file_summary) -> InspectionOutcome:
    """등록된 활성 규칙을 실행하고 규칙별 결과를 저장한다.

    실제 규칙은 DB의 inspection_rule에 등록된 항목을 기준으로 한다. 이 함수는 규칙을
    미리 생성하지 않으며, 규칙이 없으면 설정 오류로 처리한다.
    """
    rules = list(DownloadReviewRule.objects.filter(enabled=True).order_by("sort_order", "name", "id"))
    if not rules:
        raise DownloadReviewInspectionError("활성화된 점검규칙이 없습니다.")

    context = _build_rule_context(project)
    evaluations = []
    for sequence, rule in enumerate(rules, start=1):
        evaluation = _evaluate_rule(rule, sequence, project, context, verify_result, file_summary)
        evaluations.append(evaluation)
        _collect_evaluation_variables(context, evaluation)

    DownloadReviewRuleResult.objects.filter(job_project=project).delete()
    DownloadReviewRuleResult.objects.bulk_create(
        [
            DownloadReviewRuleResult(
                job_project=project,
                rule=evaluation.rule,
                rule_code=evaluation.rule.code,
                rule_name=evaluation.rule.name,
                sequence=evaluation.sequence,
                file_path=evaluation.file_path,
                file_name=evaluation.file_name,
                status=evaluation.status,
                expected=evaluation.expected,
                actual=evaluation.actual,
                message=evaluation.message,
                raw_detail_json=evaluation.raw_detail or {},
            )
            for evaluation in evaluations
        ]
    )

    failed_count = sum(
        1
        for evaluation in evaluations
        if evaluation.status in (DownloadReviewRuleStatus.FAIL, DownloadReviewRuleStatus.ERROR)
    )
    passed_count = len(evaluations) - failed_count
    artifact_results = _artifact_results_from_evaluations(evaluations)

    if failed_count:
        return InspectionOutcome(
            project_review_status=DownloadReviewProjectReviewStatus.NEEDS_FIX,
            reference_review="X",
            artifact_results=artifact_results,
            passed_count=passed_count,
            failed_count=failed_count,
            result_count=len(evaluations),
        )

    return InspectionOutcome(
        project_review_status=DownloadReviewProjectReviewStatus.COMPLETED,
        reference_review="O",
        artifact_results=artifact_results,
        passed_count=passed_count,
        failed_count=0,
        result_count=len(evaluations),
    )


def cleanup_download_dir(project, download_dir=None) -> CleanupOutcome:
    """프로젝트 다운로드 폴더를 삭제한다.

    삭제 대상은 AGENT_DOWNLOAD_BASE_DIR 아래에 있고 폴더명에 프로젝트번호가 포함된
    디렉터리로 제한한다.
    """
    raw_path = str(download_dir or project.download_dir or "").strip()
    if not raw_path:
        return CleanupOutcome(deleted=False, message="삭제할 다운로드 폴더가 없습니다.")

    target = Path(raw_path).resolve()
    base_dir = Path(getattr(settings, "AGENT_DOWNLOAD_BASE_DIR")).resolve()
    _validate_cleanup_target(project.project_number, base_dir, target)

    if not target.exists():
        project.zip_deleted_at = timezone.now()
        project.save(update_fields=["zip_deleted_at", "updated_at"])
        return CleanupOutcome(deleted=False, message="다운로드 폴더가 이미 없습니다.")

    file_count = sum(1 for item in target.rglob("*") if item.is_file())
    shutil.rmtree(target)

    project.zip_deleted_at = timezone.now()
    project.save(update_fields=["zip_deleted_at", "updated_at"])
    return CleanupOutcome(
        deleted=True,
        message="다운로드 폴더를 삭제했습니다.",
        file_count=file_count,
    )


def get_rule_output_variables(job_project):
    """이미 저장된 규칙 결과에서 후속 규칙용 산출 변수를 모은다."""
    variables = {}
    results = DownloadReviewRuleResult.objects.filter(job_project=job_project).order_by("sequence", "id")
    for result in results:
        variables.update(_raw_detail_variables(result.raw_detail_json or {}))
    return variables


def _evaluate_rule(rule, sequence, project, context, verify_result, file_summary):
    rule_type = (rule.rule_type or "").strip()
    if rule_type == "min_file_count":
        return _evaluate_min_file_count(rule, sequence, project, verify_result)
    if rule_type == "filename_contains_project_number":
        return _evaluate_filename_contains_project_number(rule, sequence, project, verify_result)
    if rule_type == "required_extension":
        return _evaluate_required_extension(rule, sequence, project, verify_result)
    if rule_type == "required_file_name_contains":
        return _evaluate_required_file_name_contains(rule, sequence, project, verify_result)
    if rule_type == "required_artifact_file":
        return _evaluate_required_artifact_file(rule, sequence, project, context, verify_result)
    if rule_type == "downloadable_artifact_check":
        return _evaluate_downloadable_artifact_check(rule, sequence, project, context, verify_result)
    if rule_type == "document_artifact_check":
        return _evaluate_document_artifact_check(rule, sequence, project, context, verify_result)
    if rule_type == "all_files_non_empty":
        return _evaluate_all_files_non_empty(rule, sequence, project, verify_result)
    if rule_type == "excel_feature_list_check":
        return _evaluate_excel_feature_list_check(rule, sequence, project, context, verify_result)
    if rule_type == "test_plan_document_check":
        return _evaluate_test_plan_document_check(rule, sequence, project, context, verify_result)
    if rule_type == "image_screenshot_folder_date_check":
        return _evaluate_image_screenshot_folder_date_check(rule, sequence, project, context, verify_result)
    if rule_type == "test_case_check":
        return _evaluate_test_case_check(rule, sequence, project, context, verify_result)
    if rule_type == "rawdata_folder_structure_check":
        return _evaluate_rawdata_folder_structure_check(rule, sequence, project, verify_result)
    if rule_type == "test_report_document_check":
        return _evaluate_test_report_document_check(rule, sequence, project, context, verify_result)
    if rule_type == "defect_report_check":
        return _evaluate_defect_report_check(rule, sequence, project, context, verify_result)
    if rule_type == "inspection_checklist_check":
        return _evaluate_inspection_checklist_check(rule, sequence, project, context, verify_result)
    if rule_type == "quality_inspection_table_check":
        return _evaluate_quality_inspection_table_check(rule, sequence, project, context, verify_result)
    if rule_type == "quality_evaluation_report_check":
        return _evaluate_quality_evaluation_report_check(rule, sequence, project, context, verify_result)

    raise DownloadReviewInspectionError(f"지원하지 않는 점검규칙 유형입니다: {rule_type or '(비어 있음)'}")


def _collect_evaluation_variables(context, evaluation):
    if not evaluation.raw_detail:
        return
    context.derived_variables.update(_raw_detail_variables(evaluation.raw_detail))


def _raw_detail_variables(raw_detail):
    if not isinstance(raw_detail, dict):
        return {}
    variables = raw_detail.get("variables")
    if not isinstance(variables, dict):
        return {}
    return {
        str(key).strip(): value
        for key, value in variables.items()
        if str(key).strip()
    }


def _evaluate_min_file_count(rule, sequence, project, verify_result):
    config = rule.config_json or {}
    min_count = int(config.get("min_count") or 1)
    files = _matching_files(rule, verify_result)
    status = DownloadReviewRuleStatus.PASS if len(files) >= min_count else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"{min_count}개 이상",
        actual=f"{len(files)}개",
        message="파일 개수 기준을 충족했습니다." if status == DownloadReviewRuleStatus.PASS else "파일 개수가 부족합니다.",
        file_path=_representative_path(files, project.project_number),
        file_name=_representative_name(files),
        raw_detail={"matched_file_count": len(files), "min_count": min_count},
    )


def _evaluate_filename_contains_project_number(rule, sequence, project, verify_result):
    files = _matching_files(rule, verify_result)
    matched = [file_info for file_info in files if project.project_number in file_info.name]
    status = DownloadReviewRuleStatus.PASS if matched else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"파일명에 {project.project_number} 포함",
        actual=", ".join(file_info.name for file_info in matched[:5]) if matched else "일치 파일 없음",
        message="프로젝트번호가 포함된 파일을 확인했습니다." if matched else "프로젝트번호가 포함된 파일을 찾지 못했습니다.",
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail={"matched_file_count": len(matched), "total_file_count": len(files)},
    )


def _evaluate_required_extension(rule, sequence, project, verify_result):
    config = rule.config_json or {}
    extension = str(config.get("extension") or _extension_from_file_type(rule.target_file_type)).lower()
    if not extension.startswith("."):
        extension = f".{extension}"

    files = _matching_files(rule, verify_result, ignore_target_file_type=True)
    matched = [file_info for file_info in files if file_info.extension.lower() == extension]
    status = DownloadReviewRuleStatus.PASS if matched else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"{extension} 파일 존재",
        actual=", ".join(file_info.name for file_info in matched[:5]) if matched else "일치 파일 없음",
        message=f"{extension} 파일을 확인했습니다." if matched else f"{extension} 파일을 찾지 못했습니다.",
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail={"extension": extension, "matched_file_count": len(matched)},
    )


def _evaluate_required_file_name_contains(rule, sequence, project, verify_result):
    config = rule.config_json or {}
    contains = str(config.get("contains") or "").strip()
    if not contains:
        raise DownloadReviewInspectionError(f"{rule.name} 규칙의 contains 설정이 없습니다.")

    files = _matching_files(rule, verify_result)
    matched = [file_info for file_info in files if contains in file_info.name]
    status = DownloadReviewRuleStatus.PASS if matched else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"파일명에 {contains} 포함",
        actual=", ".join(file_info.name for file_info in matched[:5]) if matched else "일치 파일 없음",
        message=f"{contains} 파일을 확인했습니다." if matched else f"{contains} 파일을 찾지 못했습니다.",
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail={"contains": contains, "matched_file_count": len(matched)},
    )


def _evaluate_required_artifact_file(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)

    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    extensions = _configured_extensions(config, rule.target_file_type)
    matched = [
        file_info
        for file_info in files
        if (not name_keywords or _name_contains_all(file_info.name, name_keywords))
        and _extension_matches(file_info.extension, extensions)
    ]

    exact_count = config.get("exact_count")
    min_count = int(config.get("min_count") or 1)
    if exact_count is not None:
        expected_count = int(exact_count)
        passed = len(matched) == expected_count
        expected = f"{expected_count}개"
    else:
        passed = len(matched) >= min_count
        expected = f"{min_count}개 이상"

    status = DownloadReviewRuleStatus.PASS if passed else DownloadReviewRuleStatus.FAIL
    message = _artifact_file_message(config, status, len(matched), exact_count)
    expected_parts = [
        expected,
    ]
    if name_keywords:
        expected_parts.insert(0, "파일명에 " + ", ".join(name_keywords) + " 포함")
    if extensions:
        expected_parts.append("확장자 " + ", ".join(extensions))

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=" / ".join(expected_parts),
        actual=_matched_files_actual(matched),
        message=message,
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail={
            "matched_file_count": len(matched),
            "folder_keyword_chain": config.get("folder_keyword_chain") or [],
            "selected_folder": selected_folder,
            "filename_keywords": name_keywords,
            "extensions": extensions,
            "matched_files": [
                _display_path(file_info.path, project.project_number)
                for file_info in matched[:20]
            ],
        },
    )


def _evaluate_downloadable_artifact_check(rule, sequence, project, context, verify_result):
    """파일 존재 여부만 자동 판정하고, 찾은 파일을 다운로드형 산출물로 제공한다.

    14번 시험기록서처럼 내용 검사 없이 사용자가 직접 받아 확인하는 규칙용이다.
    """
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    extensions = _configured_extensions(config, rule.target_file_type)
    matched = [
        file_info
        for file_info in files
        if (not name_keywords or _name_contains_all(file_info.name, name_keywords))
        and _extension_matches(file_info.extension, extensions)
    ]

    exact_count = config.get("exact_count")
    min_count = int(config.get("min_count") or 1)
    if exact_count is not None:
        passed = len(matched) == int(exact_count)
        expected_count = f"{int(exact_count)}개"
    else:
        passed = len(matched) >= min_count
        expected_count = f"{min_count}개 이상"

    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "extensions": extensions,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }

    expected_parts = []
    if name_keywords:
        expected_parts.append("파일명에 " + ", ".join(name_keywords) + " 포함")
    if extensions:
        expected_parts.append("확장자 " + ", ".join(extensions))
    expected_parts.append(expected_count)
    expected_text = " / ".join(expected_parts)

    if not passed:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected=expected_text,
            actual=_matched_files_actual(matched),
            message=config.get("missing_message") or "파일 확인 불가",
            file_path=_representative_path(matched or files, project.project_number),
            file_name=_representative_name(matched or files),
            raw_detail=raw_detail,
        )

    base_id = _safe_artifact_id(config.get("artifact_id") or "download")
    base_label = config.get("artifact_label") or "다운로드 파일"
    artifacts = []
    try:
        for index, file_info in enumerate(matched):
            single = len(matched) == 1
            artifacts.append(
                _store_pdf_download_artifact(
                    project,
                    rule,
                    file_info,
                    artifact_id=base_id if single else f"{base_id}_{index + 1}",
                    label=base_label if single else f"{base_label} {index + 1}",
                )
            )
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="다운로드 산출물 저장 가능",
            actual=str(exc),
            message=str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=_representative_name(matched),
            raw_detail=raw_detail,
        )

    raw_detail["artifacts"] = artifacts
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected=expected_text,
        actual=_matched_files_actual(matched),
        message=config.get("pass_message") or "파일을 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _evaluate_document_artifact_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    if not name_keywords:
        raise DownloadReviewInspectionError(f"{rule.name} 규칙의 filename_keywords 설정이 없습니다.")

    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
    ]
    file_check = _evaluate_required_file_specs(config, matched)
    content_check = _evaluate_content_checks(config, matched, context) if file_check["passed"] else {
        "passed": True,
        "expected": [],
        "actual": [],
        "message": "",
        "details": [],
    }
    artifact_check = _evaluate_configured_artifacts(config, matched, project, rule)

    if artifact_check["error"]:
        status = DownloadReviewRuleStatus.ERROR
    else:
        passed = file_check["passed"] and content_check["passed"]
        status = DownloadReviewRuleStatus.PASS if passed else DownloadReviewRuleStatus.FAIL
    message = (
        config.get("pass_message")
        if status == DownloadReviewRuleStatus.PASS
        else artifact_check["message"] or file_check["message"] or content_check["message"] or config.get("missing_message")
    ) or ("문서 내용을 확인했습니다." if status == DownloadReviewRuleStatus.PASS else "문서 내용이 기준과 일치하지 않습니다.")
    expected = " / ".join([*file_check["expected"], *content_check["expected"]])
    actual = " / ".join([*file_check["actual"], *content_check["actual"]]) or _matched_files_actual(matched)

    raw_detail = {
        "matched_file_count": len(matched),
        "folder_keyword_chain": config.get("folder_keyword_chain") or [],
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "file_checks": file_check["details"],
        "content_checks": content_check["details"],
        "matched_files": [
            _display_path(file_info.path, project.project_number)
            for file_info in matched[:20]
        ],
    }
    if artifact_check["artifacts"]:
        raw_detail["artifacts"] = artifact_check["artifacts"]
    if artifact_check["details"]:
        raw_detail["artifact_checks"] = artifact_check["details"]

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail=raw_detail,
    )


def _evaluate_all_files_non_empty(rule, sequence, project, verify_result):
    files = _matching_files(rule, verify_result)
    empty_files = [file_info for file_info in files if file_info.size == 0]
    status = (
        DownloadReviewRuleStatus.PASS
        if files and not empty_files
        else DownloadReviewRuleStatus.FAIL
    )
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected="모든 파일 크기 1 byte 이상",
        actual="0 byte 파일 없음" if status == DownloadReviewRuleStatus.PASS else ", ".join(file_info.name for file_info in empty_files[:5]) or "검사 대상 파일 없음",
        message="빈 파일이 없습니다." if status == DownloadReviewRuleStatus.PASS else "빈 파일이 있거나 검사 대상 파일이 없습니다.",
        file_path=_representative_path(empty_files or files, project.project_number),
        file_name=_representative_name(empty_files or files),
        raw_detail={"empty_file_count": len(empty_files), "matched_file_count": len(files)},
    )


def _evaluate_configured_artifacts(config, matched_files, project, rule):
    artifacts = []
    details = []
    message = ""
    error = False

    for spec in config.get("artifacts") or []:
        artifact_type = str(spec.get("type") or "").strip()
        extensions = _configured_extensions(spec, "any")
        files = [
            file_info
            for file_info in matched_files
            if _extension_matches(file_info.extension, extensions)
        ]
        detail = {
            "type": artifact_type,
            "extensions": extensions,
            "matched_file_count": len(files),
        }
        if not files:
            detail["passed"] = not spec.get("required", False)
            details.append(detail)
            if spec.get("required", False) and not error:
                error = True
                message = spec.get("missing_message") or "산출물 대상 파일이 없습니다."
            continue

        try:
            if artifact_type == "pdf_first_page":
                artifact = _store_pdf_first_page_artifact(
                    project,
                    rule,
                    files[0],
                    artifact_id=spec.get("id") or "pdf_first_page",
                    label=spec.get("label") or "PDF 1페이지",
                )
            else:
                raise DownloadReviewInspectionError(f"지원하지 않는 산출물 유형입니다: {artifact_type or '(비어 있음)'}")
        except DownloadReviewInspectionError as exc:
            detail["passed"] = False
            detail["message"] = str(exc)
            details.append(detail)
            if not error:
                error = True
                message = str(exc)
            continue

        detail["passed"] = True
        detail["artifact_id"] = artifact["id"]
        details.append(detail)
        artifacts.append(artifact)

    return {
        "artifacts": artifacts,
        "details": details,
        "error": error,
        "message": message,
    }


def _evaluate_excel_feature_list_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _extension_matches(file_info.extension, [".xlsx", ".xls"])
    ]

    if len(matched) != 1:
        status = DownloadReviewRuleStatus.FAIL
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=status,
            expected="파일명에 " + ", ".join(name_keywords) + " 포함 / Excel 파일 1개",
            actual=_matched_files_actual(matched),
            message=(
                config.get("multiple_message")
                if len(matched) > 1
                else config.get("missing_message")
            ) or "파일이 없습니다.",
            file_path=_representative_path(matched or files, project.project_number),
            file_name=_representative_name(matched or files),
            raw_detail={
                "selected_folder": selected_folder,
                "matched_file_count": len(matched),
                "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched],
            },
        )

    file_info = matched[0]
    details = {
        "selected_folder": selected_folder,
        "matched_file": _display_path(file_info.path, project.project_number),
    }
    try:
        workbook = _read_excel_workbook(file_info)
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="Excel 파일 파싱 가능",
            actual=str(exc),
            message=str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=details,
        )

    details["sheet_names"] = [sheet.name for sheet in workbook.sheets]
    if len(workbook.sheets) != 1:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected="시트 1개",
            actual=f"시트 {len(workbook.sheets)}개",
            message=config.get("sheet_count_message") or "불필요한 시트가 존재",
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=details,
        )

    sheet = workbook.sheets[0]
    title = _resolve_rule_value(config.get("title_text") or "{프로젝트번호} 기능리스트", context)
    title_cell = _find_cell_containing(sheet.rows, title)
    author_label = str(config.get("author_label") or "작성자")
    author_cell = _find_cell_containing(sheet.rows, author_label)
    author_ok = bool(author_cell and context.pl and context.pl in author_cell["value"])

    details.update({
        "title_text": title,
        "title_cell": title_cell or {},
        "author_label": author_label,
        "author_cell": author_cell or {},
    })
    if not title_cell or not author_ok:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected=f"{title} 포함 / {author_label} 셀에 {context.pl} 포함",
            actual=f"제목={'있음' if title_cell else '없음'} / 작성자={author_cell['value'] if author_cell else '없음'}",
            message=config.get("content_message") or "시험번호 또는 작성자가 잘못 작성됨",
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=details,
        )

    category_label = str(config.get("capture_anchor") or "대분류")
    capture_area = _excel_area_from_column_anchor(sheet.rows, category_label, column_index=0)
    details["capture_area"] = capture_area or {}
    if not capture_area:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected=f"A열에서 {category_label} 기준 영역 확인",
            actual="캡처 기준 셀 없음",
            message=config.get("capture_message") or "기능리스트 캡처 영역을 찾지 못했습니다.",
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=details,
        )

    try:
        artifact = _store_excel_area_artifact(
            project,
            rule,
            sheet,
            capture_area,
            artifact_id=config.get("capture_artifact_id") or "feature_list_area",
            label=config.get("capture_artifact_label") or "기능리스트 영역",
            source_file=_display_path(file_info.path, project.project_number),
        )
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="기능리스트 영역 이미지 생성 가능",
            actual=str(exc),
            message=str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=details,
        )

    details["artifacts"] = [artifact]
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected=f"시트 1개 / {title} 포함 / 작성자 {context.pl} 포함",
        actual=f"{sheet.name} / 캡처 영역 {capture_area['range']}",
        message=config.get("pass_message") or "기능리스트를 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=file_info.name,
        raw_detail=details,
    )


def _evaluate_test_plan_document_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
    ]
    docx_files = [file_info for file_info in matched if _extension_matches(file_info.extension, [".docx"])]
    pdf_files = [file_info for file_info in matched if _extension_matches(file_info.extension, [".pdf"])]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
        "docx_count": len(docx_files),
        "pdf_count": len(pdf_files),
        "checks": [],
    }

    if len(docx_files) != 1 or len(pdf_files) != 1:
        return _test_plan_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected=".docx 1개 / .pdf 1개",
            actual=f".docx {len(docx_files)}개 / .pdf {len(pdf_files)}개",
            message=config.get("missing_message") or "파일이 없습니다.",
        )

    docx_file = docx_files[0]
    pdf_file = pdf_files[0]
    try:
        tables = _docx_tables(docx_file)
        footer_text = _docx_footer_text(docx_file)
        plan_spec_table = _docx_first_table_after_text(docx_file, config.get("spec_marker") or "<세부사양>")
        artifact = _store_pdf_first_page_artifact(
            project,
            rule,
            pdf_file,
            artifact_id="pdf_first_page",
            label=config.get("pdf_artifact_label") or "시험계획서 1페이지",
        )
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="시험계획서 docx/pdf 파싱 가능",
            actual=str(exc),
            message=str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=_representative_name(matched),
            raw_detail=raw_detail,
        )

    raw_detail["table_count"] = len(tables)
    raw_detail["footer_text"] = footer_text
    raw_detail["spec_table"] = plan_spec_table or []
    raw_detail["artifacts"] = [artifact]

    if len(tables) < 2:
        return _test_plan_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="시험계획서 docx 표 2개 이상",
            actual=f"표 {len(tables)}개",
            message=config.get("product_message") or "제품정보가 틀림",
        )

    first_table = tables[0]
    first_table_checks = _test_plan_first_table_checks(first_table, config, context)
    raw_detail["checks"].extend(first_table_checks)
    failed_first = next((check for check in first_table_checks if not check["passed"]), None)
    if failed_first:
        return _test_plan_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=failed_first["expected"],
            actual=failed_first["actual"],
            message=failed_first["message"],
        )

    second_table = tables[1]
    product_checks = _test_plan_product_checks(second_table, config, context)
    raw_detail["checks"].extend(product_checks)
    failed_product = next((check for check in product_checks if not check["passed"]), None)
    if failed_product:
        return _test_plan_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=failed_product["expected"],
            actual=failed_product["actual"],
            message=failed_product["message"],
        )

    configuration_table = _docx_first_table_after_text(
        docx_file,
        config.get("configuration_marker") or "5.1 형상항목 식별 규칙",
    )
    configuration_check = _test_plan_configuration_id_check(configuration_table, config, context)
    raw_detail["checks"].append(configuration_check)
    if not configuration_check["passed"]:
        return _test_plan_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=configuration_check["expected"],
            actual=configuration_check["actual"],
            message=configuration_check["message"],
        )

    schedule_table = _docx_first_table_after_text(
        docx_file,
        config.get("schedule_marker") or "2.2 시험일정",
    )
    schedule_check = _test_plan_schedule_check(schedule_table, config, context)
    raw_detail["checks"].append(schedule_check)
    if not schedule_check["passed"]:
        return _test_plan_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=schedule_check["expected"],
            actual=schedule_check["actual"],
            message=schedule_check["message"],
        )

    footer_expected = _resolve_rule_value(config.get("footer_text") or "Copyright {연도} TTA", context)
    footer_check = {
        "name": "footer_copyright",
        "passed": footer_expected in footer_text,
        "expected": footer_expected,
        "actual": footer_text or "바닥글 없음",
        "message": config.get("footer_message") or "바닥글 Copyright가 잘못 작성됨",
    }
    raw_detail["checks"].append(footer_check)
    if not footer_check["passed"]:
        return _test_plan_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=footer_check["expected"],
            actual=footer_check["actual"],
            message=footer_check["message"],
        )

    spec_check = _test_plan_spec_table_check(plan_spec_table, config, context)
    raw_detail["checks"].append(spec_check)
    if not spec_check["passed"]:
        return _test_plan_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=spec_check["expected"],
            actual=spec_check["actual"],
            message=spec_check["message"],
        )

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected="시험계획서 docx/pdf / 표 값 / 형상항목 ID / WD / 바닥글 / 세부사양",
        actual=f".docx {docx_file.name} / .pdf {pdf_file.name}",
        message=config.get("pass_message") or "시험계획서를 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _test_plan_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _test_plan_first_table_checks(table, config, context):
    manager_expected = _resolve_rule_value(str(config.get("manager_expected") or "김진영"), context)
    return [
        {
            "name": "first_table_start_date",
            "passed": _table_cell(table, 1, 2) == context.start_date,
            "expected": f"1행 2열 = {context.start_date}",
            "actual": _table_cell(table, 1, 2) or "값 없음",
            "message": config.get("date_message") or "시험계획서 날짜가 잘못 작성됨",
        },
        {
            "name": "first_table_manager",
            "passed": manager_expected in _table_cell(table, 3, 2),
            "expected": f"3행 2열에 {manager_expected} 포함",
            "actual": _table_cell(table, 3, 2) or "값 없음",
            "message": config.get("manager_message") or "시험계획서 담당자가 잘못 작성됨",
        },
        {
            "name": "first_table_pl",
            "passed": bool(context.pl and context.pl in _table_cell(table, 4, 2)),
            "expected": f"4행 2열에 {context.pl} 포함",
            "actual": _table_cell(table, 4, 2) or "값 없음",
            "message": config.get("pl_message") or "시험계획서 PL이 잘못 작성됨",
        },
    ]


def _test_plan_product_checks(table, config, context):
    product_label = str(config.get("product_name_label") or "소프트웨어 명")
    version_label = str(config.get("version_label") or "버전")
    application_label = str(config.get("application_number_label") or "시험신청번호")
    product_actual = _find_next_cell_by_label(table, product_label)
    version_actual = _find_next_cell_by_label(table, version_label)
    application_actual = _find_next_cell_by_label(table, application_label)
    product_message = config.get("product_message") or "제품정보가 틀림"
    return [
        {
            "name": "version_exists",
            "passed": bool(context.version),
            "expected": "제품명에서 버전 파싱 가능",
            "actual": context.product_raw or "제품명 없음",
            "message": config.get("version_missing_message") or "버전을 찾을 수 없음",
        },
        {
            "name": "product_name",
            "passed": _normalize_spaces(product_actual) == context.product,
            "expected": f"{product_label} 오른쪽 셀 = {context.product}",
            "actual": product_actual or "값 없음",
            "message": product_message,
        },
        {
            "name": "product_version",
            "passed": _normalize_spaces(version_actual) == context.version,
            "expected": f"{version_label} 오른쪽 셀 = {context.version}",
            "actual": version_actual or "값 없음",
            "message": product_message,
        },
        {
            "name": "application_number",
            "passed": _normalize_spaces(application_actual) == context.project_number,
            "expected": f"{application_label} 오른쪽 셀 = {context.project_number}",
            "actual": application_actual or "값 없음",
            "message": product_message,
        },
    ]


def _test_plan_configuration_id_check(table, config, context):
    header = str(config.get("configuration_header") or "형상항목 ID")
    header_cell = _find_cell_containing(table, header)
    if not header_cell:
        return {
            "name": "configuration_id",
            "passed": False,
            "expected": f"{header} 열 존재",
            "actual": "헤더 없음",
            "message": config.get("configuration_message") or "형상항목 ID가 잘못 작성됨",
        }

    values = []
    invalid_values = []
    for row_index in range(header_cell["row"] + 1, len(table) + 1):
        value = _table_cell(table, row_index, header_cell["column"])
        if not value:
            continue
        values.append({"row": row_index, "value": value})
        if context.project_number not in value:
            invalid_values.append({"row": row_index, "value": value})

    return {
        "name": "configuration_id",
        "passed": bool(values) and not invalid_values,
        "expected": f"{header} 열 값에 {context.project_number} 포함",
        "actual": " / ".join(item["value"] for item in invalid_values[:5]) if invalid_values else f"{len(values)}개 값 확인",
        "message": config.get("configuration_message") or "형상항목 ID가 잘못 작성됨",
        "header_cell": header_cell,
        "values": values,
        "invalid_values": invalid_values,
    }


def _test_plan_schedule_check(table, config, context):
    header = str(config.get("schedule_header") or "WD")
    header_cell = _find_cell_containing(table, header)
    column = header_cell["column"] if header_cell else 2
    start_row = header_cell["row"] + 1 if header_cell else 1
    actual_values = []
    for row_index in range(start_row, len(table) + 1):
        value = _table_cell(table, row_index, column)
        if value:
            actual_values.append({"row": row_index, "value": value})
        if len(actual_values) >= 4:
            break

    wd = _context_wd_int(context)
    expected_values = ["1", "1", str(wd - 3) if wd is not None else "{WD}-3", "1"]
    normalized_actual = [_normalize_number_text(item["value"]) for item in actual_values]
    passed = wd is not None and normalized_actual == expected_values
    return {
        "name": "schedule_wd",
        "passed": passed,
        "expected": ", ".join(expected_values),
        "actual": ", ".join(item["value"] for item in actual_values) or "WD 값 없음",
        "message": config.get("schedule_message") or "시험일정 WD가 틀림",
        "header_cell": header_cell or {},
        "values": actual_values,
    }


def _test_plan_spec_table_check(plan_table, config, context):
    report_table = _context_variable(context, config.get("report_spec_variable") or "시험성적서_세부사양표")
    normalized_plan = _normalize_docx_table(plan_table)
    normalized_report = _normalize_docx_table(report_table if isinstance(report_table, list) else [])
    mismatches = _matrix_mismatches(
        normalized_plan,
        normalized_report,
        left_origin=(1, 1),
        right_origin=(1, 1),
    )[:20]
    passed = bool(normalized_plan) and normalized_plan == normalized_report
    return {
        "name": "spec_table",
        "passed": passed,
        "expected": "시험성적서 <세부사양> 표와 일치",
        "actual": "일치" if passed else f"불일치 {len(mismatches)}건" if normalized_plan and normalized_report else "비교 대상 표 없음",
        "message": config.get("spec_message") or "시험환경 세부사양 표가 결과서와 다름",
        "plan_table": normalized_plan,
        "report_table": normalized_report,
        "mismatches": mismatches,
    }


def _table_cell(table, row, column):
    row_index = row - 1
    col_index = column - 1
    if row_index < 0 or col_index < 0 or row_index >= len(table or []):
        return ""
    row_values = table[row_index]
    if col_index >= len(row_values):
        return ""
    return _normalize_spaces(row_values[col_index])


def _context_wd_int(context):
    try:
        return int(str(context.wd).strip())
    except (TypeError, ValueError):
        return None


def _normalize_number_text(value):
    text = _normalize_spaces(value)
    if re.fullmatch(r"-?\d+(?:\.0+)?", text):
        return str(int(float(text)))
    return text


def _normalize_docx_table(table):
    rows = [
        [_normalize_spaces(cell) for cell in row]
        for row in (table or [])
    ]
    return _trim_empty_edges(rows)


def _evaluate_test_case_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    extensions = _configured_extensions(config, rule.target_file_type)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _extension_matches(file_info.extension, extensions)
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "extensions": extensions,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }

    if len(matched) != 1:
        return _test_case_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="테스트케이스 Excel 파일 1개",
            actual=_matched_files_actual(matched),
            message=config.get("missing_message") or "파일이 존재하지 않음",
        )

    file_info = matched[0]
    try:
        workbook = _read_excel_workbook(file_info)
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="테스트케이스 Excel 파일 파싱 가능",
            actual=str(exc),
            message=str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=raw_detail,
        )

    raw_detail["sheet_names"] = [sheet.name for sheet in workbook.sheets]
    if len(workbook.sheets) != 1:
        return _test_case_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="시트 1개",
            actual=f"시트 {len(workbook.sheets)}개",
            message=config.get("sheet_count_message") or "테스트케이스 시트가 1개 이상임",
        )

    sheet = workbook.sheets[0]
    title_text = _resolve_rule_value(config.get("title_text") or "{project_number} 테스트케이스", context)
    title_cell = _find_cell_containing(sheet.rows, title_text)
    raw_detail["title_check"] = {
        "expected": title_text,
        "matched_cell": title_cell or {},
    }
    if not title_cell:
        return _test_case_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"문서 내 '{title_text}' 포함",
            actual="일치 문장 없음",
            message=config.get("project_number_message") or "프로젝트 번호가 잘못 작성됨",
        )

    author_label = str(config.get("author_label") or "작성자:")
    reviewer_label = str(config.get("reviewer_label") or "검토자:")
    reviewer_expected = _resolve_rule_value(str(config.get("reviewer_expected") or "김진영"), context)
    author_cell = _find_cell_with_all(sheet.rows, [author_label, context.pl])
    reviewer_cell = {}
    reviewer_ok = False
    if author_cell:
        reviewer_value = _sheet_cell(sheet, author_cell["row"] + 1, author_cell["column"])
        reviewer_cell = {
            "row": author_cell["row"] + 1,
            "column": author_cell["column"],
            "value": reviewer_value,
        }
        reviewer_ok = bool(
            reviewer_value
            and reviewer_label in reviewer_value
            and reviewer_expected in reviewer_value
        )
    raw_detail["author_reviewer_check"] = {
        "author_label": author_label,
        "pl": context.pl,
        "author_cell": author_cell or {},
        "reviewer_label": reviewer_label,
        "reviewer_expected": reviewer_expected,
        "reviewer_cell": reviewer_cell,
    }
    if not author_cell or not reviewer_ok:
        return _test_case_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"{author_label} {context.pl} / 아래 셀 {reviewer_label} {reviewer_expected}",
            actual=f"작성자={author_cell['value'] if author_cell else '없음'} / 검토자={reviewer_cell.get('value') or '없음'}",
            message=config.get("author_message") or "작성자 또는 검토자가 잘못 작성됨",
        )

    date_text = _resolve_rule_value(config.get("date_text") or "작성일: {시작일} ~ {종료일}", context)
    date_cell = _find_cell_normalized_contains_all(sheet.rows, [date_text])
    raw_detail["date_check"] = {
        "expected": date_text,
        "matched_cell": date_cell or {},
    }
    if not date_cell:
        return _test_case_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"문서 내 '{date_text}' 포함(공백 제거 후 비교)",
            actual="일치 작성일 없음",
            message=config.get("date_message") or "작성일이 잘못 작성됨",
        )

    residual_expected = _context_int(context, "잔여결함수")
    result_header = str(config.get("result_header") or "상세 테스트 결과")
    result_header_cell = _find_cell_containing(sheet.rows, result_header)
    failed_rows = []
    if result_header_cell:
        failed_rows = _test_case_failed_result_rows(
            sheet,
            start_row=result_header_cell["row"] + 1,
            column=result_header_cell["column"],
        )
    residual_check = {
        "result_header": result_header,
        "header_cell": result_header_cell or {},
        "expected_count": residual_expected,
        "actual_count": len(failed_rows),
        "failed_rows": failed_rows,
    }
    raw_detail["residual_defect_check"] = residual_check
    if residual_expected is None or not result_header_cell or len(failed_rows) != residual_expected:
        actual_parts = []
        if residual_expected is None:
            actual_parts.append("{잔여결함수} 없음")
        if not result_header_cell:
            actual_parts.append("상세 테스트 결과 열 없음")
        actual_parts.append(f"F {len(failed_rows)}개")
        return _test_case_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"{result_header} 열의 F 개수 = {residual_expected if residual_expected is not None else '{잔여결함수}'}",
            actual=" / ".join(actual_parts),
            message=config.get("residual_message") or "잔여 결함이 작성되지 않음",
        )

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected=f"시트 1개 / {title_text} / 작성자·검토자 / 작성일 / F {residual_expected}개",
        actual=f"{sheet.name} / F {len(failed_rows)}개",
        message=config.get("pass_message") or "테스트케이스를 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=file_info.name,
        raw_detail=raw_detail,
    )


def _test_case_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _test_case_failed_result_rows(sheet, *, start_row, column):
    failed_rows = []
    for row in range(start_row, len(sheet.rows) + 1):
        value = _sheet_cell(sheet, row, column)
        if _normalize_no_space(value).upper() == "F":
            failed_rows.append(row)
    return failed_rows


def _evaluate_image_screenshot_folder_date_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    min_images = int(config.get("min_images_per_folder") or 5)
    required_folder_count = int(config.get("required_candidate_folder_count") or 2)
    image_files = [
        file_info
        for file_info in files
        if file_info.extension.lower() in IMAGE_EXTENSIONS
    ]
    folders = {}
    for file_info in image_files:
        key = tuple(_folder_segments(file_info.path))
        folders.setdefault(key, []).append(file_info)

    candidate_folders = {
        folder: folder_files
        for folder, folder_files in folders.items()
        if len(folder_files) >= min_images
    }
    parent_candidates = {}
    for folder, folder_files in candidate_folders.items():
        if not folder:
            continue
        parent_candidates.setdefault(folder[:-1], []).append((folder, folder_files))

    selected_parent = None
    selected_candidates = []
    for parent, candidates in sorted(parent_candidates.items(), key=lambda item: "/".join(item[0])):
        if len(candidates) >= required_folder_count:
            selected_parent = parent
            selected_candidates = sorted(candidates, key=lambda item: "/".join(item[0]))[:required_folder_count]
            break

    raw_detail = {
        "selected_folder": selected_folder,
        "min_images_per_folder": min_images,
        "required_candidate_folder_count": required_folder_count,
        "candidate_folders": [
            {"folder": "/".join(folder), "image_count": len(folder_files)}
            for folder, folder_files in sorted(candidate_folders.items())
        ],
    }
    if not selected_candidates:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected=f"이미지 {min_images}개 이상 폴더 {required_folder_count}개",
            actual=f"후보 폴더 {len(candidate_folders)}개",
            message=config.get("folder_message") or "제품 스크린샷 폴더를 찾을 수 없음",
            raw_detail=raw_detail,
        )

    start_dt = _date_range_start(context.start_date)
    end_dt = _date_range_end(context.end_date)
    if not (start_dt and end_dt):
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected="{시작일} ~ {종료일} 기준정보",
            actual=f"{context.start_date or '(없음)'} ~ {context.end_date or '(없음)'}",
            message=config.get("date_message") or "제품 스크린샷 생성일이 시험기간과 다름",
            raw_detail={**raw_detail, "selected_parent": "/".join(selected_parent or ())},
        )

    selected_files = [file_info for _folder, folder_files in selected_candidates for file_info in folder_files]
    out_of_range = [
        file_info
        for file_info in selected_files
        if not (file_info.modified_at and start_dt <= file_info.modified_at <= end_dt)
    ]
    selected_folders = [
        {"folder": "/".join(folder), "image_count": len(folder_files)}
        for folder, folder_files in selected_candidates
    ]
    raw_detail.update({
        "selected_parent": "/".join(selected_parent or ()),
        "selected_candidate_folders": selected_folders,
        "date_range": {"start": context.start_date, "end": context.end_date},
        "out_of_range_files": [
            {
                "path": _display_path(file_info.path, project.project_number),
                "modified_at": file_info.modified_at.isoformat() if file_info.modified_at else "",
            }
            for file_info in out_of_range[:20]
        ],
    })
    status = DownloadReviewRuleStatus.PASS if not out_of_range else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"이미지 수정일자 {context.start_date} ~ {context.end_date}",
        actual="범위 밖 파일 없음" if not out_of_range else f"범위 밖 {len(out_of_range)}개",
        message=(
            config.get("pass_message")
            if status == DownloadReviewRuleStatus.PASS
            else config.get("date_message")
        ) or ("제품 스크린샷을 확인했습니다." if status == DownloadReviewRuleStatus.PASS else "제품 스크린샷 생성일이 시험기간과 다름"),
        file_path="/".join((selected_parent or ())),
        file_name="",
        raw_detail=raw_detail,
    )


def _evaluate_rawdata_folder_structure_check(rule, sequence, project, verify_result):
    config = rule.config_json or {}
    folders, file_folders = _inspection_folder_tree(verify_result)
    base = _find_folder_by_keyword_chain(folders, config.get("folder_keyword_chain") or ["수행"])
    raw_detail = {"base_folder": "/".join(base or ())}
    if not base:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected="수행 폴더",
            actual="없음",
            message=config.get("base_message") or "수행 폴더 확인 불가",
            raw_detail=raw_detail,
        )

    checks = []
    passed = True
    first_message = ""
    for folder_check in config.get("folder_checks") or []:
        result = _run_folder_check(base, folders, file_folders, folder_check)
        checks.append(result)
        if not result["passed"] and passed:
            passed = False
            first_message = result["message"]

    raw_detail["checks"] = checks
    status = DownloadReviewRuleStatus.PASS if passed else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected="rawdata 폴더 구조 충족",
        actual="정상" if passed else first_message,
        message=config.get("pass_message") if passed else first_message,
        file_path="/".join(base),
        raw_detail=raw_detail,
    )


def _evaluate_test_report_document_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
    ]
    docx_files = [file_info for file_info in matched if file_info.extension.lower() == ".docx"]
    pdf_files = [file_info for file_info in matched if file_info.extension.lower() == ".pdf"]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
        "docx_count": len(docx_files),
        "pdf_count": len(pdf_files),
    }

    if len(docx_files) != 1 or len(pdf_files) != 1:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected=".docx 1개 / .pdf 1개",
            actual=f".docx {len(docx_files)}개 / .pdf {len(pdf_files)}개",
            message=config.get("missing_message") or "파일이 없습니다.",
            file_path=_representative_path(matched or files, project.project_number),
            file_name=_representative_name(matched or files),
            raw_detail=raw_detail,
        )

    docx_file = docx_files[0]
    pdf_file = pdf_files[0]
    try:
        rounds = _docx_defect_report_round_dates(docx_file)
        spec_table = _docx_first_table_after_text(docx_file, config.get("spec_marker") or "<세부사양>")
        artifact = _store_pdf_first_page_artifact(
            project,
            rule,
            pdf_file,
            artifact_id="pdf_first_page",
            label=config.get("pdf_artifact_label") or "시험성적서 1페이지",
        )
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="시험성적서 docx/pdf 파싱 가능",
            actual=str(exc),
            message=str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=_representative_name(matched),
            raw_detail=raw_detail,
        )

    raw_detail.update({
        "variables": {
            "결함차수": len(rounds),
            **rounds,
            "시험성적서_세부사양표": spec_table or [],
        },
        "spec_table": spec_table or [],
        "artifacts": [artifact],
    })
    if "1차" not in rounds or "2차" not in rounds:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected="결함리포트 송부 표에 1차/2차 보고일자",
            actual=", ".join(f"{key}: {value}" for key, value in rounds.items()) or "날짜 없음",
            message=config.get("round_date_message") or "결함리포트 송부 정보 확인 불가",
            file_path=_representative_path(matched, project.project_number),
            file_name=docx_file.name,
            raw_detail=raw_detail,
        )

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected=".docx 1개 / .pdf 1개 / 결함리포트 송부 날짜",
        actual=f".docx {docx_file.name} / .pdf {pdf_file.name} / 결함차수 {len(rounds)}",
        message=config.get("pass_message") or "시험성적서를 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _evaluate_defect_report_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _extension_matches(file_info.extension, [".xlsx", ".xls"])
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }

    defect_round_count = _context_int(context, "결함차수")
    if defect_round_count is None:
        return _defect_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="13번 시험성적서에서 {결함차수} 산출",
            actual="{결함차수} 없음",
            message=config.get("count_mismatch_message") or "시험성적서의 결함 차수와 결함리포트 개수가 다름",
        )

    expected_file_count = defect_round_count + 1
    raw_detail["expected_file_count"] = expected_file_count
    if len(matched) != expected_file_count:
        return _defect_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"결함리포트 Excel 파일 {expected_file_count}개",
            actual=f"결함리포트 Excel 파일 {len(matched)}개",
            message=config.get("count_mismatch_message") or "시험성적서의 결함 차수와 결함리포트 개수가 다름",
        )

    versioned_files = _defect_report_versioned_files(matched, config)
    raw_detail["versions"] = {
        str(version): _display_path(file_info.path, project.project_number)
        for version, file_info in versioned_files.items()
    }
    expected_versions = set(range(1, expected_file_count + 1))
    if set(versioned_files) != expected_versions:
        return _defect_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="버전 " + ", ".join(f"v{version}.0" for version in sorted(expected_versions)),
            actual="버전 " + ", ".join(f"v{version}.0" for version in sorted(versioned_files)) if versioned_files else "버전 없음",
            message=config.get("filename_message") or "결함리포트 파일명이 잘못됨",
        )

    workbook_by_version = {}
    for version, file_info in versioned_files.items():
        try:
            workbook_by_version[version] = _read_excel_workbook(file_info)
        except DownloadReviewInspectionError as exc:
            return RuleEvaluation(
                rule=rule,
                sequence=sequence,
                status=DownloadReviewRuleStatus.ERROR,
                expected="결함리포트 Excel 파일 파싱 가능",
                actual=f"{file_info.name}: {exc}",
                message=str(exc),
                file_path=_representative_path(matched, project.project_number),
                file_name=file_info.name,
                raw_detail=raw_detail,
            )

    sheet_check = _check_defect_report_sheets(workbook_by_version, versioned_files, defect_round_count)
    raw_detail["sheet_checks"] = sheet_check["details"]
    if not sheet_check["passed"]:
        return _defect_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=sheet_check["expected"],
            actual=sheet_check["actual"],
            message=_format_config_message(config.get("sheet_message"), file_name=sheet_check["file_name"]) or f"{sheet_check['file_name']}에 시트가 잘못 작성됨",
        )

    environment_check = _check_defect_report_environment(workbook_by_version)
    raw_detail["environment_check"] = environment_check
    if not environment_check["passed"]:
        return _defect_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="모든 시트 1~3행 시험환경 값 동일",
            actual=environment_check["actual"],
            message=config.get("environment_message") or "시험환경 정보 잘못 작성됨",
        )

    report_date_check = _check_defect_report_dates(workbook_by_version, context, defect_round_count)
    raw_detail["report_date_checks"] = report_date_check["details"]
    if not report_date_check["passed"]:
        return _defect_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=report_date_check["expected"],
            actual=report_date_check["actual"],
            message=config.get("report_date_message") or "프로젝트 번호, 결함 차시, 보고일자 중 잘못된 값이 작성됨",
        )

    final_workbook = workbook_by_version[expected_file_count]
    variables = {
        "잔여결함수": _defect_residual_count(final_workbook),
        "H": _defect_analysis_value(final_workbook, "High", offset_rows=0, offset_cols=1),
        "R": _defect_analysis_value(final_workbook, "수정전", offset_rows=5, offset_cols=0),
    }
    raw_detail["variables"] = variables

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected=f"결함리포트 파일 {expected_file_count}개 / v1.0~v{expected_file_count}.0 / 시트와 보고일자 정상",
        actual=f"결함차수 {defect_round_count} / 잔여결함수 {variables['잔여결함수']} / H {variables['H']} / R {variables['R']}",
        message=config.get("pass_message") or "결함리포트를 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _defect_report_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _defect_report_versioned_files(files, config):
    pattern = re.compile(str(config.get("version_pattern") or r"(?i)v(\d+)\.0"))
    versioned = {}
    duplicates = set()
    for file_info in files:
        match = pattern.search(file_info.name)
        if not match:
            continue
        version = int(match.group(1))
        if version in versioned:
            duplicates.add(version)
        versioned[version] = file_info
    if duplicates:
        versioned[-1] = files[0]
    return versioned


def _check_defect_report_sheets(workbook_by_version, versioned_files, defect_round_count):
    details = []
    final_version = defect_round_count + 1
    for version in sorted(workbook_by_version):
        workbook = workbook_by_version[version]
        actual_names = [sheet.name for sheet in workbook.sheets]
        expected_names = _expected_defect_report_sheet_names(version, final_version)
        missing = [name for name in expected_names if name not in actual_names]
        extra = [name for name in actual_names if name not in expected_names]
        detail = {
            "version": version,
            "file_name": versioned_files[version].name,
            "expected_sheets": expected_names,
            "actual_sheets": actual_names,
            "missing_sheets": missing,
            "extra_sheets": extra,
        }
        details.append(detail)
        if missing or extra:
            return {
                "passed": False,
                "file_name": versioned_files[version].name,
                "expected": ", ".join(expected_names),
                "actual": ", ".join(actual_names) or "시트 없음",
                "details": details,
            }
    return {"passed": True, "details": details}


def _expected_defect_report_sheet_names(version, final_version):
    if version == final_version:
        return [
            *[f"{round_no}차 결함리포트" for round_no in range(1, final_version)],
            "최종결함리포트",
            "시험분석자료",
        ]
    return [f"{round_no}차 결함리포트" for round_no in range(1, version + 1)]


def _check_defect_report_environment(workbook_by_version):
    values = []
    for version, workbook in sorted(workbook_by_version.items()):
        for sheet in workbook.sheets:
            value = _sheet_top_rows_cell_containing(sheet, "시험환경")
            if not value:
                return {
                    "passed": False,
                    "actual": f"v{version}.0 {sheet.name}: 시험환경 없음",
                    "values": values,
                }
            values.append({"version": version, "sheet": sheet.name, "value": value})

    normalized = {_normalize_compare(value["value"], {"remove_whitespace": True}) for value in values}
    if len(normalized) != 1:
        return {
            "passed": False,
            "actual": "; ".join(f"v{item['version']}.0 {item['sheet']}: {item['value']}" for item in values),
            "values": values,
        }
    return {"passed": True, "actual": values[0]["value"] if values else "", "values": values}


def _check_defect_report_dates(workbook_by_version, context, defect_round_count):
    details = []
    final_version = defect_round_count + 1
    expected_dates = {
        f"{round_no}차 결함리포트": _context_variable(context, f"{round_no}차")
        for round_no in range(1, defect_round_count + 1)
    }
    expected_dates["최종결함리포트"] = context.end_date
    expected_dates["시험분석자료"] = context.end_date

    for version, workbook in sorted(workbook_by_version.items()):
        required_sheets = _expected_defect_report_sheet_names(version, final_version)
        for sheet_name in required_sheets:
            sheet = _workbook_sheet(workbook, sheet_name)
            actual_text = _sheet_top_rows_cell_containing(sheet, f"{context.project_number} {sheet_name}") if sheet else ""
            report_date = _extract_korean_report_date(actual_text)
            expected_date = expected_dates.get(sheet_name, "")
            passed = bool(
                actual_text
                and context.project_number in actual_text
                and sheet_name in actual_text
                and _same_date_text(report_date, expected_date)
            )
            detail = {
                "version": version,
                "sheet": sheet_name,
                "expected_date": expected_date,
                "actual_text": actual_text,
                "actual_date": report_date,
                "passed": passed,
            }
            details.append(detail)
            if not passed:
                return {
                    "passed": False,
                    "expected": f"{context.project_number} {sheet_name} / 보고일자 {expected_date}",
                    "actual": actual_text or f"{sheet_name} 상단 문구 없음",
                    "details": details,
                }
    return {"passed": True, "details": details}


def _sheet_top_rows_cell_containing(sheet, keyword):
    for row in sheet.rows[:3]:
        for value in row:
            if keyword in value:
                return value
    return ""


def _workbook_sheet(workbook, sheet_name):
    for sheet in workbook.sheets:
        if sheet.name == sheet_name:
            return sheet
    return None


def _extract_korean_report_date(value):
    match = re.search(r"보고일자\s*[:：]\s*(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일", str(value or ""))
    if not match:
        return ""
    year, month, day = match.groups()
    return f"{int(year):04d}.{int(month):02d}.{int(day):02d}."


def _same_date_text(left, right):
    return _format_dot_date(left) == _format_dot_date(right)


def _defect_residual_count(workbook):
    sheet = _workbook_sheet(workbook, "최종결함리포트")
    if not sheet:
        return 0
    count = 0
    for row in sheet.rows[4:]:
        value = row[1] if len(row) > 1 else ""
        if value:
            count += 1
    return count


def _defect_analysis_value(workbook, keyword, *, offset_rows, offset_cols):
    sheet = _workbook_sheet(workbook, "시험분석자료")
    if not sheet:
        return ""
    for row_index, row in enumerate(sheet.rows):
        for col_index, value in enumerate(row[2:6], start=2):
            if keyword not in value:
                continue
            target_row = row_index + offset_rows
            target_col = col_index + offset_cols
            if target_row < len(sheet.rows) and target_col < len(sheet.rows[target_row]):
                return sheet.rows[target_row][target_col]
            return ""
    return ""


def _context_int(context, key):
    value = _context_variable(context, key)
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _context_variable(context, key):
    return context.derived_variables.get(key, "")


def _format_config_message(template, **values):
    if not template:
        return ""
    message = str(template)
    for key, value in values.items():
        message = message.replace("{" + key + "}", str(value))
    return message


def _evaluate_inspection_checklist_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
    ]
    excel_files = [
        file_info
        for file_info in matched
        if _extension_matches(file_info.extension, [".xlsx", ".xls"])
    ]
    pdf_files = [
        file_info
        for file_info in matched
        if _extension_matches(file_info.extension, [".pdf"])
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
        "excel_count": len(excel_files),
        "pdf_count": len(pdf_files),
    }

    if len(excel_files) != 1:
        return _checklist_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="점검표 Excel 파일 1개",
            actual=f"점검표 Excel 파일 {len(excel_files)}개",
            message=config.get("missing_message") or "파일이 존재하지 않음",
        )
    if len(pdf_files) != 1:
        return _checklist_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="점검표 PDF 파일 1개",
            actual=f"점검표 PDF 파일 {len(pdf_files)}개",
            message=config.get("pdf_missing_message") or "점검표 pdf 파일이 없음",
        )

    excel_file = excel_files[0]
    pdf_file = pdf_files[0]
    try:
        workbook = _read_excel_workbook(excel_file)
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="점검표 Excel 파일 파싱 가능",
            actual=str(exc),
            message=str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=excel_file.name,
            raw_detail=raw_detail,
        )

    raw_detail["sheet_names"] = [sheet.name for sheet in workbook.sheets]

    header_check = _check_checklist_headers(workbook, context.project_number)
    raw_detail["header_check"] = header_check
    if not header_check["passed"]:
        return _checklist_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"모든 시트 머리글에 프로젝트번호: {context.project_number}",
            actual=header_check["actual"],
            message=config.get("header_message") or "머리글(프로젝트번호)이 잘못 작성됨",
        )

    cover_sheet = _workbook_sheet(workbook, config.get("cover_sheet") or "표지")
    if not cover_sheet:
        return _checklist_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="표지 시트",
            actual="표지 시트 없음",
            message=config.get("cover_title_message") or "표지 제목이 잘못 작성됨",
        )

    cover_check = _check_checklist_cover(cover_sheet, context, config)
    raw_detail["cover_check"] = cover_check
    if not cover_check["passed"]:
        return _checklist_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=cover_check["expected"],
            actual=cover_check["actual"],
            message=cover_check["message"],
        )

    feature_sheet = _workbook_sheet(workbook, config.get("feature_sheet") or "기능별 점검표")
    suitability_sheet = _workbook_sheet(workbook, config.get("suitability_sheet") or "2. 기능적합성")
    reliability_sheet = _workbook_sheet(workbook, config.get("reliability_sheet") or "6. 신뢰성")
    score_sheet = _workbook_sheet(workbook, config.get("score_sheet") or "측정항목별 점수표")

    feature_check = _check_checklist_feature_sheet(feature_sheet)
    raw_detail["feature_sheet_check"] = feature_check
    if not feature_check["passed"]:
        return _checklist_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=feature_check["expected"],
            actual=feature_check["actual"],
            message=config.get("feature_blank_message") or "기능별 점검표 시트에 빈 셀이 확인됨",
        )

    suitability_compare = _check_checklist_suitability_table(feature_sheet, suitability_sheet)
    raw_detail["suitability_compare"] = suitability_compare
    if not suitability_compare["passed"]:
        return _checklist_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=suitability_compare["expected"],
            actual=suitability_compare["actual"],
            message=config.get("suitability_compare_message") or "기능적합성 시트의 2.3번 기능표 내용 확인 필요함",
        )

    suitability_result = _check_checklist_suitability_results(suitability_sheet)
    raw_detail["suitability_result_check"] = suitability_result
    if not suitability_result["passed"]:
        return _checklist_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=suitability_result["expected"],
            actual=suitability_result["actual"],
            message=config.get("suitability_result_message") or "기능적합성 시트의 기능표 결과값 미작성",
        )

    reliability_check = _check_checklist_reliability(reliability_sheet, context)
    raw_detail["reliability_check"] = reliability_check
    if not reliability_check["passed"]:
        return _checklist_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=reliability_check["expected"],
            actual=reliability_check["actual"],
            message=reliability_check["message"],
        )

    score_values = _checklist_score_values(score_sheet)
    raw_detail["variables"] = {"측정항목별점수표": score_values}
    try:
        artifact = _store_pdf_first_page_artifact(
            project,
            rule,
            pdf_file,
            artifact_id=config.get("pdf_artifact_id") or "pdf_first_page",
            label=config.get("pdf_artifact_label") or "점검표 1페이지",
        )
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="점검표 PDF 1페이지 캡처 가능",
            actual=str(exc),
            message=str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=pdf_file.name,
            raw_detail=raw_detail,
        )
    raw_detail["artifacts"] = [artifact]

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected="점검표 Excel/PDF, 머리글, 표지, 기능표, 신뢰성 값 정상",
        actual=f"{excel_file.name} / {pdf_file.name} / 측정항목별점수표 {len(score_values)}개",
        message=config.get("pass_message") or "점검표를 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name([excel_file, pdf_file]),
        raw_detail=raw_detail,
    )


def _checklist_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _check_checklist_headers(workbook, project_number):
    expected = f"프로젝트번호: {project_number}"
    failures = []
    details = []
    for sheet in workbook.sheets:
        candidates = [sheet.header_text, *_top_rows_texts(sheet, limit=3)]
        passed = any(_clean_excel_header_text(candidate).find(expected) >= 0 for candidate in candidates)
        details.append({
            "sheet": sheet.name,
            "header_text": sheet.header_text,
            "passed": passed,
        })
        if not passed:
            failures.append(sheet.name)
    return {
        "passed": not failures,
        "expected": expected,
        "actual": "오류 시트: " + ", ".join(failures) if failures else "정상",
        "details": details,
    }


def _check_checklist_cover(sheet, context, config):
    title_cell = _find_cell_with_all(sheet.rows, [context.project_number, "점검표"])
    if not title_cell:
        return {
            "passed": False,
            "expected": f"{context.project_number} 및 점검표 포함 셀",
            "actual": "일치 셀 없음",
            "message": config.get("cover_title_message") or "표지 제목이 잘못 작성됨",
        }

    period = f"{context.start_date} ~ {context.end_date}"
    # 표지 날짜는 yyyy-mm-dd, yyyy.mm.dd. 등 형식이 섞일 수 있으므로 날짜를
    # 정규화해 {시작일}/{종료일}과 같은 날짜인지로 비교한다. (다른 날짜 검사와 일관)
    date_cell = _find_cell_with_date_range(sheet.rows, context.start_date, context.end_date)
    if not date_cell:
        return {
            "passed": False,
            "expected": period,
            "actual": "일치 셀 없음",
            "message": config.get("cover_date_message") or "표지 날짜가 잘못 작성됨",
        }

    author_expected = str(config.get("cover_author") or "김진영")
    # 검토자(김진영)와 작성자({PL})는 보통 다른 행/셀에 적히므로 각각 다른 셀에서 찾는다.
    reviewer_cell = _find_cell_normalized_contains_all(sheet.rows, [author_expected])
    pl_cell = _find_cell_normalized_contains_all(sheet.rows, [context.pl]) if context.pl else None
    if not reviewer_cell or not pl_cell:
        return {
            "passed": False,
            "expected": f"{author_expected}, {context.pl}",
            "actual": (
                f"검토자({author_expected})={'있음' if reviewer_cell else '없음'} / "
                f"작성자({context.pl})={'있음' if pl_cell else '없음'}"
            ),
            "message": config.get("cover_author_message") or "표지 작성자가 잘못 작성됨",
        }

    return {
        "passed": True,
        "expected": f"{context.project_number} 점검표 / {period} / {author_expected}, {context.pl}",
        "actual": {
            "title_cell": title_cell,
            "date_cell": date_cell,
            "reviewer_cell": reviewer_cell,
            "author_cell": pl_cell,
        },
    }


def _check_checklist_feature_sheet(sheet):
    if not sheet:
        return {
            "passed": False,
            "expected": "기능별 점검표 시트",
            "actual": "시트 없음",
        }
    last_row = _numeric_sequence_last_row(sheet, start_row=8, column=1)
    if not last_row:
        return {
            "passed": False,
            "expected": "A8부터 연속 숫자",
            "actual": "A8 숫자 없음",
        }
    blank_cells = []
    for row in range(8, last_row + 1):
        for column in range(5, 35):
            if not _sheet_cell(sheet, row, column):
                blank_cells.append(f"{_excel_column_name(column)}{row}")
    return {
        "passed": not blank_cells,
        "expected": f"E8:AI{last_row} 모든 셀 입력",
        "actual": "빈 셀 없음" if not blank_cells else ", ".join(blank_cells[:20]),
        "last_row": last_row,
        "blank_cells": blank_cells,
    }


def _check_checklist_suitability_table(feature_sheet, suitability_sheet):
    if not (feature_sheet and suitability_sheet):
        return {
            "passed": False,
            "expected": "기능별 점검표/2. 기능적합성 시트",
            "actual": "필수 시트 없음",
        }
    last_row = _last_non_empty_row_in_column(suitability_sheet, column=3, start_row=16)
    if not last_row:
        return {
            "passed": False,
            "expected": "2. 기능적합성 C16부터 기능표",
            "actual": "비교 대상 없음",
        }
    row_count = last_row - 16 + 1
    suitability_values = _sheet_range_values(suitability_sheet, 16, 1, last_row, 3)
    feature_values = _sheet_range_values(feature_sheet, 8, 2, 8 + row_count - 1, 4)
    mismatches = _matrix_mismatches(feature_values, suitability_values, left_origin=(8, 2), right_origin=(16, 1))
    return {
        "passed": not mismatches,
        "expected": f"기능별 점검표 B8:D{8 + row_count - 1}",
        "actual": f"2. 기능적합성 A16:C{last_row}" if not mismatches else str(mismatches[0]),
        "last_row": last_row,
        "mismatches": mismatches,
    }


def _check_checklist_suitability_results(sheet):
    if not sheet:
        return {
            "passed": False,
            "expected": "2. 기능적합성 시트",
            "actual": "시트 없음",
        }
    last_row = _last_non_empty_row_in_column(sheet, column=3, start_row=16)
    if not last_row:
        return {
            "passed": False,
            "expected": "D16부터 결과값",
            "actual": "비교 대상 없음",
        }
    blank_cells = [
        f"D{row}"
        for row in range(16, last_row + 1)
        if not _sheet_cell(sheet, row, 4)
    ]
    return {
        "passed": not blank_cells,
        "expected": f"D16:D{last_row} 모든 셀 입력",
        "actual": "빈 셀 없음" if not blank_cells else ", ".join(blank_cells[:20]),
        "blank_cells": blank_cells,
    }


def _check_checklist_reliability(sheet, context):
    if not sheet:
        return {
            "passed": False,
            "expected": "6. 신뢰성 시트",
            "actual": "시트 없음",
            "message": "신뢰성 시트의 WD 잘못 작성함",
        }

    wd_actual = _sheet_cell(sheet, 5, 3)
    if not _same_excel_text(wd_actual, context.wd):
        return {
            "passed": False,
            "expected": f"C5={context.wd}",
            "actual": f"C5={wd_actual}",
            "message": "신뢰성 시트의 WD 잘못 작성함",
        }

    high_expected = _variable_to_text(_context_variable(context, "H"))
    before_expected = _variable_to_text(_context_variable(context, "R"))
    high_actual = _sheet_cell(sheet, 11, 3)
    before_actual = _sheet_cell(sheet, 11, 5)
    if not (_same_excel_text(high_actual, high_expected) and _same_excel_text(before_actual, before_expected)):
        return {
            "passed": False,
            "expected": f"C11={high_expected}, E11={before_expected}",
            "actual": f"C11={high_actual}, E11={before_actual}",
            "message": "신뢰성 시트의 결함 개수가 잘못 작성됨",
        }

    return {
        "passed": True,
        "expected": f"C5={context.wd}, C11={high_expected}, E11={before_expected}",
        "actual": f"C5={wd_actual}, C11={high_actual}, E11={before_actual}",
    }


def _checklist_score_values(sheet):
    if not sheet:
        return []
    return [_sheet_cell(sheet, row, 4) for row in range(7, 91)]


def _top_rows_texts(sheet, *, limit):
    values = []
    for row in sheet.rows[:limit]:
        text = " ".join(value for value in row if value)
        if text:
            values.append(text)
    return values


def _clean_excel_header_text(value):
    text = str(value or "")
    text = re.sub(r"&\"[^\"]+\"", "", text)
    text = re.sub(r"&[A-Za-z0-9]+", "", text)
    return _normalize_spaces(text)


def _find_cell_with_all(rows, keywords):
    needles = [str(keyword or "") for keyword in keywords if str(keyword or "")]
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            if all(needle in value for needle in needles):
                return {"row": row_index, "column": col_index, "value": value}
    return None


def _find_cell_normalized_contains_all(rows, keywords):
    needles = [_normalize_no_space(keyword) for keyword in keywords if str(keyword or "")]
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            normalized = _normalize_no_space(value)
            if all(needle in normalized for needle in needles):
                return {"row": row_index, "column": col_index, "value": value}
    return None


def _numeric_sequence_last_row(sheet, *, start_row, column):
    # start_row부터 1,2,3...로 이어지는 연속 숫자의 마지막 행을 찾는다.
    # 번호가 끝난 뒤 합계/요약 같은 비숫자 행이 와도 마지막 숫자 행을 반환한다.
    # (start_row 셀이 "1"이 아니면 0을 반환해 호출부가 실패로 처리한다.)
    expected = 1
    last_row = 0
    row = start_row
    while True:
        value = _sheet_cell(sheet, row, column)
        if value != str(expected):
            break
        last_row = row
        expected += 1
        row += 1
    return last_row


def _last_non_empty_row_in_column(sheet, *, column, start_row=1):
    last_row = 0
    for row in range(start_row, len(sheet.rows) + 1):
        if _sheet_cell(sheet, row, column):
            last_row = row
    return last_row


def _sheet_cell(sheet, row, column):
    row_index = row - 1
    col_index = column - 1
    if not sheet or row_index < 0 or col_index < 0 or row_index >= len(sheet.rows):
        return ""
    row_values = sheet.rows[row_index]
    if col_index >= len(row_values):
        return ""
    return row_values[col_index]


def _sheet_range_values(sheet, start_row, start_col, end_row, end_col):
    return [
        [_sheet_cell(sheet, row, col) for col in range(start_col, end_col + 1)]
        for row in range(start_row, end_row + 1)
    ]


def _matrix_mismatches(left, right, *, left_origin, right_origin):
    mismatches = []
    row_count = max(len(left), len(right))
    for row_index in range(row_count):
        left_row = left[row_index] if row_index < len(left) else []
        right_row = right[row_index] if row_index < len(right) else []
        col_count = max(len(left_row), len(right_row))
        for col_index in range(col_count):
            left_value = left_row[col_index] if col_index < len(left_row) else ""
            right_value = right_row[col_index] if col_index < len(right_row) else ""
            if left_value != right_value:
                mismatches.append({
                    "left_cell": f"{_excel_column_name(left_origin[1] + col_index)}{left_origin[0] + row_index}",
                    "right_cell": f"{_excel_column_name(right_origin[1] + col_index)}{right_origin[0] + row_index}",
                    "left": left_value,
                    "right": right_value,
                })
    return mismatches


def _same_excel_text(left, right):
    return _normalize_spaces(left) == _normalize_spaces(right)


def _normalize_no_space(value):
    return re.sub(r"\s+", "", str(value or ""))


def _evaluate_quality_inspection_table_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _extension_matches(file_info.extension, [".xlsx", ".xls"])
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }

    if len(matched) != 1:
        return _quality_table_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="품질검사표 Excel 파일 1개",
            actual=f"품질검사표 Excel 파일 {len(matched)}개",
            message=config.get("missing_message") or "품질검사표 파일 확인 불가",
        )

    file_info = matched[0]
    try:
        workbook = _read_excel_workbook(file_info)
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="품질검사표 Excel 파일 파싱 가능",
            actual=str(exc),
            message=str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=raw_detail,
        )

    expected_sheet_name = _resolve_rule_value(config.get("sheet_name") or "{프로젝트번호} 품질검사표", context)
    actual_sheet_names = [sheet.name for sheet in workbook.sheets]
    raw_detail["sheet_names"] = actual_sheet_names
    if actual_sheet_names != [expected_sheet_name]:
        return _quality_table_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"시트 1개: {expected_sheet_name}",
            actual=", ".join(actual_sheet_names) or "시트 없음",
            message=config.get("sheet_message") or "품질검사표 시트명 확인 필요",
        )

    sheet = workbook.sheets[0]
    expected_scores = _context_variable(context, "측정항목별점수표")
    if not isinstance(expected_scores, list):
        return _quality_table_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="11번 점검표 산출 변수 {측정항목별점수표}",
            actual="{측정항목별점수표} 없음",
            message=config.get("score_message") or "측정항목별 점수표가 점검표와 상이함",
        )

    actual_scores = [_sheet_cell(sheet, row, 4) for row in range(4, 88)]
    raw_detail["score_compare"] = {
        "expected_count": len(expected_scores),
        "actual_count": len(actual_scores),
        "mismatches": _list_mismatches(expected_scores, actual_scores, start_index=4),
    }
    if raw_detail["score_compare"]["mismatches"]:
        return _quality_table_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="11번 점검표 D7:D90 값과 품질검사표 D4:D87 값 동일",
            actual=str(raw_detail["score_compare"]["mismatches"][0]),
            message=config.get("score_message") or "측정항목별 점수표가 점검표와 상이함",
        )

    quality_values_raw = [_sheet_cell(sheet, row, 5) for row in range(4, 86)]
    quality_values = [value for value in quality_values_raw if value]
    raw_detail["quality_sub_characteristic_values"] = {
        "source_range": "E4:E85",
        "raw_value_count": len(quality_values),
        "raw_values": quality_values,
    }
    expected_quality_count = int(config.get("quality_value_count") or 33)
    if len(quality_values) != expected_quality_count:
        return _quality_table_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"E4:E85 실제 값 {expected_quality_count}개",
            actual=f"E4:E85 실제 값 {len(quality_values)}개",
            message=config.get("quality_value_message") or "품질검사표의 품질부특성 측정값 확인 필요",
        )

    rotated_values = quality_values[3:] + quality_values[:3]
    raw_detail["quality_sub_characteristic_values"]["rotated_values"] = rotated_values
    raw_detail["variables"] = {
        "품질부특성측정값": rotated_values,
    }

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected=f"{expected_sheet_name} 단일 시트 / D4:D87 점수표 일치 / 품질부특성측정값 {expected_quality_count}개",
        actual=f"{file_info.name} / 품질부특성측정값 {len(rotated_values)}개",
        message=config.get("pass_message") or "품질검사표를 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=file_info.name,
        raw_detail=raw_detail,
    )


def _quality_table_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _list_mismatches(expected_values, actual_values, *, start_index):
    mismatches = []
    max_length = max(len(expected_values), len(actual_values))
    for index in range(max_length):
        expected = _variable_to_text(expected_values[index]) if index < len(expected_values) else ""
        actual = _variable_to_text(actual_values[index]) if index < len(actual_values) else ""
        if expected != actual:
            mismatches.append({
                "row": start_index + index,
                "expected": expected,
                "actual": actual,
            })
    return mismatches


def _evaluate_quality_evaluation_report_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _extension_matches(file_info.extension, [".docx"])
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }

    if len(matched) != 1:
        return _quality_report_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="품질평가보고서 docx 파일 1개",
            actual=f"품질평가보고서 docx 파일 {len(matched)}개",
            message=config.get("missing_message") or "품질평가보고서 파일 확인 불가",
        )

    file_info = matched[0]
    try:
        text = _docx_all_text(file_info)
        tables = _docx_tables(file_info)
    except Exception as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="품질평가보고서 docx 파싱 가능",
            actual=str(exc),
            message="품질평가보고서 파일 확인 불가",
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=raw_detail,
        )

    project_count = text.count(context.project_number)
    expected_count = int(config.get("project_number_count") or 6)
    raw_detail["project_number_count"] = project_count
    if project_count != expected_count:
        return _quality_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"{context.project_number} {expected_count}회",
            actual=f"{context.project_number} {project_count}회",
            message=config.get("project_number_message") or "프로젝트 번호 확인 필요",
        )

    signature_check = _quality_report_signature_check(text, config)
    raw_detail["signature_check"] = signature_check
    if not signature_check["passed"]:
        return _quality_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=signature_check["expected"],
            actual=signature_check["actual"],
            message=config.get("signature_message") or "서명란 이름 확인 필요",
        )

    table_rows = [row for table in tables for row in table]
    company_value = _find_next_cell_by_label(table_rows, str(config.get("company_label") or "회사(기관)명"))
    raw_detail["company_check"] = {"expected": context.company, "actual": company_value}
    if not _same_excel_text(company_value, context.company):
        return _quality_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"회사명 {context.company}",
            actual=company_value or "회사명 값 없음",
            message=config.get("company_message") or "1. 신청 회사 현황 표 값 확인 필요",
        )

    date_checks = [
        ("request_date", "신청일자", context.request_date, config.get("request_date_message") or "신청일자가 잘못 작성됨"),
        ("contract_date", "계약일자", context.contract_date, config.get("contract_date_message") or "계약일자가 잘못 작성됨"),
        ("committee_date", "품질인증심의위원회", context.certification_committee_date, config.get("committee_date_message") or "인증위 날짜가 잘못 작성됨"),
    ]
    raw_detail["date_checks"] = []
    for key, label, expected_date, message in date_checks:
        actual_date = _extract_labeled_korean_date(text, label)
        detail = {
            "key": key,
            "label": label,
            "expected": expected_date,
            "actual": actual_date,
        }
        raw_detail["date_checks"].append(detail)
        if not _same_date_text(actual_date, expected_date):
            return _quality_report_failure(
                rule,
                sequence,
                matched,
                project,
                raw_detail,
                expected=f"{label}: {expected_date}",
                actual=f"{label}: {actual_date or '날짜 없음'}",
                message=message,
            )

    period = _extract_labeled_korean_period(text, "제품시험평가")
    raw_detail["period_check"] = {
        "expected_start": context.start_date,
        "expected_end": context.end_date,
        "actual_start": period[0],
        "actual_end": period[1],
    }
    if not (_same_date_text(period[0], context.start_date) and _same_date_text(period[1], context.end_date)):
        return _quality_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"제품시험평가: {context.start_date} ~ {context.end_date}",
            actual=f"제품시험평가: {period[0] or '시작일 없음'} ~ {period[1] or '종료일 없음'}",
            message=config.get("period_message") or "시험기간이 잘못 작성됨",
        )

    quality_table = _docx_first_table_after_text(file_info, config.get("quality_marker") or "<품질특성별 세부 평가결과>")
    quality_check = _quality_report_table_check(quality_table, context)
    raw_detail["quality_value_check"] = quality_check
    if not quality_check["passed"]:
        return _quality_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=quality_check["expected"],
            actual=quality_check["actual"],
            message=(
                (config.get("na_message") or "NA 해당사항 없음 작성 오류")
                if quality_check.get("na_error")
                else config.get("quality_value_message") or "품질검사표의 품질부특성 측정값과 상이함"
            ),
        )

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected="프로젝트번호 6회 / 서명 / 회사명 / 날짜 / 품질부특성 측정값 정상",
        actual=f"{file_info.name} / 품질부특성 측정값 {len(quality_check['actual_values'])}개",
        message=config.get("pass_message") or "품질평가보고서를 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=file_info.name,
        raw_detail=raw_detail,
    )


def _quality_report_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _quality_report_signature_check(text, config):
    primary = str(config.get("primary_signer") or "성  명 : 김  성  희")
    secondary = str(config.get("secondary_signer") or "정  성  룡     (서명)")
    normalized_text = _normalize_no_space(text)
    primary_ok = _normalize_no_space(primary) in normalized_text
    secondary_ok = _normalize_no_space(secondary) in normalized_text
    return {
        "passed": primary_ok and secondary_ok,
        "expected": f"{primary} / {secondary}",
        "actual": f"primary={'있음' if primary_ok else '없음'}, secondary={'있음' if secondary_ok else '없음'}",
    }


def _extract_labeled_korean_date(text, label):
    pattern = re.compile(
        re.escape(label)
        + r"\s*[:：]\s*(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일"
    )
    match = pattern.search(str(text or ""))
    if not match:
        return ""
    year, month, day = match.groups()
    return f"{int(year):04d}.{int(month):02d}.{int(day):02d}."


def _extract_labeled_korean_period(text, label):
    date_pattern = r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일"
    pattern = re.compile(
        re.escape(label)
        + r"\s*[:：]\s*"
        + date_pattern
        + r"\s*[~\-]\s*"
        + date_pattern
    )
    match = pattern.search(str(text or ""))
    if not match:
        return "", ""
    sy, sm, sd, ey, em, ed = match.groups()
    return (
        f"{int(sy):04d}.{int(sm):02d}.{int(sd):02d}.",
        f"{int(ey):04d}.{int(em):02d}.{int(ed):02d}.",
    )


def _quality_report_table_check(table, context):
    expected_values = _context_variable(context, "품질부특성측정값")
    if not isinstance(expected_values, list):
        return {
            "passed": False,
            "expected": "16번 품질검사표 산출 변수 {품질부특성측정값}",
            "actual": "{품질부특성측정값} 없음",
            "actual_values": [],
        }
    if not table:
        return {
            "passed": False,
            "expected": "<품질특성별 세부 평가결과> 다음 표",
            "actual": "표 없음",
            "actual_values": [],
        }
    header_value = _table_cell(table, 1, 3)
    if "평가결과" not in header_value:
        return {
            "passed": False,
            "expected": "1행 3열 평가결과",
            "actual": header_value or "1행 3열 값 없음",
            "actual_values": [],
        }
    actual_values = _trim_trailing_empty_values([
        _table_cell(table, row, 3)
        for row in range(2, len(table) + 1)
    ])
    mismatches = _list_mismatches(expected_values, actual_values, start_index=2)
    if mismatches:
        return {
            "passed": False,
            "expected": "품질검사표 품질부특성 측정값과 동일",
            "actual": str(mismatches[0]),
            "actual_values": actual_values,
            "mismatches": mismatches,
        }
    na_errors = []
    for offset, value in enumerate(actual_values, start=2):
        if _is_na_value(value) and _table_cell(table, offset, 4) != "해당사항 없음":
            na_errors.append({
                "row": offset,
                "value": value,
                "right_cell": _table_cell(table, offset, 4),
            })
    if na_errors:
        return {
            "passed": False,
            "expected": "NA/N/A 오른쪽 칸 해당사항 없음",
            "actual": str(na_errors[0]),
            "actual_values": actual_values,
            "na_errors": na_errors,
            "na_error": True,
        }
    return {
        "passed": True,
        "expected": f"품질부특성 측정값 {len(expected_values)}개",
        "actual": f"품질부특성 측정값 {len(actual_values)}개",
        "actual_values": actual_values,
    }


def _table_cell(table, row, column):
    row_index = row - 1
    col_index = column - 1
    if row_index < 0 or col_index < 0 or row_index >= len(table):
        return ""
    row_values = table[row_index]
    if col_index >= len(row_values):
        return ""
    return row_values[col_index]


def _is_na_value(value):
    normalized = str(value or "").strip().upper()
    return normalized in {"NA", "N/A"}


def _trim_trailing_empty_values(values):
    trimmed = list(values)
    while trimmed and not trimmed[-1]:
        trimmed.pop()
    return trimmed


def _matching_files(rule, verify_result, *, ignore_target_file_type=False):
    files = _inspection_files(verify_result)
    pattern = (rule.target_file_pattern or "").strip()
    file_type = (rule.target_file_type or "any").strip().lower()

    if pattern:
        files = [file_info for file_info in files if fnmatch.fnmatch(file_info.name, pattern)]

    if not ignore_target_file_type and file_type and file_type != "any":
        extension = _extension_from_file_type(file_type)
        if extension:
            files = [file_info for file_info in files if file_info.extension.lower() == extension]

    return files


def _inspection_files(verify_result):
    files = list(verify_result.files or [])
    expanded_files = list(files)
    for file_info in files:
        if file_info.extension.lower() != ".zip":
            continue
        try:
            expanded_files.extend(_zip_entry_files(file_info))
        except (BadZipFile, OSError) as exc:
            raise DownloadReviewInspectionError(
                f"zip 파일을 읽을 수 없습니다: {file_info.name}"
            ) from exc
    return expanded_files


def _zip_entry_files(zip_file_info):
    entries = []
    with ZipFile(zip_file_info.path) as zip_file:
        for entry in zip_file.infolist():
            if entry.is_dir():
                continue
            inner_path = entry.filename.replace("\\", "/")
            inner_name = PurePosixPath(inner_path).name
            entries.append(
                FileInfo(
                    name=inner_name,
                    path=f"{zip_file_info.path}::{inner_path}",
                    size=entry.file_size,
                    extension=PurePosixPath(inner_name).suffix.lower(),
                    modified_at=datetime(*entry.date_time),
                )
            )
    return entries


def _read_file_bytes(file_info):
    raw_path = str(file_info.path or "")
    if "::" not in raw_path:
        return Path(raw_path).read_bytes()

    zip_path, inner_path = raw_path.split("::", 1)
    with ZipFile(zip_path) as zip_file:
        return zip_file.read(inner_path)


def _files_in_configured_folder(rule, verify_result):
    config = rule.config_json or {}
    files = _matching_files(rule, verify_result)
    folder_keyword_chain = [
        str(item).strip()
        for item in config.get("folder_keyword_chain") or []
        if str(item).strip()
    ]
    if not folder_keyword_chain:
        return files, ""

    selected_folder_segments = None
    for file_info in files:
        folder_segments = _folder_segments(file_info.path)
        matched_indices = _folder_keyword_chain_indices(folder_segments, folder_keyword_chain)
        if matched_indices:
            selected_folder_segments = folder_segments[: matched_indices[-1] + 1]
            break

    if not selected_folder_segments:
        return [], ""

    selected_files = [
        file_info
        for file_info in files
        if _folder_segments(file_info.path)[: len(selected_folder_segments)] == selected_folder_segments
    ]
    return selected_files, "/".join(selected_folder_segments)


def _folder_segments(path):
    normalized = str(path or "").replace("\\", "/")
    if "::" in normalized:
        normalized = normalized.split("::", 1)[1]
    parts = [part for part in normalized.split("/") if part]
    if not parts:
        return []
    return parts[:-1]


def _folder_keyword_chain_indices(folder_segments, keyword_chain):
    indices = []
    start = 0
    for keyword in keyword_chain:
        found = None
        for index in range(start, len(folder_segments)):
            if keyword in folder_segments[index]:
                found = index
                break
        if found is None:
            return []
        indices.append(found)
        start = found + 1
    return indices


def _build_rule_context(project):
    ecm_row = project.ecm_row_json or {}
    product_raw = _first_line(
        ecm_row.get("product")
        or ecm_row.get("제품명")
        or ""
    )
    product, version = _split_product_and_version(product_raw)
    start_date, end_date = _reference_start_end_dates(project)
    return RuleContext(
        project_number=project.project_number,
        product_raw=product_raw,
        product=product,
        version=version,
        company=_first_line(ecm_row.get("company") or ecm_row.get("회사명") or ""),
        pl=ecm_row.get("pl") or ecm_row.get("시험PL") or "",
        wd=ecm_row.get("wd") or ecm_row.get("WD") or "",
        start_date=start_date,
        end_date=end_date,
        year=_project_year(project.project_number),
        request_date=ecm_row.get("request_date") or ecm_row.get("신청일") or "",
        contract_date=ecm_row.get("contract_date") or ecm_row.get("계약일") or "",
        certification_committee_date=ecm_row.get("cert_date") or ecm_row.get("인증일자") or "",
        derived_variables={},
    )


def _split_product_and_version(product_name):
    value = _normalize_spaces(product_name)
    if not value:
        return "", ""

    version_match = re.search(
        r"(?i)(?:^|\s)((?:v|ver|version)\s*\.?\s*[\w.\-]+)$",
        value,
    )
    if version_match:
        version = version_match.group(1).strip()
        product = value[: version_match.start(1)].strip()
        return product, version

    if " " not in value:
        return value, ""

    product, version = value.rsplit(" ", 1)
    return product.strip(), version.strip()


def _project_year(project_number):
    match = re.search(r"TTA-(\d{2})-", project_number or "", re.IGNORECASE)
    if not match:
        return ""
    return f"20{match.group(1)}"


def _reference_start_end_dates(project):
    db_path = Path(getattr(settings, "DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH", "")) if getattr(settings, "DOWNLOAD_REVIEW_REFERENCE_MASTER_DB_PATH", None) else Path(settings.BASE_DIR) / "main" / "data" / "reference.db"
    table_name = getattr(settings, "DOWNLOAD_REVIEW_REFERENCE_MASTER_TABLE", "sw_data")
    if not db_path.exists():
        return "", ""

    conn = None
    try:
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        columns = _sqlite_columns(conn, table_name)
        project_column = _first_existing_column(columns, ["프로젝트번호", "시험번호"])
        start_column = _first_existing_column(columns, ["시작일자", "시작일"])
        end_column = _first_existing_column(columns, ["종료일자", "종료일"])
        if not (project_column and start_column and end_column):
            return "", ""

        row = conn.execute(
            (
                f"SELECT {_quote_sqlite_identifier(start_column)}, {_quote_sqlite_identifier(end_column)} "
                f"FROM {_quote_sqlite_identifier(table_name)} "
                f"WHERE {_quote_sqlite_identifier(project_column)} = ? "
                "LIMIT 1"
            ),
            [project.project_number],
        ).fetchone()
    except sqlite3.Error:
        return "", ""
    finally:
        if conn is not None:
            conn.close()

    if not row:
        return "", ""
    return _format_dot_date(row[start_column]), _format_dot_date(row[end_column])


def _sqlite_columns(conn, table_name):
    try:
        rows = conn.execute(f"PRAGMA table_info({_quote_sqlite_identifier(table_name)})").fetchall()
    except sqlite3.Error:
        return set()
    return {row[1] for row in rows}


def _first_existing_column(columns, candidates):
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return ""


def _quote_sqlite_identifier(identifier):
    return '"' + str(identifier).replace('"', '""') + '"'


def _format_dot_date(value):
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.replace("년", ".").replace("월", ".").replace("일", "")
    normalized = normalized.replace("/", ".").replace("-", ".")
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", normalized)
    if not match:
        return text
    year, month, day = match.groups()
    return f"{int(year):04d}.{int(month):02d}.{int(day):02d}."


def _resolved_keywords(keywords, context):
    resolved = []
    for keyword in keywords:
        value = _resolve_rule_value(str(keyword), context)
        if value:
            resolved.append(value)
    return resolved


def _resolve_rule_value(value, context):
    replacements = {
        "{project_number}": context.project_number,
        "{프로젝트번호}": context.project_number,
        "{product}": context.product,
        "{제품명}": context.product,
        "{company}": context.company,
        "{회사명}": context.company,
        "{버전}": context.version,
        "{pl}": context.pl,
        "{PL}": context.pl,
        "{wd}": context.wd,
        "{WD}": context.wd,
        "{시작일}": context.start_date,
        "{종료일}": context.end_date,
        "{연도}": context.year,
        "{신청일}": context.request_date,
        "{계약일}": context.contract_date,
        "{인증위}": context.certification_committee_date,
    }
    for key, replacement in context.derived_variables.items():
        replacements.setdefault(f"{{{key}}}", _variable_to_text(replacement))
    for placeholder, replacement in replacements.items():
        value = value.replace(placeholder, str(replacement))
    return value.strip()


def _variable_to_text(value):
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _configured_extensions(config, target_file_type):
    raw_extensions = config.get("extensions")
    if raw_extensions is None:
        extension = _extension_from_file_type(target_file_type)
        return [extension] if extension else []
    extensions = []
    for extension in raw_extensions:
        normalized = _normalize_extension(extension)
        if normalized:
            extensions.append(normalized)
    return extensions


def _normalize_extension(extension):
    value = str(extension or "").strip().lower()
    if not value or value == "any":
        return ""
    if value.startswith("."):
        return value
    return f".{value}"


def _name_contains_all(file_name, keywords):
    return all(keyword in file_name for keyword in keywords)


def _extension_matches(extension, extensions):
    if not extensions:
        return True
    return str(extension or "").lower() in extensions


def _artifact_file_message(config, status, matched_count, exact_count):
    if status == DownloadReviewRuleStatus.PASS:
        return config.get("pass_message") or "대상 파일을 확인했습니다."
    if exact_count is not None and matched_count > int(exact_count):
        return config.get("multiple_message") or "대상 파일이 여러개 존재합니다."
    return config.get("missing_message") or "파일이 없습니다."


def _matched_files_actual(files):
    if not files:
        return "일치 파일 없음"
    return ", ".join(file_info.name for file_info in files[:5])


def _evaluate_required_file_specs(config, matched_files):
    specs = config.get("required_files") or []
    if not specs:
        specs = [
            {
                "extensions": config.get("extensions") or [],
                "exact_count": config.get("exact_count"),
                "min_count": config.get("min_count") or 1,
            }
        ]

    passed = True
    expected = []
    actual = []
    details = []
    message = ""
    for spec in specs:
        extensions = _configured_extensions(spec, "any")
        files = [
            file_info
            for file_info in matched_files
            if _extension_matches(file_info.extension, extensions)
        ]
        exact_count = spec.get("exact_count")
        min_count = int(spec.get("min_count") or 1)
        if exact_count is not None:
            count_expected = int(exact_count)
            spec_passed = len(files) == count_expected
            count_text = f"{count_expected}개"
        else:
            spec_passed = len(files) >= min_count
            count_text = f"{min_count}개 이상"

        extension_text = ", ".join(extensions) if extensions else "확장자 무관"
        expected.append(f"{extension_text} {count_text}")
        actual.append(f"{extension_text} {len(files)}개")
        details.append({
            "extensions": extensions,
            "expected": count_text,
            "actual_count": len(files),
            "matched_files": [file_info.name for file_info in files[:10]],
            "passed": spec_passed,
        })
        if not spec_passed:
            passed = False
            message = spec.get("message") or config.get("missing_message") or "필요한 파일이 없습니다."

    return {
        "passed": passed,
        "expected": expected,
        "actual": actual,
        "message": message,
        "details": details,
    }


def _evaluate_content_checks(config, matched_files, context):
    passed = True
    expected = []
    actual = []
    details = []
    message = ""

    for check in config.get("content_checks") or []:
        extensions = _configured_extensions(check, "any")
        files = [
            file_info
            for file_info in matched_files
            if _extension_matches(file_info.extension, extensions)
        ]
        if not files:
            check_result = {
                "passed": False,
                "expected": _content_check_expected(check, context),
                "actual": "검사 대상 파일 없음",
                "message": check.get("missing_message") or "검사 대상 파일이 없습니다.",
                "detail": {"extensions": extensions},
            }
        else:
            check_result = _run_content_check(check, files[0], context)

        expected.append(check_result["expected"])
        actual.append(check_result["actual"])
        details.append({
            **check_result.get("detail", {}),
            "type": check.get("type"),
            "file_name": files[0].name if files else "",
            "passed": check_result["passed"],
        })
        if not check_result["passed"] and passed:
            passed = False
            message = check_result["message"]

    return {
        "passed": passed,
        "expected": expected,
        "actual": actual,
        "message": message,
        "details": details,
    }


def _run_content_check(check, file_info, context):
    check_type = str(check.get("type") or "").strip()
    if check_type == "docx_table_next_cell_equals":
        return _check_docx_table_next_cell_equals(check, file_info, context)
    if check_type == "pdf_first_page_label_value_contains":
        return _check_pdf_first_page_label_value_contains(check, file_info, context)
    if check_type == "docx_text_contains":
        return _check_docx_text_contains(check, file_info, context)
    if check_type == "docx_next_paragraph_matches":
        return _check_docx_next_paragraph_matches(check, file_info, context)
    raise DownloadReviewInspectionError(f"지원하지 않는 문서 내용 검사 유형입니다: {check_type or '(비어 있음)'}")


def _check_docx_table_next_cell_equals(check, file_info, context):
    label = str(check.get("label") or "").strip()
    expected_value = _resolve_rule_value(str(check.get("expected") or ""), context)
    rows = _docx_table_rows(file_info)
    actual_value = _find_next_cell_by_label(rows, label)
    passed = _normalize_compare(actual_value, check) == _normalize_compare(expected_value, check)
    return _content_result(
        check,
        passed,
        expected=f"{label} 오른쪽 셀 = {expected_value}",
        actual=f"{label} 오른쪽 셀 = {actual_value or '(없음)'}",
        detail={"label": label, "expected_value": expected_value, "actual_value": actual_value},
    )


def _check_pdf_first_page_label_value_contains(check, file_info, context):
    label = str(check.get("label") or "").strip()
    expected_value = _resolve_rule_value(str(check.get("expected") or ""), context)
    text = _pdf_page_text(file_info, page_index=int(check.get("page_index") or 0))
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    window_size = int(check.get("line_window") or 3)
    actual_value = ""
    for index, line in enumerate(lines):
        if label not in line:
            continue
        window = lines[index:index + window_size + 1]
        actual_value = " ".join(window)
        if expected_value in actual_value:
            break
    passed = bool(actual_value and expected_value in actual_value)
    return _content_result(
        check,
        passed,
        expected=f"PDF 1페이지 {label} 주변에 {expected_value} 포함",
        actual=actual_value or f"PDF 1페이지에서 {label} 없음",
        detail={"label": label, "expected_value": expected_value, "text_excerpt": lines[:20]},
    )


def _check_docx_text_contains(check, file_info, context):
    expected_text = _resolve_rule_value(str(check.get("text") or ""), context)
    paragraphs = _docx_paragraphs(file_info)
    matched = _find_matching_paragraph(paragraphs, expected_text, check)
    passed = matched is not None
    return _content_result(
        check,
        passed,
        expected=f"문서에 '{expected_text}' 포함",
        actual=matched or "일치 문장 없음",
        detail={"expected_text": expected_text, "matched_text": matched or ""},
    )


def _check_docx_next_paragraph_matches(check, file_info, context):
    after_text = _resolve_rule_value(str(check.get("after_text") or ""), context)
    pattern = str(check.get("regex") or "").strip()
    paragraphs = _docx_paragraphs(file_info)
    matched_index = _find_matching_paragraph_index(paragraphs, after_text, check)
    next_text = ""
    if matched_index is not None:
        for paragraph in paragraphs[matched_index + 1:]:
            if paragraph.strip():
                next_text = paragraph.strip()
                break
    passed = bool(next_text and re.search(pattern, next_text))
    return _content_result(
        check,
        passed,
        expected=f"'{after_text}' 다음 문단이 {pattern} 형식",
        actual=next_text or "다음 문단 없음",
        detail={"after_text": after_text, "regex": pattern, "actual_text": next_text},
    )


def _content_result(check, passed, *, expected, actual, detail):
    return {
        "passed": passed,
        "expected": expected,
        "actual": actual,
        "message": (
            check.get("pass_message")
            if passed
            else check.get("failure_message")
        ) or ("문서 내용을 확인했습니다." if passed else "문서 내용이 기준과 일치하지 않습니다."),
        "detail": detail,
    }


def _content_check_expected(check, context):
    if "expected" in check:
        return _resolve_rule_value(str(check.get("expected") or ""), context)
    if "text" in check:
        return _resolve_rule_value(str(check.get("text") or ""), context)
    return str(check.get("type") or "")


def _read_excel_workbook(file_info):
    extension = file_info.extension.lower()
    data = _read_file_bytes(file_info)
    if extension == ".xlsx":
        return _read_xlsx_workbook(data)
    if extension == ".xls":
        return _read_xls_workbook(data)
    raise DownloadReviewInspectionError(f"지원하지 않는 Excel 확장자입니다: {extension or '(없음)'}")


def _read_xlsx_workbook(data):
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=False)
    except Exception as exc:
        raise DownloadReviewInspectionError("xlsx 파일을 읽을 수 없습니다.") from exc

    sheets = []
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows():
            rows.append([_excel_cell_text(cell.value) for cell in row])
        sheets.append(
            ExcelSheet(
                name=worksheet.title,
                rows=_trim_empty_edges(rows),
                header_text=_worksheet_header_text(worksheet),
            )
        )
    workbook.close()
    return ExcelWorkbook(sheets=sheets)


def _worksheet_header_text(worksheet):
    parts = []
    for header in (worksheet.oddHeader, worksheet.evenHeader, worksheet.firstHeader):
        for section in (header.left, header.center, header.right):
            text = getattr(section, "text", "") or ""
            if text:
                parts.append(text)
    return _normalize_spaces(" ".join(parts))


def _read_xls_workbook(data):
    try:
        import xlrd
    except ImportError as exc:
        raise DownloadReviewInspectionError("xls 파일을 읽으려면 xlrd 패키지가 필요합니다.") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=data)
    except Exception as exc:
        raise DownloadReviewInspectionError("xls 파일을 읽을 수 없습니다.") from exc

    # xlrd는 인쇄 머리글(BIFF HEADER 레코드)을 노출하지 않으므로 직접 파싱한다.
    headers_by_sheet = _xls_print_headers(data)

    sheets = []
    for worksheet in workbook.sheets():
        rows = []
        for row_index in range(worksheet.nrows):
            rows.append([
                _excel_cell_text(worksheet.cell_value(row_index, col_index))
                for col_index in range(worksheet.ncols)
            ])
        sheets.append(
            ExcelSheet(
                name=worksheet.name,
                rows=_trim_empty_edges(rows),
                header_text=headers_by_sheet.get(worksheet.name, ""),
            )
        )
    return ExcelWorkbook(sheets=sheets)


def _xls_print_headers(data):
    """`.xls` 워크북 스트림을 직접 파싱해 시트별 인쇄 머리글을 추출한다.

    xlrd는 BIFF HEADER(0x14) 레코드를 노출하지 않으므로, OLE2 컨테이너에서
    Workbook 스트림을 꺼낸 뒤 BOUNDSHEET(0x85)로 시트별 substream 위치를 찾고
    각 substream 첫 HEADER 레코드를 읽는다. 실패하면 빈 매핑을 반환한다.
    """
    try:
        from io import StringIO

        from xlrd.compdoc import CompDoc

        compdoc = CompDoc(data, logfile=StringIO())
        mem, base, size = compdoc.locate_named_stream("Workbook")
        if mem is None or not size:
            return {}
        stream = mem[base:base + size]
    except Exception:
        return {}

    try:
        boundsheets = []
        pos = 0
        total = len(stream)
        while pos + 4 <= total:
            opcode, length = struct.unpack("<HH", stream[pos:pos + 4])
            body = stream[pos + 4:pos + 4 + length]
            if opcode == 0x0085 and len(body) >= 8:  # BOUNDSHEET
                ply_pos = struct.unpack("<I", body[0:4])[0]
                name, _ = _xls_unicode_string(body, 6, 1)
                boundsheets.append((ply_pos, name))
            pos += 4 + length

        headers = {}
        for ply_pos, name in boundsheets:
            headers[name] = _xls_header_at(stream, ply_pos)
        return headers
    except Exception:
        return {}


def _xls_header_at(stream, start):
    pos = start
    total = len(stream)
    depth = 0
    while pos + 4 <= total:
        opcode, length = struct.unpack("<HH", stream[pos:pos + 4])
        body = stream[pos + 4:pos + 4 + length]
        if opcode == 0x0809:  # BOF
            depth += 1
        elif opcode == 0x000A:  # EOF
            depth -= 1
            if depth <= 0:
                return ""
        elif opcode == 0x0014:  # HEADER
            if length == 0:
                return ""
            text, _ = _xls_unicode_string(body, 0, 2)
            return text
        pos += 4 + length
    return ""


def _xls_unicode_string(body, offset, cch_size):
    if cch_size == 1:
        cch = body[offset]
        cursor = offset + 1
    else:
        cch = struct.unpack("<H", body[offset:offset + 2])[0]
        cursor = offset + 2
    grbit = body[cursor]
    cursor += 1
    if grbit & 0x01:
        text = body[cursor:cursor + cch * 2].decode("utf-16-le", "replace")
    else:
        text = body[cursor:cursor + cch].decode("cp949", "replace")
    return text, cursor


def _excel_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return _normalize_spaces(value)


def _trim_empty_edges(rows):
    trimmed = [list(row) for row in rows]
    while trimmed and not any(cell for cell in trimmed[-1]):
        trimmed.pop()
    max_width = 0
    for row in trimmed:
        for index, cell in enumerate(row):
            if cell:
                max_width = max(max_width, index + 1)
    if not max_width:
        return []
    return [row[:max_width] + [""] * max(0, max_width - len(row)) for row in trimmed]


def _find_cell_containing(rows, text):
    needle = str(text or "")
    if not needle:
        return None
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            if needle in value:
                return {"row": row_index, "column": col_index, "value": value}
    return None


def _find_cell_with_date_range(rows, start_date, end_date):
    """`{시작일} ~ {종료일}` 기간이 적힌 셀을 날짜 정규화 기준으로 찾는다.

    셀 안 날짜 구분자(`.`/`-`/`/`/`년월일`)가 달라도 같은 날짜면 인정한다.
    """
    start_norm = _format_dot_date(start_date)
    end_norm = _format_dot_date(end_date)
    if not (start_norm and end_norm):
        return None
    date_pattern = r"\d{4}\s*[.\-/년]\s*\d{1,2}\s*[.\-/월]\s*\d{1,2}\s*일?\.?"
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            found = [_format_dot_date(token) for token in re.findall(date_pattern, str(value or ""))]
            if start_norm in found and end_norm in found:
                return {"row": row_index, "column": col_index, "value": value}
    return None


def _excel_area_from_column_anchor(rows, label, *, column_index):
    for row_index, row in enumerate(rows):
        if column_index >= len(row) or label not in row[column_index]:
            continue

        last_row = row_index
        last_col = column_index
        for scan_row_index in range(row_index, len(rows)):
            scan_row = rows[scan_row_index]
            populated_columns = [
                col_index
                for col_index, value in enumerate(scan_row[column_index:], start=column_index)
                if value
            ]
            if populated_columns:
                last_row = scan_row_index
                last_col = max(last_col, max(populated_columns))
        return {
            "start_row": row_index + 1,
            "start_column": column_index + 1,
            "end_row": last_row + 1,
            "end_column": last_col + 1,
            "range": f"{_excel_column_name(column_index + 1)}{row_index + 1}:{_excel_column_name(last_col + 1)}{last_row + 1}",
        }
    return None


def _excel_column_name(index):
    name = ""
    value = index
    while value:
        value, remainder = divmod(value - 1, 26)
        name = chr(ord("A") + remainder) + name
    return name


def _date_range_start(value):
    date_value = _parse_date(value)
    if not date_value:
        return None
    return datetime.combine(date_value, time.min)


def _date_range_end(value):
    date_value = _parse_date(value)
    if not date_value:
        return None
    return datetime.combine(date_value, time.max)


def _parse_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    normalized = text.replace("년", ".").replace("월", ".").replace("일", "")
    normalized = normalized.replace("/", ".").replace("-", ".")
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", normalized)
    if not match:
        return None
    try:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()
    except ValueError:
        return None


def _inspection_folder_tree(verify_result):
    folders = set()
    file_folders = set()
    for file_info in _inspection_files(verify_result):
        segments = tuple(_folder_segments(file_info.path))
        file_folders.add(segments)
        for index in range(1, len(segments) + 1):
            folders.add(segments[:index])

    for folder in _local_download_folders(getattr(verify_result, "download_dir", "")):
        folders.add(folder)

    return folders, file_folders


def _local_download_folders(download_dir):
    base = Path(download_dir or "")
    if not base.is_dir():
        return set()
    folders = set()
    for path in base.rglob("*"):
        if not path.is_dir():
            continue
        parts = tuple(path.relative_to(base).parts)
        for index in range(1, len(parts) + 1):
            folders.add(parts[:index])
    return folders


def _find_folder_by_keyword_chain(folders, keyword_chain):
    chain = [str(item).strip() for item in keyword_chain if str(item).strip()]
    for folder in sorted(folders, key=lambda item: (len(item), "/".join(item))):
        if _folder_keyword_chain_indices(list(folder), chain):
            return folder
    return ()


def _run_folder_check(base, folders, file_folders, folder_check):
    keyword = str(folder_check.get("keyword") or "").strip()
    failure_message = str(folder_check.get("failure_message") or "폴더 구조 확인 불가")
    folder = _find_immediate_descendant_folder(base, folders, keyword)
    if not folder:
        return {
            "keyword": keyword,
            "folder": "",
            "passed": False,
            "message": failure_message,
            "actual": "폴더 없음",
        }

    exact_child_folders = folder_check.get("exact_child_folders")
    child_folders = _immediate_child_folders(folder, folders)
    child_files = _immediate_child_file_folders(folder, file_folders)
    if exact_child_folders is not None and len(child_folders) != int(exact_child_folders):
        return {
            "keyword": keyword,
            "folder": "/".join(folder),
            "passed": False,
            "message": failure_message,
            "actual": f"하위 폴더 {len(child_folders)}개",
        }

    if folder_check.get("each_child_has_entry"):
        empty_children = [
            child
            for child in child_folders
            if not _immediate_child_folders(child, folders)
            and not _immediate_child_file_folders(child, file_folders)
        ]
        if empty_children:
            return {
                "keyword": keyword,
                "folder": "/".join(folder),
                "passed": False,
                "message": failure_message,
                "actual": "빈 하위 폴더: " + ", ".join("/".join(child) for child in empty_children[:5]),
            }

    min_entries = folder_check.get("min_entries")
    if min_entries is not None:
        entry_count = len(child_folders) + len(child_files)
        if entry_count < int(min_entries):
            return {
                "keyword": keyword,
                "folder": "/".join(folder),
                "passed": False,
                "message": failure_message,
                "actual": f"항목 {entry_count}개",
            }

    return {
        "keyword": keyword,
        "folder": "/".join(folder),
        "passed": True,
        "message": "정상",
        "actual": {
            "child_folder_count": len(child_folders),
            "child_file_count": len(child_files),
        },
    }


def _find_immediate_descendant_folder(base, folders, keyword):
    candidates = [
        folder
        for folder in folders
        if len(folder) == len(base) + 1
        and folder[: len(base)] == base
        and keyword in folder[-1]
    ]
    return sorted(candidates, key=lambda item: "/".join(item))[0] if candidates else ()


def _immediate_child_folders(parent, folders):
    return {
        folder
        for folder in folders
        if len(folder) == len(parent) + 1
        and folder[: len(parent)] == parent
    }


def _immediate_child_file_folders(parent, file_folders):
    return {
        folder
        for folder in file_folders
        if folder == parent
    }


def _store_excel_area_artifact(project, rule, sheet, area, *, artifact_id, label, source_file):
    values = _excel_area_values(sheet.rows, area)
    if not values:
        raise DownloadReviewInspectionError("Excel 영역 이미지 대상 값이 없습니다.")
    png_bytes = _render_excel_area_png(values, title=f"{sheet.name} {area.get('range') or ''}".strip())
    return _store_artifact_bytes(
        project,
        rule,
        artifact_id=artifact_id,
        label=label,
        file_suffix=".png",
        content_type="image/png",
        content_bytes=png_bytes,
        kind="image",
        source_file=source_file,
        download=False,
    )


def _excel_area_values(rows, area):
    start_row = max(int(area.get("start_row") or 1) - 1, 0)
    end_row = max(int(area.get("end_row") or 0) - 1, start_row)
    start_col = max(int(area.get("start_column") or 1) - 1, 0)
    end_col = max(int(area.get("end_column") or 0) - 1, start_col)
    values = []
    for row_index in range(start_row, end_row + 1):
        row = rows[row_index] if row_index < len(rows) else []
        values.append([
            row[col_index] if col_index < len(row) else ""
            for col_index in range(start_col, end_col + 1)
        ])
    return values


def _render_excel_area_png(values, *, title):
    try:
        import fitz
    except Exception as exc:
        raise DownloadReviewInspectionError("Excel 영역 이미지를 생성하려면 PyMuPDF가 필요합니다.") from exc

    try:
        margin = 18
        title_height = 24 if title else 0
        row_height = 25
        font_size = 8.5
        col_widths = _excel_render_column_widths(values)
        width = max(sum(col_widths) + margin * 2, 240)
        height = max(len(values) * row_height + margin * 2 + title_height, 120)

        document = fitz.open()
        page = document.new_page(width=width, height=height)
        font_name = _insert_artifact_font(page)

        if title:
            page.insert_textbox(
                fitz.Rect(margin, margin - 2, width - margin, margin + title_height),
                title,
                fontsize=10,
                fontname=font_name,
                color=(0.1, 0.1, 0.1),
            )

        y = margin + title_height
        for row_index, row in enumerate(values):
            x = margin
            fill = (0.95, 0.96, 0.98) if row_index == 0 else None
            for col_index, cell in enumerate(row):
                cell_width = col_widths[col_index]
                rect = fitz.Rect(x, y, x + cell_width, y + row_height)
                page.draw_rect(rect, color=(0.68, 0.72, 0.78), fill=fill, width=0.5)
                page.insert_textbox(
                    rect + (4, 4, -4, -3),
                    _artifact_cell_text(cell),
                    fontsize=font_size,
                    fontname=font_name,
                    color=(0.08, 0.09, 0.11),
                )
                x += cell_width
            y += row_height

        pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        png_bytes = pixmap.tobytes("png")
        document.close()
        return png_bytes
    except DownloadReviewInspectionError:
        raise
    except Exception as exc:
        raise DownloadReviewInspectionError("Excel 영역 이미지를 생성할 수 없습니다.") from exc


def _excel_render_column_widths(values):
    column_count = max((len(row) for row in values), default=0)
    widths = []
    for col_index in range(column_count):
        max_length = max(
            len(str(row[col_index])) if col_index < len(row) else 0
            for row in values
        )
        widths.append(min(max(70, max_length * 7 + 18), 220))
    return widths


def _insert_artifact_font(page):
    font_path = _artifact_font_path()
    if not font_path:
        return "helv"
    try:
        page.insert_font(fontname="gscert_cjk", fontfile=str(font_path))
        return "gscert_cjk"
    except Exception:
        return "helv"


def _artifact_font_path():
    candidates = [
        Path("C:/Windows/Fonts/malgun.ttf"),
        Path("C:/Windows/Fonts/malgunbd.ttf"),
        Path("/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
    ]
    for path in candidates:
        if path.is_file():
            return path
    return None


def _artifact_cell_text(value):
    text = str(value or "").replace("\r", " ").replace("\n", " ")
    return _normalize_spaces(text)


def _store_pdf_first_page_artifact(project, rule, file_info, *, artifact_id, label):
    data = _read_file_bytes(file_info)
    try:
        import fitz

        with fitz.open(stream=data, filetype="pdf") as document:
            if document.page_count < 1:
                raise DownloadReviewInspectionError("PDF 첫 페이지가 없습니다.")
            pixmap = document[0].get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            png_bytes = pixmap.tobytes("png")
    except DownloadReviewInspectionError:
        raise
    except Exception as exc:
        raise DownloadReviewInspectionError("PDF 1페이지 캡처를 생성할 수 없습니다.") from exc

    return _store_artifact_bytes(
        project,
        rule,
        artifact_id=artifact_id,
        label=label,
        file_suffix=".png",
        content_type="image/png",
        content_bytes=png_bytes,
        kind="image",
        source_file=_display_path(file_info.path, project.project_number),
        download=False,
    )


def _store_pdf_download_artifact(project, rule, file_info, *, artifact_id, label):
    """PDF 원본을 그대로 저장해 사용자가 버튼으로 직접 다운로드하게 한다."""
    data = _read_file_bytes(file_info)
    return _store_artifact_bytes(
        project,
        rule,
        artifact_id=artifact_id,
        label=label,
        file_suffix=".pdf",
        content_type="application/pdf",
        content_bytes=data,
        kind="file",
        source_file=_display_path(file_info.path, project.project_number),
        download=True,
    )


def _store_artifact_bytes(
    project,
    rule,
    *,
    artifact_id,
    label,
    file_suffix,
    content_type,
    content_bytes,
    kind,
    source_file,
    download,
):
    safe_artifact_id = _safe_artifact_id(artifact_id)
    safe_suffix = _safe_artifact_suffix(file_suffix)
    file_name = f"{rule.code}_{safe_artifact_id}{safe_suffix}"
    base_dir = _artifact_base_dir()
    relative_path = Path(str(project.id)) / file_name
    target_path = (base_dir / relative_path).resolve()
    try:
        target_path.relative_to(base_dir)
    except ValueError as exc:
        raise DownloadReviewInspectionError("산출물 저장 경로가 올바르지 않습니다.") from exc
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(content_bytes)

    return {
        "id": safe_artifact_id,
        "label": label,
        "kind": kind,
        "content_type": content_type,
        "file_name": file_name,
        "relative_path": str(relative_path).replace("\\", "/"),
        "source_file": source_file,
        "download": download,
    }


def _artifact_base_dir():
    return Path(
        getattr(
            settings,
            "DOWNLOAD_REVIEW_ARTIFACT_DIR",
            Path(settings.BASE_DIR) / "main" / "data" / "download_review_artifacts",
        )
    ).resolve()


def _safe_artifact_id(value):
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "").strip())
    return cleaned.strip("._-") or "artifact"


def _safe_artifact_suffix(value):
    suffix = str(value or "").strip().lower()
    if re.fullmatch(r"\.[a-z0-9]{1,12}", suffix):
        return suffix
    return ".bin"


def _docx_defect_report_round_dates(file_info):
    tables = _docx_tables(file_info)
    target_text = ""
    for table in tables:
        flattened = " ".join(cell for row in table for cell in row if cell)
        if "결함리포트 송부" in flattened:
            target_text = flattened
            break
    if not target_text:
        raise DownloadReviewInspectionError("결함리포트 송부 표를 찾을 수 없습니다.")

    rounds = {}
    for match in re.finditer(r"([1-9]\d*)\s*차\s*[:：]\s*(\d{4})\s*[./-]\s*(\d{1,2})\s*[./-]\s*(\d{1,2})", target_text):
        round_no, year, month, day = match.groups()
        rounds[f"{int(round_no)}차"] = f"{int(year):04d}.{int(month):02d}.{int(day):02d}."
    return rounds


def _docx_first_table_after_text(file_info, marker):
    root = _docx_root(file_info)
    ns = _word_ns()
    marker_text = str(marker or "")
    marker_seen = False
    body = root.find(".//w:body", namespaces=ns)
    if body is None:
        return []

    for child in body:
        tag_name = etree.QName(child).localname
        if tag_name == "tbl":
            rows = _word_table_rows(child, ns)
            flattened = " ".join(cell for row in rows for cell in row if cell)
            if marker_seen:
                return rows
            if marker_text and marker_text in flattened:
                marker_seen = True
            continue

        text = _word_element_text(child, ns)
        if marker_text and marker_text in text:
            marker_seen = True
    return []


def _docx_tables(file_info):
    root = _docx_root(file_info)
    ns = _word_ns()
    return [
        _word_table_rows(table, ns)
        for table in root.xpath(".//w:tbl", namespaces=ns)
    ]


def _word_table_rows(table, ns):
    rows = []
    for row in table.xpath("./w:tr", namespaces=ns):
        cells = []
        for cell in row.xpath("./w:tc", namespaces=ns):
            cells.append(_word_cell_text(cell, ns))
        rows.append(cells)
    return rows


def _word_element_text(element, ns):
    return _normalize_spaces(" ".join(element.xpath(".//w:t/text()", namespaces=ns)))


def _docx_root(file_info):
    data = _read_file_bytes(file_info)
    with ZipFile(BytesIO(data)) as docx_file:
        return etree.fromstring(docx_file.read("word/document.xml"))


def _docx_paragraphs(file_info):
    root = _docx_root(file_info)
    ns = _word_ns()
    paragraphs = []
    for paragraph in root.xpath(".//w:p", namespaces=ns):
        text = "".join(paragraph.xpath(".//w:t/text()", namespaces=ns)).strip()
        if text:
            paragraphs.append(_normalize_spaces(text))
    return paragraphs


def _docx_all_text(file_info):
    data = _read_file_bytes(file_info)
    ns = _word_ns()
    paragraphs = []
    try:
        with ZipFile(BytesIO(data)) as docx_file:
            target_names = [
                name
                for name in docx_file.namelist()
                if name.startswith("word/")
                and name.endswith(".xml")
                and (
                    name == "word/document.xml"
                    or PurePosixPath(name).name.startswith("header")
                    or PurePosixPath(name).name.startswith("footer")
                )
            ]
            for name in sorted(target_names):
                try:
                    root = etree.fromstring(docx_file.read(name))
                except etree.XMLSyntaxError:
                    continue
                # 한 문단 안의 run(w:t)은 공백 없이 이어붙인다. Word가 숫자/단어를
                # 여러 run으로 쪼개므로 공백 join을 쓰면 "2026"이 "20 2 6"이 된다.
                for paragraph in root.xpath(".//w:p", namespaces=ns):
                    text = "".join(paragraph.xpath(".//w:t/text()", namespaces=ns))
                    if text.strip():
                        paragraphs.append(text)
    except (KeyError, BadZipFile) as exc:
        raise DownloadReviewInspectionError("docx 본문/머리말/바닥글을 읽을 수 없습니다.") from exc
    return _normalize_spaces(" ".join(paragraphs))


def _docx_footer_text(file_info):
    data = _read_file_bytes(file_info)
    ns = _word_ns()
    texts = []
    try:
        with ZipFile(BytesIO(data)) as docx_file:
            footer_names = sorted(
                name
                for name in docx_file.namelist()
                if name.startswith("word/footer") and name.endswith(".xml")
            )
            for footer_name in footer_names:
                root = etree.fromstring(docx_file.read(footer_name))
                # 문단 안 run은 공백 없이 이어붙여 "2026"이 쪼개지지 않게 한다.
                for paragraph in root.xpath(".//w:p", namespaces=ns):
                    text = _normalize_spaces("".join(paragraph.xpath(".//w:t/text()", namespaces=ns)))
                    if text:
                        texts.append(text)
    except (KeyError, BadZipFile, etree.XMLSyntaxError) as exc:
        raise DownloadReviewInspectionError("docx 바닥글을 읽을 수 없습니다.") from exc
    return _normalize_spaces(" ".join(texts))


def _docx_table_rows(file_info):
    root = _docx_root(file_info)
    ns = _word_ns()
    rows = []
    for table in root.xpath(".//w:tbl", namespaces=ns):
        for row in table.xpath("./w:tr", namespaces=ns):
            cells = []
            for cell in row.xpath("./w:tc", namespaces=ns):
                cells.append(_word_cell_text(cell, ns))
            rows.append(cells)
    return rows


def _word_cell_text(cell, ns):
    paragraphs = []
    for paragraph in cell.xpath("./w:p", namespaces=ns):
        text = "".join(paragraph.xpath(".//w:t/text()", namespaces=ns)).strip()
        if text:
            paragraphs.append(text)
    if not paragraphs:
        paragraphs = ["".join(cell.xpath(".//w:t/text()", namespaces=ns)).strip()]
    return _normalize_spaces(" ".join(paragraphs))


def _word_ns():
    return {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def _find_next_cell_by_label(rows, label):
    target = _normalize_label(label)
    for row in rows:
        for index, cell in enumerate(row[:-1]):
            if _normalize_label(cell) == target:
                return row[index + 1].strip()
    return ""


def _pdf_page_text(file_info, *, page_index=0):
    import fitz

    data = _read_file_bytes(file_info)
    with fitz.open(stream=data, filetype="pdf") as document:
        if document.page_count <= page_index:
            return ""
        return document[page_index].get_text("text")


def _find_matching_paragraph(paragraphs, expected_text, check):
    index = _find_matching_paragraph_index(paragraphs, expected_text, check)
    return paragraphs[index] if index is not None else None


def _find_matching_paragraph_index(paragraphs, expected_text, check):
    expected = _normalize_content(expected_text, check)
    for index, paragraph in enumerate(paragraphs):
        actual = _normalize_content(paragraph, check)
        if expected and expected in actual:
            return index
    return None


def _normalize_content(value, check):
    text = str(value or "")
    if check.get("remove_whitespace"):
        return re.sub(r"\s+", "", text)
    if check.get("normalize_whitespace", True):
        return _normalize_spaces(text)
    return text


def _normalize_compare(value, check):
    text = str(value or "")
    if check.get("remove_whitespace"):
        return re.sub(r"\s+", "", text)
    return _normalize_spaces(text) if check.get("normalize_whitespace") else text.strip()


def _normalize_label(value):
    return re.sub(r"[\s:：]+", "", str(value or "")).lower()


def _normalize_spaces(value):
    # None\ub9cc \ube48 \ubb38\uc790\uc5f4\ub85c \ubcf8\ub2e4. `value or ""`\ub97c \uc4f0\uba74 \uc815\uc218 0\u00b7False\uac00 falsy\ub77c
    # \ube48 \ubb38\uc790\uc5f4\uc774 \ub418\uc5b4 "0"\uc774 \uc0ac\ub77c\uc9c0\ub294 \ubc84\uadf8\uac00 \uc788\uc5c8\ub2e4.
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.replace("\u00a0", " ")).strip()


def _first_line(value):
    text = str(value or "")
    for separator in ("\r\n", "\n", "\r"):
        if separator in text:
            text = text.split(separator, 1)[0]
            break
    return text.strip()


def _artifact_results_from_evaluations(evaluations):
    results = {}
    for evaluation in evaluations:
        column = _artifact_column(evaluation.rule)
        if not column:
            continue
        value = "O" if evaluation.status == DownloadReviewRuleStatus.PASS else "X"
        previous = results.get(column)
        if previous == "X":
            continue
        results[column] = value
    return results


def _artifact_column(rule):
    config = rule.config_json or {}
    configured = str(config.get("artifact_column") or "").strip()
    candidates = (configured, rule.code, rule.name)
    for candidate in candidates:
        if candidate in ARTIFACT_REVIEW_COLUMNS:
            return candidate
    return ""


def _extension_from_file_type(file_type):
    value = str(file_type or "").strip().lower()
    if not value or value == "any":
        return ""
    if value.startswith("."):
        return value
    return f".{value}"


def _representative_name(files):
    if not files:
        return ""
    if len(files) == 1:
        return files[0].name
    return f"{files[0].name} 외 {len(files) - 1}개"


def _representative_path(files, project_number):
    if not files:
        return ""
    return _display_path(files[0].path, project_number) or files[0].name


def _display_path(path, project_number):
    normalized = str(path or "").replace("\\", "/")
    index = normalized.find(project_number)
    if index >= 0:
        return normalized[index:]
    return Path(normalized).name


def _validate_cleanup_target(project_number, base_dir, target):
    try:
        target.relative_to(base_dir)
    except ValueError as exc:
        raise DownloadReviewCleanupSafetyError(
            "다운로드 폴더가 허용된 기본 경로 밖에 있어 삭제하지 않았습니다."
        ) from exc

    if target == base_dir:
        raise DownloadReviewCleanupSafetyError("다운로드 기본 폴더 자체는 삭제할 수 없습니다.")
    if not target.is_dir() and target.exists():
        raise DownloadReviewCleanupSafetyError("삭제 대상이 폴더가 아닙니다.")
    if project_number not in target.name:
        raise DownloadReviewCleanupSafetyError(
            "삭제 대상 폴더명에 프로젝트번호가 없어 삭제하지 않았습니다."
        )
