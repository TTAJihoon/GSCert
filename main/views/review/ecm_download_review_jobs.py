import json
import logging
import re
import time
import zipfile
from dataclasses import dataclass
from datetime import timedelta, timezone as datetime_timezone
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from urllib.parse import quote
from zoneinfo import ZoneInfo

from django.conf import settings
from django.db import transaction
from django.db.utils import DatabaseError, OperationalError
from django.http import FileResponse, HttpResponse
from django.utils import timezone

from gscert_review_core.result_display import build_display_rows, serialize_display_row

from main.models import (
    DownloadReviewJob,
    DownloadReviewJobStatus,
    DownloadReviewLock,
    DownloadReviewLog,
    DownloadReviewLogLevel,
    DownloadReviewManualOverride,
    DownloadReviewProject,
    DownloadReviewProjectStatus,
    DownloadReviewProjectReviewStatus,
    DownloadReviewRule,
    DownloadReviewRuleResult,
    DownloadReviewRuleStatus,
)
from main.views.review.ecm_change_note import change_note_payload, change_note_summary
from main.views.review.ecm_manual_override import (
    apply_manual_override_to_result,
    manual_override_public,
    mark_overrides_applied,
)
from main.views.review.ecm_reference_db import (
    ARTIFACT_REVIEW_COLUMNS,
    get_projects_by_numbers,
    is_completed_review_value,
    write_project_review_result,
)
from main.views.review.ecm_download_review_centers import center_label, normalize_center_code


logger = logging.getLogger(__name__)

PROJECT_NUMBER_RE = re.compile(r"^TTA-\d{2}-\d{5}$")
ACTIVE_JOB_STATUSES = (
    DownloadReviewJobStatus.SCHEDULED,
    DownloadReviewJobStatus.QUEUED,
    DownloadReviewJobStatus.RUNNING,
)
CANCELABLE_JOB_STATUSES = (
    DownloadReviewJobStatus.SCHEDULED,
    DownloadReviewJobStatus.QUEUED,
)
JOB_LIST_PARAM_NAMES = {"status", "limit", "offset", "center"}
JOB_LIST_STATUS_FILTERS = {
    "all": None,
    "finished": (
        DownloadReviewJobStatus.COMPLETED,
        DownloadReviewJobStatus.FAILED,
        DownloadReviewJobStatus.CANCELED,
    ),
    DownloadReviewJobStatus.SCHEDULED: (DownloadReviewJobStatus.SCHEDULED,),
    DownloadReviewJobStatus.QUEUED: (DownloadReviewJobStatus.QUEUED,),
    DownloadReviewJobStatus.RUNNING: (DownloadReviewJobStatus.RUNNING,),
    DownloadReviewJobStatus.COMPLETED: (DownloadReviewJobStatus.COMPLETED,),
    DownloadReviewJobStatus.FAILED: (DownloadReviewJobStatus.FAILED,),
    DownloadReviewJobStatus.CANCELED: (DownloadReviewJobStatus.CANCELED,),
}
DEFAULT_JOB_LIST_LIMIT = 20
MAX_JOB_LIST_LIMIT = 100


class DownloadReviewJobRequestError(ValueError):
    error_code = "invalid_job_request"
    status_code = 400

    def __init__(self, message, *, details=None):
        super().__init__(message)
        self.details = details or {}


class DownloadReviewDuplicateProjectError(DownloadReviewJobRequestError):
    error_code = "active_project_conflict"
    status_code = 409


class DownloadReviewQueueFullError(DownloadReviewJobRequestError):
    error_code = "queue_full"
    status_code = 409


class DownloadReviewJobCancelError(DownloadReviewJobRequestError):
    error_code = "job_cancel_not_allowed"
    status_code = 409


class DownloadReviewCompletedProjectError(DownloadReviewJobRequestError):
    error_code = "completed_project_not_allowed"
    status_code = 400


class DownloadReviewNotFoundError(LookupError):
    error_code = "not_found"
    status_code = 404

    def __init__(self, message):
        super().__init__(message)
        self.details = {}


@dataclass(frozen=True)
class JobSchedule:
    status: str
    available_after: object
    queued_at: object


def create_download_review_job(payload, request_ip=None, now=None):
    center_code = parse_center_code(payload.get("center"))
    project_numbers = parse_project_numbers(payload)
    projects = get_projects_by_numbers(project_numbers, center_code=center_code)
    _validate_projects_found(project_numbers, projects)
    _validate_not_completed(projects)

    workflow_alias = getattr(settings, "WORKFLOW_DATABASE_ALIAS", "workflow")
    with transaction.atomic(using=workflow_alias):
        _validate_active_job_limit()
        _validate_no_active_project(project_numbers, center_code)

        schedule = build_job_schedule(now=now)
        job = DownloadReviewJob.objects.create(
            center_code=center_code,
            status=schedule.status,
            available_after=schedule.available_after,
            queued_at=schedule.queued_at,
            selected_projects_json=project_numbers,
            requested_project_count=len(project_numbers),
            requested_ip=request_ip,
            progress_message=_initial_progress_message(schedule.status),
        )
        DownloadReviewProject.objects.bulk_create(
            [
                DownloadReviewProject(
                    job=job,
                    center_code=center_code,
                    project_number=project["project_number"],
                    ecm_row_json={**project, "center_code": center_code, "center_label": center_label(center_code)},
                    status=DownloadReviewProjectStatus.QUEUED,
                )
                for project in projects
            ]
        )

    return {
        "success": True,
        "job_id": str(job.id),
        "center_code": job.center_code,
        "center_label": center_label(job.center_code),
        "status": job.status,
        "status_label": job_status_label(job.status),
        "requested_project_count": job.requested_project_count,
        "available_after": job.available_after.isoformat() if job.available_after else None,
        "message": job_request_message(job.status),
    }


def get_active_job_payload():
    job = find_active_job()
    active_count = DownloadReviewJob.objects.filter(status__in=ACTIVE_JOB_STATUSES).count()
    if job is None:
        return {
            "success": True,
            "active_job": None,
            "active_job_count": active_count,
            "polling": {
                "should_poll": False,
                "recommended_interval_ms": None,
                "wake_at": None,
            },
        }

    return {
        "success": True,
        "active_job": serialize_job(job),
        "active_job_count": active_count,
        "polling": polling_hint(job),
    }


def get_jobs_payload(query_params):
    query = parse_job_list_query(query_params)
    qs = DownloadReviewJob.objects.all()
    if query["center"]:
        qs = qs.filter(center_code=query["center"])
    statuses = JOB_LIST_STATUS_FILTERS[query["status"]]
    if statuses:
        qs = qs.filter(status__in=statuses)

    total = qs.count()
    jobs = qs.order_by("-requested_at", "-created_at", "-id")[
        query["offset"]:query["offset"] + query["limit"]
    ]
    return {
        "success": True,
        "items": [serialize_job(job) for job in jobs],
        "pagination": {
            "total": total,
            "limit": query["limit"],
            "offset": query["offset"],
            "has_more": query["offset"] + len(jobs) < total,
        },
        "status": query["status"],
        "center": query["center"],
    }


def attach_active_project_states(projects_payload):
    items = projects_payload.get("items") or []
    project_keys = [
        (item.get("center_code") or normalize_center_code(None), item.get("project_number"))
        for item in items
        if item.get("project_number")
    ]
    project_numbers = [project_number for _, project_number in project_keys]
    if not project_numbers:
        return projects_payload

    active_by_number = {}
    try:
        active_projects = (
            DownloadReviewProject.objects
            .select_related("job")
            .filter(project_number__in=project_numbers, job__status__in=ACTIVE_JOB_STATUSES)
            .order_by("job__requested_at", "job__created_at", "created_at", "id")
        )
        for project in active_projects:
            active_by_number.setdefault((project.center_code, project.project_number), project)
    except DatabaseError:
        logger.info("Failed to read active download-review project states; continuing without active state.")
        _attach_empty_active_project_states(items)
        return projects_payload

    for item in items:
        item_center = item.get("center_code") or normalize_center_code(None)
        active_project = active_by_number.get((item_center, item.get("project_number")))
        if not active_project:
            _attach_empty_active_project_state(item)
            continue

        job_status = active_project.job.status
        item.update(
            {
                "active_job_id": str(active_project.job_id),
                "active_job_status": job_status,
                "active_job_status_label": job_status_label(job_status),
                "active_project_status": active_project.status,
                "active_project_status_label": project_status_label(active_project.status),
                "active_state_label": active_project_state_label(job_status, active_project.status),
                "selectable": False,
            }
        )

    _attach_latest_failed_project_states(items)
    return projects_payload


def _attach_empty_active_project_states(items):
    for item in items:
        _attach_empty_active_project_state(item)


def _attach_empty_active_project_state(item):
    completed = is_completed_review_value(item.get("review_raw") or item.get("review"))
    item.update(
        {
            "active_job_id": None,
            "active_job_status": "",
            "active_job_status_label": "",
            "active_project_status": "",
            "active_project_status_label": "",
            "active_state_label": "완료" if completed else "",
        }
    )


def _attach_latest_failed_project_states(items):
    project_keys = [
        (item.get("center_code") or normalize_center_code(None), item.get("project_number"))
        for item in items
        if item.get("project_number")
    ]
    project_numbers = [project_number for _, project_number in project_keys]
    if not project_numbers:
        return

    try:
        finished_projects = (
            DownloadReviewProject.objects
            .filter(project_number__in=project_numbers)
            .exclude(job__status__in=ACTIVE_JOB_STATUSES)
            .exclude(job__status=DownloadReviewJobStatus.CANCELED)
            .order_by("-completed_at", "-updated_at", "-created_at", "-id")
        )
        latest_by_number = {}
        for project in finished_projects:
            latest_by_number.setdefault((project.center_code, project.project_number), project)
    except DatabaseError:
        logger.info("Failed to read failed download-review project states; continuing with reference state.")
        return

    for item in items:
        item_center = item.get("center_code") or normalize_center_code(None)
        failed_project = latest_by_number.get((item_center, item.get("project_number")))
        if (
            not failed_project
            or failed_project.status != DownloadReviewProjectStatus.FAILED
            or failed_project.review_status != DownloadReviewProjectReviewStatus.HELD
        ):
            continue
        item.update(
            {
                "review": "실패",
                "review_raw": "실패",
                "latest_failed_job_project_id": str(failed_project.id),
                "latest_failed_job_id": str(failed_project.job_id),
                "latest_failed_step": failed_project.current_step,
                "latest_failed_message": failed_project.error_message,
                "latest_failed_detail": failed_project.error_detail,
            }
        )


def get_latest_project_results_payload(project_number, center_code=None):
    center_code = parse_center_code(center_code)
    project_number = str(project_number or "").strip()
    if not PROJECT_NUMBER_RE.match(project_number):
        raise DownloadReviewNotFoundError("작업 프로젝트를 찾을 수 없습니다.")

    project = (
        DownloadReviewProject.objects
        .select_related("job")
        .exclude(job__status__in=ACTIVE_JOB_STATUSES)
        .exclude(job__status=DownloadReviewJobStatus.CANCELED)
        .filter(center_code=center_code, project_number=project_number)
        .order_by("-completed_at", "-updated_at", "-created_at", "-id")
        .first()
    )
    if project is None:
        raise DownloadReviewNotFoundError("점검 이력을 찾을 수 없습니다.")

    results = list(project.rule_results.order_by("sequence", "id"))
    items = [serialize_rule_result(result) for result in results]
    project_payload = serialize_project(project)
    project_payload["change_note"] = change_note_summary(project)
    return {
        "success": True,
        "job": serialize_job(project.job),
        "project": project_payload,
        "items": items,
        "display_items": _display_items_for_results(results, items),
    }


def get_bulk_projects_zip_response(project_numbers, center_code=None):
    center_code = parse_center_code(center_code)
    valid_numbers = [
        n for n in project_numbers
        if PROJECT_NUMBER_RE.match(str(n or "").strip())
    ]
    if not valid_numbers:
        raise DownloadReviewJobRequestError("유효한 프로젝트번호가 없습니다.")

    projects = (
        DownloadReviewProject.objects
        .select_related("job")
        .exclude(job__status__in=ACTIVE_JOB_STATUSES)
        .exclude(job__status=DownloadReviewJobStatus.CANCELED)
        .filter(center_code=center_code, project_number__in=valid_numbers)
        .order_by("project_number", "-completed_at", "-updated_at", "-created_at", "-id")
    )

    latest_by_number = {}
    for project in projects:
        if project.project_number not in latest_by_number:
            latest_by_number[project.project_number] = project

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for number in valid_numbers:
            project = latest_by_number.get(number)
            if project is None:
                continue
            results = list(project.rule_results.order_by("sequence", "id"))
            xlsx_bytes = _xlsx_project_bytes(project, results)
            zf.writestr(f"{number}_점검결과.xlsx", xlsx_bytes)

    zip_buffer.seek(0)
    content = zip_buffer.getvalue()
    response = HttpResponse(content, content_type="application/zip")
    zip_filename = "점검결과_일괄다운로드.zip"
    response["Content-Disposition"] = (
        f"attachment; filename=\"bulk_inspection_results.zip\"; filename*=UTF-8''{quote(zip_filename)}"
    )
    response["Cache-Control"] = "no-store"
    return response


def _xlsx_bytes(rows, sheet_title):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise DownloadReviewJobRequestError("엑셀 파일 생성을 위해 openpyxl이 필요합니다.") from exc

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = str(sheet_title or "점검결과")[:31]
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="E7EEF8")
    header_font = Font(bold=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if _is_header_row(row):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)
    worksheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _split_slash(value):
    if not value:
        return []
    return [p.strip() for p in str(value).split(" / ")]


def _rule_result_sub_rows(result):
    """Return [(expected, actual, passed_or_None), ...] for sub-rows, or [] for a single row."""
    raw_detail = result.raw_detail_json or {}
    subs = raw_detail.get("sub_checks")
    if isinstance(subs, list) and len(subs) > 1:
        return [
            (str(s.get("expected") or ""), str(s.get("actual") or ""), s.get("passed"))
            for s in subs
        ]
    exp_parts = _split_slash(result.expected)
    act_parts = _split_slash(result.actual)
    count = max(len(exp_parts), len(act_parts))
    if count <= 1:
        return []
    return [
        (
            exp_parts[i] if i < len(exp_parts) else "",
            act_parts[i] if i < len(act_parts) else "",
            None,
        )
        for i in range(count)
    ]


def _write_project_to_ws(ws, project, results):
    """Write project inspection results to worksheet, splitting ' / ' values into merged sub-rows."""
    from openpyxl.styles import Alignment, Font, PatternFill

    top_wrap = Alignment(vertical="top", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="E7EEF8")
    header_font = Font(bold=True)

    ecm_row = project.ecm_row_json or {}
    for row_data in [
        ["프로젝트번호", project.project_number],
        ["회사명", ecm_row.get("company", "")],
        ["제품명", ecm_row.get("product", "")],
        ["작업상태", project_status_label(project.status)],
        ["점검결과", review_status_label(project.review_status)],
        ["현재단계", project.current_step],
        ["오류", project.error_message],
        [],
    ]:
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.alignment = top_wrap

    ws.append(_rule_result_excel_header())
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = top_wrap

    if not results:
        ws.append(["", "", "", "", "", "", "", "생성된 규칙 결과가 없습니다.", ""])
        for cell in ws[ws.max_row]:
            cell.alignment = top_wrap
        return

    result_list = list(results)
    for row in build_display_rows(result_list):
        parent = result_list[row.parent_index - 1]
        ws.append(_rule_display_excel_row(project, row, parent))
        for cell in ws[ws.max_row]:
            cell.alignment = top_wrap


def _xlsx_project_bytes(project, results):
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise DownloadReviewJobRequestError("엑셀 파일 생성을 위해 openpyxl이 필요합니다.") from exc

    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "점검결과"
    _write_project_to_ws(ws, project, results)

    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(cell.value or "")) for cell in col_cells if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)
    ws.freeze_panes = "A10"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _xlsx_project_response(filename, project, results):
    content = _xlsx_project_bytes(project, results)
    response = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=\"download_review_results.xlsx\"; filename*=UTF-8''{quote(filename)}"
    )
    response["Cache-Control"] = "no-store"
    return response


def cancel_download_review_job(job_id):
    workflow_alias = getattr(settings, "WORKFLOW_DATABASE_ALIAS", "workflow")
    with transaction.atomic(using=workflow_alias):
        try:
            job = DownloadReviewJob.objects.select_for_update().get(id=job_id)
        except DownloadReviewJob.DoesNotExist as exc:
            raise DownloadReviewNotFoundError("작업을 찾을 수 없습니다.") from exc

        if job.status not in CANCELABLE_JOB_STATUSES:
            raise DownloadReviewJobCancelError(
                "예약됨 또는 대기중인 작업만 취소할 수 있습니다.",
                details={
                    "job_id": str(job.id),
                    "status": job.status,
                    "status_label": job_status_label(job.status),
                },
            )

        current = timezone.now()
        job.status = DownloadReviewJobStatus.CANCELED
        job.canceled_at = current
        job.progress_message = "사용자가 작업을 취소했습니다."
        job.save(update_fields=["status", "canceled_at", "progress_message", "updated_at"])

        job.projects.exclude(
            status__in=(
                DownloadReviewProjectStatus.COMPLETED,
                DownloadReviewProjectStatus.FAILED,
                DownloadReviewProjectStatus.SKIPPED,
            )
        ).update(
            status=DownloadReviewProjectStatus.SKIPPED,
            current_step="사용자 취소",
            completed_at=current,
        )
        DownloadReviewLog.objects.create(
            job=job,
            level=DownloadReviewLogLevel.INFO,
            event_code="job_canceled",
            message="사용자가 예약/대기 작업을 취소했습니다.",
        )

    return {
        "success": True,
        "job": serialize_job(job),
        "message": "예약된 작업을 취소했습니다.",
    }


def force_stop_download_review_jobs(job_id=None):
    """진행중(RUNNING 포함) 작업을 강제 종료하고 워커 락을 해제한다.

    워커가 비정상 종료되어 작업이 RUNNING 상태로 멈추고 락(DownloadReviewLock)이
    잠긴 채 남아 새 작업을 시작할 수 없을 때 사용한다.
    cancel_download_review_job 과 달리 RUNNING 작업도 종료하고 락까지 해제한다.

    Args:
        job_id: 특정 작업만 종료. None 이면 활성(SCHEDULED/QUEUED/RUNNING) 작업 전체.

    주의: 이미 실행 중인 워커 OS 프로세스나 브라우저는 종료하지 않는다.
    DB 상태(작업/프로젝트/락)만 정리한다. 워커 프로세스가 살아 있다면 먼저
    중지(stop_worker)해야 한다.
    """
    workflow_alias = getattr(settings, "WORKFLOW_DATABASE_ALIAS", "workflow")
    current = timezone.now()
    stopped_ids = []

    with transaction.atomic(using=workflow_alias):
        jobs_qs = DownloadReviewJob.objects.select_for_update()
        if job_id is not None:
            jobs_qs = jobs_qs.filter(id=job_id)
        else:
            jobs_qs = jobs_qs.filter(status__in=ACTIVE_JOB_STATUSES)
        jobs = list(jobs_qs)

        if job_id is not None and not jobs:
            raise DownloadReviewNotFoundError("작업을 찾을 수 없습니다.")

        for job in jobs:
            if job.status not in ACTIVE_JOB_STATUSES:
                # 이미 종료된 작업은 건너뛴다(특정 job_id 지정 시).
                continue
            job.status = DownloadReviewJobStatus.CANCELED
            job.canceled_at = current
            job.completed_at = job.completed_at or current
            job.progress_message = "사용자가 작업을 강제 종료했습니다."
            job.last_error_message = "사용자 강제 종료"
            job.save(update_fields=[
                "status", "canceled_at", "completed_at",
                "progress_message", "last_error_message", "updated_at",
            ])
            job.projects.exclude(
                status__in=(
                    DownloadReviewProjectStatus.COMPLETED,
                    DownloadReviewProjectStatus.FAILED,
                    DownloadReviewProjectStatus.SKIPPED,
                )
            ).update(
                status=DownloadReviewProjectStatus.SKIPPED,
                current_step="사용자 강제 종료",
                completed_at=current,
                updated_at=current,
            )
            DownloadReviewLog.objects.create(
                job=job,
                level=DownloadReviewLogLevel.WARNING,
                event_code="job_force_stopped",
                message="사용자가 진행중 작업을 강제 종료했습니다.",
            )
            stopped_ids.append(job.id)

        # 워커 락 해제 — 강제 종료의 핵심. 락이 종료 대상 작업을 가리키거나,
        # 전체 강제 종료이거나, 락이 가리키는 작업이 더 이상 활성이 아니면 해제한다.
        lock = DownloadReviewLock.objects.select_for_update().filter(id=1).first()
        lock_released = False
        if lock and lock.locked:
            should_release = (
                job_id is None
                or lock.job_id in stopped_ids
                or lock.job_id is None
            )
            if not should_release and lock.job_id is not None:
                locked_job = DownloadReviewJob.objects.filter(id=lock.job_id).first()
                if locked_job is None or locked_job.status not in ACTIVE_JOB_STATUSES:
                    should_release = True
            if should_release:
                lock.locked = False
                lock.owner = ""
                lock.job = None
                lock.locked_at = None
                lock.heartbeat_at = None
                lock.note = ""
                lock.save(update_fields=[
                    "locked", "owner", "job", "locked_at",
                    "heartbeat_at", "note", "updated_at",
                ])
                lock_released = True

    return {
        "success": True,
        "stopped_count": len(stopped_ids),
        "lock_released": lock_released,
        "message": (
            f"진행중 작업 {len(stopped_ids)}건을 강제 종료했습니다."
            + (" 워커 락을 해제했습니다." if lock_released else "")
            if stopped_ids or lock_released
            else "강제 종료할 진행중 작업이 없습니다."
        ),
    }


def get_job_detail_payload(job_id):
    job = get_job_or_raise(job_id)
    return {
        "success": True,
        "job": serialize_job(job),
        "polling": polling_hint(job),
    }


def get_job_projects_payload(job_id):
    job = get_job_or_raise(job_id)
    projects = job.projects.order_by("created_at", "id")
    return {
        "success": True,
        "job": serialize_job(job),
        "items": [serialize_project(project) for project in projects],
    }


def get_project_results_payload(job_project_id):
    try:
        project = (
            DownloadReviewProject.objects
            .select_related("job")
            .get(id=job_project_id)
        )
    except DownloadReviewProject.DoesNotExist as exc:
        raise DownloadReviewNotFoundError("작업 프로젝트를 찾을 수 없습니다.") from exc

    results = list(project.rule_results.order_by("sequence", "id"))
    items = [serialize_rule_result(result) for result in results]
    project_payload = serialize_project(project)
    project_payload["change_note"] = change_note_summary(project)
    return {
        "success": True,
        "job": serialize_job(project.job),
        "project": project_payload,
        "items": items,
        "display_items": _display_items_for_results(results, items),
    }


def mark_rule_result_manual_pass(result_id, memo, *, requested_by=""):
    memo = str(memo or "").strip()
    if not memo:
        raise DownloadReviewJobRequestError("수동 적합 처리 사유를 입력해야 합니다.")

    workflow_alias = getattr(settings, "WORKFLOW_DATABASE_ALIAS", "workflow")
    with transaction.atomic(using=workflow_alias):
        try:
            result = _get_rule_result_for_manual_pass(result_id, for_update=True)
        except DownloadReviewRuleResult.DoesNotExist as exc:
            raise DownloadReviewNotFoundError("점검 결과를 찾을 수 없습니다.") from exc

        if not result.rule_code:
            raise DownloadReviewJobRequestError("규칙 코드가 없는 결과는 수동 적합 처리할 수 없습니다.")

        context = _manual_pass_context(result, memo, requested_by)

    override, override_persisted = _manual_override_for_context(context)

    with transaction.atomic(using=workflow_alias):
        try:
            result = _get_rule_result_for_manual_pass(result_id, for_update=True)
        except DownloadReviewRuleResult.DoesNotExist as exc:
            raise DownloadReviewNotFoundError("점검 결과를 찾을 수 없습니다.") from exc

        apply_manual_override_to_result(result, override, save=True)
        if override_persisted:
            _safe_mark_overrides_applied([override])

    _safe_recalculate_project_review_after_manual_override(context["job_project_id"])
    _safe_create_download_log(
        job_id=context.get("job_id"),
        job_project_id=context.get("job_project_id"),
        level=DownloadReviewLogLevel.INFO,
        event_code="manual_pass_override",
        message=f"{context.get('project_number')} {context.get('rule_name')} 수동 적합 처리",
        detail_json={
            "center_code": context.get("center_code"),
            "project_number": context.get("project_number"),
            "rule_code": context.get("rule_code"),
            "rule_name": context.get("rule_name"),
            "memo": context.get("memo"),
            "requested_by": context.get("requested_by"),
            "override_persisted": override_persisted,
        },
    )
    return _manual_pass_response_payload(context["job_project_id"], result_id)


def _get_rule_result_for_manual_pass(result_id, *, for_update=False):
    queryset = (
        DownloadReviewRuleResult.objects
        .select_related("job_project", "job_project__job")
    )
    if for_update:
        queryset = queryset.select_for_update()
    return queryset.get(id=result_id)


def _manual_pass_context(result, memo, requested_by):
    project = result.job_project
    return {
        "job_id": project.job_id,
        "job_project_id": project.id,
        "center_code": str(project.center_code or "").strip(),
        "project_number": str(project.project_number or "").strip(),
        "rule_code": str(result.rule_code or "").strip(),
        "rule_name": result.rule_name,
        "memo": memo,
        "requested_by": requested_by or "",
    }


def _manual_override_for_context(context):
    for attempt in range(3):
        try:
            override, _created = DownloadReviewManualOverride.objects.update_or_create(
                center_code=context["center_code"],
                project_number=context["project_number"],
                rule_code=context["rule_code"],
                defaults={
                    "rule_name": context["rule_name"],
                    "memo": context["memo"],
                    "created_by": context["requested_by"],
                },
            )
            return override, True
        except OperationalError as exc:
            if _is_sqlite_locked_error(exc) and attempt < 2:
                time_to_sleep = 0.15 * (attempt + 1)
                time.sleep(time_to_sleep)
                continue
            if _is_manual_override_storage_unavailable(exc):
                logger.exception("Manual override table unavailable; using log fallback: %s", exc)
                return _ephemeral_manual_override(context), False
            raise
        except DatabaseError as exc:
            if _is_manual_override_storage_unavailable(exc):
                logger.exception("Manual override table unavailable; using log fallback: %s", exc)
                return _ephemeral_manual_override(context), False
            raise
    return _ephemeral_manual_override(context), False


def _ephemeral_manual_override(context):
    return SimpleNamespace(
        id=f"log:{context['center_code']}:{context['project_number']}:{context['rule_code']}",
        memo=context["memo"],
        rule_code=context["rule_code"],
        rule_name=context["rule_name"],
        updated_at=timezone.now(),
    )


def _is_sqlite_locked_error(exc):
    return "database is locked" in str(exc).lower()


def _is_manual_override_storage_unavailable(exc):
    message = str(exc).lower()
    return (
        "inspection_manual_override" in message
        and (
            "no such table" in message
            or "no such column" in message
            or "has no column" in message
            or "does not exist" in message
            or "undefinedtable" in message
        )
    )


def _safe_mark_overrides_applied(overrides):
    try:
        mark_overrides_applied(overrides)
    except Exception as exc:
        logger.warning("Manual override applied timestamp update failed: %s", exc, exc_info=True)


def get_project_change_note_payload(job_project_id):
    try:
        project = (
            DownloadReviewProject.objects
            .select_related("job")
            .get(id=job_project_id)
        )
    except DownloadReviewProject.DoesNotExist as exc:
        raise DownloadReviewNotFoundError("작업 프로젝트를 찾을 수 없습니다.") from exc

    note = change_note_payload(project)
    if not note.get("available"):
        raise DownloadReviewNotFoundError("수정 내용 파일을 찾을 수 없습니다.")

    return {
        "success": True,
        "job": serialize_job(project.job),
        "project": serialize_project(project),
        "change_note": note,
    }


def get_project_results_excel_response(job_project_id):
    try:
        project = (
            DownloadReviewProject.objects
            .select_related("job")
            .get(id=job_project_id)
        )
    except DownloadReviewProject.DoesNotExist as exc:
        raise DownloadReviewNotFoundError("작업 프로젝트를 찾을 수 없습니다.") from exc

    results = list(project.rule_results.order_by("sequence", "id"))
    filename = f"{project.project_number}_점검결과.xlsx"
    return _xlsx_project_response(filename, project, results)


def get_job_results_excel_response(job_id):
    job = get_job_or_raise(job_id)
    projects = (
        job.projects
        .order_by("created_at", "id")
        .prefetch_related("rule_results")
    )
    rows = [_project_summary_excel_header()]
    for project in projects:
        project_results = sorted(project.rule_results.all(), key=lambda item: (item.sequence, str(item.id)))
        rows.extend(_project_summary_excel_rows(project, project_results))
    filename = f"download_review_{job.id}_전체점검결과.xlsx"
    return _xlsx_response(filename, "전체점검결과", rows)


def find_active_job():
    running = (
        DownloadReviewJob.objects
        .filter(status=DownloadReviewJobStatus.RUNNING)
        .order_by("started_at", "requested_at", "id")
        .first()
    )
    if running:
        return running

    queued = (
        DownloadReviewJob.objects
        .filter(status=DownloadReviewJobStatus.QUEUED)
        .order_by("queued_at", "requested_at", "id")
        .first()
    )
    if queued:
        return queued

    return (
        DownloadReviewJob.objects
        .filter(status=DownloadReviewJobStatus.SCHEDULED)
        .order_by("available_after", "requested_at", "id")
        .first()
    )


def parse_json_body(request):
    try:
        raw_body = request.body.decode("utf-8") if request.body else "{}"
        payload = json.loads(raw_body)
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise DownloadReviewJobRequestError("JSON 요청 본문이 올바르지 않습니다.") from exc

    if not isinstance(payload, dict):
        raise DownloadReviewJobRequestError("JSON 요청 본문은 객체여야 합니다.")
    return payload


def parse_center_code(value):
    try:
        return normalize_center_code(value)
    except ValueError as exc:
        raise DownloadReviewJobRequestError(str(exc)) from exc


def parse_job_list_query(query_params):
    unknown = set(query_params.keys()) - JOB_LIST_PARAM_NAMES
    if unknown:
        names = ", ".join(sorted(unknown))
        raise DownloadReviewJobRequestError(f"지원하지 않는 작업 조회 조건입니다: {names}")

    status = str(query_params.get("status") or "all").strip()
    if status not in JOB_LIST_STATUS_FILTERS:
        raise DownloadReviewJobRequestError(f"지원하지 않는 작업 상태 필터입니다: {status}")
    center = None
    if query_params.get("center"):
        center = parse_center_code(query_params.get("center"))

    limit = _parse_int(query_params.get("limit"), DEFAULT_JOB_LIST_LIMIT, "limit")
    offset = _parse_int(query_params.get("offset"), 0, "offset")
    if limit < 1 or limit > MAX_JOB_LIST_LIMIT:
        raise DownloadReviewJobRequestError(f"limit은 1부터 {MAX_JOB_LIST_LIMIT} 사이여야 합니다.")
    if offset < 0:
        raise DownloadReviewJobRequestError("offset은 0 이상이어야 합니다.")

    return {
        "status": status,
        "center": center,
        "limit": limit,
        "offset": offset,
    }


def parse_project_numbers(payload):
    project_numbers = payload.get("project_numbers")
    if not isinstance(project_numbers, list) or not project_numbers:
        raise DownloadReviewJobRequestError("project_numbers는 1개 이상의 프로젝트번호 배열이어야 합니다.")

    max_projects = getattr(settings, "DOWNLOAD_REVIEW_MAX_PROJECTS_PER_JOB", 100)
    if len(project_numbers) > max_projects:
        raise DownloadReviewJobRequestError(f"한 작업에는 최대 {max_projects}개 프로젝트만 요청할 수 있습니다.")

    normalized = []
    duplicates = []
    seen = set()
    for number in project_numbers:
        value = str(number).strip() if number is not None else ""
        if not PROJECT_NUMBER_RE.match(value):
            raise DownloadReviewJobRequestError(
                "프로젝트번호 형식이 올바르지 않습니다.",
                details={"project_number": value},
            )
        if value in seen:
            duplicates.append(value)
            continue
        seen.add(value)
        normalized.append(value)

    if duplicates:
        raise DownloadReviewJobRequestError(
            "중복된 프로젝트번호가 포함되어 있습니다.",
            details={"duplicates": duplicates},
        )
    return normalized


def build_job_schedule(now=None):
    current = now or timezone.now()
    local_tz = ZoneInfo(getattr(settings, "DOWNLOAD_REVIEW_TIME_ZONE", "Asia/Seoul"))
    local_now = current.astimezone(local_tz)

    if is_start_window(local_now):
        return JobSchedule(
            status=DownloadReviewJobStatus.QUEUED,
            available_after=current,
            queued_at=current,
        )

    next_start_local = local_now.replace(
        hour=getattr(settings, "DOWNLOAD_REVIEW_START_HOUR", 20),
        minute=0,
        second=0,
        microsecond=0,
    )
    if local_now.hour >= getattr(settings, "DOWNLOAD_REVIEW_START_HOUR", 20):
        next_start_local = next_start_local + timedelta(days=1)

    return JobSchedule(
        status=DownloadReviewJobStatus.SCHEDULED,
        available_after=next_start_local.astimezone(datetime_timezone.utc),
        queued_at=None,
    )


def is_start_window(local_time):
    if getattr(settings, "DOWNLOAD_REVIEW_IGNORE_TIME_WINDOW", False):
        return True

    start_hour = getattr(settings, "DOWNLOAD_REVIEW_START_HOUR", 20)
    end_hour = getattr(settings, "DOWNLOAD_REVIEW_END_HOUR", 7)
    hour = local_time.hour
    if start_hour > end_hour:
        return hour >= start_hour or hour < end_hour
    return start_hour <= hour < end_hour


def job_status_label(status):
    labels = {
        DownloadReviewJobStatus.SCHEDULED: "예약됨",
        DownloadReviewJobStatus.QUEUED: "대기중",
        DownloadReviewJobStatus.RUNNING: "진행중",
        DownloadReviewJobStatus.COMPLETED: "완료",
        DownloadReviewJobStatus.FAILED: "실패",
        DownloadReviewJobStatus.CANCELED: "취소",
    }
    return labels.get(status, status)


def project_status_label(status):
    labels = {
        DownloadReviewProjectStatus.QUEUED: "대기중",
        DownloadReviewProjectStatus.RUNNING: "진행중",
        DownloadReviewProjectStatus.DOWNLOADED: "다운로드완료",
        DownloadReviewProjectStatus.INSPECTING: "검사중",
        DownloadReviewProjectStatus.COMPLETED: "완료",
        DownloadReviewProjectStatus.FAILED: "실패",
        DownloadReviewProjectStatus.SKIPPED: "건너뜀",
    }
    return labels.get(status, status)


def active_project_state_label(job_status, project_status):
    if job_status == DownloadReviewJobStatus.SCHEDULED:
        return "예약중"
    if job_status == DownloadReviewJobStatus.QUEUED:
        return "대기중"
    if job_status == DownloadReviewJobStatus.RUNNING:
        return project_status_label(project_status)
    return job_status_label(job_status)


def review_status_label(status):
    labels = {
        DownloadReviewProjectReviewStatus.UNREVIEWED: "미점검",
        DownloadReviewProjectReviewStatus.COMPLETED: "완료",
        DownloadReviewProjectReviewStatus.NEEDS_FIX: "수정 필요",
        DownloadReviewProjectReviewStatus.HELD: "보류",
    }
    return labels.get(status, status)


def rule_status_label(status):
    labels = {
        DownloadReviewRuleStatus.PASS: "정상",
        DownloadReviewRuleStatus.FAIL: "부적합",
        DownloadReviewRuleStatus.WARNING: "경고",
        DownloadReviewRuleStatus.ERROR: "오류",
    }
    return labels.get(status, status)


def _recalculate_project_review_after_manual_override(project):
    results = list(project.rule_results.order_by("sequence", "id"))
    failed_count = sum(
        1
        for result in results
        if result.status in (DownloadReviewRuleStatus.FAIL, DownloadReviewRuleStatus.ERROR)
    )
    total_count = len(results)
    passed_count = total_count - failed_count
    project.review_status = (
        DownloadReviewProjectReviewStatus.NEEDS_FIX
        if failed_count
        else DownloadReviewProjectReviewStatus.COMPLETED
    )
    project.current_step = f"점검 완료: 정상 {passed_count}건, 부적합 {failed_count}건"
    project.save(update_fields=["review_status", "current_step", "updated_at"])

    reference_review = "X" if failed_count else "O"
    try:
        artifact_results = _artifact_results_from_rule_results(results)
        write_project_review_result(
            project.project_number,
            reference_review,
            artifact_results=artifact_results,
            inspected_at=timezone.now(),
            center_code=project.center_code,
        )
    except Exception as exc:
        logger.warning("Manual override reference writeback failed: %s (%s)", project.project_number, exc)
        _safe_create_download_log(
            job_id=project.job_id,
            job_project_id=project.id,
            level=DownloadReviewLogLevel.WARNING,
            event_code="manual_pass_reference_write_failed",
            message=f"{project.project_number} 수동 적합 기준 DB 반영 실패",
            detail_json={"reference_review": reference_review},
            admin_only=True,
        )


def _safe_recalculate_project_review_after_manual_override(project_id):
    try:
        project = (
            DownloadReviewProject.objects
            .select_related("job")
            .get(id=project_id)
        )
        _recalculate_project_review_after_manual_override(project)
    except Exception as exc:
        logger.exception("Manual override project recalculation failed: %s (%s)", project_id, exc)
        _safe_create_download_log(
            job_project_id=project_id,
            level=DownloadReviewLogLevel.WARNING,
            event_code="manual_pass_recalculate_failed",
            message="수동 적합 후 프로젝트 상태 재계산 실패",
            detail_json={},
            admin_only=True,
        )


def _safe_create_download_log(**kwargs):
    try:
        DownloadReviewLog.objects.create(**kwargs)
    except Exception as exc:
        logger.warning("Download review log write failed: %s", exc, exc_info=True)


def _manual_pass_response_payload(project_id, result_id):
    try:
        return get_project_results_payload(project_id)
    except Exception as exc:
        logger.exception("Manual override response refresh failed: %s (%s)", project_id, exc)
        try:
            result = (
                DownloadReviewRuleResult.objects
                .select_related("job_project", "job_project__job")
                .get(id=result_id)
            )
            project = result.job_project
            item = serialize_rule_result(result)
            return {
                "success": True,
                "message": "수동 적합 처리는 저장되었습니다. 결과 목록을 다시 조회해 주세요.",
                "job": serialize_job(project.job),
                "project": serialize_project(project),
                "items": [item],
                "display_items": _display_items_for_results([result], [item]),
            }
        except Exception:
            logger.exception("Manual override fallback payload failed: %s", result_id)
            return {
                "success": True,
                "message": "수동 적합 처리는 저장되었습니다. 페이지를 새로고침해 결과를 확인해 주세요.",
                "items": [],
                "display_items": [],
            }


def _artifact_results_from_rule_results(results):
    rules_by_code = {
        rule.code: rule
        for rule in DownloadReviewRule.objects.filter(
            code__in=sorted({result.rule_code for result in results if result.rule_code})
        )
    }
    artifact_results = {}
    for result in results:
        column = _artifact_column_for_result(result, rules_by_code.get(result.rule_code))
        if not column:
            continue
        value = "O" if result.status == DownloadReviewRuleStatus.PASS else "X"
        if artifact_results.get(column) == "X":
            continue
        artifact_results[column] = value
    return artifact_results


def _artifact_column_for_result(result, rule):
    raw_detail = result.raw_detail_json or {}
    configured = ""
    if isinstance(raw_detail, dict):
        configured = str(raw_detail.get("artifact_column") or "").strip()
    if not configured and rule is not None:
        configured = str((rule.config_json or {}).get("artifact_column") or "").strip()
    candidates = (configured, result.rule_code, result.rule_name)
    for candidate in candidates:
        if candidate in ARTIFACT_REVIEW_COLUMNS:
            return candidate
    return ""


def polling_hint(job):
    status = job.status
    if status in (DownloadReviewJobStatus.RUNNING, DownloadReviewJobStatus.QUEUED):
        return {
            "should_poll": True,
            "recommended_interval_ms": 3000,
            "wake_at": None,
        }
    if status == DownloadReviewJobStatus.SCHEDULED:
        return {
            "should_poll": False,
            "recommended_interval_ms": None,
            "wake_at": _iso(job.available_after),
        }
    return {
        "should_poll": False,
        "recommended_interval_ms": None,
        "wake_at": None,
    }


def job_request_message(status):
    if status == DownloadReviewJobStatus.SCHEDULED:
        return "작업 요청이 예약되었습니다. 시작 가능 시간이 되면 요청 순서대로 진행합니다."
    return "작업 요청이 대기열에 등록되었습니다. 요청 순서대로 진행합니다."


def get_job_or_raise(job_id):
    try:
        return DownloadReviewJob.objects.get(id=job_id)
    except DownloadReviewJob.DoesNotExist as exc:
        raise DownloadReviewNotFoundError("작업을 찾을 수 없습니다.") from exc


def serialize_job(job):
    total = job.requested_project_count or job.projects.count()
    completed = job.completed_project_count
    failed = job.failed_project_count
    progress_percent = int(((completed + failed) / total) * 100) if total else 0
    return {
        "id": str(job.id),
        "center_code": job.center_code,
        "center_label": center_label(job.center_code),
        "status": job.status,
        "status_label": job_status_label(job.status),
        "requested_at": _iso(job.requested_at),
        "queued_at": _iso(job.queued_at),
        "started_at": _iso(job.started_at),
        "completed_at": _iso(job.completed_at),
        "canceled_at": _iso(job.canceled_at),
        "available_after": _iso(job.available_after),
        "progress_message": job.progress_message,
        "requested_project_count": total,
        "completed_project_count": completed,
        "failed_project_count": failed,
        "progress_percent": progress_percent,
        "selected_project_numbers": job.selected_projects_json or [],
        "last_error_message": job.last_error_message,
        "worker": {
            "pid": job.worker_pid,
            "host": job.worker_host,
            "heartbeat_at": _iso(job.worker_heartbeat_at),
        },
    }


def serialize_project(project):
    ecm_row = project.ecm_row_json or {}
    center_code = project.center_code or ecm_row.get("center_code") or normalize_center_code(None)
    return {
        "id": str(project.id),
        "job_id": str(project.job_id),
        "center_code": center_code,
        "center_label": center_label(center_code),
        "project_number": project.project_number,
        "cert_date": ecm_row.get("cert_date", ""),
        "company": ecm_row.get("company", ""),
        "product": ecm_row.get("product", ""),
        "pl": ecm_row.get("pl", ""),
        "wd": ecm_row.get("wd", ""),
        "request_date": ecm_row.get("request_date", ""),
        "contract_date": ecm_row.get("contract_date", ""),
        "status": project.status,
        "status_label": project_status_label(project.status),
        "review_status": project.review_status,
        "review_status_label": review_status_label(project.review_status),
        "current_step": project.current_step,
        "error_message": project.error_message,
        "error_detail": project.error_detail,
        "retry_count": project.retry_count,
        "zip_file_name": project.zip_file_name,
        "download_dir": _display_path(project.download_dir, project.project_number),
        "started_at": _iso(project.started_at),
        "completed_at": _iso(project.completed_at),
    }


def serialize_rule_result(result):
    project_number = result.job_project.project_number
    raw_detail = result.raw_detail_json or {}
    manual_override = manual_override_public(raw_detail)
    return {
        "id": str(result.id),
        "job_project_id": str(result.job_project_id),
        "rule_code": result.rule_code,
        "rule_name": result.rule_name,
        "sequence": result.sequence,
        "file_path": _display_path(result.file_path, project_number),
        "file_name": result.file_name,
        "status": result.status,
        "status_label": rule_status_label(result.status),
        "expected": result.expected,
        "actual": result.actual,
        "message": result.message,
        "artifacts": _serialize_artifacts(raw_detail),
        "manual_override": manual_override,
        "raw_detail": _public_raw_detail(raw_detail),
        "created_at": _iso(result.created_at),
    }


def _display_items_for_results(results, serialized_items=None):
    serialized_items = serialized_items or [serialize_rule_result(result) for result in results]
    display_items = []
    for row in build_display_rows(results):
        item = serialize_display_row(row, status_labeler=rule_status_label)
        parent = serialized_items[row.parent_index - 1] if row.parent_index - 1 < len(serialized_items) else {}
        item.update({
            "id": parent.get("id", ""),
            "job_project_id": parent.get("job_project_id", ""),
            "artifacts": parent.get("artifacts", []),
            "manual_override": parent.get("manual_override"),
            "created_at": parent.get("created_at", ""),
        })
        display_items.append(item)
    return display_items


def _project_excel_rows(project, results):
    rows = [
        ["프로젝트번호", project.project_number],
        ["회사명", (project.ecm_row_json or {}).get("company", "")],
        ["제품명", (project.ecm_row_json or {}).get("product", "")],
        ["작업상태", project_status_label(project.status)],
        ["점검결과", review_status_label(project.review_status)],
        ["현재단계", project.current_step],
        ["오류", project.error_message],
        [],
        _rule_result_excel_header(),
    ]
    result_list = list(results)
    for row in build_display_rows(result_list):
        parent = result_list[row.parent_index - 1]
        rows.append(_rule_display_excel_row(project, row, parent))
    if not result_list:
        rows.append(["", "", "", "", "", "", "", "생성된 규칙 결과가 없습니다.", ""])
    return rows


def _project_summary_excel_header():
    return [
        "프로젝트번호",
        "회사명",
        "제품명",
        "작업상태",
        "점검결과",
        "현재단계",
        "번호",
        "점검항목",
        "결과",
        "파일명",
        "기대값",
        "실제값",
        "메시지",
    ]


def _rule_result_excel_header():
    return [
        "번호",
        "점검항목",
        "결과",
        "파일명",
        "기대값",
        "실제값",
        "메시지",
        "파일경로",
        "생성일시",
    ]


def _project_summary_excel_rows(project, results):
    ecm_row = project.ecm_row_json or {}
    base = [
        project.project_number,
        ecm_row.get("company", ""),
        ecm_row.get("product", ""),
        project_status_label(project.status),
        review_status_label(project.review_status),
        project.current_step,
    ]
    if not results:
        return [base + ["", "", "", "", "", "", project.error_message or "생성된 규칙 결과가 없습니다."]]
    result_list = list(results)
    return [
        base + [
            row.display_number,
            row.rule_name,
            rule_status_label(row.status),
            row.file_name,
            row.expected,
            row.actual,
            row.message,
        ]
        for row in build_display_rows(result_list)
    ]


def _rule_result_excel_row(project, result):
    return [
        result.sequence,
        result.rule_name,
        rule_status_label(result.status),
        result.file_name,
        result.expected,
        result.actual,
        result.message,
        _display_path(result.file_path, project.project_number),
        _iso(result.created_at) or "",
    ]


def _rule_display_excel_row(project, row, parent_result):
    return [
        row.display_number,
        row.rule_name,
        rule_status_label(row.status),
        row.file_name,
        row.expected,
        row.actual,
        row.message,
        _display_path(row.file_path or parent_result.file_path, project.project_number),
        _iso(parent_result.created_at) or "",
    ]


def _xlsx_response(filename, sheet_title, rows):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise DownloadReviewJobRequestError("엑셀 파일 생성을 위해 openpyxl이 필요합니다.") from exc

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = str(sheet_title or "점검결과")[:31]
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="E7EEF8")
    header_font = Font(bold=True)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if _is_header_row(row):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)
    worksheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    content = output.getvalue()
    response = HttpResponse(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    ascii_filename = "download_review_results.xlsx"
    response["Content-Disposition"] = (
        f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{quote(filename)}"
    )
    response["Cache-Control"] = "no-store"
    return response


def _is_header_row(row):
    values = [str(cell.value or "") for cell in row]
    return "점검항목" in values and "결과" in values


def get_rule_result_artifact_response(result_id, artifact_id):
    try:
        result = DownloadReviewRuleResult.objects.select_related("job_project").get(id=result_id)
    except DownloadReviewRuleResult.DoesNotExist as exc:
        raise DownloadReviewNotFoundError("점검 산출물을 찾을 수 없습니다.") from exc

    artifact = _find_artifact(result.raw_detail_json or {}, artifact_id)
    if not artifact:
        raise DownloadReviewNotFoundError("점검 산출물을 찾을 수 없습니다.")

    base_dir = _artifact_base_dir()
    relative_path = str(artifact.get("relative_path") or "").strip()
    if not relative_path:
        raise DownloadReviewNotFoundError("점검 산출물을 찾을 수 없습니다.")

    file_path = (base_dir / relative_path).resolve()
    try:
        file_path.relative_to(base_dir)
    except ValueError as exc:
        raise DownloadReviewNotFoundError("점검 산출물을 찾을 수 없습니다.") from exc
    if not file_path.is_file():
        raise DownloadReviewNotFoundError("점검 산출물 파일이 없습니다.")

    response = FileResponse(
        file_path.open("rb"),
        content_type=artifact.get("content_type") or "application/octet-stream",
        as_attachment=bool(artifact.get("download")),
        filename=artifact.get("file_name") or file_path.name,
    )
    response["Cache-Control"] = "no-store"
    return response


def _serialize_artifacts(raw_detail):
    items = raw_detail.get("artifacts") if isinstance(raw_detail, dict) else []
    if not isinstance(items, list):
        return []

    public_items = []
    for item in items:
        if not isinstance(item, dict):
            continue
        artifact_id = str(item.get("id") or "").strip()
        if not artifact_id:
            continue
        public_items.append({
            "id": artifact_id,
            "label": str(item.get("label") or item.get("file_name") or "산출물"),
            "kind": str(item.get("kind") or "file"),
            "file_name": str(item.get("file_name") or ""),
            "content_type": str(item.get("content_type") or "application/octet-stream"),
            "download": bool(item.get("download")),
        })
    return public_items


def _public_raw_detail(raw_detail):
    if not isinstance(raw_detail, dict):
        return {}
    public = dict(raw_detail)
    if "artifacts" in public:
        public["artifacts"] = _serialize_artifacts(raw_detail)
    return public


def _find_artifact(raw_detail, artifact_id):
    artifacts = raw_detail.get("artifacts") if isinstance(raw_detail, dict) else []
    if not isinstance(artifacts, list):
        return None

    target = str(artifact_id or "").strip()
    for item in artifacts:
        if isinstance(item, dict) and str(item.get("id") or "").strip() == target:
            return item
    return None


def _artifact_base_dir():
    return Path(
        getattr(
            settings,
            "DOWNLOAD_REVIEW_ARTIFACT_DIR",
            Path(settings.BASE_DIR) / "main" / "data" / "download_review_artifacts",
        )
    ).resolve()


def _validate_projects_found(project_numbers, projects):
    missing = [
        project_number
        for project_number, project in zip(project_numbers, projects)
        if project is None
    ]
    if missing:
        raise DownloadReviewJobRequestError(
            "ecmlist.db에서 찾을 수 없는 프로젝트번호가 포함되어 있습니다.",
            details={"missing_project_numbers": missing},
        )


def _validate_not_completed(projects):
    completed = [
        project["project_number"]
        for project in projects
        if is_completed_review_value(project.get("review_raw") or project.get("review"))
    ]
    if completed:
        raise DownloadReviewCompletedProjectError(
            "이미 점검 완료된 프로젝트는 작업 요청할 수 없습니다.",
            details={"completed_project_numbers": completed},
        )


def _validate_active_job_limit():
    active_limit = getattr(settings, "DOWNLOAD_REVIEW_ACTIVE_JOB_LIMIT", 5)
    active_count = DownloadReviewJob.objects.filter(status__in=ACTIVE_JOB_STATUSES).count()
    if active_count >= active_limit:
        raise DownloadReviewQueueFullError(
            f"등록 가능한 작업 수({active_limit}개)를 초과했습니다.",
            details={"active_job_count": active_count, "active_job_limit": active_limit},
        )


def _validate_no_active_project(project_numbers, center_code):
    conflicts = list(
        DownloadReviewProject.objects.filter(
            center_code=center_code,
            project_number__in=project_numbers,
            job__status__in=ACTIVE_JOB_STATUSES,
        )
        .select_related("job")
        .values("project_number", "job_id", "job__status")
        .order_by("project_number")
    )
    if conflicts:
        raise DownloadReviewDuplicateProjectError(
            "이미 예약됨, 대기중 또는 진행중인 프로젝트가 포함되어 있습니다.",
            details={
                "conflicts": [
                    {
                        "project_number": item["project_number"],
                        "job_id": str(item["job_id"]),
                        "job_status": item["job__status"],
                        "job_status_label": job_status_label(item["job__status"]),
                    }
                    for item in conflicts
                ]
            },
        )


def _initial_progress_message(status):
    if status == DownloadReviewJobStatus.SCHEDULED:
        return "시작 가능 시간까지 예약됨"
    return "대기열 등록 완료"


def _iso(value):
    if not value:
        return None
    return value.isoformat()


def _display_path(path, project_number):
    if not path:
        return ""
    normalized = str(path).replace("\\", "/")
    marker = project_number
    index = normalized.find(marker)
    if index >= 0:
        return normalized[index:]
    return ""


def _parse_int(value, default, name):
    if value in (None, ""):
        return default
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise DownloadReviewJobRequestError(f"{name}은 숫자여야 합니다.") from exc
