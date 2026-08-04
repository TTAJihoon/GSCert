"""GSCert 점검 엔진 (Django 비종속, 웹/로컬 공용).

서버 main/views/review/ecm_download_review_inspection.py 의 점검 로직을 그대로
이식했다. Django 결합부(오케스트레이션·컨텍스트 생성·산출물 저장·reference.db
조회)는 어댑터로 분리하고, 여기서는 순수 평가 로직만 둔다.

진입점: evaluate_rules(rules, context, files, project=None, sink=None)
"""

import contextvars
import fnmatch
import json
import os
import re
import struct
import unicodedata
from dataclasses import dataclass, field, replace
from datetime import datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import BadZipFile, ZipFile

from lxml import etree

from .artifacts import NoOpArtifactSink


class DownloadReviewRuleStatus:
    """Django DownloadReviewRuleStatus 와 동일한 문자열 값 shim."""
    PASS = "pass"
    FAIL = "fail"
    WARNING = "warning"
    ERROR = "error"
    UNSUPPORTED = "unsupported"


@dataclass(frozen=True)
class FileInfo:
    """점검 대상 파일. path 는 실제 디스크 경로(또는 zip 내부는 'zip::inner')."""
    name: str
    path: str
    size: int = 0
    extension: str = ""
    modified_at: "datetime | None" = None


# 산출물 sink — 동시 점검 안전을 위해 contextvar 사용(스레드/컨텍스트별 격리).
# 어댑터가 set_artifact_sink 로 주입. 기본은 no-op.
_ARTIFACT_SINK_VAR = contextvars.ContextVar("gscert_artifact_sink", default=None)


def _current_sink():
    sink = _ARTIFACT_SINK_VAR.get()
    return sink if sink is not None else NoOpArtifactSink()


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".gif"}
WORD_EXTENSIONS = (".docx", ".docm")


class DownloadReviewInspectionError(RuntimeError):
    """검사 규칙을 실행할 수 없을 때 발생한다."""


class DownloadReviewCleanupSafetyError(RuntimeError):
    """다운로드 폴더 삭제 대상이 안전하지 않을 때 발생한다."""


@dataclass(frozen=True)
class RuleEvaluation:
    rule: Any
    sequence: int
    status: str
    expected: str
    actual: str
    message: str
    file_path: str = ""
    file_name: str = ""
    raw_detail: dict | None = None


@dataclass(frozen=True)
class RuleContext:
    project_number: str
    product_raw: str
    product: str
    version: str
    company: str
    pl: str
    wd: str
    start_date: str
    end_date: str
    year: str
    request_date: str
    contract_date: str
    certification_committee_date: str
    derived_variables: dict[str, object]
    center: str = ""


@dataclass(frozen=True)
class InspectionOutcome:
    project_review_status: str
    reference_review: str
    artifact_results: dict
    passed_count: int
    failed_count: int
    result_count: int


@dataclass(frozen=True)
class CleanupOutcome:
    deleted: bool
    message: str
    file_count: int = 0


@dataclass(frozen=True)
class ExcelSheet:
    name: str
    rows: list[list[str]]
    header_text: str = ""
    footer_text: str = ""


@dataclass(frozen=True)
class ExcelWorkbook:
    sheets: list[ExcelSheet]


# 엔진이 평가할 수 있는 rule_type 목록(단일 진실 소스).
# _evaluate_rule 의 분기와 일치해야 하며, 로컬 앱은 이 집합을 import 해 미지원 판정을 한다
# (앱이 별도 하드코딩 목록을 두면 엔진에 규칙 추가 시 멀쩡한 규칙이 '미지원'으로 오판됨).
SUPPORTED_RULE_TYPES = frozenset({
    "min_file_count",
    "filename_contains_project_number",
    "required_extension",
    "required_file_name_contains",
    "required_artifact_file",
    "downloadable_artifact_check",
    "document_artifact_check",
    "all_files_non_empty",
    "excel_feature_list_check",
    "test_plan_document_check",
    "image_screenshot_folder_date_check",
    "test_case_check",
    "rawdata_folder_structure_check",
    "test_report_document_check",
    "defect_report_check",
    "inspection_checklist_check",
    "quality_inspection_table_check",
    "quality_evaluation_report_check",
})


def _evaluate_rule(rule, sequence, project, context, verify_result, file_summary):
    rule_type = (rule.rule_type or "").strip()
    if rule_type == "min_file_count":
        return _evaluate_min_file_count(rule, sequence, project, verify_result)
    if rule_type == "filename_contains_project_number":
        return _evaluate_filename_contains_project_number(rule, sequence, project, verify_result)
    if rule_type == "required_extension":
        return _evaluate_required_extension(rule, sequence, project, verify_result)
    if rule_type == "required_file_name_contains":
        return _evaluate_required_file_name_contains(rule, sequence, project, verify_result)
    if rule_type == "required_artifact_file":
        return _evaluate_required_artifact_file(rule, sequence, project, context, verify_result)
    if rule_type == "downloadable_artifact_check":
        return _evaluate_downloadable_artifact_check(rule, sequence, project, context, verify_result)
    if rule_type == "document_artifact_check":
        return _evaluate_document_artifact_check(rule, sequence, project, context, verify_result)
    if rule_type == "all_files_non_empty":
        return _evaluate_all_files_non_empty(rule, sequence, project, verify_result)
    if rule_type == "excel_feature_list_check":
        return _evaluate_excel_feature_list_check(rule, sequence, project, context, verify_result)
    if rule_type == "test_plan_document_check":
        return _evaluate_test_plan_document_check(rule, sequence, project, context, verify_result)
    if rule_type == "image_screenshot_folder_date_check":
        return _evaluate_image_screenshot_folder_date_check(rule, sequence, project, context, verify_result)
    if rule_type == "test_case_check":
        return _evaluate_test_case_check(rule, sequence, project, context, verify_result)
    if rule_type == "rawdata_folder_structure_check":
        return _evaluate_rawdata_folder_structure_check(rule, sequence, project, verify_result)
    if rule_type == "test_report_document_check":
        return _evaluate_test_report_document_check(rule, sequence, project, context, verify_result)
    if rule_type == "defect_report_check":
        return _evaluate_defect_report_check(rule, sequence, project, context, verify_result)
    if rule_type == "inspection_checklist_check":
        return _evaluate_inspection_checklist_check(rule, sequence, project, context, verify_result)
    if rule_type == "quality_inspection_table_check":
        return _evaluate_quality_inspection_table_check(rule, sequence, project, context, verify_result)
    if rule_type == "quality_evaluation_report_check":
        return _evaluate_quality_evaluation_report_check(rule, sequence, project, context, verify_result)

    raise DownloadReviewInspectionError(f"지원하지 않는 점검규칙 유형입니다: {rule_type or '(비어 있음)'}")


def _collect_evaluation_variables(context, evaluation):
    if not evaluation.raw_detail:
        return
    context.derived_variables.update(_raw_detail_variables(evaluation.raw_detail))


def _apply_disabled_sub_checks(rule, evaluation):
    """rule.config_json["disabled_sub_checks"](세부항목 1-based 순번 목록, 문자열/숫자
    혼용 허용)에 담긴 항목을 sub_checks에서 완전히 제거하고, 남은 항목만으로
    status/expected/actual을 다시 계산한다. 제외된 항목은 화면에 아예 표시되지
    않고 전체 규칙 판정에도 영향을 주지 않는다.
    """
    if evaluation.status not in (DownloadReviewRuleStatus.PASS, DownloadReviewRuleStatus.FAIL):
        return evaluation
    config = getattr(rule, "config_json", None) or {}
    disabled = config.get("disabled_sub_checks")
    if not disabled:
        return evaluation
    raw_detail = evaluation.raw_detail
    if not isinstance(raw_detail, dict):
        return evaluation
    sub_checks = raw_detail.get("sub_checks")
    if not isinstance(sub_checks, list) or not sub_checks:
        return evaluation
    disabled_positions = {str(item).strip() for item in disabled if str(item or "").strip()}
    if not disabled_positions:
        return evaluation
    kept = [
        item
        for index, item in enumerate(sub_checks, start=1)
        if str(index) not in disabled_positions
    ]
    if len(kept) == len(sub_checks):
        return evaluation

    new_raw_detail = dict(raw_detail)
    new_raw_detail["sub_checks"] = kept
    new_raw_detail["disabled_sub_check_count"] = len(sub_checks) - len(kept)
    if not kept:
        return replace(evaluation, raw_detail=new_raw_detail)

    all_passed = all(item.get("passed") for item in kept)
    first_fail = next((item for item in kept if not item.get("passed")), None)
    return replace(
        evaluation,
        status=DownloadReviewRuleStatus.PASS if all_passed else DownloadReviewRuleStatus.FAIL,
        expected=" / ".join(str(item.get("expected", "")) for item in kept),
        actual=" / ".join(str(item.get("actual", "")) for item in kept),
        message="" if all_passed else (first_fail.get("message") or evaluation.message),
        raw_detail=new_raw_detail,
    )


def _raw_detail_variables(raw_detail):
    if not isinstance(raw_detail, dict):
        return {}
    variables = raw_detail.get("variables")
    if not isinstance(variables, dict):
        return {}
    return {
        str(key).strip(): value
        for key, value in variables.items()
        if str(key).strip()
    }


def _evaluate_min_file_count(rule, sequence, project, verify_result):
    config = rule.config_json or {}
    min_count = int(config.get("min_count") or 1)
    files = _matching_files(rule, verify_result)
    status = DownloadReviewRuleStatus.PASS if len(files) >= min_count else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"{min_count}개 이상",
        actual=f"{len(files)}개",
        message="파일 개수 기준을 충족했습니다." if status == DownloadReviewRuleStatus.PASS else "파일 개수가 부족합니다.",
        file_path=_representative_path(files, project.project_number),
        file_name=_representative_name(files),
        raw_detail={"matched_file_count": len(files), "min_count": min_count},
    )


def _evaluate_filename_contains_project_number(rule, sequence, project, verify_result):
    files = _matching_files(rule, verify_result)
    matched = [file_info for file_info in files if project.project_number in file_info.name]
    status = DownloadReviewRuleStatus.PASS if matched else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"파일명에 {project.project_number} 포함",
        actual=", ".join(file_info.name for file_info in matched[:5]) if matched else "일치 파일 없음",
        message="프로젝트번호가 포함된 파일을 확인했습니다." if matched else "프로젝트번호가 포함된 파일을 찾지 못했습니다.",
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail={"matched_file_count": len(matched), "total_file_count": len(files)},
    )


def _evaluate_required_extension(rule, sequence, project, verify_result):
    config = rule.config_json or {}
    extension = str(config.get("extension") or _extension_from_file_type(rule.target_file_type)).lower()
    if not extension.startswith("."):
        extension = f".{extension}"

    files = _matching_files(rule, verify_result, ignore_target_file_type=True)
    matched = [file_info for file_info in files if file_info.extension.lower() == extension]
    status = DownloadReviewRuleStatus.PASS if matched else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"{extension} 파일 존재",
        actual=", ".join(file_info.name for file_info in matched[:5]) if matched else "일치 파일 없음",
        message=f"{extension} 파일을 확인했습니다." if matched else f"{extension} 파일을 찾지 못했습니다.",
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail={"extension": extension, "matched_file_count": len(matched)},
    )


def _evaluate_required_file_name_contains(rule, sequence, project, verify_result):
    config = rule.config_json or {}
    contains = str(config.get("contains") or "").strip()
    if not contains:
        raise DownloadReviewInspectionError(f"{rule.name} 규칙의 contains 설정이 없습니다.")

    files = _matching_files(rule, verify_result)
    matched = [file_info for file_info in files if contains in file_info.name]
    status = DownloadReviewRuleStatus.PASS if matched else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"파일명에 {contains} 포함",
        actual=", ".join(file_info.name for file_info in matched[:5]) if matched else "일치 파일 없음",
        message=f"{contains} 파일을 확인했습니다." if matched else f"{contains} 파일을 찾지 못했습니다.",
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail={"contains": contains, "matched_file_count": len(matched)},
    )


def _evaluate_required_artifact_file(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)

    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    extensions = _configured_extensions(config, rule.target_file_type)
    matched = [
        file_info
        for file_info in files
        if (not name_keywords or _name_contains_all(file_info.name, name_keywords))
        and _extension_matches(file_info.extension, extensions)
    ]

    exact_count = config.get("exact_count")
    min_count = int(config.get("min_count") or 1)
    if exact_count is not None:
        expected_count = int(exact_count)
        count_ok = len(matched) == expected_count
        expected = f"{expected_count}개"
    else:
        count_ok = len(matched) >= min_count
        expected = f"{min_count}개 이상"

    # 금지 키워드: 파일명에 포함되면 안 되는 단어(예: '예시'). 대상은 매칭된 파일들.
    forbidden_keywords = [
        str(keyword).strip()
        for keyword in (config.get("forbidden_filename_keywords") or [])
        if str(keyword).strip()
    ]
    forbidden_matches = [
        file_info
        for file_info in matched
        if any(keyword in file_info.name for keyword in forbidden_keywords)
    ] if forbidden_keywords else []

    passed = count_ok and not forbidden_matches

    status = DownloadReviewRuleStatus.PASS if passed else DownloadReviewRuleStatus.FAIL
    if forbidden_matches:
        message = (
            config.get("forbidden_message")
            or f"파일명에 {', '.join(forbidden_keywords)} 포함된 파일이 있습니다."
        )
    elif status == DownloadReviewRuleStatus.PASS:
        message = config.get("pass_message") or "대상 파일을 확인했습니다."
    else:
        message = _artifact_failure_message(
            rule,
            config,
            verify_result,
            matched=matched,
            selected_folder=selected_folder,
            name_keywords=name_keywords,
            exact_count=exact_count,
        )
    expected_parts = [
        expected,
    ]
    if name_keywords:
        expected_parts.insert(0, "파일명에 " + ", ".join(name_keywords) + " 포함")
    if extensions:
        expected_parts.append("확장자 " + _extensions_label(extensions))
    if forbidden_keywords:
        expected_parts.append("파일명에 " + ", ".join(forbidden_keywords) + " 미포함")

    actual_text = _matched_files_actual(matched)
    if forbidden_matches:
        actual_text = (
            f"{actual_text} / 금지어 포함: "
            + ", ".join(file_info.name for file_info in forbidden_matches[:5])
        )

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=" / ".join(expected_parts),
        actual=actual_text,
        message=message,
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail={
            "matched_file_count": len(matched),
            "folder_keyword_chain": config.get("folder_keyword_chain") or [],
            "selected_folder": selected_folder,
            "filename_keywords": name_keywords,
            "extensions": extensions,
            "forbidden_filename_keywords": forbidden_keywords,
            "forbidden_matches": [file_info.name for file_info in forbidden_matches[:20]],
            "matched_files": [
                _display_path(file_info.path, project.project_number)
                for file_info in matched[:20]
            ],
        },
    )


def _evaluate_downloadable_artifact_check(rule, sequence, project, context, verify_result):
    """파일 존재 여부만 자동 판정하고, 찾은 파일을 다운로드형 산출물로 제공한다.

    14번 시험기록서처럼 내용 검사 없이 사용자가 직접 받아 확인하는 규칙용이다.
    """
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    extensions = _configured_extensions(config, rule.target_file_type)
    matched = [
        file_info
        for file_info in files
        if (not name_keywords or _name_contains_all(file_info.name, name_keywords))
        and _extension_matches(file_info.extension, extensions)
    ]

    exact_count = config.get("exact_count")
    min_count = int(config.get("min_count") or 1)
    if exact_count is not None:
        passed = len(matched) == int(exact_count)
        expected_count = f"{int(exact_count)}개"
    else:
        passed = len(matched) >= min_count
        expected_count = f"{min_count}개 이상"

    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "extensions": extensions,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }

    expected_parts = []
    if name_keywords:
        expected_parts.append("파일명에 " + ", ".join(name_keywords) + " 포함")
    if extensions:
        expected_parts.append("확장자 " + _extensions_label(extensions))
    expected_parts.append(expected_count)
    expected_text = " / ".join(expected_parts)

    if not passed:
        # 매칭 파일이 없을 때 무관한 파일을 대표 파일로 표시하지 않는다("... 외 N개" 방지).
        failure_message = _artifact_failure_message(
            rule,
            config,
            verify_result,
            matched=matched,
            selected_folder=selected_folder,
            name_keywords=name_keywords,
            exact_count=exact_count,
        )
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected=expected_text,
            actual=_matched_files_actual(matched) if matched else "일치 파일 없음",
            message=failure_message,
            file_path=_representative_path(matched, project.project_number),
            file_name=_representative_name(matched),
            raw_detail=raw_detail,
        )

    base_id = _safe_artifact_id(config.get("artifact_id") or "download")
    base_label = config.get("artifact_label") or "다운로드 파일"
    # artifact_first_page=True 이면 PDF 1페이지 스크린샷 이미지를 산출물로 제공한다.
    use_first_page = bool(config.get("artifact_first_page"))
    artifacts = []
    try:
        for index, file_info in enumerate(matched):
            single = len(matched) == 1
            artifact_id = base_id if single else f"{base_id}_{index + 1}"
            artifact_label = base_label if single else f"{base_label} {index + 1}"
            if use_first_page and file_info.extension.lower() == ".pdf":
                artifacts.append(
                    _store_pdf_first_page_artifact(
                        project,
                        rule,
                        file_info,
                        artifact_id=artifact_id,
                        label=artifact_label,
                    )
                )
            else:
                artifacts.append(
                    _store_pdf_download_artifact(
                        project,
                        rule,
                        file_info,
                        artifact_id=artifact_id,
                        label=artifact_label,
                    )
                )
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="다운로드 산출물 저장 가능",
            actual=str(exc),
            message=config.get("artifact_error_message") or str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=_representative_name(matched),
            raw_detail=raw_detail,
        )

    raw_detail["artifacts"] = artifacts
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS,
        expected=expected_text,
        actual=_matched_files_actual(matched),
        message=config.get("pass_message") or "파일을 확인했습니다.",
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _evaluate_document_artifact_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    if not name_keywords:
        raise DownloadReviewInspectionError(f"{rule.name} 규칙의 filename_keywords 설정이 없습니다.")

    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
    ]
    file_check = _evaluate_required_file_specs(config, matched)
    content_check = _evaluate_content_checks(config, matched, context) if file_check["passed"] else {
        "passed": True,
        "expected": [],
        "actual": [],
        "message": "",
        "details": [],
    }
    artifact_check = _evaluate_configured_artifacts(config, matched, project, rule)

    if artifact_check["error"]:
        status = DownloadReviewRuleStatus.ERROR
    else:
        passed = file_check["passed"] and content_check["passed"]
        status = DownloadReviewRuleStatus.PASS if passed else DownloadReviewRuleStatus.FAIL
    if status == DownloadReviewRuleStatus.PASS:
        message = config.get("pass_message") or "문서 내용을 확인했습니다."
    elif status == DownloadReviewRuleStatus.ERROR:
        message = artifact_check["message"] or "산출물 처리 중 오류가 발생했습니다."
    else:
        message = _document_artifact_failure_message(
            rule,
            config,
            verify_result,
            name_keywords=name_keywords,
            file_check=file_check,
            content_check=content_check,
        )
    expected = " / ".join([*file_check["expected"], *content_check["expected"]])
    actual = " / ".join([*file_check["actual"], *content_check["actual"]]) or _matched_files_actual(matched)

    raw_detail = {
        "matched_file_count": len(matched),
        "folder_keyword_chain": config.get("folder_keyword_chain") or [],
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "file_checks": file_check["details"],
        "content_checks": content_check["details"],
        "matched_files": [
            _display_path(file_info.path, project.project_number)
            for file_info in matched[:20]
        ],
    }
    if artifact_check["artifacts"]:
        raw_detail["artifacts"] = artifact_check["artifacts"]
    if artifact_check["details"]:
        raw_detail["artifact_checks"] = artifact_check["details"]

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched or files, project.project_number),
        file_name=_representative_name(matched or files),
        raw_detail=raw_detail,
    )


def _evaluate_all_files_non_empty(rule, sequence, project, verify_result):
    files = _matching_files(rule, verify_result)
    empty_files = [file_info for file_info in files if file_info.size == 0]
    status = (
        DownloadReviewRuleStatus.PASS
        if files and not empty_files
        else DownloadReviewRuleStatus.FAIL
    )
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected="모든 파일 크기 1 byte 이상",
        actual="0 byte 파일 없음" if status == DownloadReviewRuleStatus.PASS else ", ".join(file_info.name for file_info in empty_files[:5]) or "검사 대상 파일 없음",
        message="빈 파일이 없습니다." if status == DownloadReviewRuleStatus.PASS else "빈 파일이 있거나 검사 대상 파일이 없습니다.",
        file_path=_representative_path(empty_files or files, project.project_number),
        file_name=_representative_name(empty_files or files),
        raw_detail={"empty_file_count": len(empty_files), "matched_file_count": len(files)},
    )


def _evaluate_configured_artifacts(config, matched_files, project, rule):
    artifacts = []
    details = []
    message = ""
    error = False

    for spec in config.get("artifacts") or []:
        artifact_type = str(spec.get("type") or "").strip()
        extensions = _configured_extensions(spec, "any")
        files = [
            file_info
            for file_info in matched_files
            if _extension_matches(file_info.extension, extensions)
        ]
        detail = {
            "type": artifact_type,
            "extensions": extensions,
            "matched_file_count": len(files),
        }
        if not files:
            detail["passed"] = not spec.get("required", False)
            details.append(detail)
            if spec.get("required", False) and not error:
                error = True
                message = spec.get("missing_message") or "산출물 대상 파일이 없습니다."
            continue

        try:
            if artifact_type == "pdf_first_page":
                artifact = _store_pdf_first_page_artifact(
                    project,
                    rule,
                    files[0],
                    artifact_id=spec.get("id") or "pdf_first_page",
                    label=spec.get("label") or "PDF 1페이지",
                )
            else:
                raise DownloadReviewInspectionError(f"지원하지 않는 산출물 유형입니다: {artifact_type or '(비어 있음)'}")
        except DownloadReviewInspectionError as exc:
            fail_message = spec.get("error_message") or str(exc)
            detail["passed"] = False
            detail["message"] = fail_message
            details.append(detail)
            if not error:
                error = True
                message = fail_message
            continue

        detail["passed"] = True
        detail["artifact_id"] = artifact["id"]
        details.append(detail)
        artifacts.append(artifact)

    return {
        "artifacts": artifacts,
        "details": details,
        "error": error,
        "message": message,
    }


def _evaluate_excel_feature_list_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _extension_matches(file_info.extension, [".xlsx", ".xls"])
    ]

    if len(matched) != 1:
        status = DownloadReviewRuleStatus.FAIL
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=status,
            expected="파일명에 " + ", ".join(name_keywords) + " 포함 / Excel 파일 1개",
            actual=_matched_files_actual(matched),
            message=_artifact_failure_message(
                rule,
                config,
                verify_result,
                matched=matched,
                selected_folder=selected_folder,
                name_keywords=name_keywords,
                exact_count=config.get("exact_count"),
            ),
            file_path=_representative_path(matched or files, project.project_number),
            file_name=_representative_name(matched or files),
            raw_detail={
                "selected_folder": selected_folder,
                "matched_file_count": len(matched),
                "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched],
            },
        )

    file_info = matched[0]
    details = {
        "selected_folder": selected_folder,
        "matched_file": _display_path(file_info.path, project.project_number),
    }
    try:
        workbook = _read_excel_workbook(file_info)
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="Excel 파일 파싱 가능",
            actual=str(exc),
            message=config.get("parse_error_message") or str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=details,
        )

    # 첫 실패에서 멈추지 않고 모든 세부 점검을 수행한다.
    details["sheet_names"] = [sheet.name for sheet in workbook.sheets]
    sheet = workbook.sheets[0] if workbook.sheets else None
    checks = []  # {expected, actual, passed, message}

    # 1) 시트 개수
    checks.append({
        "expected": "[시트] 1개",
        "actual": f"시트 {len(workbook.sheets)}개",
        "passed": len(workbook.sheets) == 1,
        "message": config.get("sheet_count_message") or "불필요한 시트가 존재",
    })

    title = _resolve_rule_value(config.get("title_text") or "{프로젝트번호} 기능리스트", context)
    author_label = str(config.get("author_label") or "작성자")
    title_cell = _find_cell_containing(sheet.rows, title) if sheet else None
    author_cell = _find_cell_containing(sheet.rows, author_label) if sheet else None
    author_ok = bool(author_cell and context.pl and context.pl in author_cell["value"])
    details.update({
        "title_text": title,
        "title_cell": title_cell or {},
        "author_label": author_label,
        "author_cell": author_cell or {},
    })

    # 2) 제목(프로젝트번호)
    checks.append({
        "expected": f"[제목] '{title}' 포함",
        "actual": (title_cell.get("value") if title_cell else "없음"),
        "passed": bool(title_cell),
        "message": config.get("content_message") or "시험번호가 잘못 작성됨",
    })
    # 3) 작성자
    checks.append({
        "expected": f"[작성자] {author_label} 셀에 {context.pl} 포함",
        "actual": (author_cell.get("value") if author_cell else "없음"),
        "passed": author_ok,
        "message": config.get("content_message") or "작성자가 잘못 작성됨",
    })

    # 4) 캡처 영역 (+ 영역 이미지 산출물)
    category_label = str(config.get("capture_anchor") or "대분류")
    capture_area = _excel_area_from_column_anchor(sheet.rows, category_label, column_index=0) if sheet else None
    details["capture_area"] = capture_area or {}
    checks.append({
        "expected": f"[캡처영역] A열 {category_label} 기준 영역",
        "actual": (capture_area["range"] if capture_area else "캡처 기준 셀 없음"),
        "passed": bool(capture_area),
        "message": config.get("capture_message") or "기능리스트 캡처 영역을 찾지 못했습니다.",
    })
    if capture_area and sheet:
        try:
            artifact = _store_excel_area_artifact(
                project, rule, sheet, capture_area,
                artifact_id=config.get("capture_artifact_id") or "feature_list_area",
                label=config.get("capture_artifact_label") or "기능리스트 영역",
                source_file=_display_path(file_info.path, project.project_number),
            )
            details["artifacts"] = [artifact]
        except DownloadReviewInspectionError:
            pass

    # 5) 머리글/바닥글 금지 문자열 (예: 머리글의 프로젝트번호, 바닥글의 서식번호)
    header_text = _clean_excel_header_text(sheet.header_text) if sheet else ""
    footer_text = _clean_excel_header_text(sheet.footer_text) if sheet else ""
    details["header_text"] = header_text
    details["footer_text"] = footer_text
    forbidden_header = [str(t) for t in (config.get("forbidden_header_texts") or []) if str(t).strip()]
    if forbidden_header:
        header_hits = [t for t in forbidden_header if t in header_text]
        checks.append({
            "expected": f"[머리글] {', '.join(forbidden_header)} 미포함",
            "actual": header_text or "머리글 없음",
            "passed": not header_hits,
            "message": config.get("forbidden_header_message") or "머리글에 금지된 문자열이 있습니다.",
        })
    forbidden_footer = [str(t) for t in (config.get("forbidden_footer_texts") or []) if str(t).strip()]
    if forbidden_footer:
        footer_hits = [t for t in forbidden_footer if t in footer_text]
        checks.append({
            "expected": f"[바닥글] {', '.join(forbidden_footer)} 미포함",
            "actual": footer_text or "바닥글 없음",
            "passed": not footer_hits,
            "message": config.get("forbidden_footer_message") or "바닥글에 금지된 문자열이 있습니다.",
        })

    details["sub_checks"] = [
        {"expected": c["expected"], "actual": c["actual"], "passed": c["passed"],
         "message": c.get("message", "")} for c in checks
    ]
    all_passed = all(c["passed"] for c in checks)
    first_fail = next((c for c in checks if not c["passed"]), None)
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS if all_passed else DownloadReviewRuleStatus.FAIL,
        expected=f"시트 1개 / {title} 포함 / 작성자 {context.pl} 포함 / 캡처영역",
        actual=f"{sheet.name if sheet else '-'}",
        message=(config.get("pass_message") or "기능리스트를 확인했습니다.") if all_passed else (_append_current_value(first_fail["message"], first_fail.get("actual")) if first_fail else "기능리스트 확인 필요"),
        file_path=_representative_path(matched, project.project_number),
        file_name=file_info.name,
        raw_detail=details,
    )


def _evaluate_test_plan_document_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
    ]
    word_files = [file_info for file_info in matched if _is_word_file(file_info)]
    pdf_files = [file_info for file_info in matched if _extension_matches(file_info.extension, [".pdf"])]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
        "word_count": len(word_files),
        "pdf_count": len(pdf_files),
        "checks": [],
    }

    if len(word_files) != 1 or len(pdf_files) != 1:
        return _test_plan_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="Word 파일 1개 / PDF 파일 1개",
            actual=f"Word 파일 {len(word_files)}개 / PDF 파일 {len(pdf_files)}개",
            message=_test_plan_file_failure_message(
                config,
                matched=matched,
                selected_folder=selected_folder,
                word_files=word_files,
                pdf_files=pdf_files,
            ),
        )

    docx_file = word_files[0]
    pdf_file = pdf_files[0]
    try:
        tables = _docx_tables(docx_file)
        footer_text = _docx_footer_text(docx_file)
        plan_spec_table = _docx_first_table_after_text(docx_file, config.get("spec_marker") or "<세부사양>")
        artifact = _store_pdf_first_page_artifact(
            project,
            rule,
            pdf_file,
            artifact_id="pdf_first_page",
            label=config.get("pdf_artifact_label") or "시험계획서 1페이지",
        )
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="시험계획서 Word/PDF 파싱 가능",
            actual=str(exc),
            message=config.get("parse_error_message") or str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=_representative_name(matched),
            raw_detail=raw_detail,
        )

    raw_detail["table_count"] = len(tables)
    raw_detail["footer_text"] = footer_text
    raw_detail["spec_table"] = plan_spec_table or []
    raw_detail["artifacts"] = [artifact]

    # 첫 실패에서 멈추지 않고 모든 세부 점검을 수행한다(선행 실패가 후행 검사를 건너뛰지 않도록).
    checks = raw_detail.setdefault("checks", [])

    # 1) 첫 번째 표 (시작일/담당자/PL)
    if len(tables) >= 1:
        checks.extend(_test_plan_first_table_checks(tables[0], config, context))
    else:
        checks.append({
            "name": "first_table", "passed": False,
            "expected": "시험계획서 첫 번째 표", "actual": "표 없음",
            "message": config.get("date_message") or "시험계획서 표를 찾을 수 없음",
        })

    # 2) 두 번째 표 (제품정보)
    if len(tables) >= 2:
        checks.extend(_test_plan_product_checks(tables[1], config, context))
    else:
        checks.append({
            "name": "product_table", "passed": False,
            "expected": "제품정보 표(두 번째 표)", "actual": f"표 {len(tables)}개",
            "message": config.get("product_message") or "제품정보가 틀림",
        })

    # 3) 형상항목 ID
    configuration_table = _docx_first_table_after_text(
        docx_file, config.get("configuration_marker") or "5.1 형상항목 식별 규칙",
    )
    checks.append(_test_plan_configuration_id_check(configuration_table, config, context))

    # 4) 시험일정 WD
    schedule_table = _docx_first_table_after_text(
        docx_file, config.get("schedule_marker") or "2.2 시험일정",
    )
    checks.append(_test_plan_schedule_check(schedule_table, config, context))

    # 5) 바닥글 Copyright
    footer_expected = _resolve_rule_value(config.get("footer_text") or "Copyright {연도} TTA", context)
    checks.append({
        "name": "footer_copyright",
        "passed": footer_expected in footer_text,
        "expected": f"바닥글에 {footer_expected} 포함",
        "actual": footer_text or "바닥글 없음",
        "message": config.get("footer_message") or "바닥글 Copyright가 잘못 작성됨",
    })

    # 6) 바닥글 금지어 (설정된 경우)
    forbidden_footer_check = _check_forbidden_text_terms(
        footer_text,
        config.get("forbidden_footer_terms") or [],
        context,
        subject="시험계획서 바닥글",
        default_message="시험계획서 바닥글에 잘못된 단어가 작성됨",
    )
    if forbidden_footer_check["details"]:
        checks.append(forbidden_footer_check)

    # 7) 세부사양 표 (시험성적서와 비교)
    checks.append(_test_plan_spec_table_check(plan_spec_table, config, context))

    raw_detail["sub_checks"] = _checks_to_sub_checks(checks)
    all_passed = all(check.get("passed") for check in checks)
    first_fail = next((check for check in checks if not check.get("passed")), None)
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS if all_passed else DownloadReviewRuleStatus.FAIL,
        expected="시험계획서 Word/PDF / 표 값 / 형상항목 ID / WD / 바닥글 / 세부사양",
        actual=f"Word {docx_file.name} / PDF {pdf_file.name}",
        message=(config.get("pass_message") or "시험계획서를 확인했습니다.") if all_passed else (_append_current_value(first_fail.get("message"), first_fail.get("actual")) if first_fail else "시험계획서 확인 필요"),
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _checks_to_sub_checks(checks):
    """{name,passed,expected,actual} 리스트를 팝업 표시용 sub_checks로 변환한다."""
    sub_checks = []
    for check in checks or []:
        if not isinstance(check, dict):
            continue
        sub_checks.append({
            "expected": check.get("expected", ""),
            "actual": check.get("actual", ""),
            "passed": check.get("passed"),
            "message": check.get("message", ""),
        })
    return sub_checks


def _test_plan_file_failure_message(config, *, matched, selected_folder, word_files, pdf_files):
    """시험계획서 파일 구성 실패 사유를 (폴더 → 파일명 → 개수초과 → Word/PDF 누락) 순으로 구분한다."""
    if config.get("folder_keyword_chain") and not selected_folder:
        return config.get("folder_missing_message") or config.get("missing_message") or "대상 폴더를 찾을 수 없습니다."
    if not matched:
        return config.get("missing_message") or "대상 파일을 찾을 수 없습니다."
    if len(word_files) > 1 or len(pdf_files) > 1:
        return config.get("multiple_message") or "대상 파일이 여러개 존재합니다."
    if len(word_files) < 1:
        return config.get("word_missing_message") or "Word 파일을 찾을 수 없습니다"
    if len(pdf_files) < 1:
        return config.get("pdf_missing_message") or "pdf 파일을 찾을 수 없습니다"
    return config.get("missing_message") or "필요한 파일을 찾을 수 없습니다."


def _test_plan_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    # 지금까지 수행한 세부 점검(checks)을 항목별 적합/부적합 행으로 표시한다.
    sub_checks = _checks_to_sub_checks(raw_detail.get("checks"))
    if sub_checks:
        raw_detail["sub_checks"] = sub_checks
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _resolve_center_expected(config, context, key, default=""):
    """센터별 기대값 매핑('{key}_by_center')이 있으면 센터에 맞는 값을, 없으면 config[key] 사용.

    센터 코드(분당=bundang, 상암=sangam, 영남=yeongnam) 기준으로 담당자/검토자/작성자
    이름을 다르게 검사할 때 쓴다.
    """
    by_center = config.get(f"{key}_by_center") or {}
    center = str(getattr(context, "center", "") or "").strip()
    if center and center in by_center:
        return _resolve_rule_value(str(by_center[center]), context)
    return _resolve_rule_value(str(config.get(key) or default), context)


def _resolve_manager_expected(config, context):
    """센터별 담당자명을 고른다. 센터 매핑에 없으면 manager_expected 기본값 사용."""
    return _resolve_center_expected(config, context, "manager_expected", "김진영")


def _test_plan_first_table_checks(table, config, context):
    manager_expected = _resolve_manager_expected(config, context)
    return [
        {
            "name": "first_table_start_date",
            "passed": _same_date_text(_table_cell(table, 1, 2), context.start_date),
            "expected": f"1행 2열 = {context.start_date}",
            "actual": _table_cell(table, 1, 2) or "값 없음",
            "message": config.get("date_message") or "시험계획서 날짜가 잘못 작성됨",
        },
        {
            "name": "first_table_manager",
            "passed": manager_expected in _table_cell(table, 3, 2),
            "expected": f"3행 2열에 {manager_expected} 포함",
            "actual": _table_cell(table, 3, 2) or "값 없음",
            "message": config.get("manager_message") or "시험계획서 담당자가 잘못 작성됨",
        },
        {
            "name": "first_table_pl",
            "passed": bool(context.pl and context.pl in _table_cell(table, 4, 2)),
            "expected": f"4행 2열에 {context.pl} 포함",
            "actual": _table_cell(table, 4, 2) or "값 없음",
            "message": config.get("pl_message") or "시험계획서 PL이 잘못 작성됨",
        },
    ]


def _version_matches(expected, actual):
    """버전 비교.

    - 접두사('v', 'ver' 등 숫자 앞 문자)는 있어도 없어도 무시한다.
    - 숫자 부분(예: 4.0)이 같으면 같은 버전으로 본다 - 표기 자릿수가 달라도
      (예: 'v1.0'과 실제 문서의 'v1', '2.10.0'과 '2.10') 수치가 같으면 동일
      버전으로 인정한다. 등록된 제품명이 'v1.0'이어도 문서에 'v1'로만 적는
      경우가 실제로 있어서, 문자열 그대로 비교하면(예전 방식) '1.0' != '1'로
      오판했다.
    - 숫자 없는 문자 버전(예: Enterprise)은 공백과 대소문자를 무시해 비교한다.
    예) 기대 'v4.0' vs 실제 '4.0'/'v4.0'/'ver4.0'/'4' → 모두 정상.
    """
    def core(value):
        text = _normalize_no_space(value)
        match = re.search(r"\d[\d.]*", text)
        return match.group(0) if match else ""

    expected_core = core(expected)
    actual_core = core(actual)
    if expected_core:
        if expected_core == actual_core:
            return True
        return _version_numbers_equal(expected_core, actual_core)

    def text_core(value):
        text = _normalize_no_space(value).lower()
        return re.sub(r"^(?:v|ver|version)\.?", "", text) or text

    expected_text = text_core(expected)
    actual_text = text_core(actual)
    return bool(expected_text) and expected_text == actual_text


def _version_numbers_equal(left, right):
    """'1.0'과 '1', '2.10.0'과 '2.10'처럼 끝자리 0 유무만 다른 숫자 버전을 같다고 본다.

    자리별로 정수 비교하므로 '2.10'과 '2.1'처럼 실제로 다른 버전은 여전히
    다르게 판정된다(끝에 붙는 0만 무시).
    """
    def segments(value):
        parts = [part for part in value.split(".") if part != ""]
        try:
            return [int(part) for part in parts]
        except ValueError:
            return None

    left_segments = segments(left)
    right_segments = segments(right)
    if left_segments is None or right_segments is None:
        return False
    length = max(len(left_segments), len(right_segments))
    left_segments += [0] * (length - len(left_segments))
    right_segments += [0] * (length - len(right_segments))
    return left_segments == right_segments


def _test_plan_product_checks(table, config, context):
    product_label = str(config.get("product_name_label") or "소프트웨어 명")
    version_label = str(config.get("version_label") or "버전")
    application_label = str(config.get("application_number_label") or "시험신청번호")
    product_actual = _find_next_cell_by_label(table, product_label)
    version_actual = _find_next_cell_by_label(table, version_label)
    application_actual = _find_next_cell_by_label(table, application_label)
    product_message = config.get("product_message") or "제품정보가 틀림"
    return [
        {
            "name": "version_exists",
            "passed": bool(context.version),
            "expected": "제품명에서 버전 파싱 가능",
            "actual": context.product_raw or "제품명 없음",
            "message": config.get("version_missing_message") or "버전을 찾을 수 없음",
        },
        {
            "name": "product_name",
            "passed": _normalize_spaces(product_actual) == context.product,
            "expected": f"{product_label} 오른쪽 셀 = {context.product}",
            "actual": product_actual or "값 없음",
            "message": product_message,
        },
        {
            "name": "product_version",
            "passed": _version_matches(context.version, version_actual),
            "expected": f"{version_label} 오른쪽 셀 = {context.version} (숫자 버전은 v/ver 접두사 생략 가능, 문자 버전은 대소문자 무시)",
            "actual": version_actual or "값 없음",
            "message": product_message,
        },
        {
            "name": "application_number",
            "passed": _normalize_spaces(application_actual) == context.project_number,
            "expected": f"{application_label} 오른쪽 셀 = {context.project_number}",
            "actual": application_actual or "값 없음",
            "message": product_message,
        },
    ]


def _test_plan_configuration_id_check(table, config, context):
    header = str(config.get("configuration_header") or "형상항목 ID")
    header_cell = _find_cell_containing(table, header)
    if not header_cell:
        return {
            "name": "configuration_id",
            "passed": False,
            "expected": f"{header} 열 존재",
            "actual": "헤더 없음",
            "message": config.get("configuration_message") or "형상항목 ID가 잘못 작성됨",
        }

    values = []
    invalid_values = []
    for row_index in range(header_cell["row"] + 1, len(table) + 1):
        value = _table_cell(table, row_index, header_cell["column"])
        if not value:
            continue
        values.append({"row": row_index, "value": value})
        if context.project_number not in value:
            invalid_values.append({"row": row_index, "value": value})

    return {
        "name": "configuration_id",
        "passed": bool(values) and not invalid_values,
        "expected": f"{header} 열 값에 {context.project_number} 포함",
        "actual": " / ".join(item["value"] for item in invalid_values[:5]) if invalid_values else f"{len(values)}개 값 확인",
        "message": config.get("configuration_message") or "형상항목 ID가 잘못 작성됨",
        "header_cell": header_cell,
        "values": values,
        "invalid_values": invalid_values,
    }


def _schedule_column_values(table, header, *, limit=4):
    """시험일정류 표에서 header 셀 다음 열의 값을 최대 limit개 뽑는다.

    시험계획서(2.2 시험일정, 헤더 'WD')와 시험성적서(4.4 시험일정, 헤더
    '소요일수')는 같은 일정을 서로 다른 표 모양으로 담고 있어, 헤더 텍스트만
    다르게 넘겨서 재사용한다.
    """
    header_cell = _find_cell_containing(table, header)
    column = header_cell["column"] if header_cell else 2
    start_row = header_cell["row"] + 1 if header_cell else 1
    values = []
    for row_index in range(start_row, len(table or []) + 1):
        value = _table_cell(table, row_index, column)
        if value:
            values.append(value)
        if len(values) >= limit:
            break
    return values, header_cell


def _test_plan_schedule_check(table, config, context):
    """시험일정 표의 WD 열을 확인한다.

    규칙 문서(3.7 시험계획서)는 'WD 열이 존재하고 값이 작성되어 있는지'만
    요구한다. 예전에는 표준 신규시험(환경구축1/제품분석1/시험(WD-3)/종료1)
    고정 패턴을 그대로 강제했으나, 제품명 변경 등에 따른 간소화 시험은
    반나절(0.5) 단위로 일정을 잡는 등 표준 패턴과 다르게 작성되는 경우가
    있다.

    가장 신뢰할 수 있는 기준은 같은 시험 일정을 담고 있는 시험성적서
    4.4 시험일정(소요일수 열)이므로, 13번(시험성적서) 규칙이 먼저 실행되며
    남긴 그 값과 그대로 비교한다 - 계획서와 성적서 값이 같으면 통과, 다르면
    계획서 쪽을 부적합 처리한다(<세부사양> 표 비교와 동일한 방식).
    성적서 쪽 값을 못 구한 경우에만(파싱 실패 등) WD 값이 채워져 있고
    합계가 등록된 총 WD와 같은지로 대체 검증한다.
    """
    header = str(config.get("schedule_header") or "WD")
    actual_value_items, header_cell = _schedule_column_values(table, header)
    normalized_actual = [_normalize_number_text(value) for value in actual_value_items]

    report_variable = config.get("report_schedule_variable") or "시험성적서_시험일정"
    report_values = _context_variable(context, report_variable)
    if isinstance(report_values, list) and report_values:
        normalized_report = [_normalize_number_text(value) for value in report_values]
        passed = bool(header_cell) and normalized_actual == normalized_report
        return {
            "name": "schedule_wd",
            "passed": passed,
            "expected": f"시험성적서 4.4 시험일정과 동일: {', '.join(report_values)}",
            "actual": ", ".join(actual_value_items) or "WD 값 없음",
            "message": config.get("schedule_message") or "시험일정 WD가 시험성적서와 다름",
            "header_cell": header_cell or {},
            "values": actual_value_items,
        }

    # 성적서 쪽 시험일정을 못 구했을 때의 대체 검증: 값이 채워져 있고 합계가
    # 등록된 총 WD와 일치하는지만 확인한다(고정 배분 패턴은 강제하지 않음).
    wd = _context_wd_int(context)
    numeric_values = []
    all_numeric = bool(normalized_actual)
    for text in normalized_actual:
        try:
            numeric_values.append(float(text))
        except (TypeError, ValueError):
            all_numeric = False
            break
    total = sum(numeric_values) if all_numeric else None
    passed = bool(
        header_cell
        and len(actual_value_items) >= 4
        and all_numeric
        and wd is not None
        and total is not None
        and abs(total - wd) < 1e-6
    )
    expected_total = str(wd) if wd is not None else "{WD}"
    actual_text = ", ".join(actual_value_items) or "WD 값 없음"
    if total is not None:
        actual_text = f"{actual_text} (합계 {total:g})"
    return {
        "name": "schedule_wd",
        "passed": passed,
        "expected": f"WD 열 값 4개 작성 / 합계 {expected_total}",
        "actual": actual_text,
        "message": config.get("schedule_message") or "시험일정 WD가 틀림",
        "header_cell": header_cell or {},
        "values": actual_value_items,
    }


def _test_plan_spec_table_check(plan_table, config, context):
    report_table = _context_variable(context, config.get("report_spec_variable") or "시험성적서_세부사양표")
    normalized_plan = _normalize_docx_table(plan_table)
    normalized_report = _normalize_docx_table(report_table if isinstance(report_table, list) else [])
    comparison_plan = _normalize_docx_table(plan_table, remove_whitespace=True)
    comparison_report = _normalize_docx_table(
        report_table if isinstance(report_table, list) else [],
        remove_whitespace=True,
    )
    mismatches = _matrix_mismatches(
        normalized_plan,
        normalized_report,
        left_origin=(1, 1),
        right_origin=(1, 1),
    )[:20]
    passed = bool(comparison_plan) and comparison_plan == comparison_report
    return {
        "name": "spec_table",
        "passed": passed,
        "expected": "시험성적서 <세부사양> 표와 일치",
        "actual": (
            "일치" if passed
            else _format_mismatch_summary(mismatches) if normalized_plan and normalized_report
            else "비교 대상 표 없음"
        ),
        "message": config.get("spec_message") or "시험환경 세부사양 표가 결과서와 다름",
        "plan_table": normalized_plan,
        "report_table": normalized_report,
        "comparison_mode": "ignore_whitespace",
        "mismatches": mismatches,
    }


def _format_mismatch_summary(mismatches, limit=5):
    """표 비교 불일치를 '불일치 N건' 대신 실제로 어떤 셀이 어떻게 다른지 나열한다."""
    if not mismatches:
        return "불일치 0건"
    lines = [
        f"{item['left_cell']} 계획서 '{item['left'] or '(빈 값)'}' → 성적서 '{item['right'] or '(빈 값)'}'"
        for item in mismatches[:limit]
    ]
    if len(mismatches) > limit:
        lines.append(f"외 {len(mismatches) - limit}건 더")
    return " / ".join(lines)


def _table_cell(table, row, column):
    row_index = row - 1
    col_index = column - 1
    if row_index < 0 or col_index < 0 or row_index >= len(table or []):
        return ""
    row_values = table[row_index]
    if col_index >= len(row_values):
        return ""
    return _normalize_spaces(row_values[col_index])


def _context_wd_int(context):
    try:
        return int(str(context.wd).strip())
    except (TypeError, ValueError):
        return None


def _normalize_number_text(value):
    text = _normalize_spaces(value)
    if re.fullmatch(r"-?\d+(?:\.0+)?", text):
        return str(int(float(text)))
    return text


def _normalize_docx_table(table, *, remove_whitespace=False):
    normalizer = _normalize_no_space if remove_whitespace else _normalize_spaces
    rows = [
        [normalizer(cell) for cell in row]
        for row in (table or [])
    ]
    return _trim_empty_edges(rows)


def _evaluate_test_case_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    extensions = _configured_extensions(config, rule.target_file_type)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _extension_matches(file_info.extension, extensions)
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "extensions": extensions,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }
    raw_detail["_expected_sub_check_templates"] = _test_case_expected_sub_check_templates(config, context)

    if len(matched) != 1:
        return _test_case_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="테스트케이스 Excel 파일 1개",
            actual=_matched_files_actual(matched),
            message=_artifact_failure_message(
                rule,
                config,
                verify_result,
                matched=matched,
                selected_folder=selected_folder,
                name_keywords=name_keywords,
                exact_count=config.get("exact_count"),
            ),
        )

    file_info = matched[0]
    try:
        workbook = _read_excel_workbook(file_info)
    except DownloadReviewInspectionError as exc:
        raw_detail["sub_checks"] = _complete_expected_sub_checks(
            [],
            raw_detail.get("_expected_sub_check_templates"),
            actual=str(exc),
            message=config.get("parse_error_message") or str(exc),
        )
        raw_detail.pop("_expected_sub_check_templates", None)
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="테스트케이스 Excel 파일 파싱 가능",
            actual=str(exc),
            message=config.get("parse_error_message") or str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=raw_detail,
        )

    raw_detail["sheet_names"] = [sheet.name for sheet in workbook.sheets]
    # 첫 실패에서 멈추지 않고 모든 세부 점검을 수행한 뒤 항목별 적합/부적합을 표시한다.
    sub_checks = []  # {expected, actual, passed, message}

    # 0) 시트 개수 (1개가 아니어도 멈추지 않고 첫 시트로 나머지 점검 진행)
    sub_checks.append({
        "expected": "[시트] 1개",
        "actual": f"시트 {len(workbook.sheets)}개",
        "passed": len(workbook.sheets) == 1,
        "message": config.get("sheet_count_message") or "테스트케이스 시트가 1개 이상임",
    })
    sheet = workbook.sheets[0] if workbook.sheets else None
    if sheet is None:
        raw_detail["sub_checks"] = [dict(sub_checks[0])]
        return _test_case_failure(
            rule, sequence, matched, project, raw_detail,
            expected="시트 1개", actual="시트 없음",
            message=config.get("sheet_count_message") or "테스트케이스 시트가 없음",
        )

    # 1) 바닥글 금지어 (금지어가 설정된 경우에만 점검 항목으로 추가)
    footer_check = _check_forbidden_text_terms(
        _clean_excel_header_text(sheet.footer_text),
        config.get("forbidden_footer_terms") or [],
        context,
        subject="테스트케이스 바닥글",
        default_message="테스트케이스 바닥글에 잘못된 단어가 작성됨",
    )
    if footer_check["details"]:
        raw_detail["footer_check"] = footer_check
        sub_checks.append({
            "expected": f"[바닥글] {_stringify_check_value(footer_check.get('expected', ''))}",
            "actual": _stringify_check_value(footer_check.get("actual", "")),
            "passed": bool(footer_check.get("passed")),
            "message": footer_check.get("message") or "테스트케이스 바닥글에 잘못된 단어가 작성됨",
        })

    # 1-1) 머리글 금지어 (설정된 경우에만)
    header_forbidden_check = _check_forbidden_text_terms(
        _clean_excel_header_text(sheet.header_text),
        config.get("forbidden_header_terms") or [],
        context,
        subject="테스트케이스 머리글",
        default_message="테스트케이스 머리글에 잘못된 단어가 작성됨",
    )
    if header_forbidden_check["details"]:
        raw_detail["header_forbidden_check"] = header_forbidden_check
        sub_checks.append({
            "expected": f"[머리글] {_stringify_check_value(header_forbidden_check.get('expected', ''))}",
            "actual": _stringify_check_value(header_forbidden_check.get("actual", "")),
            "passed": bool(header_forbidden_check.get("passed")),
            "message": header_forbidden_check.get("message") or "테스트케이스 머리글에 잘못된 단어가 작성됨",
        })

    # 1-2) 바닥글 필수어 (설정된 경우에만)
    footer_required_check = _check_required_text_terms(
        _clean_excel_header_text(sheet.footer_text),
        config.get("required_footer_terms") or [],
        context,
        subject="테스트케이스 바닥글",
        default_message="테스트케이스 바닥글에 필요한 단어가 누락됨",
    )
    if footer_required_check["details"]:
        raw_detail["footer_required_check"] = footer_required_check
        sub_checks.append({
            "expected": f"[바닥글 필수어] {_stringify_check_value(footer_required_check.get('expected', ''))}",
            "actual": _stringify_check_value(footer_required_check.get("actual", "")),
            "passed": bool(footer_required_check.get("passed")),
            "message": footer_required_check.get("message") or "테스트케이스 바닥글에 필요한 단어가 누락됨",
        })

    # 2) 제목(프로젝트번호)
    title_text = _resolve_rule_value(config.get("title_text") or "{project_number} 테스트케이스", context)
    title_cell = _find_cell_containing(sheet.rows, title_text)
    raw_detail["title_check"] = {"expected": title_text, "matched_cell": title_cell or {}}
    sub_checks.append({
        "expected": f"[제목] '{title_text}' 포함",
        "actual": (title_cell.get("value") if title_cell else "일치 문장 없음"),
        "passed": bool(title_cell),
        "message": config.get("project_number_message") or "프로젝트 번호가 잘못 작성됨",
    })

    # 3) 작성자 / 검토자
    author_label = str(config.get("author_label") or "작성자:")
    reviewer_label = str(config.get("reviewer_label") or "검토자:")
    reviewer_expected = _resolve_center_expected(config, context, "reviewer_expected", "김진영")
    author_cell = _find_cell_with_all(sheet.rows, [author_label, context.pl])
    reviewer_cell = {}
    reviewer_ok = False
    if author_cell:
        reviewer_value = _sheet_cell(sheet, author_cell["row"] + 1, author_cell["column"])
        reviewer_cell = {
            "row": author_cell["row"] + 1,
            "column": author_cell["column"],
            "value": reviewer_value,
        }
        reviewer_ok = bool(
            reviewer_value
            and reviewer_label in reviewer_value
            and reviewer_expected in reviewer_value
        )
    raw_detail["author_reviewer_check"] = {
        "author_label": author_label,
        "pl": context.pl,
        "author_cell": author_cell or {},
        "reviewer_label": reviewer_label,
        "reviewer_expected": reviewer_expected,
        "reviewer_cell": reviewer_cell,
    }
    sub_checks.append({
        "expected": f"[작성자/검토자] {author_label} {context.pl} / {reviewer_label} {reviewer_expected}",
        "actual": f"작성자={author_cell['value'] if author_cell else '없음'} / 검토자={reviewer_cell.get('value') or '없음'}",
        "passed": bool(author_cell and reviewer_ok),
        "message": config.get("author_message") or "작성자 또는 검토자가 잘못 작성됨",
    })

    # 4) 작성일 — 공백 제거 후, 작성일 라벨과 시작일~종료일 날짜가 맞는지 확인(형식 무관)
    date_label = str(config.get("date_label") or "작성일")
    date_cell = _find_labeled_date_range_cell(sheet.rows, date_label, context.start_date, context.end_date)
    raw_detail["date_check"] = {
        "expected": f"{date_label} {context.start_date} ~ {context.end_date}",
        "matched_cell": date_cell or {},
    }
    sub_checks.append({
        "expected": f"[작성일] {date_label} {context.start_date} ~ {context.end_date}",
        "actual": (date_cell.get("value") if date_cell else "일치 작성일 없음"),
        "passed": bool(date_cell),
        "message": config.get("date_message") or "작성일이 잘못 작성됨",
    })

    # 5) 잔여결함수 (상세 테스트 결과 열의 F 개수)
    residual_expected = _context_int(context, "잔여결함수")
    result_header = str(config.get("result_header") or "상세 테스트 결과")
    result_header_cell = _find_cell_containing(sheet.rows, result_header)
    failed_rows = []
    if result_header_cell:
        failed_rows = _test_case_failed_result_rows(
            sheet,
            start_row=result_header_cell["row"] + 1,
            column=result_header_cell["column"],
        )
    raw_detail["residual_defect_check"] = {
        "result_header": result_header,
        "header_cell": result_header_cell or {},
        "expected_count": residual_expected,
        "actual_count": len(failed_rows),
        "failed_rows": failed_rows,
    }
    residual_actual_parts = []
    if not result_header_cell:
        residual_actual_parts.append("상세 테스트 결과 열 없음")
    residual_actual_parts.append(f"F {len(failed_rows)}개")
    sub_checks.append({
        "expected": f"[잔여결함] {result_header} F 개수 = {residual_expected if residual_expected is not None else '{잔여결함수}'}",
        "actual": " / ".join(residual_actual_parts),
        "passed": bool(residual_expected is not None and result_header_cell and len(failed_rows) == residual_expected),
        "message": config.get("residual_message") or "잔여 결함이 작성되지 않음",
    })

    raw_detail["sub_checks"] = [
        {"expected": item["expected"], "actual": item["actual"], "passed": item["passed"],
         "message": item.get("message", "")}
        for item in sub_checks
    ]
    raw_detail.pop("_expected_sub_check_templates", None)
    all_passed = all(item["passed"] for item in sub_checks)
    first_fail = next((item for item in sub_checks if not item["passed"]), None)
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS if all_passed else DownloadReviewRuleStatus.FAIL,
        expected=" / ".join(item["expected"] for item in sub_checks),
        actual=" / ".join(item["actual"] for item in sub_checks),
        message=(config.get("pass_message") or "테스트케이스를 확인했습니다.") if all_passed else (_append_current_value(first_fail["message"], first_fail.get("actual")) if first_fail else "테스트케이스 확인 필요"),
        file_path=_representative_path(matched, project.project_number),
        file_name=file_info.name,
        raw_detail=raw_detail,
    )


def _test_case_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    existing = raw_detail.get("sub_checks") if isinstance(raw_detail, dict) else None
    raw_detail["sub_checks"] = _complete_expected_sub_checks(
        existing if isinstance(existing, list) else [],
        raw_detail.get("_expected_sub_check_templates"),
        actual=actual,
        message=message,
    )
    raw_detail.pop("_expected_sub_check_templates", None)
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _test_case_expected_sub_check_templates(config, context):
    templates = [
        {
            "expected": "[sheet] one worksheet",
            "message": config.get("sheet_count_message") or "Worksheet count check failed",
        },
    ]
    if config.get("forbidden_footer_terms"):
        templates.append({
            "expected": "[footer] forbidden text not included",
            "message": config.get("footer_message") or "Footer forbidden text check could not run",
        })
    if config.get("forbidden_header_terms"):
        templates.append({
            "expected": "[header] forbidden text not included",
            "message": config.get("header_message") or "Header forbidden text check could not run",
        })
    if config.get("required_footer_terms"):
        templates.append({
            "expected": "[footer required] required text included",
            "message": config.get("footer_required_message") or "Footer required text check could not run",
        })

    title_text = _resolve_rule_value(config.get("title_text") or "{project_number} test case", context)
    author_label = str(config.get("author_label") or "author")
    reviewer_label = str(config.get("reviewer_label") or "reviewer")
    reviewer_expected = _resolve_center_expected(config, context, "reviewer_expected", "")
    date_label = str(config.get("date_label") or "date")
    result_header = str(config.get("result_header") or "test result")
    residual_expected = _context_int(context, "잔여결함수")

    templates.extend([
        {
            "expected": f"[title] contains {title_text}",
            "message": config.get("project_number_message") or "Title check could not run",
        },
        {
            "expected": f"[author/reviewer] {author_label} {context.pl} / {reviewer_label} {reviewer_expected}",
            "message": config.get("author_message") or "Author/reviewer check could not run",
        },
        {
            "expected": f"[date] {date_label} {context.start_date} ~ {context.end_date}",
            "message": config.get("date_message") or "Date check could not run",
        },
        {
            "expected": f"[residual defect] {result_header} F count = {residual_expected if residual_expected is not None else '{residual_defects}'}",
            "message": config.get("residual_message") or "Residual defect check could not run",
        },
    ])
    return templates


def _complete_expected_sub_checks(existing, templates, *, actual, message):
    templates = [item for item in (templates or []) if isinstance(item, dict)]
    rows = []
    for index, item in enumerate(existing or [], start=1):
        if not isinstance(item, dict):
            rows.append(item)
            continue
        row = dict(item)
        row.setdefault("sub_check_key", f"sub-{index}")
        rows.append(row)

    for index in range(len(rows), len(templates)):
        template = templates[index]
        rows.append({
            "sub_check_key": f"sub-{index + 1}",
            "expected": template.get("expected") or "-",
            "actual": actual or "Not checked",
            "passed": False,
            "message": message or template.get("message") or "Not checked because a prerequisite check failed",
            "blocked_by_prerequisite": True,
        })
    return rows


def _test_case_failed_result_rows(sheet, *, start_row, column):
    failed_rows = []
    for row in range(start_row, len(sheet.rows) + 1):
        value = _sheet_cell(sheet, row, column)
        if _normalize_no_space(value).upper() == "F":
            failed_rows.append(row)
    return failed_rows


def _files_under_keyword_folder(files, keywords):
    """경로의 어떤 폴더명이든 keywords 중 하나를 포함하면 그 파일을 선택한다(OR 매칭).

    반환: (선택된 파일 목록, 처음 매칭된 폴더 경로 라벨).
    '설계' 폴더가 없을 때 '스크린샷'/'형상' 등 대체 폴더로 폴백하는 용도.
    """
    needles = [str(k).strip() for k in (keywords or []) if str(k).strip()]
    if not needles:
        return [], ""
    selected = []
    label = ""
    for file_info in files:
        segments = _folder_segments(file_info.path)
        for index, segment in enumerate(segments):
            if any(needle in segment for needle in needles):
                selected.append(file_info)
                if not label:
                    label = "/".join(segments[: index + 1])
                break
    return selected, label


def _evaluate_image_screenshot_folder_date_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    all_files = _matching_files(rule, verify_result)
    min_images = int(config.get("min_images_per_folder") or 5)
    required_folder_count = int(config.get("required_candidate_folder_count") or 2)

    def _find_candidates(files):
        selected_files, selected_folder = _select_folder_chain_files(files, config.get("folder_keyword_chain"))
        # '설계' 폴더가 없으면 '스크린샷'/'형상' 등 대체 폴더로 폴백한다.
        if not selected_folder:
            selected_files, selected_folder = _files_under_keyword_folder(
                files, config.get("fallback_folder_keywords")
            )
        image_files = [
            file_info
            for file_info in selected_files
            if file_info.extension.lower() in IMAGE_EXTENSIONS
        ]
        folders = {}
        for file_info in image_files:
            key = tuple(_folder_segments(file_info.path))
            folders.setdefault(key, []).append(file_info)

        candidate_folders = {
            folder: folder_files
            for folder, folder_files in folders.items()
            if len(folder_files) >= min_images
        }

        # 후보 폴더(이미지 min_images개 이상)들을 상위 폴더 기준으로 묶어
        # required_folder_count개 이상 형제가 있는지 찾는다. 대부분은 같은
        # 바로 위 폴더(부모) 아래 나란히 있지만('형상/최초형상', '형상/최종형상'),
        # '패치 전/기존 제품 형상', '패치 후/재인증 제품 형상'처럼 각 후보 폴더가
        # 서로 다른 한 단계 상위 폴더 아래 하나씩만 있는 구조도 실제로 쓰인다.
        # 이런 경우를 놓치지 않도록 조상 단계를 부모→조부모 순으로 넓혀가며 찾는다.
        max_depth = max((len(folder) for folder in candidate_folders), default=0)
        for ancestor_level in range(1, max_depth + 1):
            ancestor_candidates = {}
            for folder, folder_files in candidate_folders.items():
                if len(folder) < ancestor_level:
                    continue
                ancestor = folder[:-ancestor_level] if ancestor_level else folder
                ancestor_candidates.setdefault(ancestor, []).append((folder, folder_files))

            for ancestor, candidates in sorted(ancestor_candidates.items(), key=lambda item: "/".join(item[0])):
                if len(candidates) >= required_folder_count:
                    selected = sorted(candidates, key=lambda item: "/".join(item[0]))[:required_folder_count]
                    return selected, ancestor, selected_folder, candidate_folders
        return [], None, selected_folder, candidate_folders

    # 원래 규칙대로 먼저 전체 파일에서 직접 '설계'(또는 대체) 폴더를 찾는다.
    selected_candidates, selected_parent, selected_folder, candidate_folders = _find_candidates(all_files)
    used_rawdata_scope = False
    rawdata_files = []

    # 못 찾으면 이름에 'rawdata'가 든 폴더/zip 안에서 다시 찾는다
    # (ECM이 원시자료를 별도의 rawdata.zip으로만 내려준 경우 대비. 대소문자/공백 무시).
    # rawdata 스코프에서도 못 찾으면, 더 유용한 직접 탐색 실패 사유를 그대로 유지한다
    # (rawdata 스코프 실패 사유로 덮어쓰면 오히려 헷갈리는 메시지가 될 수 있음).
    if not selected_candidates:
        rawdata_files = [file_info for file_info in all_files if _is_rawdata_file(file_info, "rawdata")]
        if rawdata_files:
            rd_candidates, rd_parent, rd_folder, rd_candidate_folders = _find_candidates(rawdata_files)
            if rd_candidates:
                selected_candidates, selected_parent, selected_folder, candidate_folders = (
                    rd_candidates, rd_parent, rd_folder, rd_candidate_folders
                )
                used_rawdata_scope = True

    raw_detail = {
        "selected_folder": selected_folder,
        "min_images_per_folder": min_images,
        "required_candidate_folder_count": required_folder_count,
        "used_rawdata_scope": used_rawdata_scope,
        "candidate_folders": [
            {"folder": "/".join(folder), "image_count": len(folder_files)}
            for folder, folder_files in sorted(candidate_folders.items())
        ],
    }
    if not selected_candidates:
        if not selected_folder:
            if not rawdata_files:
                folder_fail_message = config.get("rawdata_missing_message") or "rawdata 폴더를 찾을 수 없습니다"
            else:
                folder_fail_message = config.get("folder_missing_message") or "설계·스크린샷·형상 폴더를 찾을 수 없습니다"
        else:
            folder_fail_message = config.get("folder_message") or "제품 스크린샷 폴더를 찾을 수 없음"
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected=f"이미지 {min_images}개 이상 폴더 {required_folder_count}개",
            actual=f"후보 폴더 {len(candidate_folders)}개",
            message=folder_fail_message,
            raw_detail=raw_detail,
        )

    start_dt = _date_range_start(context.start_date)
    end_dt = _date_range_end(context.end_date)
    if not (start_dt and end_dt):
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected="{시작일} ~ {종료일} 기준정보",
            actual=f"{context.start_date or '(없음)'} ~ {context.end_date or '(없음)'}",
            message=config.get("date_message") or "제품 스크린샷 생성일이 시험기간과 다름",
            raw_detail={**raw_detail, "selected_parent": "/".join(selected_parent or ())},
        )

    selected_files = [file_info for _folder, folder_files in selected_candidates for file_info in folder_files]
    out_of_range = [
        file_info
        for file_info in selected_files
        if not (file_info.modified_at and start_dt <= file_info.modified_at <= end_dt)
    ]
    selected_folders = [
        {"folder": "/".join(folder), "image_count": len(folder_files)}
        for folder, folder_files in selected_candidates
    ]
    raw_detail.update({
        "selected_parent": "/".join(selected_parent or ()),
        "selected_candidate_folders": selected_folders,
        "date_range": {"start": context.start_date, "end": context.end_date},
        "out_of_range_date_counts": _image_modified_date_counts(out_of_range),
        "out_of_range_files": [
            {
                "path": _display_path(file_info.path, project.project_number),
                "modified_at": file_info.modified_at.isoformat() if file_info.modified_at else "",
            }
            for file_info in out_of_range[:20]
        ],
    })
    status = DownloadReviewRuleStatus.PASS if not out_of_range else DownloadReviewRuleStatus.FAIL
    failure_message = _image_out_of_range_message(context, out_of_range)
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected=f"이미지 수정일자 {context.start_date} ~ {context.end_date}",
        actual="범위 밖 파일 없음" if not out_of_range else failure_message,
        message=(
            config.get("pass_message")
            if status == DownloadReviewRuleStatus.PASS
            else failure_message
        ) or "제품 스크린샷을 확인했습니다.",
        file_path="/".join((selected_parent or ())),
        file_name="",
        raw_detail=raw_detail,
    )


def _image_out_of_range_message(context, out_of_range):
    date_counts = _image_modified_date_counts(out_of_range)
    date_text = _join_korean_or(list(date_counts))
    return (
        f"시험기간은 {context.start_date}~{context.end_date}인데 "
        f"수정일자가 {date_text}인 이미지가 {len(out_of_range)}개 존재함"
    )


def _image_modified_date_counts(files):
    counts = {}
    for file_info in files:
        if file_info.modified_at:
            date_text = _format_dot_date(file_info.modified_at.date().isoformat())
        else:
            date_text = "수정일자 없음"
        counts[date_text] = counts.get(date_text, 0) + 1
    return dict(sorted(counts.items()))


def _join_korean_or(values):
    values = [str(value) for value in values if str(value)]
    if not values:
        return "수정일자 없음"
    if len(values) == 1:
        return values[0]
    if len(values) == 2:
        return f"{values[0]} 또는 {values[1]}"
    return ", ".join(values[:-1]) + f" 또는 {values[-1]}"


def _evaluate_rawdata_folder_structure_check(rule, sequence, project, verify_result):
    config = rule.config_json or {}
    all_files = _inspection_files(verify_result)

    def _run_checks(files):
        folders, file_folders = _folder_tree_from_files(files)
        checks = []
        passed = True
        first_message = ""
        for folder_check in config.get("folder_checks") or []:
            result = _run_folder_check(folders, file_folders, folder_check, files=files)
            checks.append(result)
            if not result["passed"] and passed:
                passed = False
                first_message = result["message"]
        return passed, first_message, checks

    # 원래 규칙대로 먼저 전체 파일에서 결함/보안/성능 폴더를 직접 찾아 점검한다.
    passed, first_message, checks = _run_checks(all_files)
    used_rawdata_scope = False
    rawdata_files = []

    # 못 찾으면 이름에 'rawdata'가 든 zip/폴더 안에서 다시 찾는다
    # (ECM이 원시자료를 별도의 rawdata.zip으로만 내려준 경우 대비).
    # rawdata 스코프에서도 통과하지 못하면, 더 유용한 직접 탐색 실패 사유를 그대로 유지한다
    # (rawdata 스코프 실패 사유로 덮어쓰면 오히려 헷갈리는 메시지가 될 수 있음).
    if not passed:
        rawdata_files = [
            file_info
            for file_info in all_files
            if _is_rawdata_file(file_info, "rawdata")
        ]
        if rawdata_files:
            rd_passed, rd_first_message, rd_checks = _run_checks(rawdata_files)
            if rd_passed:
                passed, first_message, checks = rd_passed, rd_first_message, rd_checks
                used_rawdata_scope = True

    raw_detail = {
        "rawdata_file_count": len(rawdata_files),
        "used_rawdata_scope": used_rawdata_scope,
        "checks": checks,
    }
    status = DownloadReviewRuleStatus.PASS if passed else DownloadReviewRuleStatus.FAIL
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=status,
        expected="rawdata 폴더 구조 충족",
        actual="정상" if passed else first_message,
        message=config.get("pass_message") if passed else first_message,
        file_path="rawdata",
        raw_detail=raw_detail,
    )


def _evaluate_test_report_document_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
    ]
    word_files = [file_info for file_info in matched if _is_word_file(file_info)]
    pdf_files = [file_info for file_info in matched if file_info.extension.lower() == ".pdf"]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
        "word_count": len(word_files),
        "pdf_count": len(pdf_files),
    }

    if len(word_files) != 1 or len(pdf_files) != 1:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.FAIL,
            expected="Word 파일 1개 / PDF 파일 1개",
            actual=f"Word 파일 {len(word_files)}개 / PDF 파일 {len(pdf_files)}개",
            message=_test_plan_file_failure_message(
                config,
                matched=matched,
                selected_folder=selected_folder,
                word_files=word_files,
                pdf_files=pdf_files,
            ),
            file_path=_representative_path(matched or files, project.project_number),
            file_name=_representative_name(matched or files),
            raw_detail=raw_detail,
        )

    docx_file = word_files[0]
    pdf_file = pdf_files[0]
    try:
        rounds = _docx_defect_report_round_dates(docx_file)
        spec_table = _docx_first_table_after_text(docx_file, config.get("spec_marker") or "<세부사양>")
        # 마커 기반 탐색(_docx_first_table_after_text)은 안 쓴다: '4.4 시험일정'은
        # 목차에도 그대로 나와서 목차 뒤 첫 표(회사 개요 표 등)를 잘못 집어온다.
        # 대신 헤더 셀 자체('소요일수')로 표를 직접 찾는다(문서 전체에 유일함).
        report_schedule_table = _docx_table_with_header(
            docx_file, config.get("report_schedule_header") or "소요일수",
        )
        report_schedule_values, _ = _schedule_column_values(
            report_schedule_table, config.get("report_schedule_header") or "소요일수",
        )
        artifact = _store_pdf_first_page_artifact(
            project,
            rule,
            pdf_file,
            artifact_id="pdf_first_page",
            label=config.get("pdf_artifact_label") or "시험성적서 1페이지",
        )
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="시험성적서 Word/PDF 파싱 가능",
            actual=str(exc),
            message=config.get("parse_error_message") or str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=_representative_name(matched),
            raw_detail=raw_detail,
        )

    raw_detail.update({
        "variables": {
            # rounds가 비어 있으면(최신 서식에 표 자체가 없음) 결함차수를 0으로
            # 단정하지 않고 값을 생략한다 - 결함리포트 규칙(artifact_10)이 이 경우
            # 결함리포트 파일 자체에서 차수를 추론하는 폴백을 타도록 하기 위함.
            **({"결함차수": len(rounds)} if rounds else {}),
            **rounds,
            "시험성적서_세부사양표": spec_table or [],
            # 시험계획서 2.2 시험일정 WD 열과 직접 대조하기 위한 값(4.4 시험일정의
            # '소요일수' 열). 계획서 쪽에서 값이 있으면 이걸 기준으로 비교한다.
            "시험성적서_시험일정": report_schedule_values,
        },
        "spec_table": spec_table or [],
        "artifacts": [artifact],
    })
    header_text = _docx_header_text(docx_file)
    footer_text = _docx_footer_text(docx_file)
    footer_form = _resolve_rule_value(str(config.get("footer_form_number") or ""), context)
    sub_checks = []
    # 1) 결함리포트 송부 표 차수별 보고일자
    # 이 항목은 차수별 보고일자를 변수로 뽑아 결함리포트 규칙(artifact_10)에
    # 넘겨주는 역할만 한다 - 몇 차까지 있어야 하는지(1차만/1~3차 등)는 실제
    # 결함리포트 파일 구조와 대조해야 판단할 수 있고, 그건 artifact_10의
    # 일이다. 예전에는 여기서 '1차, 2차가 둘 다 있어야 함'을 강제했는데,
    # 실제로는 1차만 있는 프로젝트도 정상이라 오탐이었다. 표를 못 찾은
    # 경우(최신 서식)도 포함해 여기서는 항상 통과 처리한다.
    sub_checks.append({
        "expected": "결함리포트 송부 표의 차수별 보고일자(있는 만큼)",
        "actual": ", ".join(f"{key}: {value}" for key, value in rounds.items()) or "표 없음(최신 서식)",
        "passed": True,
        "message": config.get("round_date_message") or "결함리포트 송부 정보 확인 불가",
    })
    # 2) 머리글에 프로젝트번호
    sub_checks.append({
        "expected": f"머리글에 프로젝트번호 {context.project_number} 포함",
        "actual": header_text or "머리글 없음",
        "passed": bool(context.project_number and context.project_number in header_text),
        "message": config.get("header_message") or "머리글에 프로젝트번호가 잘못 작성됨",
    })
    # 3) 바닥글에 서식번호 (공백 제거 후 비교)
    if footer_form:
        sub_checks.append({
            "expected": f"바닥글에 {footer_form} 포함",
            "actual": footer_text or "바닥글 없음",
            "passed": _normalize_no_space(footer_form) in _normalize_no_space(footer_text),
            "message": config.get("footer_message") or "바닥글에 서식번호가 잘못 작성됨",
        })

    raw_detail["sub_checks"] = [
        {"expected": c["expected"], "actual": c["actual"], "passed": c["passed"],
         "message": c.get("message", "")} for c in sub_checks
    ]
    all_passed = all(c["passed"] for c in sub_checks)
    first_fail = next((c for c in sub_checks if not c["passed"]), None)
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS if all_passed else DownloadReviewRuleStatus.FAIL,
        expected="Word/PDF 1개 / 1차·2차 보고일자 / 머리글 프로젝트번호 / 바닥글 서식번호",
        actual=f"Word {docx_file.name} / PDF {pdf_file.name} / 결함차수 {len(rounds)}",
        message=(config.get("pass_message") or "시험성적서를 확인했습니다.") if all_passed else _append_current_value(first_fail["message"], first_fail["actual"]),
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _evaluate_defect_report_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _extension_matches(file_info.extension, [".xlsx", ".xls"])
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }
    if not matched:
        if config.get("folder_keyword_chain") and not selected_folder:
            file_missing_message = config.get("folder_missing_message") or "수행 폴더를 찾을 수 없습니다"
        elif any(_name_contains_all(file_info.name, name_keywords) for file_info in files):
            file_missing_message = config.get("extension_mismatch_message") or "결함리포트 파일이 xlsx/xls가 아닙니다"
        else:
            file_missing_message = config.get("missing_message") or "결함리포트 파일을 찾을 수 없습니다"
        return _defect_report_failure(
            rule, sequence, matched or files, project, raw_detail,
            expected="결함리포트 Excel 파일",
            actual="결함리포트 파일 없음",
            message=file_missing_message,
        )
    versioned_files = _defect_report_versioned_files(matched, config)
    raw_detail["versions"] = {
        str(version): _display_path(file_info.path, project.project_number)
        for version, file_info in versioned_files.items()
    }
    _collect_defect_report_variables_from_latest_file(raw_detail, versioned_files)

    defect_round_count = _context_int(context, "결함차수")
    inferred_round_count = False
    if defect_round_count is None:
        # 13번(시험성적서)이 결함차수를 산출하지 못한 경우(최신 서식에 결함리포트
        # 송부 표 자체가 없음) 결함리포트 파일들의 시트 구성에서 직접 차수를
        # 추론한다. 시험성적서 파싱에만 의존하면 그 표가 없는 모든 제출물이
        # 무조건 부적합 처리되어 버린다.
        defect_round_count = _infer_defect_round_count_from_files(versioned_files)
        inferred_round_count = defect_round_count is not None
    raw_detail["defect_round_count_inferred_from_files"] = inferred_round_count
    if defect_round_count is None:
        return _defect_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="13번 시험성적서 또는 결함리포트 시트 구성에서 {결함차수} 산출",
            actual="{결함차수} 없음",
            message=config.get("count_mismatch_message") or "시험성적서의 결함 차수와 결함리포트 개수가 다름",
        )

    expected_file_count = defect_round_count + 1
    raw_detail["expected_file_count"] = expected_file_count
    if len(matched) != expected_file_count:
        return _defect_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected=f"결함리포트 Excel 파일 {expected_file_count}개",
            actual=f"결함리포트 Excel 파일 {len(matched)}개",
            message=config.get("count_mismatch_message") or "시험성적서의 결함 차수와 결함리포트 개수가 다름",
        )

    expected_versions = set(range(1, expected_file_count + 1))
    if set(versioned_files) != expected_versions:
        return _defect_report_failure(
            rule,
            sequence,
            matched,
            project,
            raw_detail,
            expected="버전 " + ", ".join(f"v{version}.x" for version in sorted(expected_versions)),
            actual=(
                "버전 " + ", ".join(
                    _defect_report_version_label(versioned_files, version)
                    for version in sorted(versioned_files)
                )
                if versioned_files else "버전 없음"
            ),
            message=config.get("filename_message") or "결함리포트 파일명이 잘못됨",
        )

    workbook_by_version = {}
    for version, file_info in versioned_files.items():
        try:
            workbook_by_version[version] = _read_excel_workbook(file_info)
        except DownloadReviewInspectionError as exc:
            return RuleEvaluation(
                rule=rule,
                sequence=sequence,
                status=DownloadReviewRuleStatus.ERROR,
                expected="결함리포트 Excel 파일 파싱 가능",
                actual=f"{file_info.name}: {exc}",
                message=config.get("parse_error_message") or str(exc),
                file_path=_representative_path(matched, project.project_number),
                file_name=file_info.name,
                raw_detail=raw_detail,
            )

    final_workbook = workbook_by_version.get(expected_file_count)
    if final_workbook:
        raw_detail["variables"] = _defect_report_variables(final_workbook)

    # 첫 실패에서 멈추지 않고 모든 세부 점검을 수행한 뒤 차시별 적합/부적합을 표시한다.
    header_forbidden_check = _check_versioned_workbook_forbidden_print_terms(
        workbook_by_version,
        config.get("forbidden_header_terms") or [],
        context,
        field="header_text",
        subject="결함리포트 머리글",
        default_message="결함리포트 머리글에 잘못된 단어가 작성됨",
    )
    footer_forbidden_check = _check_versioned_workbook_forbidden_print_terms(
        workbook_by_version,
        config.get("forbidden_footer_terms") or [],
        context,
        field="footer_text",
        subject="결함리포트 바닥글",
        default_message="결함리포트 바닥글에 잘못된 단어가 작성됨",
    )
    raw_detail["print_text_checks"] = {
        "forbidden_headers": header_forbidden_check,
        "forbidden_footers": footer_forbidden_check,
    }

    sheet_check = _check_defect_report_sheets(workbook_by_version, versioned_files, defect_round_count)
    raw_detail["sheet_checks"] = sheet_check["details"]

    environment_check = _check_defect_report_environment(workbook_by_version)
    raw_detail["environment_check"] = environment_check

    report_date_check = _check_defect_report_dates(workbook_by_version, context, defect_round_count)
    raw_detail["report_date_checks"] = report_date_check["details"]

    sub_checks = []

    header_forbidden_words = _print_terms_text(config.get("forbidden_header_terms")) or "(없음)"
    footer_forbidden_words = _print_terms_text(config.get("forbidden_footer_terms")) or "(없음)"
    header_msg = config.get("header_message") or "결함리포트 머리글에 프로젝트번호가 작성됨"
    footer_msg = config.get("footer_message") or "결함리포트 바닥글에 금지 단어가 작성됨"
    sheet_msg = config.get("sheet_message") or "결함리포트 시트 구성이 잘못됨"
    environment_msg = config.get("environment_message") or "시험환경 정보가 잘못 작성됨"

    # 1) 머리글 금지어 (차수별로 집계, 실제로 걸린 금지어의 메시지 사용)
    header_fail_msg = _defect_print_fail_message(header_forbidden_check, header_msg)
    for version, version_passed in _defect_print_pass_by_version(header_forbidden_check):
        version_label = _defect_report_version_label(versioned_files, version)
        sub_checks.append({
            "expected": f"[{version_label} 머리글] 금지어 미포함: {header_forbidden_words}",
            "actual": "정상" if version_passed else f"금지어 포함: {header_forbidden_words}",
            "passed": version_passed,
            "message": header_fail_msg,
        })
    # 2) 바닥글 금지어 (차수별)
    footer_fail_msg = _defect_print_fail_message(footer_forbidden_check, footer_msg)
    for version, version_passed in _defect_print_pass_by_version(footer_forbidden_check):
        version_label = _defect_report_version_label(versioned_files, version)
        sub_checks.append({
            "expected": f"[{version_label} 바닥글] 금지어 미포함: {footer_forbidden_words}",
            "actual": "정상" if version_passed else f"금지어 포함: {footer_forbidden_words}",
            "passed": version_passed,
            "message": footer_fail_msg,
        })
    # 3) 시트 구성 (차수별)
    for detail in sheet_check.get("details", []):
        sheet_ok = not detail.get("missing_sheets") and not detail.get("extra_sheets")
        problems = []
        if detail.get("missing_sheets"):
            problems.append("누락: " + ", ".join(detail["missing_sheets"]))
        if detail.get("extra_sheets"):
            problems.append("불필요: " + ", ".join(detail["extra_sheets"]))
        version_label = _defect_report_version_label(versioned_files, detail.get("version"))
        sub_checks.append({
            "expected": f"[{version_label} 시트] " + ", ".join(detail.get("expected_sheets", [])),
            "actual": "정상" if sheet_ok else " / ".join(problems),
            "passed": sheet_ok,
            "message": sheet_msg,
        })
    # 4) 시험환경 — 차수(버전)별로 한 줄씩 값을 표시해 구분할 수 있게 한다.
    env_values = environment_check.get("values") or []
    env_by_version = {}
    for item in env_values:
        env_by_version.setdefault(item.get("version"), item.get("value"))
    baseline_value = env_values[0]["value"] if env_values else ""
    if env_by_version:
        env_actual = "\n".join(
            f"{_defect_report_version_label(versioned_files, version)}: {value}"
            for version, value in sorted(env_by_version.items())
        )
    else:
        env_actual = _stringify_check_value(environment_check.get("actual", "")) or "시험환경 없음"
    # 기대값(기준=첫 시트 값)과 실제값(버전별)을 모두 표시해 사용자가 비교할 수 있게 한다.
    env_expected = (
        f"[시험환경] 모든 시트 동일 (기준: {baseline_value})"
        if baseline_value else "[시험환경] 모든 시트 1~3행 값 동일"
    )
    sub_checks.append({
        "expected": env_expected,
        "actual": env_actual,
        "passed": bool(environment_check.get("passed")),
        "message": environment_msg,
    })
    # 5) 보고일자 (차시·시트별)
    sub_checks.extend(report_date_check.get("sub_checks", []))
    # 6) 시험분석자료 요약표(품질특성별/결함정도별)가 실제 결함 목록과 일치하는지
    if final_workbook:
        sub_checks.extend(_defect_report_summary_consistency_checks(final_workbook))

    raw_detail["sub_checks"] = sub_checks
    variables = raw_detail.get("variables") or _defect_report_variables(final_workbook)
    raw_detail["variables"] = variables

    all_passed = all(item["passed"] for item in sub_checks)
    first_fail = next((item for item in sub_checks if not item["passed"]), None)
    fail_message = (
        (first_fail.get("message") if first_fail else None)
        or config.get("report_date_message")
        or "결함리포트 확인 필요"
    )
    if first_fail:
        fail_message = _append_current_value(fail_message, first_fail.get("actual"))
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS if all_passed else DownloadReviewRuleStatus.FAIL,
        expected=" / ".join(item["expected"] for item in sub_checks),
        actual=" / ".join(item["actual"] for item in sub_checks),
        message=(config.get("pass_message") or "결함리포트를 확인했습니다.") if all_passed else fail_message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _print_terms_text(terms):
    """머리글/바닥글 금지어·필수어 설정에서 단어 목록을 사람이 읽을 문자열로 만든다."""
    words = []
    for term in terms or []:
        word = str(term.get("text") if isinstance(term, dict) else term or "").strip()
        if word:
            words.append(word)
    return ", ".join(words)


def _defect_print_pass_by_version(check):
    """차수(version)별로 머리글/바닥글 금지어 검사 통과 여부를 집계한다."""
    by_version = {}
    for detail in check.get("details", []) if isinstance(check, dict) else []:
        version = detail.get("version")
        by_version.setdefault(version, True)
        if not detail.get("passed"):
            by_version[version] = False
    return sorted(by_version.items(), key=lambda item: (item[0] is None, item[0]))


def _defect_print_fail_message(check, default):
    """머리글/바닥글 금지어 검사에서 실제로 걸린 금지어의 메시지를 찾는다(없으면 default)."""
    for detail in check.get("details", []) if isinstance(check, dict) else []:
        for term in detail.get("checks") or []:
            if not term.get("passed"):
                return term.get("message") or default
    return default


def _defect_report_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _infer_defect_round_count_from_files(versioned_files):
    """결함리포트 파일들의 시트명에서 직접 결함차수를 추론한다.

    시험성적서에 결함리포트 송부 표가 없어 결함차수를 못 구한 경우의 폴백.
    가장 높은 버전 파일의 시트 중 'N차 결함리포트' 형태의 최대 N을 차수로 본다.
    """
    versions = [version for version in versioned_files if version > 0]
    if not versions:
        return None
    latest_file = versioned_files[max(versions)]
    try:
        workbook = _read_excel_workbook(latest_file)
    except DownloadReviewInspectionError:
        return None
    round_numbers = set()
    for sheet in workbook.sheets:
        match = re.match(r"^\s*(\d+)\s*차\s*결함리포트", sheet.name or "")
        if match:
            round_numbers.add(int(match.group(1)))
    return max(round_numbers) if round_numbers else None


def _defect_report_versioned_files(files, config):
    pattern = re.compile(str(config.get("version_pattern") or r"(?i)v(\d+)(?:[._-]\d+)*"))
    versioned = {}
    for file_info in files:
        revision = _artifact_revision_info(file_info)
        if revision:
            version = revision.major
        else:
            match = pattern.search(file_info.name)
            if not match:
                continue
            version = int(match.group(1))
        versioned[version] = file_info
    return dict(sorted(versioned.items()))


def _check_defect_report_sheets(workbook_by_version, versioned_files, defect_round_count):
    details = []
    final_version = defect_round_count + 1
    for version in sorted(workbook_by_version):
        workbook = workbook_by_version[version]
        actual_names = [sheet.name for sheet in workbook.sheets]
        expected_names = _expected_defect_report_sheet_names(version, final_version)
        expected_normalized = {_normalize_no_space(name) for name in expected_names}
        actual_normalized = {_normalize_no_space(name) for name in actual_names}
        missing = [
            name
            for name in expected_names
            if _normalize_no_space(name) not in actual_normalized
        ]
        extra = [
            name
            for name in actual_names
            if _normalize_no_space(name) not in expected_normalized
        ]
        detail = {
            "version": version,
            "file_name": versioned_files[version].name,
            "expected_sheets": expected_names,
            "actual_sheets": actual_names,
            "missing_sheets": missing,
            "extra_sheets": extra,
        }
        details.append(detail)
        if missing or extra:
            return {
                "passed": False,
                "file_name": versioned_files[version].name,
                "expected": ", ".join(expected_names),
                "actual": ", ".join(actual_names) or "시트 없음",
                "details": details,
            }
    return {"passed": True, "details": details}


def _expected_defect_report_sheet_names(version, final_version):
    if version == final_version:
        return [
            *[f"{round_no}차 결함리포트" for round_no in range(1, final_version)],
            "최종결함리포트",
            "시험분석자료",
        ]
    return [f"{round_no}차 결함리포트" for round_no in range(1, version + 1)]


def _check_defect_report_environment(workbook_by_version):
    """차수별(N차 결함리포트) 시트의 시험환경이 모든 버전에서 동일한지 확인한다.

    '최종결함리포트'/'시험분석자료'는 결과 요약용 시트로 시험환경 항목을 아예
    싣지 않는 표준 서식이라(표지+보고일자만 존재), 이 시트들까지 포함해
    시험환경 값을 요구하면 정상 제출물도 '시험환경 없음'으로 오탐한다.
    규칙 문서도 'v1.0, v2.0, v3.0' 즉 차수별 리포트 시트만 언급하므로
    차수 시트로 범위를 한정한다.
    """
    round_sheet_pattern = re.compile(r"^\s*\d+\s*차\s*결함리포트")
    values = []
    for version, workbook in sorted(workbook_by_version.items()):
        for sheet in workbook.sheets:
            if not round_sheet_pattern.match(sheet.name or ""):
                continue
            value = _sheet_top_rows_label_value(sheet, "시험환경")
            if not value:
                return {
                    "passed": False,
                    "actual": f"v{version}.0 {sheet.name}: 시험환경 없음",
                    "values": values,
                }
            values.append({"version": version, "sheet": sheet.name, "value": value})

    normalized = {_normalize_compare(value["value"], {"remove_whitespace": True}) for value in values}
    if len(normalized) != 1:
        return {
            "passed": False,
            "actual": "; ".join(f"v{item['version']}.0 {item['sheet']}: {item['value']}" for item in values),
            "values": values,
        }
    return {"passed": True, "actual": values[0]["value"] if values else "", "values": values}


def _check_defect_report_dates(workbook_by_version, context, defect_round_count):
    details = []
    final_version = defect_round_count + 1
    expected_dates = {
        f"{round_no}차 결함리포트": _context_variable(context, f"{round_no}차")
        for round_no in range(1, defect_round_count + 1)
    }
    expected_dates["최종결함리포트"] = context.end_date
    expected_dates["시험분석자료"] = context.end_date

    expected_parts = []
    actual_parts = []
    sub_checks = []
    all_passed = True
    for version, workbook in sorted(workbook_by_version.items()):
        required_sheets = _expected_defect_report_sheet_names(version, final_version)
        for sheet_name in required_sheets:
            sheet = _workbook_sheet(workbook, sheet_name)
            # 프로젝트번호, 시트명, 보고일자는 서로 다른 셀에 작성되는 양식도 허용한다.
            project_text = _sheet_top_rows_cell_containing(sheet, context.project_number) if sheet and context.project_number else ""
            sheet_text = _sheet_top_rows_cell_containing(sheet, sheet_name) if sheet else ""
            combined_text = _sheet_top_rows_cell_containing(sheet, f"{context.project_number} {sheet_name}") if sheet and context.project_number else ""
            header_text = combined_text or " / ".join(
                text for text in (project_text, sheet_text) if text
            )
            report_date = _sheet_top_rows_report_date(sheet) if sheet else ""
            actual_text = (header_text + (" / 보고일자: " + report_date if report_date else "")).strip()
            expected_date = expected_dates.get(sheet_name, "")
            # 라운드(N차) 결함리포트는 표지 제목이 'TTA-XX-XXXXX 결함리포트'처럼 적히고
            # 'N차'는 시트 탭 이름에만 존재한다. 따라서 표지 존재 여부를 시트명 그대로가
            # 아니라 'N차 ' 접두어를 뗀 기본 명칭(예: '결함리포트')으로도 확인한다.
            # (이 완화가 없으면 보고일자가 일치해도 표지 문구 불일치로 부적합 처리된다.)
            base_sheet_name = re.sub(r"^\s*\d+차\s*", "", sheet_name)
            header_found = bool(
                sheet_text
                or (
                    sheet
                    and base_sheet_name != sheet_name
                    and _sheet_top_rows_cell_containing(sheet, base_sheet_name)
                )
            )
            # expected_date가 비어 있는 건 'N차' 라운드인데 시험성적서에 결함리포트
            # 송부 표가 없어(최신 서식) 기준 날짜를 못 구한 경우다(최종결함리포트/
            # 시험분석자료는 항상 context.end_date가 있으므로 영향받지 않음).
            # 이때는 비교 기준이 없으므로 보고일자 값 자체의 존재만 확인한다.
            passed = bool(
                header_found
                and (_same_date_text(report_date, expected_date) if expected_date else bool(report_date))
            )
            detail = {
                "version": version,
                "sheet": sheet_name,
                "expected_date": expected_date,
                "actual_text": actual_text,
                "actual_date": report_date,
                "project_text": project_text,
                "sheet_text": sheet_text,
                "passed": passed,
            }
            details.append(detail)
            # 차시별로 한 줄씩 표시되도록 " / " 로 구분해 누적한다.
            label = f"v{version}.0 {sheet_name}"
            # 형식이 달라도 값으로 비교되도록, 기대/실제 날짜를 정규화 형식(YYYY.MM.DD.)으로 나란히 표시한다.
            expected_display = _format_dot_date(expected_date) or (expected_date or "(기준없음)")
            actual_display = _format_dot_date(report_date) or (report_date or "문구없음")
            sub_expected = f"{label} 기대 보고일자 {expected_display}"
            sub_actual = f"{label} 실제 보고일자 {actual_display}"
            expected_parts.append(sub_expected)
            actual_parts.append(sub_actual)
            sub_checks.append({"expected": sub_expected, "actual": sub_actual, "passed": passed})
            if not passed:
                all_passed = False

    return {
        "passed": all_passed,
        "expected": " / ".join(expected_parts),
        "actual": " / ".join(actual_parts) or "결함리포트 보고일자 확인 불가",
        "details": details,
        "sub_checks": sub_checks,
    }


def _collect_defect_report_variables_from_latest_file(raw_detail, versioned_files):
    versions = [version for version in versioned_files if version > 0]
    if not versions:
        return
    latest_version = max(versions)
    try:
        workbook = _read_excel_workbook(versioned_files[latest_version])
    except DownloadReviewInspectionError as exc:
        raw_detail["variable_error"] = str(exc)
        return
    variables = _defect_report_variables(workbook)
    if variables:
        raw_detail["variables"] = variables


def _defect_report_variables(workbook):
    variables = {}
    if _workbook_sheet(workbook, "최종결함리포트"):
        variables["잔여결함수"] = _defect_residual_count(workbook)
    if _workbook_sheet(workbook, "시험분석자료"):
        variables["H"] = _defect_high_count(workbook)
        variables["R"] = _defect_analysis_value(workbook, "수정전", offset_rows=5, offset_cols=0)
    return variables


def _sheet_top_rows_cell_containing(sheet, keyword, *, limit=4):
    """상단 행에서 keyword를 포함한 셀 값을 반환한다(공백 차이 무시)."""
    needle = _normalize_no_space(keyword)
    for row in sheet.rows[:limit]:
        for value in row:
            if needle and needle in _normalize_no_space(value):
                return value
    return ""


def _sheet_top_rows_label_value(sheet, keyword, *, limit=4):
    """상단 행의 라벨 셀 오른쪽에 실제 값이 적힌 양식도 함께 읽는다."""
    needle = _normalize_no_space(keyword)
    if not sheet or not needle:
        return ""
    for row_index, row in enumerate(sheet.rows[:limit], start=1):
        for column_index, value in enumerate(row, start=1):
            if needle not in _normalize_no_space(value):
                continue
            right_value = _sheet_cell(sheet, row_index, column_index + 1)
            if right_value and _is_label_only_cell(value, keyword):
                return _normalize_spaces(f"{value} {right_value}")
            return value
    return ""


def _is_label_only_cell(value, keyword):
    text = str(value or "")
    remainder = text.replace(str(keyword or ""), "")
    remainder = re.sub(r"[\s\[\]\(\):：]+", "", remainder)
    return not remainder


def _sheet_top_rows_report_date(sheet, *, limit=5):
    """상단 행에서 '보고일자'가 적힌 셀(헤더와 다른 셀일 수 있음)을 찾아 날짜를 추출한다."""
    if not sheet:
        return ""
    for row in sheet.rows[:limit]:
        for value in row:
            date = _extract_korean_report_date(value)
            if date:
                return date
    return ""


def _workbook_sheet(workbook, sheet_name):
    # 정확히 일치하는 시트를 우선 찾고, 없으면 공백 차이를 무시해 찾는다.
    # 예: '최종결함리포트' = '최종 결함리포트' = '최종 결함 리포트'
    for sheet in workbook.sheets:
        if sheet.name == sheet_name:
            return sheet
    target = _normalize_no_space(sheet_name)
    for sheet in workbook.sheets:
        if _normalize_no_space(sheet.name) == target:
            return sheet
    return None


def _extract_korean_report_date(value):
    # 보고일자 형식이 '2026년 4월 13일' / '2026.04.13' / '26.4.13' 등 달라도 인식한다.
    match = re.search(r"보고일자\s*[:：]?\s*" + _FLEX_DATE_PATTERN, str(value or ""))
    if not match:
        return ""
    year, month, day = match.groups()
    return f"{_full_year(year):04d}.{int(month):02d}.{int(day):02d}."


def _parse_loose_date(value):
    """다양한 형식의 날짜를 (year|None, month, day)로 파싱한다.
    연도가 없는 '5/7' 같은 값은 year=None으로 반환한다.
    """
    text = str(value or "")
    text = (
        text.replace("년", ".").replace("월", ".").replace("일", "")
        .replace("/", ".").replace("-", ".")
    )
    # 연.월.일 (연도 2~4자리)
    match = re.search(r"(?<!\d)(\d{2,4})\s*\.\s*(\d{1,2})\s*\.\s*(\d{1,2})(?!\d)", text)
    if match:
        return (_full_year(match.group(1)), int(match.group(2)), int(match.group(3)))
    # 월.일 (연도 없음)
    match = re.search(r"(?<!\d)(\d{1,2})\s*\.\s*(\d{1,2})(?!\d)", text)
    if match:
        return (None, int(match.group(1)), int(match.group(2)))
    return None


def _same_date_text(left, right):
    left_parsed = _parse_loose_date(left)
    right_parsed = _parse_loose_date(right)
    if not left_parsed or not right_parsed:
        return _format_dot_date(left) == _format_dot_date(right)
    ly, lm, ld = left_parsed
    ry, rm, rd = right_parsed
    if (lm, ld) != (rm, rd):
        return False
    # 한쪽이라도 연도가 없으면(예: ECM '5/7') 월/일만 같으면 같은 날짜로 본다.
    if ly is None or ry is None:
        return True
    return ly == ry


def _defect_residual_count(workbook):
    """'최종결함리포트' 시트에서 G열의 '품질특성' 셀을 찾아, 그 아래로 값이 연속으로
    작성된 셀의 개수를 잔여결함 개수로 본다. 바로 아래 셀이 비어 있으면 0개."""
    sheet = _workbook_sheet(workbook, "최종결함리포트")
    if not sheet:
        return 0
    header = _find_cell_containing(sheet.rows, "품질특성")
    if not header:
        return 0
    column = header["column"]
    count = 0
    for row in range(header["row"] + 1, len(sheet.rows) + 1):
        if not _sheet_cell(sheet, row, column):
            break
        count += 1
    return count


def _defect_high_count(workbook):
    sheet = _workbook_sheet(workbook, "시험분석자료")
    if not sheet:
        return "0"
    count = 0
    for row in sheet.rows:
        c_value = row[2] if len(row) > 2 else ""
        e_value = row[4] if len(row) > 4 else ""
        if e_value == "H" and c_value.strip() not in {"", "-"}:
            count += 1
    return str(count)


def _defect_analysis_value(workbook, keyword, *, offset_rows, offset_cols):
    sheet = _workbook_sheet(workbook, "시험분석자료")
    if not sheet:
        return ""
    for row_index, row in enumerate(sheet.rows):
        for col_index, value in enumerate(row[2:6], start=2):
            if keyword not in value:
                continue
            target_row = row_index + offset_rows
            target_col = col_index + offset_cols
            if target_row < len(sheet.rows) and target_col < len(sheet.rows[target_row]):
                return sheet.rows[target_row][target_col]
            return ""
    return ""


def _defect_report_column_run(sheet, *, start_row, column):
    """start_row(1-based)부터 column(0-based)에 연속으로 값이 있는 셀까지 모은다."""
    values = []
    row_index = start_row - 1
    while row_index < len(sheet.rows):
        value = sheet.rows[row_index][column] if column < len(sheet.rows[row_index]) else ""
        if not str(value or "").strip():
            break
        values.append(str(value).strip())
        row_index += 1
    return values


def _defect_report_nth_label_row(sheet, keyword, column, occurrence):
    """column(0-based)에서 keyword가 포함된 occurrence번째(1-based) 셀의 row_index(0-based)."""
    seen = 0
    for row_index, row in enumerate(sheet.rows):
        value = row[column] if column < len(row) else ""
        if keyword in str(value or ""):
            seen += 1
            if seen == occurrence:
                return row_index
    return None


def _defect_report_summary_consistency_check(sheet, values, label_row, *, keyword_labels):
    """values(문자열 목록)에서 각 키워드가 포함된 개수를 세어, label_row 아래
    1~N칸 셀 값과 일치하는지 확인한다. keyword_labels: [(keyword, 표시라벨), ...]."""
    if label_row is None:
        return None
    expected_parts = []
    actual_parts = []
    passed = True
    for offset, (keyword, label) in enumerate(keyword_labels, start=1):
        count = sum(1 for value in values if keyword in value)
        target_row = label_row + offset
        actual_value = (
            sheet.rows[target_row][4]
            if target_row < len(sheet.rows) and 4 < len(sheet.rows[target_row])
            else ""
        )
        expected_parts.append(f"{label}: {count}개")
        actual_parts.append(f"{label}: {actual_value}개")
        if str(actual_value).strip() != str(count):
            passed = False
    return {
        "expected": "\n".join(expected_parts),
        "actual": "\n".join(actual_parts),
        "passed": passed,
    }


def _defect_report_summary_consistency_checks(workbook):
    """시험분석자료 시트의 품질특성별/결함정도별 결함내역 요약표가 실제 결함
    목록(G열=품질특성, E열=결함정도)의 개수와 일치하는지 검증한다.

    결함 목록의 시작 행은 고정 행 번호가 아니라, G열에서 `품질특성` 헤더 셀을,
    E열에서 `결함정도` 헤더 셀을 찾아 그 바로 다음 행부터로 본다(서식이 바뀌어
    헤더 행 위치가 달라져도 안전하게 동작한다).
    """
    sheet = _workbook_sheet(workbook, "시험분석자료")
    if not sheet:
        return []

    quality_header_row = _defect_report_nth_label_row(sheet, "품질특성", column=6, occurrence=1)
    quality_check = None
    if quality_header_row is not None:
        quality_values = _defect_report_column_run(sheet, start_row=quality_header_row + 2, column=6)
        quality_label_row = _defect_report_nth_label_row(sheet, "수정전", column=4, occurrence=1)
        quality_check = _defect_report_summary_consistency_check(
            sheet,
            quality_values,
            quality_label_row,
            keyword_labels=[
                ("기능", "기능적합성"),
                ("성능", "성능효율성"),
                ("호환", "호환성"),
                ("사용", "사용성"),
                ("신뢰", "신뢰성"),
                ("보안", "보안성"),
                ("유지", "유지보수성"),
                ("이식", "이식성"),
                ("일반적", "일반적요구사항"),
            ],
        )
        if quality_check:
            quality_check["message"] = "시험분석자료의 품질특성별 결함내역 표가 결함 목록과 다릅니다"

    severity_header_row = _defect_report_nth_label_row(sheet, "결함정도", column=4, occurrence=1)
    severity_check = None
    if severity_header_row is not None:
        severity_values = _defect_report_column_run(sheet, start_row=severity_header_row + 2, column=4)
        severity_label_row = _defect_report_nth_label_row(sheet, "수정전", column=4, occurrence=2)
        severity_check = _defect_report_summary_consistency_check(
            sheet,
            severity_values,
            severity_label_row,
            keyword_labels=[("H", "H"), ("M", "M"), ("L", "L")],
        )
        if severity_check:
            severity_check["message"] = "시험분석자료의 결함정도별 결함내역 표가 결함 목록과 다릅니다"

    return [check for check in (quality_check, severity_check) if check]


def _context_int(context, key):
    value = _context_variable(context, key)
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _context_variable(context, key):
    return context.derived_variables.get(key, "")


def _format_config_message(template, **values):
    if not template:
        return ""
    message = str(template)
    for key, value in values.items():
        message = message.replace("{" + key + "}", str(value))
    return message


def _stringify_check_value(value):
    """expected/actual 값이 dict/list 등이면 사람이 읽을 수 있는 문자열로 변환한다."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        # 기능적합성 불일치 dict 등
        if "left_cell" in value or "right_cell" in value:
            return (
                f"{value.get('left_cell', '')}({value.get('left', '')}) ≠ "
                f"{value.get('right_cell', '')}({value.get('right', '')})"
            )
        return ", ".join(f"{k}={v}" for k, v in value.items())
    if isinstance(value, (list, tuple)):
        return ", ".join(_stringify_check_value(item) for item in value)
    return str(value)


def _append_current_value(message, actual):
    """부적합 메시지 뒤에 실제(현재) 값을 붙인다: '메시지(현재 값: …)'.

    실제 값이 없거나 '정상'처럼 의미 없는 값이면 메시지를 그대로 둔다.
    """
    message = str(message or "")
    text = _stringify_check_value(actual).strip()
    if not text or text in ("정상", "일치", "없음", "범위 밖 파일 없음", "확인됨"):
        return message
    if "현재 값" in message:
        return message
    if len(text) > 120:
        text = text[:120] + "…"
    text = " ".join(text.split())
    return f"{message}(현재 값: {text})"


def _evaluate_inspection_checklist_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
    ]
    excel_files = [
        file_info
        for file_info in matched
        if _extension_matches(file_info.extension, [".xlsx", ".xls"])
    ]
    pdf_files = [
        file_info
        for file_info in matched
        if _extension_matches(file_info.extension, [".pdf"])
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
        "excel_count": len(excel_files),
        "pdf_count": len(pdf_files),
    }

    if not matched:
        if config.get("folder_keyword_chain") and not selected_folder:
            file_missing_message = config.get("folder_missing_message") or "설계 폴더를 찾을 수 없습니다"
        else:
            file_missing_message = config.get("missing_message") or "점검표 파일을 찾을 수 없습니다"
        return _checklist_failure(
            rule, sequence, matched or files, project, raw_detail,
            expected="점검표 파일", actual="점검표 파일 없음",
            message=file_missing_message,
        )

    if len(excel_files) != 1:
        return _checklist_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="점검표 Excel 파일 1개",
            actual=f"점검표 Excel 파일 {len(excel_files)}개",
            message=config.get("excel_missing_message") or "점검표 Excel 파일을 찾을 수 없습니다",
        )
    if len(pdf_files) != 1:
        return _checklist_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="점검표 PDF 파일 1개",
            actual=f"점검표 PDF 파일 {len(pdf_files)}개",
            message=config.get("pdf_missing_message") or "점검표 pdf 파일이 없음",
        )

    excel_file = excel_files[0]
    pdf_file = pdf_files[0]
    try:
        workbook = _read_excel_workbook(excel_file)
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="점검표 Excel 파일 파싱 가능",
            actual=str(exc),
            message=config.get("parse_error_message") or str(exc),
            file_path=_representative_path(matched, project.project_number),
            file_name=excel_file.name,
            raw_detail=raw_detail,
        )

    raw_detail["sheet_names"] = [sheet.name for sheet in workbook.sheets]

    # 측정항목별 점수표 값은 이후 점검 항목의 통과/실패와 무관하게 항상 변수로 채운다.
    # (품질검사표 규칙이 이 변수를 사용하므로, 점검표가 다른 항목에서 실패해도 값이 있어야 한다.)
    score_sheet = _workbook_sheet(workbook, config.get("score_sheet") or "측정항목별 점수표")
    score_values = _checklist_score_values(score_sheet)
    raw_detail["variables"] = {"측정항목별점수표": score_values}

    # 첫 실패에서 멈추지 않고 모든 세부 점검을 수행한 뒤 항목별 적합/부적합을 표시한다.
    # sub_results: (라벨, check_dict, expected_override, message_override)
    sub_results = []

    header_check = _check_checklist_headers(workbook, context.project_number)
    raw_detail["header_check"] = header_check
    sub_results.append((
        "머리글", header_check,
        f"모든 시트 머리글에 프로젝트번호: {context.project_number}",
        config.get("header_message") or "머리글(프로젝트번호)이 잘못 작성됨",
    ))

    footer_forbidden_check = _check_workbook_forbidden_print_terms(
        workbook,
        config.get("forbidden_footer_terms") or [],
        context,
        field="footer_text",
        subject="점검표 바닥글",
        default_message="점검표 바닥글에 잘못된 단어가 작성됨",
    )
    footer_required_check = _check_workbook_required_print_terms(
        workbook,
        config.get("required_footer_terms") or [],
        context,
        field="footer_text",
        subject="점검표 바닥글",
        default_message="점검표 바닥글에 필요한 단어가 누락됨",
    )
    raw_detail["footer_checks"] = {
        "forbidden": footer_forbidden_check,
        "required": footer_required_check,
    }
    forbidden_footer_words = _print_terms_text(config.get("forbidden_footer_terms")) or "(없음)"
    required_footer_words = _print_terms_text(config.get("required_footer_terms")) or "(없음)"
    sub_results.append((
        "바닥글(금지어)", footer_forbidden_check,
        f"모든 시트 바닥글 금지어 미포함: {forbidden_footer_words}", None,
    ))
    sub_results.append((
        "바닥글(필수어)", footer_required_check,
        f"모든 시트 바닥글 필수어 포함: {required_footer_words}", None,
    ))

    cover_sheet = _workbook_sheet(workbook, config.get("cover_sheet") or "표지")
    if not cover_sheet:
        cover_check = {
            "passed": False,
            "expected": "표지 시트",
            "actual": "표지 시트 없음",
            "message": config.get("cover_title_message") or "표지 제목이 잘못 작성됨",
        }
    else:
        cover_check = _check_checklist_cover(cover_sheet, context, config)
    raw_detail["cover_check"] = cover_check
    sub_results.append(("표지", cover_check, None, None))

    feature_sheet = _workbook_sheet(workbook, config.get("feature_sheet") or "기능별 점검표")
    suitability_sheet = _workbook_sheet(workbook, config.get("suitability_sheet") or "2. 기능적합성")
    reliability_sheet = _workbook_sheet(workbook, config.get("reliability_sheet") or "6. 신뢰성")

    feature_check = _check_checklist_feature_sheet(feature_sheet)
    raw_detail["feature_sheet_check"] = feature_check
    sub_results.append((
        "기능별 점검표 빈셀", feature_check, None,
        config.get("feature_blank_message") or "기능별 점검표 시트에 빈 셀이 확인됨",
    ))

    suitability_result = _check_checklist_suitability_results(suitability_sheet)
    raw_detail["suitability_result_check"] = suitability_result
    sub_results.append((
        "기능적합성 결과값", suitability_result, None,
        config.get("suitability_result_message") or "기능적합성 시트의 기능표 결과값 미작성",
    ))

    reliability_check = _check_checklist_reliability(reliability_sheet, context)
    raw_detail["reliability_check"] = reliability_check
    sub_results.append(("신뢰성", reliability_check, None, None))

    # PDF 1페이지 산출물은 통과/실패와 무관하게 저장한다.
    try:
        artifact = _store_pdf_first_page_artifact(
            project,
            rule,
            pdf_file,
            artifact_id=config.get("pdf_artifact_id") or "pdf_first_page",
            label=config.get("pdf_artifact_label") or "점검표 1페이지",
        )
        raw_detail["artifacts"] = [artifact]
    except DownloadReviewInspectionError:
        pass

    sub_checks = []
    for label, check, expected_override, _message_override in sub_results:
        expected_value = expected_override if expected_override is not None else check.get("expected", "")
        sub_checks.append({
            "expected": f"[{label}] {_stringify_check_value(expected_value)}",
            "actual": _stringify_check_value(check.get("actual", "")),
            "passed": bool(check.get("passed")),
            "message": _message_override if _message_override is not None else check.get("message", ""),
        })
    raw_detail["sub_checks"] = sub_checks

    all_passed = all(item["passed"] for item in sub_checks)
    first_fail = next(
        ((label, check, message_override) for label, check, _e, message_override in sub_results if not check.get("passed")),
        None,
    )
    if all_passed:
        message = config.get("pass_message") or "점검표를 확인했습니다."
    else:
        fail_label, fail_check, fail_message = first_fail
        message = _append_current_value(
            fail_message or fail_check.get("message") or f"{fail_label} 확인 필요",
            fail_check.get("actual"),
        )

    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS if all_passed else DownloadReviewRuleStatus.FAIL,
        expected=" / ".join(item["expected"] for item in sub_checks),
        actual=" / ".join(item["actual"] for item in sub_checks),
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name([excel_file, pdf_file]),
        raw_detail=raw_detail,
    )


def _checklist_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _check_checklist_headers(workbook, project_number):
    expected = f"프로젝트번호: {project_number}"
    failures = []
    details = []
    for sheet in workbook.sheets:
        candidates = [sheet.header_text, *_top_rows_texts(sheet, limit=3)]
        passed = any(_clean_excel_header_text(candidate).find(expected) >= 0 for candidate in candidates)
        details.append({
            "sheet": sheet.name,
            "header_text": sheet.header_text,
            "passed": passed,
        })
        if not passed:
            failures.append(sheet.name)
    return {
        "passed": not failures,
        "expected": expected,
        "actual": "오류 시트: " + ", ".join(failures) if failures else "정상",
        "details": details,
    }


def _check_workbook_forbidden_print_terms(workbook, terms, context, *, field, subject, default_message):
    if not terms:
        return {"passed": True, "expected": "", "actual": "검사 조건 없음", "message": "", "details": []}

    details = []
    failures = []
    for sheet in workbook.sheets:
        text = _clean_excel_header_text(getattr(sheet, field, ""))
        check = _check_forbidden_text_terms(
            text,
            terms,
            context,
            subject=f"{subject}({sheet.name})",
            default_message=default_message,
        )
        details.append({
            "sheet": sheet.name,
            "text": text,
            "passed": check["passed"],
            "checks": check["details"],
        })
        if not check["passed"]:
            failures.append({"sheet": sheet.name, "actual": check["actual"], "message": check["message"]})

    return {
        "passed": not failures,
        "expected": f"모든 시트 {subject} 금지어 없음",
        "actual": (f"{failures[0]['sheet']}: {failures[0]['actual']}" if failures else "정상"),
        "message": failures[0]["message"] if failures else "",
        "details": details,
    }


def _check_workbook_required_print_terms(workbook, terms, context, *, field, subject, default_message):
    if not terms:
        return {"passed": True, "expected": "", "actual": "검사 조건 없음", "message": "", "details": []}

    details = []
    failures = []
    for sheet in workbook.sheets:
        text = _clean_excel_header_text(getattr(sheet, field, ""))
        check = _check_required_text_terms(
            text,
            terms,
            context,
            subject=f"{subject}({sheet.name})",
            default_message=default_message,
        )
        details.append({
            "sheet": sheet.name,
            "text": text,
            "passed": check["passed"],
            "checks": check["details"],
        })
        if not check["passed"]:
            failures.append({"sheet": sheet.name, "actual": check["actual"], "message": check["message"]})

    return {
        "passed": not failures,
        "expected": f"모든 시트 {subject} 필수어 포함",
        "actual": (f"{failures[0]['sheet']}: {failures[0]['actual']}" if failures else "정상"),
        "message": failures[0]["message"] if failures else "",
        "details": details,
    }


def _check_versioned_workbook_forbidden_print_terms(workbook_by_version, terms, context, *, field, subject, default_message):
    if not terms:
        return {"passed": True, "expected": "", "actual": "검사 조건 없음", "message": "", "details": []}

    details = []
    failures = []
    for version in sorted(workbook_by_version):
        workbook = workbook_by_version[version]
        for sheet in workbook.sheets:
            text = _clean_excel_header_text(getattr(sheet, field, ""))
            check = _check_forbidden_text_terms(
                text,
                terms,
                context,
                subject=f"{subject}(v{version}.0 {sheet.name})",
                default_message=default_message,
            )
            details.append({
                "version": version,
                "sheet": sheet.name,
                "text": text,
                "passed": check["passed"],
                "checks": check["details"],
            })
            if not check["passed"]:
                failures.append({
                    "version": version,
                    "sheet": sheet.name,
                    "actual": check["actual"],
                    "message": check["message"],
                })

    return {
        "passed": not failures,
        "expected": f"모든 파일·시트 {subject} 금지어 없음",
        "actual": (
            "오류 위치: "
            + ", ".join(f"v{failure['version']}.0 {failure['sheet']}" for failure in failures)
            if failures
            else "정상"
        ),
        "message": failures[0]["message"] if failures else "",
        "details": details,
    }


def _check_forbidden_text_terms(actual_text, terms, context, *, subject, default_message):
    actual_text = str(actual_text or "")
    term_items = _term_items(terms, context)
    details = []
    failures = []
    for item in term_items:
        term = item["text"]
        passed = term not in actual_text
        details.append({
            "term": term,
            "passed": passed,
            "message": item["message"] or default_message,
        })
        if not passed:
            failures.append(item)

    return {
        "passed": not failures,
        "expected": f"{subject}에 금지어 없음: " + ", ".join(item["text"] for item in term_items),
        "actual": (actual_text.strip() or "없음") if failures else "정상",
        "message": (failures[0]["message"] if failures else "") or default_message,
        "details": details,
    }


def _check_required_text_terms(actual_text, terms, context, *, subject, default_message):
    actual_text = str(actual_text or "")
    term_items = _term_items(terms, context)
    details = []
    failures = []
    for item in term_items:
        term = item["text"]
        passed = term in actual_text
        details.append({
            "term": term,
            "passed": passed,
            "message": item["message"] or default_message,
        })
        if not passed:
            failures.append(item)

    return {
        "passed": not failures,
        "expected": f"{subject}에 필수어 포함: " + ", ".join(item["text"] for item in term_items),
        "actual": (actual_text.strip() or "없음") if failures else "정상",
        "message": (failures[0]["message"] if failures else "") or default_message,
        "details": details,
    }


def _term_items(terms, context):
    items = []
    for term_config in terms or []:
        if isinstance(term_config, dict):
            text = _resolve_rule_value(str(term_config.get("text") or ""), context)
            message = str(term_config.get("message") or "")
        else:
            text = _resolve_rule_value(str(term_config or ""), context)
            message = ""
        if text:
            items.append({"text": text, "message": message})
    return items


def _check_checklist_cover(sheet, context, config):
    title_cell = _find_cell_with_all(sheet.rows, [context.project_number, "점검표"])
    if not title_cell:
        return {
            "passed": False,
            "expected": f"{context.project_number} 및 점검표 포함 셀",
            "actual": "일치 셀 없음",
            "message": config.get("cover_title_message") or "표지 제목이 잘못 작성됨",
        }

    period = f"{context.start_date} ~ {context.end_date}"
    # 표지 날짜는 yyyy-mm-dd, yyyy.mm.dd. 등 형식이 섞일 수 있으므로 날짜를
    # 정규화해 {시작일}/{종료일}과 같은 날짜인지로 비교한다. (다른 날짜 검사와 일관)
    date_cell = _find_cell_with_date_range(sheet.rows, context.start_date, context.end_date)
    if not date_cell:
        return {
            "passed": False,
            "expected": period,
            "actual": "일치 셀 없음",
            "message": config.get("cover_date_message") or "표지 날짜가 잘못 작성됨",
        }

    author_expected = _resolve_center_expected(config, context, "cover_author", "김진영")
    # 검토자(센터별 담당자)와 작성자({PL})는 보통 다른 행/셀에 적히므로 각각 다른 셀에서 찾는다.
    reviewer_cell = _find_cell_normalized_contains_all(sheet.rows, [author_expected])
    pl_cell = _find_cell_normalized_contains_all(sheet.rows, [context.pl]) if context.pl else None
    if not reviewer_cell or not pl_cell:
        return {
            "passed": False,
            "expected": f"{author_expected}, {context.pl}",
            "actual": (
                f"검토자({author_expected})={'있음' if reviewer_cell else '없음'} / "
                f"작성자({context.pl})={'있음' if pl_cell else '없음'}"
            ),
            "message": config.get("cover_author_message") or "표지 작성자가 잘못 작성됨",
        }

    return {
        "passed": True,
        "expected": f"{context.project_number} 점검표 / {period} / {author_expected}, {context.pl}",
        "actual": (
            f"제목='{(title_cell or {}).get('value', '')}' / "
            f"기간='{(date_cell or {}).get('value', '')}' / "
            f"검토자·작성자='{(reviewer_cell or {}).get('value', '') or (pl_cell or {}).get('value', '')}'"
        ),
        "title_cell": title_cell,
        "date_cell": date_cell,
        "reviewer_cell": reviewer_cell,
        "author_cell": pl_cell,
    }


def _check_checklist_feature_sheet(sheet):
    if not sheet:
        return {
            "passed": False,
            "expected": "기능별 점검표 시트",
            "actual": "시트 없음",
        }
    last_row = _numeric_sequence_last_row(sheet, start_row=8, column=1)
    if not last_row:
        return {
            "passed": False,
            "expected": "A8부터 연속 숫자",
            "actual": "A8 숫자 없음",
        }
    blank_cells = []
    for row in range(8, last_row + 1):
        for column in range(5, 35):
            if not _sheet_cell(sheet, row, column):
                blank_cells.append(f"{_excel_column_name(column)}{row}")
    return {
        "passed": not blank_cells,
        "expected": f"E8:AI{last_row} 모든 셀 입력",
        "actual": "빈 셀 없음" if not blank_cells else ", ".join(blank_cells[:20]),
        "last_row": last_row,
        "blank_cells": blank_cells,
    }


def _check_checklist_suitability_table(feature_sheet, suitability_sheet):
    if not (feature_sheet and suitability_sheet):
        return {
            "passed": False,
            "expected": "기능별 점검표/2. 기능적합성 시트",
            "actual": "필수 시트 없음",
        }
    last_row = _last_non_empty_row_in_column(suitability_sheet, column=3, start_row=16)
    if not last_row:
        return {
            "passed": False,
            "expected": "2. 기능적합성 C16부터 기능표",
            "actual": "비교 대상 없음",
        }
    row_count = last_row - 16 + 1
    suitability_values = _sheet_range_values(suitability_sheet, 16, 1, last_row, 3)
    feature_values = _sheet_range_values(feature_sheet, 8, 2, 8 + row_count - 1, 4)
    mismatches = _matrix_mismatches(feature_values, suitability_values, left_origin=(8, 2), right_origin=(16, 1))
    return {
        "passed": not mismatches,
        "expected": f"기능별 점검표 B8:D{8 + row_count - 1} = 2.기능적합성 표",
        "actual": f"2. 기능적합성 A16:C{last_row} 일치" if not mismatches else _stringify_check_value(mismatches[0]),
        "last_row": last_row,
        "mismatches": mismatches,
    }


def _check_checklist_suitability_results(sheet):
    if not sheet:
        return {
            "passed": False,
            "expected": "2. 기능적합성 시트",
            "actual": "시트 없음",
        }
    last_row = _last_non_empty_row_in_column(sheet, column=3, start_row=16)
    if not last_row:
        return {
            "passed": False,
            "expected": "D16부터 결과값",
            "actual": "비교 대상 없음",
        }
    blank_cells = [
        f"D{row}"
        for row in range(16, last_row + 1)
        if not _sheet_cell(sheet, row, 4)
    ]
    return {
        "passed": not blank_cells,
        "expected": f"D16:D{last_row} 모든 셀 입력",
        "actual": "빈 셀 없음" if not blank_cells else ", ".join(blank_cells[:20]),
        "blank_cells": blank_cells,
    }


def _check_checklist_reliability(sheet, context):
    if not sheet:
        return {
            "passed": False,
            "expected": "6. 신뢰성 시트",
            "actual": "시트 없음",
            "message": "신뢰성 시트의 WD 잘못 작성함",
        }

    wd_actual = _sheet_cell(sheet, 5, 3)
    if not _same_excel_text(wd_actual, context.wd):
        return {
            "passed": False,
            "expected": f"C5={context.wd}",
            "actual": f"C5={wd_actual}",
            "message": "신뢰성 시트의 WD 잘못 작성함",
        }

    high_expected = _variable_to_text(_context_variable(context, "H"))
    before_expected = _variable_to_text(_context_variable(context, "R"))
    high_actual = _sheet_cell(sheet, 11, 3)
    before_actual = _sheet_cell(sheet, 11, 5)
    if not (_same_excel_text(high_actual, high_expected) and _same_excel_text(before_actual, before_expected)):
        return {
            "passed": False,
            "expected": f"C11={high_expected}, E11={before_expected}",
            "actual": f"C11={high_actual}, E11={before_actual}",
            "message": "신뢰성 시트의 결함 개수가 잘못 작성됨",
        }

    return {
        "passed": True,
        "expected": f"C5={context.wd}, C11={high_expected}, E11={before_expected}",
        "actual": f"C5={wd_actual}, C11={high_actual}, E11={before_actual}",
    }


def _checklist_score_values(sheet):
    if not sheet:
        return []
    return [_sheet_cell(sheet, row, 4) for row in range(7, 91)]


def _top_rows_texts(sheet, *, limit):
    values = []
    for row in sheet.rows[:limit]:
        text = " ".join(value for value in row if value)
        if text:
            values.append(text)
    return values


def _clean_excel_header_text(value):
    """Excel 인쇄 머리글/바닥글의 필드 코드(&L/&C/&R, &P, &N, &"font", &12, &Kxxxxxx 등)만 제거한다.

    이전에는 `&[A-Za-z0-9]+` 로 지웠는데, 이 패턴이 탐욕적으로 매치돼서
    `&RTTA`(오른쪽 섹션 + "TTA" 텍스트)처럼 필드 코드 바로 뒤에 공백 없이
    실제 텍스트가 붙으면(엑셀에서 흔한 표기) "TTA"까지 통째로 지워버렸다.
    필드 코드는 실제로 &+한 글자(섹션/페이지번호 등) 이거나 &+숫자(글꼴 크기),
    &K+16진수 6자리(글꼴 색) 형태뿐이므로, 그 범위로만 좁혀서 뒤따르는 실제
    텍스트는 지우지 않는다.
    """
    text = str(value or "")
    text = re.sub(r"&\"[^\"]+\"", "", text)
    text = re.sub(r"&K[0-9A-Fa-f]{6}", "", text)
    text = re.sub(r"&\d+", "", text)
    text = re.sub(r"&[LCRPNDTFAZGBIUSXYlcrpndtfazgbiusxy]", "", text)
    return _normalize_spaces(text)


def _find_cell_with_all(rows, keywords):
    needles = [str(keyword or "") for keyword in keywords if str(keyword or "")]
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            if all(needle in value for needle in needles):
                return {"row": row_index, "column": col_index, "value": value}
    return None


def _find_cell_normalized_contains_all(rows, keywords):
    needles = [_normalize_no_space(keyword) for keyword in keywords if str(keyword or "")]
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            normalized = _normalize_no_space(value)
            if all(needle in normalized for needle in needles):
                return {"row": row_index, "column": col_index, "value": value}
    return None


def _numeric_sequence_last_row(sheet, *, start_row, column):
    # start_row부터 1,2,3...로 이어지는 연속 숫자의 마지막 행을 찾는다.
    # 번호가 끝난 뒤 합계/요약 같은 비숫자 행이 와도 마지막 숫자 행을 반환한다.
    # (start_row 셀이 "1"이 아니면 0을 반환해 호출부가 실패로 처리한다.)
    expected = 1
    last_row = 0
    row = start_row
    while True:
        value = _sheet_cell(sheet, row, column)
        if value != str(expected):
            break
        last_row = row
        expected += 1
        row += 1
    return last_row


def _last_non_empty_row_in_column(sheet, *, column, start_row=1):
    last_row = 0
    for row in range(start_row, len(sheet.rows) + 1):
        if _sheet_cell(sheet, row, column):
            last_row = row
    return last_row


def _sheet_cell(sheet, row, column):
    row_index = row - 1
    col_index = column - 1
    if not sheet or row_index < 0 or col_index < 0 or row_index >= len(sheet.rows):
        return ""
    row_values = sheet.rows[row_index]
    if col_index >= len(row_values):
        return ""
    return row_values[col_index]


def _sheet_range_values(sheet, start_row, start_col, end_row, end_col):
    return [
        [_sheet_cell(sheet, row, col) for col in range(start_col, end_col + 1)]
        for row in range(start_row, end_row + 1)
    ]


def _matrix_mismatches(left, right, *, left_origin, right_origin):
    mismatches = []
    row_count = max(len(left), len(right))
    for row_index in range(row_count):
        left_row = left[row_index] if row_index < len(left) else []
        right_row = right[row_index] if row_index < len(right) else []
        col_count = max(len(left_row), len(right_row))
        for col_index in range(col_count):
            left_value = left_row[col_index] if col_index < len(left_row) else ""
            right_value = right_row[col_index] if col_index < len(right_row) else ""
            if left_value != right_value:
                mismatches.append({
                    "left_cell": f"{_excel_column_name(left_origin[1] + col_index)}{left_origin[0] + row_index}",
                    "right_cell": f"{_excel_column_name(right_origin[1] + col_index)}{right_origin[0] + row_index}",
                    "left": left_value,
                    "right": right_value,
                })
    return mismatches


def _same_excel_text(left, right):
    return _normalize_spaces(left) == _normalize_spaces(right)


def _normalize_no_space(value):
    return re.sub(r"\s+", "", str(value or ""))


def _evaluate_quality_inspection_table_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _extension_matches(file_info.extension, [".xlsx", ".xls"])
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }

    if len(matched) != 1:
        return _quality_table_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="품질검사표 Excel 파일 1개",
            actual=f"품질검사표 Excel 파일 {len(matched)}개",
            message=_artifact_failure_message(
                rule,
                config,
                verify_result,
                matched=matched,
                selected_folder=selected_folder,
                name_keywords=name_keywords,
                exact_count=1,
            ),
        )

    file_info = matched[0]
    try:
        workbook = _read_excel_workbook(file_info)
    except DownloadReviewInspectionError as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="품질검사표 Excel 파일 파싱 가능",
            actual=str(exc),
            message=config.get("parse_error_message") or "품질검사표 파일을 읽을 수 없습니다",
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=raw_detail,
        )

    # 첫 실패에서 멈추지 않고 모든 세부 점검을 수행한다.
    sheet = workbook.sheets[0]
    checks = []  # {expected, actual, passed, message}

    # 1) 시트명
    expected_sheet_name = _resolve_rule_value(config.get("sheet_name") or "{프로젝트번호} 품질검사표", context)
    actual_sheet_names = [s.name for s in workbook.sheets]
    raw_detail["sheet_names"] = actual_sheet_names
    checks.append({
        "expected": f"[시트명] 시트 1개: {expected_sheet_name}",
        "actual": ", ".join(actual_sheet_names) or "시트 없음",
        "passed": actual_sheet_names == [expected_sheet_name],
        "message": config.get("sheet_message") or "품질검사표 시트명 확인 필요",
    })

    # 2) 품질부특성 측정값 개수 (E4:E85)
    quality_values_raw = [_sheet_cell(sheet, row, 5) for row in range(4, 86)]
    quality_values = [value for value in quality_values_raw if value]
    raw_detail["quality_sub_characteristic_values"] = {
        "source_range": "E4:E85",
        "raw_value_count": len(quality_values),
        "raw_values": quality_values,
    }
    expected_quality_count = int(config.get("quality_value_count") or 33)
    checks.append({
        "expected": f"[측정값 개수] E4:E85 {expected_quality_count}개",
        "actual": f"E4:E85 {len(quality_values)}개",
        "passed": len(quality_values) == expected_quality_count,
        "message": config.get("quality_value_message") or "품질검사표의 품질부특성 측정값 확인 필요",
    })

    # 품질부특성측정값 변수는 이후 규칙(품질평가보고서)에서 쓰므로 항상 산출한다.
    excluded_indices = _quality_value_excluded_indices(config)
    rotated_values = _quality_sub_characteristic_output_values(quality_values, excluded_indices)
    raw_detail["quality_sub_characteristic_values"]["excluded_source_indices"] = sorted(excluded_indices)
    raw_detail["quality_sub_characteristic_values"]["output_value_count"] = len(rotated_values)
    raw_detail["quality_sub_characteristic_values"]["rotated_values"] = rotated_values
    raw_detail["variables"] = {"품질부특성측정값": rotated_values}

    # 3) 점검표 점수표(11번 산출 변수)와 비교
    expected_scores = _context_variable(context, "측정항목별점수표")
    if not isinstance(expected_scores, list):
        checks.append({
            "expected": "[점수표 비교] 11번 점검표 D7:D90 = 품질검사표 D4:D87",
            "actual": "{측정항목별점수표} 없음(점검표가 먼저 통과해야 비교 가능)",
            "passed": False,
            "message": config.get("score_message") or "측정항목별 점수표가 점검표와 상이함",
        })
    else:
        actual_scores = [_sheet_cell(sheet, row, 4) for row in range(4, 88)]
        mismatches = _list_mismatches(expected_scores, actual_scores, start_index=4)
        total_score_count = max(len(expected_scores), len(actual_scores))
        raw_detail["score_compare"] = {
            "expected_count": len(expected_scores),
            "actual_count": len(actual_scores),
            "total_count": total_score_count,
            "mismatch_count": len(mismatches),
            "mismatches": mismatches,
        }
        mismatch_message = (
            f"총 {total_score_count}개 중 {len(mismatches)}개 값이 다름"
            if mismatches else "일치"
        )
        checks.append({
            "expected": "[점수표 비교] 11번 점검표 D7:D90 = 품질검사표 D4:D87",
            "actual": mismatch_message,
            "passed": not mismatches,
            "message": (
                f"점검표와 품질검사표의 품질부특성 값이 총 {total_score_count}개의 값 중에 {len(mismatches)}개의 값이 다름"
                if mismatches else (config.get("score_message") or "")
            ),
        })

    # 4) 바닥글 금지어/필수어 (설정된 경우에만)
    footer_forbidden_check = _check_workbook_forbidden_print_terms(
        workbook,
        config.get("forbidden_footer_terms") or [],
        context,
        field="footer_text",
        subject="품질검사표 바닥글",
        default_message="품질검사표 바닥글에 잘못된 단어가 작성됨",
    )
    forbidden_footer_words = _print_terms_text(config.get("forbidden_footer_terms")) or "(없음)"
    required_footer_words = _print_terms_text(config.get("required_footer_terms")) or "(없음)"
    if footer_forbidden_check.get("details"):
        checks.append({
            "expected": f"[바닥글 금지어] 모든 시트 바닥글 금지어 미포함: {forbidden_footer_words}",
            "actual": _stringify_check_value(footer_forbidden_check.get("actual", "")),
            "passed": bool(footer_forbidden_check.get("passed")),
            "message": footer_forbidden_check.get("message") or "품질검사표 바닥글에 잘못된 단어가 작성됨",
        })
    footer_required_check = _check_workbook_required_print_terms(
        workbook,
        config.get("required_footer_terms") or [],
        context,
        field="footer_text",
        subject="품질검사표 바닥글",
        default_message="품질검사표 바닥글에 필요한 단어가 누락됨",
    )
    if footer_required_check.get("details"):
        checks.append({
            "expected": f"[바닥글 필수어] 모든 시트 바닥글 필수어 포함: {required_footer_words}",
            "actual": _stringify_check_value(footer_required_check.get("actual", "")),
            "passed": bool(footer_required_check.get("passed")),
            "message": footer_required_check.get("message") or "품질검사표 바닥글에 필요한 단어가 누락됨",
        })

    raw_detail["sub_checks"] = [
        {"expected": c["expected"], "actual": c["actual"], "passed": c["passed"],
         "message": c.get("message", "")} for c in checks
    ]
    all_passed = all(c["passed"] for c in checks)
    first_fail = next((c for c in checks if not c["passed"]), None)
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS if all_passed else DownloadReviewRuleStatus.FAIL,
        expected=f"{expected_sheet_name} 단일 시트 / 측정값 {expected_quality_count}개 / 점수표 일치",
        actual=f"{file_info.name} / 품질부특성측정값 {len(rotated_values)}개",
        message=(config.get("pass_message") or "품질검사표를 확인했습니다.") if all_passed else (_append_current_value(first_fail["message"], first_fail.get("actual")) if first_fail else "품질검사표 확인 필요"),
        file_path=_representative_path(matched, project.project_number),
        file_name=file_info.name,
        raw_detail=raw_detail,
    )


def _quality_table_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _quality_value_excluded_indices(config):
    raw_indices = config.get("quality_value_excluded_indices")
    if raw_indices is None:
        raw_indices = [27]
    indices = set()
    for raw_index in raw_indices:
        try:
            index = int(raw_index)
        except (TypeError, ValueError):
            continue
        if index > 0:
            indices.add(index)
    return indices


def _quality_sub_characteristic_output_values(values, excluded_indices):
    ordered_indices = [
        *range(4, len(values) + 1),
        *range(1, min(3, len(values)) + 1),
    ]
    return [
        values[index - 1]
        for index in ordered_indices
        if index not in excluded_indices
    ]


def _list_mismatches(expected_values, actual_values, *, start_index):
    mismatches = []
    max_length = max(len(expected_values), len(actual_values))
    for index in range(max_length):
        expected = _variable_to_text(expected_values[index]) if index < len(expected_values) else ""
        actual = _variable_to_text(actual_values[index]) if index < len(actual_values) else ""
        if not _same_rule_value(expected, actual):
            mismatches.append({
                "row": start_index + index,
                "expected": expected,
                "actual": actual,
            })
    return mismatches


def _same_rule_value(left, right):
    left_decimal = _decimal_text(left)
    right_decimal = _decimal_text(right)
    if left_decimal is not None and right_decimal is not None:
        return left_decimal == right_decimal
    return left == right


def _decimal_text(value):
    text = _normalize_spaces(value)
    if not text:
        return None
    normalized = text.replace(",", "")
    if not re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)", normalized):
        return None
    try:
        decimal_value = Decimal(normalized)
    except InvalidOperation:
        return None
    return decimal_value if decimal_value.is_finite() else None


def _evaluate_quality_evaluation_report_check(rule, sequence, project, context, verify_result):
    config = rule.config_json or {}
    name_keywords = _resolved_keywords(config.get("filename_keywords") or [], context)
    files, selected_folder = _files_in_configured_folder(rule, verify_result)
    matched = [
        file_info
        for file_info in files
        if _name_contains_all(file_info.name, name_keywords)
        and _is_word_file(file_info)
    ]
    raw_detail = {
        "selected_folder": selected_folder,
        "filename_keywords": name_keywords,
        "matched_file_count": len(matched),
        "matched_files": [_display_path(file_info.path, project.project_number) for file_info in matched[:20]],
    }

    if len(matched) != 1:
        # 품질평가보고서 파일 자체가 없어도(예: 제품명변경 등 간소화 심사에서
        # '인증제품 변경 검토 보고서'처럼 다른 이름의 보고서로 대체 제출되는
        # 경우), 같은 폴더에 '보고서'가 포함된 문서가 있으면 그 파일을 대체
        # 산출물로 인정해 적합 처리한다.
        if not matched:
            alt_keyword = str(config.get("alternate_report_keyword") or "보고서")
            alt_candidates = [
                file_info
                for file_info in files
                if alt_keyword in file_info.name and _is_word_file(file_info)
            ]
            if alt_candidates:
                alt_file = alt_candidates[0]
                raw_detail["alternate_report_file"] = _display_path(alt_file.path, project.project_number)
                return RuleEvaluation(
                    rule=rule,
                    sequence=sequence,
                    status=DownloadReviewRuleStatus.PASS,
                    expected="품질평가보고서 Word 파일 1개",
                    actual=alt_file.name,
                    message=f"품질평가보고서 대신 {alt_file.name}이 존재함",
                    file_path=_representative_path([alt_file], project.project_number),
                    file_name=alt_file.name,
                    raw_detail=raw_detail,
                )
        return _quality_report_failure(
            rule,
            sequence,
            matched or files,
            project,
            raw_detail,
            expected="품질평가보고서 Word 파일 1개",
            actual=f"품질평가보고서 Word 파일 {len(matched)}개",
            message=_artifact_failure_message(
                rule,
                config,
                verify_result,
                matched=matched,
                selected_folder=selected_folder,
                name_keywords=name_keywords,
                exact_count=config.get("exact_count"),
            ),
        )

    file_info = matched[0]
    try:
        text = _docx_all_text(file_info)
        tables = _docx_tables(file_info)
    except Exception as exc:
        return RuleEvaluation(
            rule=rule,
            sequence=sequence,
            status=DownloadReviewRuleStatus.ERROR,
            expected="품질평가보고서 Word 파일 파싱 가능",
            actual=str(exc),
            message=config.get("parse_error_message") or "품질평가보고서 파일을 읽을 수 없습니다",
            file_path=_representative_path(matched, project.project_number),
            file_name=file_info.name,
            raw_detail=raw_detail,
        )

    # 첫 실패에서 멈추지 않고 모든 세부 점검을 수행한 뒤 항목별 적합/부적합을 표시한다.
    sub_checks = []  # {expected, actual, passed, message}

    # 0) 프로젝트번호 횟수 — 공백으로 쪼개진 'TTA-26- 00492'도 세도록 공백 제거 후 카운트
    project_count = _normalize_no_space(text).count(_normalize_no_space(context.project_number))
    expected_count = int(config.get("project_number_count") or 6)
    raw_detail["project_number_count"] = project_count
    sub_checks.append({
        "expected": f"[프로젝트번호] {context.project_number} {expected_count}회",
        "actual": f"{context.project_number} {project_count}회",
        "passed": project_count == expected_count,
        "message": config.get("project_number_message") or "프로젝트 번호 확인 필요",
    })

    # 1) 서명
    signature_check = _quality_report_signature_check(text, config)
    raw_detail["signature_check"] = signature_check
    sub_checks.append({
        "expected": f"[서명] {_stringify_check_value(signature_check.get('expected', ''))}",
        "actual": _stringify_check_value(signature_check.get("actual", "")),
        "passed": bool(signature_check.get("passed")),
        "message": config.get("signature_message") or "서명란 이름 확인 필요",
    })

    # 2) 회사명
    table_rows = [row for table in tables for row in table]
    company_value = _find_next_cell_by_label(table_rows, str(config.get("company_label") or "회사(기관)명"))
    raw_detail["company_check"] = {"expected": context.company, "actual": company_value}
    sub_checks.append({
        "expected": f"[회사명] {context.company}",
        "actual": company_value or "회사명 값 없음",
        "passed": _same_excel_text(company_value, context.company),
        "message": config.get("company_message") or "1. 신청 회사 현황 표 값 확인 필요",
    })

    # 3) 날짜 (신청/계약/인증위) — 형식 무관 비교
    date_checks = [
        ("request_date", "신청일자", context.request_date, config.get("request_date_message") or "신청일자가 잘못 작성됨"),
        ("contract_date", "계약일자", context.contract_date, config.get("contract_date_message") or "계약일자가 잘못 작성됨"),
        ("committee_date", "품질인증심의위원회", context.certification_committee_date, config.get("committee_date_message") or "인증위 날짜가 잘못 작성됨"),
    ]
    raw_detail["date_checks"] = []
    for key, label, expected_date, message in date_checks:
        actual_date = _extract_labeled_korean_date(text, label)
        raw_detail["date_checks"].append({
            "key": key, "label": label, "expected": expected_date, "actual": actual_date,
        })
        sub_checks.append({
            "expected": f"[{label}] {expected_date}",
            "actual": f"{label}: {actual_date or '날짜 없음'}",
            "passed": _same_date_text(actual_date, expected_date),
            "message": message,
        })

    # 4) 제품시험평가 기간
    period = _extract_labeled_korean_period(text, "제품시험평가")
    raw_detail["period_check"] = {
        "expected_start": context.start_date,
        "expected_end": context.end_date,
        "actual_start": period[0],
        "actual_end": period[1],
    }
    sub_checks.append({
        "expected": f"[제품시험평가] {context.start_date} ~ {context.end_date}",
        "actual": f"{period[0] or '시작일 없음'} ~ {period[1] or '종료일 없음'}",
        "passed": bool(_same_date_text(period[0], context.start_date) and _same_date_text(period[1], context.end_date)),
        "message": config.get("period_message") or "시험기간이 잘못 작성됨",
    })

    # 5) 품질부특성 측정값
    quality_table = _docx_last_table_with_first_cell(
        tables,
        config.get("quality_table_first_cell_keyword") or "품질특성",
    )
    quality_check = _quality_report_table_check(quality_table, context)
    raw_detail["quality_value_check"] = quality_check
    sub_checks.append({
        "expected": f"[품질부특성] {_stringify_check_value(quality_check.get('expected', ''))}",
        "actual": _stringify_check_value(quality_check.get("actual", "")),
        "passed": bool(quality_check.get("passed")),
        "message": (
            (config.get("na_message") or "NA 해당사항 없음 작성 오류")
            if quality_check.get("na_error")
            else config.get("quality_value_message") or "품질검사표의 품질부특성 측정값과 상이함"
        ),
    })

    raw_detail["sub_checks"] = [
        {"expected": item["expected"], "actual": item["actual"], "passed": item["passed"],
         "message": item.get("message", "")}
        for item in sub_checks
    ]
    all_passed = all(item["passed"] for item in sub_checks)
    first_fail = next((item for item in sub_checks if not item["passed"]), None)
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.PASS if all_passed else DownloadReviewRuleStatus.FAIL,
        expected=" / ".join(item["expected"] for item in sub_checks),
        actual=" / ".join(item["actual"] for item in sub_checks),
        message=(config.get("pass_message") or "품질평가보고서를 확인했습니다.") if all_passed else (_append_current_value(first_fail["message"], first_fail.get("actual")) if first_fail else "품질평가보고서 확인 필요"),
        file_path=_representative_path(matched, project.project_number),
        file_name=file_info.name,
        raw_detail=raw_detail,
    )


def _quality_report_failure(rule, sequence, matched, project, raw_detail, *, expected, actual, message):
    return RuleEvaluation(
        rule=rule,
        sequence=sequence,
        status=DownloadReviewRuleStatus.FAIL,
        expected=expected,
        actual=actual,
        message=message,
        file_path=_representative_path(matched, project.project_number),
        file_name=_representative_name(matched),
        raw_detail=raw_detail,
    )


def _quality_report_signature_check(text, config):
    primary = str(config.get("primary_signer") or "성  명 : 김  성  희")
    secondary = str(config.get("secondary_signer") or "정  성  룡     (서명)")
    normalized_text = _normalize_no_space(text)
    primary_ok = _normalize_no_space(primary) in normalized_text
    secondary_ok = _normalize_no_space(secondary) in normalized_text
    return {
        "passed": primary_ok and secondary_ok,
        "expected": f"{primary} / {secondary}",
        "actual": f"primary={'있음' if primary_ok else '없음'}, secondary={'있음' if secondary_ok else '없음'}",
    }


# 포맷/연도자리수 무관 날짜 패턴: 2026년 6월 13일 / 2026.06.13 / 26.01.27 / 26/01/27 / 2026-06-13
_FLEX_DATE_PATTERN = r"(\d{2,4})\s*(?:년|[.\-/])\s*(\d{1,2})\s*(?:월|[.\-/])\s*(\d{1,2})\s*일?\.?"


def _extract_labeled_korean_date(text, label):
    pattern = re.compile(
        re.escape(label)
        + r"\s*[:：]?\s*"
        + _FLEX_DATE_PATTERN
    )
    match = pattern.search(str(text or ""))
    if not match:
        return ""
    year, month, day = match.groups()
    return f"{_full_year(year):04d}.{int(month):02d}.{int(day):02d}."


def _extract_labeled_korean_period(text, label):
    pattern = re.compile(
        re.escape(label)
        + r"\s*[:：]?\s*"
        + _FLEX_DATE_PATTERN
        + r"\s*[~]\s*"
        + _FLEX_DATE_PATTERN
    )
    match = pattern.search(str(text or ""))
    if not match:
        return "", ""
    sy, sm, sd, ey, em, ed = match.groups()
    return (
        f"{_full_year(sy):04d}.{int(sm):02d}.{int(sd):02d}.",
        f"{_full_year(ey):04d}.{int(em):02d}.{int(ed):02d}.",
    )


def _docx_last_table_with_first_cell(tables, keyword):
    keyword_text = str(keyword or "")
    for table in reversed(tables or []):
        first_cell = _table_cell(table, 1, 1)
        if keyword_text and keyword_text in first_cell:
            return table
    return []


def _quality_report_table_check(table, context):
    expected_values = _context_variable(context, "품질부특성측정값")
    if not isinstance(expected_values, list):
        return {
            "passed": False,
            "expected": "16번 품질검사표 산출 변수 {품질부특성측정값}",
            "actual": "{품질부특성측정값} 없음",
            "actual_values": [],
        }
    if not table:
        return {
            "passed": False,
            "expected": "1행 1열에 품질특성 포함 표",
            "actual": "표 없음",
            "actual_values": [],
        }
    header_value = _table_cell(table, 1, 3)
    if "평가결과" not in header_value:
        return {
            "passed": False,
            "expected": "1행 3열 평가결과",
            "actual": header_value or "1행 3열 값 없음",
            "actual_values": [],
        }
    actual_values = _trim_trailing_empty_values([
        _table_cell(table, row, 3)
        for row in range(2, len(table) + 1)
    ])
    mismatches = _list_mismatches(expected_values, actual_values, start_index=2)
    if mismatches:
        total_count = max(len(expected_values), len(actual_values))
        mismatch_message = (
            f"품질검사표와 품질평가보고서의 품질부특성 값이 총 {total_count}개의 값 중에 "
            f"{len(mismatches)}개의 값이 다름"
        )
        return {
            "passed": False,
            "expected": "품질검사표 품질부특성 측정값과 동일",
            "actual": mismatch_message,
            "actual_values": actual_values,
            "total_count": total_count,
            "mismatch_count": len(mismatches),
            "mismatches": mismatches,
        }
    na_errors = []
    for offset, value in enumerate(actual_values, start=2):
        if _is_na_value(value) and _table_cell(table, offset, 4) != "해당사항 없음":
            na_errors.append({
                "row": offset,
                "value": value,
                "right_cell": _table_cell(table, offset, 4),
            })
    if na_errors:
        return {
            "passed": False,
            "expected": "NA/N/A 오른쪽 칸 해당사항 없음",
            "actual": str(na_errors[0]),
            "actual_values": actual_values,
            "na_errors": na_errors,
            "na_error": True,
        }
    return {
        "passed": True,
        "expected": f"품질부특성 측정값 {len(expected_values)}개",
        "actual": f"품질부특성 측정값 {len(actual_values)}개",
        "actual_values": actual_values,
    }


def _table_cell(table, row, column):
    row_index = row - 1
    col_index = column - 1
    if row_index < 0 or col_index < 0 or row_index >= len(table):
        return ""
    row_values = table[row_index]
    if col_index >= len(row_values):
        return ""
    return row_values[col_index]


def _is_na_value(value):
    normalized = str(value or "").strip().upper()
    return normalized in {"NA", "N/A"}


def _trim_trailing_empty_values(values):
    trimmed = list(values)
    while trimmed and not trimmed[-1]:
        trimmed.pop()
    return trimmed


def _matching_files(rule, verify_result, *, ignore_target_file_type=False):
    files = _inspection_files(verify_result)
    pattern = (rule.target_file_pattern or "").strip()
    file_type = (rule.target_file_type or "any").strip().lower()

    if pattern:
        files = [file_info for file_info in files if fnmatch.fnmatch(file_info.name, pattern)]

    if not ignore_target_file_type and file_type and file_type != "any":
        extension = _extension_from_file_type(file_type)
        if extension:
            # docx 대상은 구형 .doc도 통과시킨다(이후 .docx로 변환해 점검).
            accepted = {extension}
            if extension == ".docx":
                accepted.add(".doc")
            files = [file_info for file_info in files if file_info.extension.lower() in accepted]

    return files


def _is_ignorable_file(name):
    """점검 대상에서 제외할 파일.
    - MS Office 임시/잠금 파일(~$ 접두사): 문서가 열린 채 압축되면 zip에 포함된다.
    - 맥/숨김 메타파일(.DS_Store, ~ 접두사 백업).
    """
    base = str(name or "").strip()
    if not base:
        return True
    if base.startswith("~$") or base.startswith("~"):
        return True
    if base.lower() in ("thumbs.db", ".ds_store"):
        return True
    return False


def _inspection_files(verify_result):
    cached_files = getattr(verify_result, "_inspection_files_cache", None)
    if cached_files is not None:
        return cached_files

    files = list(verify_result.files or [])
    expanded_files = list(files)
    zip_errors = []
    pending_files = list(files)
    expanded_zip_paths = set()
    while pending_files:
        file_info = pending_files.pop(0)
        if file_info.extension.lower() != ".zip":
            continue
        raw_path = str(file_info.path or "")
        if raw_path in expanded_zip_paths:
            continue
        expanded_zip_paths.add(raw_path)
        try:
            zip_entries = _zip_entry_files(file_info)
        except (BadZipFile, KeyError, OSError, RuntimeError) as exc:
            zip_errors.append({
                "file_name": file_info.name,
                "path": raw_path,
                "message": str(exc),
            })
            continue
        expanded_files.extend(zip_entries)
        pending_files.extend(zip_entries)
    # ~$ 임시 파일 등 점검에 방해되는 파일 제외
    inspection_files = [
        file_info for file_info in expanded_files if not _is_ignorable_file(file_info.name)
    ]
    setattr(verify_result, "_inspection_files_cache", inspection_files)
    setattr(verify_result, "_inspection_zip_errors", zip_errors)
    return inspection_files


def _zip_entry_files(zip_file_info):
    entries = []
    zip_bytes = _read_path_bytes(str(zip_file_info.path or ""))
    with ZipFile(BytesIO(zip_bytes)) as zip_file:
        for entry in zip_file.infolist():
            if entry.is_dir():
                continue
            inner_path = entry.filename.replace("\\", "/")
            inner_name = PurePosixPath(inner_path).name
            entries.append(
                FileInfo(
                    name=inner_name,
                    path=f"{zip_file_info.path}::{inner_path}",
                    size=entry.file_size,
                    extension=PurePosixPath(inner_name).suffix.lower(),
                    modified_at=datetime(*entry.date_time),
                )
            )
    return entries


# 구형 .doc(OLE) → .docx 변환 결과 캐시(run_download_inspection 시작 시 비움).
_DOC_CONVERT_CACHE = {}


def _soffice_executable():
    """LibreOffice soffice 실행 파일 경로를 찾는다(환경변수/PATH/기본 설치 경로)."""
    import shutil as _shutil
    candidates = []
    env_path = os.environ.get("AGENT_SOFFICE_PATH", "")
    if env_path:
        candidates.append(env_path)
    for name in ("soffice", "soffice.exe"):
        found = _shutil.which(name)
        if found:
            candidates.append(found)
    candidates += [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    return ""


def _convert_doc_to_docx_bytes(doc_bytes):
    """구형 .doc 바이트를 .docx 로 변환해 바이트를 반환한다.

    우선순위: MS Word(win32com) → 없거나 실패 시 LibreOffice(soffice) 폴백.
    """
    converted = _convert_doc_via_msword(doc_bytes)
    if converted is not None:
        return converted

    soffice = _soffice_executable()
    if soffice:
        return _convert_doc_via_soffice(doc_bytes, soffice)

    raise DownloadReviewInspectionError(
        ".doc 파일은 MS Word 또는 LibreOffice 가 있어야 변환·점검할 수 있습니다. "
        "MS Word 가 설치된 환경에서 실행하거나 LibreOffice 설치(또는 AGENT_SOFFICE_PATH 지정) "
        "하세요. (또는 .docx 로 재제출 요청)"
    )


def _convert_doc_via_soffice(doc_bytes, soffice):
    """LibreOffice headless 로 .doc → .docx 변환."""
    import subprocess
    import tempfile
    import shutil as _shutil

    tmpdir = tempfile.mkdtemp(prefix="docconv_")
    try:
        doc_path = os.path.join(tmpdir, "input.doc")
        with open(doc_path, "wb") as handle:
            handle.write(doc_bytes)
        try:
            proc = subprocess.run(
                [soffice, "--headless", "--convert-to", "docx", "--outdir", tmpdir, doc_path],
                capture_output=True,
                timeout=180,
            )
        except (OSError, subprocess.SubprocessError) as exc:
            raise DownloadReviewInspectionError(f".doc 변환 실행 실패: {exc}") from exc
        out_path = os.path.join(tmpdir, "input.docx")
        if not os.path.exists(out_path):
            stderr = (proc.stderr or b"").decode("utf-8", "replace")[:200]
            raise DownloadReviewInspectionError(
                f".doc → .docx 변환 실패: {stderr} "
                "(LibreOffice 변환 시 프로필이 격리되지 않아 발생할 수 있는 문제입니다. "
                "동일 파일에서 반복되면 서버의 LibreOffice 사용자 프로필 초기화가 필요합니다.)"
            )
        with open(out_path, "rb") as handle:
            return handle.read()
    finally:
        _shutil.rmtree(tmpdir, ignore_errors=True)


def _convert_doc_via_msword(doc_bytes):
    """MS Word(win32com) 로 .doc → .docx 변환. Word 미설치/미지원 시 None 반환.

    LibreOffice 가 없는 로컬 앱(Office 설치 PC)에서 .doc 를 처리하기 위한 폴백.
    Windows + pywin32 + MS Word 가 있어야 동작한다.
    """
    try:
        import pythoncom
        import win32com.client as win32
    except Exception:
        return None

    import os as _os
    import tempfile
    import shutil as _shutil

    tmpdir = tempfile.mkdtemp(prefix="docword_")
    doc_path = _os.path.join(tmpdir, "input.doc")
    out_path = _os.path.join(tmpdir, "input.docx")
    word = None
    coinit = False
    try:
        with open(doc_path, "wb") as handle:
            handle.write(doc_bytes)
        try:
            pythoncom.CoInitialize()
            coinit = True
        except Exception:
            coinit = False
        word = win32.DispatchEx("Word.Application")
        try:
            word.Visible = False
            word.DisplayAlerts = 0
        except Exception:
            pass
        document = word.Documents.Open(doc_path, ReadOnly=True)
        document.SaveAs2(out_path, FileFormat=16)  # 16 = wdFormatXMLDocument(.docx)
        document.Close(False)
        if _os.path.exists(out_path):
            with open(out_path, "rb") as handle:
                return handle.read()
        return None
    except Exception:
        return None
    finally:
        try:
            if word is not None:
                word.Quit()
        except Exception:
            pass
        if coinit:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
        _shutil.rmtree(tmpdir, ignore_errors=True)


def _read_file_bytes(file_info):
    raw_path = str(file_info.path or "")
    data = _read_path_bytes(raw_path)

    # 구형 .doc(OLE)는 docx 파서가 읽을 수 있도록 .docx로 변환해 반환한다.
    # (.doc는 Word 문서 점검 규칙에서만 읽히므로 다른 파서에는 영향이 없다.)
    if str(getattr(file_info, "extension", "")).lower() == ".doc" and data[:2] != b"PK":
        if raw_path not in _DOC_CONVERT_CACHE:
            _DOC_CONVERT_CACHE[raw_path] = _convert_doc_to_docx_bytes(data)
        return _DOC_CONVERT_CACHE[raw_path]
    return data


def _read_path_bytes(raw_path):
    path_parts = str(raw_path or "").split("::")
    if not path_parts or not path_parts[0]:
        raise OSError("파일 경로가 비어 있습니다.")
    data = Path(path_parts[0]).read_bytes()
    for inner_path in path_parts[1:]:
        with ZipFile(BytesIO(data)) as zip_file:
            data = zip_file.read(inner_path)
    return data


@dataclass(frozen=True)
class _ArtifactRevision:
    major: int
    minor: tuple[int, ...]
    base_stem: str
    label: str


_PREFIXED_ARTIFACT_REVISION_RE = re.compile(
    r"(?i)(?:^|[\s_\-\(\)\[\]\{\}])(?P<label>(?:v|ver|version)\s*\.?\s*(?P<number>\d+(?:[._-]\d+)*))\s*$"
)
_DOTTED_ARTIFACT_REVISION_RE = re.compile(
    r"(?:^|[\s_\-\(\)\[\]\{\}])(?P<label>(?P<number>\d+\.\d+(?:[._-]\d+)*))\s*$"
)


def _artifact_revision_info(file_info):
    """파일명 끝의 산출물 개정 버전을 파싱한다.

    ECM에서는 같은 산출물이 v1.0, v1.1처럼 뒤늦게 추가될 수 있다. 제품명 안의
    ISM3.0 같은 숫자는 버전으로 오인하지 않도록 파일명 끝 토큰만 본다.
    """
    stem = PurePosixPath(str(file_info.name or "")).stem
    for pattern in (_PREFIXED_ARTIFACT_REVISION_RE, _DOTTED_ARTIFACT_REVISION_RE):
        match = pattern.search(stem)
        if not match:
            continue
        number = match.group("number")
        parts = [part for part in re.split(r"[._-]", number) if part != ""]
        if not parts:
            continue
        try:
            major = int(parts[0])
            minor = tuple(int(part) for part in parts[1:]) or (0,)
        except ValueError:
            continue
        if major > 99:
            continue
        base_stem = stem[: match.start()].strip(" _-()[]{}")
        if not base_stem:
            continue
        return _ArtifactRevision(
            major=major,
            minor=minor,
            base_stem=base_stem,
            label=" ".join(match.group("label").split()),
        )
    return None


def _artifact_revision_identity(file_info):
    revision = _artifact_revision_info(file_info)
    stem = revision.base_stem if revision else PurePosixPath(str(file_info.name or "")).stem
    normalized_stem = unicodedata.normalize("NFKC", stem).casefold()
    normalized_stem = re.sub(r"[\s_\-\(\)\[\]\{\}.]+", "", normalized_stem)
    return (normalized_stem, _artifact_revision_extension_family(file_info))


def _artifact_revision_extension_family(file_info):
    extension = str(getattr(file_info, "extension", "") or "").lower()
    if extension in (".xlsx", ".xlsm", ".xls"):
        return "excel"
    if extension in (".docx", ".docm", ".doc"):
        return "word"
    return extension or "any"


def _artifact_revision_modified_value(file_info):
    modified_at = getattr(file_info, "modified_at", None)
    if isinstance(modified_at, datetime):
        return modified_at.timestamp()
    return 0


def _file_identity_key(file_info):
    return str(getattr(file_info, "path", "") or getattr(file_info, "name", ""))


def _expand_revision_related_files(selected_files, all_files):
    selected = list(selected_files or [])
    if not selected:
        return selected

    selected_identities = {_artifact_revision_identity(file_info) for file_info in selected}
    if not selected_identities:
        return selected

    seen = {_file_identity_key(file_info) for file_info in selected}
    expanded = list(selected)
    for file_info in all_files or []:
        key = _file_identity_key(file_info)
        if key in seen:
            continue
        if _artifact_revision_info(file_info) and _artifact_revision_identity(file_info) in selected_identities:
            expanded.append(file_info)
            seen.add(key)
    return expanded


def _latest_revision_files(files):
    indexed_files = list(enumerate(files or []))
    groups = {}
    for index, file_info in indexed_files:
        groups.setdefault(_artifact_revision_identity(file_info), []).append((index, file_info))

    selected_keys = set()
    for group in groups.values():
        versioned = [
            (index, file_info, _artifact_revision_info(file_info))
            for index, file_info in group
            if _artifact_revision_info(file_info)
        ]
        if not versioned:
            selected_keys.update(_file_identity_key(file_info) for index, file_info in group)
            continue

        latest_by_major = {}
        for index, file_info, revision in versioned:
            major = revision.major
            candidate_key = (
                revision.minor,
                _artifact_revision_modified_value(file_info),
                str(file_info.path or file_info.name),
            )
            current = latest_by_major.get(major)
            if current is None or candidate_key > current[0]:
                latest_by_major[major] = (candidate_key, index, file_info)
        selected_keys.update(
            _file_identity_key(file_info)
            for _candidate_key, _index, file_info in latest_by_major.values()
        )

    return [
        file_info
        for _index, file_info in indexed_files
        if _file_identity_key(file_info) in selected_keys
    ]


def _files_in_configured_folder(rule, verify_result, *, ignore_target_file_type=False):
    config = rule.config_json or {}
    files = _matching_files(rule, verify_result, ignore_target_file_type=ignore_target_file_type)
    # rawdata zip(예: '..._RAWDATA.zip')의 파일은 rawdata 전용 규칙에서만 사용한다.
    # 일반 규칙이 rawdata의 스크린샷 이미지 폴더를 제출물 폴더로 잘못 선택하지 않도록 제외한다.
    # (최초/최종형상RawData 규칙은 _files_in_configured_folder를 쓰지 않고 직접 rawdata를 필터링한다.)
    files = [file_info for file_info in files if not _is_rawdata_file(file_info, "rawdata")]
    selected_files, selected_folder = _select_folder_chain_files(files, config.get("folder_keyword_chain"))
    selected_files = _expand_revision_related_files(selected_files, files)
    selected_files = _latest_revision_files(selected_files)
    return selected_files, selected_folder


def _defect_report_version_label(versioned_files, version):
    file_info = versioned_files.get(version) if isinstance(versioned_files, dict) else None
    revision = _artifact_revision_info(file_info) if file_info else None
    return revision.label if revision else f"v{version}.0"


def _select_folder_chain_files(files, folder_keyword_chain_raw):
    folder_keyword_chain = [
        str(item).strip()
        for item in folder_keyword_chain_raw or []
        if str(item).strip()
    ]
    if not folder_keyword_chain:
        return files, ""

    selected_folder_segments = None
    for file_info in files:
        folder_segments = _folder_segments(file_info.path)
        matched_indices = _folder_keyword_chain_indices(folder_segments, folder_keyword_chain)
        if matched_indices:
            selected_folder_segments = folder_segments[: matched_indices[-1] + 1]
            break

    if not selected_folder_segments:
        return [], ""

    selected_files = [
        file_info
        for file_info in files
        if _folder_segments(file_info.path)[: len(selected_folder_segments)] == selected_folder_segments
    ]
    return selected_files, "/".join(selected_folder_segments)


def _zip_name_contains(file_info, keyword):
    """파일이 이름에 keyword를 포함하는 zip에서 추출됐는지 확인한다(대소문자/공백 무시)."""
    raw_path = str(file_info.path or "")
    if "::" not in raw_path:
        return False
    needle = _normalize_zip_keyword(keyword)
    if not needle:
        return False
    return any(
        needle in _normalize_zip_keyword(zip_name)
        for zip_name in _source_zip_names(raw_path)
    )


def _is_rawdata_file(file_info, keyword="rawdata"):
    """파일이 'rawdata' 산출물에 속하는지 판정한다(폴더·zip 공통).

    원래 rawdata 점검은 폴더/파일 기준이었으나, 메인 zip 밖에 별도 rawdata.zip이
    존재하는 경우를 처리하려고 zip 이름 기준(_zip_name_contains)으로만 좁혀졌었다.
    그 결과 압축을 푼 'rawdata' 폴더는 인식하지 못했다. 여기서는 경로의 어떤
    세그먼트(상위 폴더명 또는 zip 파일명)든 keyword를 포함하면 rawdata로 본다.
    - 압축 해제 폴더: .../rawdata/결함/a.png        → 폴더명 'rawdata' 매치
    - 별도 rawdata.zip: RAWDATA.zip::결함/a.png      → zip명 'rawdata' 매치
    - zip 내부 rawdata 폴더: x.zip::수행/rawdata/...  → 내부 폴더명 매치
    (대소문자/공백/_/- 무시. 마지막 세그먼트=파일명 자체는 판정에서 제외)
    """
    needle = _normalize_zip_keyword(keyword)
    if not needle:
        return False
    raw_path = str(file_info.path or "").replace("\\", "/")
    segments = [seg for seg in raw_path.replace("::", "/").split("/") if seg][:-1]
    for seg in segments:
        if seg.lower().endswith(".zip"):
            seg = seg[:-4]
        if needle in _normalize_zip_keyword(seg):
            return True
    return False


def _source_zip_names(raw_path):
    zip_names = []
    for path_part in str(raw_path or "").split("::")[:-1]:
        zip_name = PurePosixPath(path_part.replace("\\", "/")).name
        if zip_name.lower().endswith(".zip"):
            zip_names.append(zip_name)
    return zip_names


def _normalize_zip_keyword(value):
    return re.sub(r"[\s_-]+", "", str(value or "")).lower()


def _folder_segments(path):
    normalized = str(path or "").replace("\\", "/")
    if "::" in normalized:
        normalized = normalized.rsplit("::", 1)[1]
    parts = [part for part in normalized.split("/") if part]
    if not parts:
        return []
    return parts[:-1]


def _folder_keyword_chain_indices(folder_segments, keyword_chain):
    indices = []
    start = 0
    for keyword in keyword_chain:
        found = None
        for index in range(start, len(folder_segments)):
            if keyword in folder_segments[index]:
                found = index
                break
        if found is None:
            return []
        indices.append(found)
        start = found + 1
    return indices


def _first_pl_name(pl_value):
    """시험PL 값이 '우수진, 석민경'처럼 쉼표로 구분된 경우 첫 번째 이름만 사용한다."""
    text = str(pl_value or "").strip()
    if not text:
        return ""
    return text.split(",")[0].strip()


def _split_product_and_version(product_name):
    value = _normalize_spaces(product_name)
    if not value:
        return "", ""

    version_match = re.search(
        r"(?i)(?:^|\s)((?:v|ver|version)\s*\.?\s*[\w.\-]+)$",
        value,
    )
    if version_match:
        version = version_match.group(1).strip()
        product = value[: version_match.start(1)].strip()
        return product, version

    product_prefix = ""
    final_token = value
    if " " in value:
        product_prefix, final_token = value.rsplit(" ", 1)

    numeric_suffix = re.search(r"(\d+(?:[._-]\d+)*)$", final_token)
    if numeric_suffix:
        version = numeric_suffix.group(1)
        token_prefix = final_token[: numeric_suffix.start()].strip()
        product = _normalize_spaces(f"{product_prefix} {token_prefix}".strip())
        return product, version

    if product_prefix:
        return product_prefix.strip(), final_token.strip()

    return value, ""


def _project_year(project_number):
    match = re.search(r"TTA-(\d{2})-", project_number or "", re.IGNORECASE)
    if not match:
        return ""
    return f"20{match.group(1)}"


def _full_year(year_text):
    """2자리 연도(예: 26)는 2000년대(2026)로 보정한다."""
    value = int(year_text)
    return value + 2000 if value < 100 else value


def _format_dot_date(value):
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.replace("년", ".").replace("월", ".").replace("일", "")
    normalized = normalized.replace("/", ".").replace("-", ".")
    # 연도는 2자리 또는 4자리 모두 허용 (26 → 2026)
    match = re.search(r"(\d{2,4})\D+(\d{1,2})\D+(\d{1,2})", normalized)
    if not match:
        return text
    year, month, day = match.groups()
    return f"{_full_year(year):04d}.{int(month):02d}.{int(day):02d}."


def _resolved_keywords(keywords, context):
    resolved = []
    for keyword in keywords:
        value = _resolve_rule_value(str(keyword), context)
        if value:
            resolved.append(value)
    return resolved


def _resolve_rule_value(value, context):
    replacements = {
        "{project_number}": context.project_number,
        "{프로젝트번호}": context.project_number,
        "{product}": context.product,
        "{제품명}": context.product,
        "{company}": context.company,
        "{회사명}": context.company,
        "{버전}": context.version,
        "{pl}": context.pl,
        "{PL}": context.pl,
        "{wd}": context.wd,
        "{WD}": context.wd,
        "{시작일}": context.start_date,
        "{종료일}": context.end_date,
        "{연도}": context.year,
        "{신청일}": context.request_date,
        "{계약일}": context.contract_date,
        "{인증위}": context.certification_committee_date,
    }
    for key, replacement in context.derived_variables.items():
        replacements.setdefault(f"{{{key}}}", _variable_to_text(replacement))
    for placeholder, replacement in replacements.items():
        value = value.replace(placeholder, str(replacement))
    return value.strip()


def _variable_to_text(value):
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _configured_extensions(config, target_file_type):
    raw_extensions = config.get("extensions")
    if raw_extensions is None:
        extension = _extension_from_file_type(target_file_type)
        return [extension] if extension else []
    extensions = []
    for extension in raw_extensions:
        normalized = _normalize_extension(extension)
        if normalized:
            extensions.append(normalized)
    return extensions


def _extensions_label(extensions):
    normalized = {_normalize_extension(extension) for extension in extensions}
    if normalized and normalized <= set(WORD_EXTENSIONS):
        return "Word 파일"
    return ", ".join(extension for extension in extensions if extension)


def _is_word_file(file_info):
    return _extension_matches(file_info.extension, WORD_EXTENSIONS)


def _normalize_extension(extension):
    value = str(extension or "").strip().lower()
    if not value or value == "any":
        return ""
    if value.startswith("."):
        return value
    return f".{value}"


def _name_contains_all(file_name, keywords):
    # 파일명/키워드의 '-'와 '_' 차이를 무시한다(예: 'TTA_26_00492' = 'TTA-26-00492').
    name = str(file_name).replace("_", "-")
    return all(str(keyword).replace("_", "-") in name for keyword in keywords)


def _extension_matches(extension, extensions):
    if not extensions:
        return True
    ext = str(extension or "").lower()
    if ext in extensions:
        return True
    # 구형 .doc는 변환 후 점검하므로 .docx 요구 조건을 충족하는 것으로 본다.
    if ext == ".doc" and ".docx" in extensions:
        return True
    return False


def _artifact_failure_message(
    rule,
    config,
    verify_result,
    *,
    matched,
    selected_folder,
    name_keywords,
    exact_count,
):
    """required_artifact_file 규칙의 실패 사유를 세분화해 메시지를 고른다.

    실패 원인을 (폴더 없음 → 개수 초과 → 파일명 없음 → 확장자 불일치) 순서로
    판정한다. 확장자 필터 때문에 원인 파악이 왜곡되지 않도록, 진단 단계에서는
    target_file_type 확장자 필터를 무시하고 폴더/파일명 기준으로 후보를 다시 찾는다.
    각 원인별 메시지 config 키가 없으면 기존 동작(missing_message)으로 되돌아간다.
    """
    default_missing = config.get("missing_message") or "파일이 없습니다."
    folder_chain = config.get("folder_keyword_chain")

    files_any_ext, selected_folder_any = _files_in_configured_folder(
        rule, verify_result, ignore_target_file_type=True
    )

    # 1) 대상 폴더 자체를 찾지 못함
    if folder_chain and not selected_folder_any:
        return config.get("folder_missing_message") or default_missing

    # 2) 개수 초과(요구 개수보다 많이 발견)
    if exact_count is not None and len(matched) > int(exact_count):
        return config.get("multiple_message") or "대상 파일이 여러개 존재합니다."

    name_matched = [
        file_info
        for file_info in files_any_ext
        if (not name_keywords or _name_contains_all(file_info.name, name_keywords))
    ]

    # 3) 폴더는 있으나 파일명 조건을 만족하는 파일이 아예 없음
    if not name_matched:
        return default_missing

    # 4) 파일명은 맞지만 요구 확장자가 아님
    if not matched:
        return config.get("extension_mismatch_message") or default_missing

    return default_missing


def _extension_missing_label(extensions):
    """실패 메시지용 확장자 라벨. Word 계열은 'Word', 그 외는 'pdf', 'png/pptx' 형태."""
    normalized = [ext for ext in (_normalize_extension(e) for e in extensions) if ext]
    if not normalized:
        return "대상"
    if set(normalized) <= set(WORD_EXTENSIONS):
        return "Word"
    return "/".join(ext.lstrip(".") for ext in normalized)


def _document_artifact_failure_message(
    rule,
    config,
    verify_result,
    *,
    name_keywords,
    file_check,
    content_check,
):
    """document_artifact_check 실패 사유를 (폴더 → 파일명 → 확장자 → 내용) 순서로 구분한다.

    확장자 필터가 원인 파악을 왜곡하지 않도록 진단 시 target_file_type 필터를
    무시하고 폴더/파일명 기준으로 후보를 다시 찾는다.
    """
    folder_chain = config.get("folder_keyword_chain")
    files_any_ext, selected_folder_any = _files_in_configured_folder(
        rule, verify_result, ignore_target_file_type=True
    )

    # 1) 대상 폴더 자체를 찾지 못함
    if folder_chain and not selected_folder_any:
        return config.get("folder_missing_message") or config.get("missing_message") or "대상 폴더를 찾을 수 없습니다."

    # 2) 폴더는 있으나 파일명 조건을 만족하는 파일이 없음
    name_matched = [
        file_info
        for file_info in files_any_ext
        if (not name_keywords or _name_contains_all(file_info.name, name_keywords))
    ]
    if not name_matched:
        return config.get("missing_message") or "대상 파일을 찾을 수 없습니다."

    # 3) 파일명은 맞지만 필요한 확장자 파일이 없음(_evaluate_required_file_specs 가 확장자별 메시지 생성)
    if not file_check["passed"]:
        return file_check["message"] or config.get("missing_message") or "필요한 파일을 찾을 수 없습니다."

    # 4) 내용 검사 실패
    return content_check["message"] or "문서 내용이 기준과 일치하지 않습니다."


def _matched_files_actual(files):
    if not files:
        return "일치 파일 없음"
    return ", ".join(file_info.name for file_info in files[:5])


def _evaluate_required_file_specs(config, matched_files):
    specs = config.get("required_files") or []
    if not specs:
        specs = [
            {
                "extensions": config.get("extensions") or [],
                "exact_count": config.get("exact_count"),
                "min_count": config.get("min_count") or 1,
            }
        ]

    passed = True
    expected = []
    actual = []
    details = []
    message = ""
    for spec in specs:
        extensions = _configured_extensions(spec, "any")
        files = [
            file_info
            for file_info in matched_files
            if _extension_matches(file_info.extension, extensions)
        ]
        exact_count = spec.get("exact_count")
        min_count = int(spec.get("min_count") or 1)
        if exact_count is not None:
            count_expected = int(exact_count)
            spec_passed = len(files) == count_expected
            count_text = f"{count_expected}개"
        else:
            spec_passed = len(files) >= min_count
            count_text = f"{min_count}개 이상"

        extension_text = _extensions_label(extensions) if extensions else "확장자 무관"
        expected.append(f"{extension_text} {count_text}")
        actual.append(f"{extension_text} {len(files)}개")
        details.append({
            "extensions": extensions,
            "expected": count_text,
            "actual_count": len(files),
            "matched_files": [file_info.name for file_info in files[:10]],
            "passed": spec_passed,
        })
        if not spec_passed:
            passed = False
            if spec.get("message"):
                message = spec["message"]
            elif extensions:
                message = f"{_extension_missing_label(extensions)} 파일을 찾을 수 없습니다"
            else:
                message = config.get("missing_message") or "필요한 파일을 찾을 수 없습니다"

    return {
        "passed": passed,
        "expected": expected,
        "actual": actual,
        "message": message,
        "details": details,
    }


def _evaluate_content_checks(config, matched_files, context):
    passed = True
    expected = []
    actual = []
    details = []
    message = ""

    for check in config.get("content_checks") or []:
        extensions = _configured_extensions(check, "any")
        files = [
            file_info
            for file_info in matched_files
            if _extension_matches(file_info.extension, extensions)
        ]
        if not files:
            check_result = {
                "passed": False,
                "expected": _content_check_expected(check, context),
                "actual": "검사 대상 파일 없음",
                "message": check.get("missing_message") or "검사 대상 파일이 없습니다.",
                "detail": {"extensions": extensions},
            }
        else:
            check_result = _run_content_check(check, files[0], context)

        expected.append(check_result["expected"])
        actual.append(check_result["actual"])
        details.append({
            **check_result.get("detail", {}),
            "type": check.get("type"),
            "file_name": files[0].name if files else "",
            "passed": check_result["passed"],
        })
        if not check_result["passed"] and passed:
            passed = False
            message = check_result["message"]

    return {
        "passed": passed,
        "expected": expected,
        "actual": actual,
        "message": message,
        "details": details,
    }


def _run_content_check(check, file_info, context):
    check_type = str(check.get("type") or "").strip()
    if check_type == "docx_table_next_cell_equals":
        return _check_docx_table_next_cell_equals(check, file_info, context)
    if check_type == "pdf_first_page_label_value_contains":
        return _check_pdf_first_page_label_value_contains(check, file_info, context)
    if check_type == "docx_text_contains":
        return _check_docx_text_contains(check, file_info, context)
    if check_type == "docx_header_contains":
        return _check_docx_part_contains(check, file_info, context, part="header")
    if check_type == "docx_footer_contains":
        return _check_docx_part_contains(check, file_info, context, part="footer")
    if check_type == "docx_header_not_contains":
        return _check_docx_part_not_contains(check, file_info, context, part="header")
    if check_type == "docx_footer_not_contains":
        return _check_docx_part_not_contains(check, file_info, context, part="footer")
    if check_type == "docx_next_paragraph_matches":
        return _check_docx_next_paragraph_matches(check, file_info, context)
    raise DownloadReviewInspectionError(f"지원하지 않는 문서 내용 검사 유형입니다: {check_type or '(비어 있음)'}")


def _check_docx_part_not_contains(check, file_info, context, *, part):
    """docx 머리글/바닥글에 금지 문자열이 없어야 통과한다(texts 중 하나라도 있으면 실패)."""
    raw_texts = check.get("texts") or ([check.get("text")] if check.get("text") else [])
    forbidden = [
        _resolve_rule_value(str(text), context)
        for text in raw_texts
        if str(text).strip()
    ]
    if part == "header":
        actual_text = _docx_header_text(file_info)
        label = "머리글"
    else:
        actual_text = _docx_footer_text(file_info)
        label = "바닥글"
    hits = [text for text in forbidden if text in actual_text]
    return _content_result(
        check,
        not hits,
        expected=f"{label}에 금지어 미포함: {', '.join(forbidden)}",
        actual=("포함된 금지어: " + ", ".join(hits)) if hits else (actual_text or f"{label} 없음"),
        detail={"part": part, "forbidden": forbidden, "hits": hits},
    )


def _check_docx_table_next_cell_equals(check, file_info, context):
    label = str(check.get("label") or "").strip()
    expected_value = _resolve_rule_value(str(check.get("expected") or ""), context)
    rows = _docx_table_rows(file_info)
    actual_value = _find_next_cell_by_label(rows, label)
    passed = _normalize_compare(actual_value, check) == _normalize_compare(expected_value, check)
    return _content_result(
        check,
        passed,
        expected=f"{label} 오른쪽 셀 = {expected_value}",
        actual=f"{label} 오른쪽 셀 = {actual_value or '(없음)'}",
        detail={"label": label, "expected_value": expected_value, "actual_value": actual_value},
    )


def _check_pdf_first_page_label_value_contains(check, file_info, context):
    label = str(check.get("label") or "").strip()
    expected_value = _resolve_rule_value(str(check.get("expected") or ""), context)
    text = _pdf_page_text(file_info, page_index=int(check.get("page_index") or 0))
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    window_size = int(check.get("line_window") or 3)
    actual_value = ""
    for index, line in enumerate(lines):
        if label not in line:
            continue
        window = lines[index:index + window_size + 1]
        actual_value = " ".join(window)
        if expected_value in actual_value:
            break
    passed = bool(actual_value and expected_value in actual_value)
    return _content_result(
        check,
        passed,
        expected=f"PDF 1페이지 {label} 주변에 {expected_value} 포함",
        actual=actual_value or f"PDF 1페이지에서 {label} 없음",
        detail={"label": label, "expected_value": expected_value, "text_excerpt": lines[:20]},
    )


def _check_docx_text_contains(check, file_info, context):
    expected_texts = [
        _resolve_rule_value(str(text), context)
        for text in check.get("texts") or []
        if str(text).strip()
    ]
    expected_text = _resolve_rule_value(str(check.get("text") or ""), context)
    paragraphs = _docx_paragraphs(file_info)
    if expected_texts:
        matched = _find_matching_paragraph_with_all(paragraphs, expected_texts, check)
        expected = "문서에 " + ", ".join(f"'{text}'" for text in expected_texts) + " 포함"
    else:
        matched = _find_matching_paragraph(paragraphs, expected_text, check)
        expected = f"문서에 '{expected_text}' 포함"
    passed = matched is not None
    return _content_result(
        check,
        passed,
        expected=expected,
        actual=matched or "일치 문장 없음",
        detail={
            "expected_text": expected_text,
            "expected_texts": expected_texts,
            "matched_text": matched or "",
        },
    )


def _check_docx_part_contains(check, file_info, context, *, part):
    expected_text = _resolve_rule_value(str(check.get("text") or ""), context)
    if part == "header":
        actual_text = _docx_header_text(file_info)
        label = "머리글"
    else:
        actual_text = _docx_footer_text(file_info)
        label = "바닥글"
    passed = expected_text in actual_text
    return _content_result(
        check,
        passed,
        expected=f"{label}에 '{expected_text}' 포함",
        actual=actual_text or f"{label} 없음",
        detail={"part": part, "expected_text": expected_text, "actual_text": actual_text},
    )


def _check_docx_next_paragraph_matches(check, file_info, context):
    after_texts = [
        _resolve_rule_value(str(text), context)
        for text in check.get("after_texts") or []
        if str(text).strip()
    ]
    after_text = _resolve_rule_value(str(check.get("after_text") or ""), context)
    pattern = str(check.get("regex") or "").strip()
    paragraphs = _docx_paragraphs(file_info)
    if after_texts:
        matched_index = _find_matching_paragraph_index_with_all(paragraphs, after_texts, check)
        after_label = ", ".join(after_texts)
    else:
        matched_index = _find_matching_paragraph_index(paragraphs, after_text, check)
        after_label = after_text
    next_text = ""
    if matched_index is not None:
        for paragraph in paragraphs[matched_index + 1:]:
            if paragraph.strip():
                next_text = paragraph.strip()
                break
    passed = bool(next_text and re.search(pattern, next_text))
    return _content_result(
        check,
        passed,
        expected=f"'{after_label}' 다음 문단이 {pattern} 형식",
        actual=next_text or "다음 문단 없음",
        detail={
            "after_text": after_text,
            "after_texts": after_texts,
            "regex": pattern,
            "actual_text": next_text,
        },
    )


def _content_result(check, passed, *, expected, actual, detail):
    return {
        "passed": passed,
        "expected": expected,
        "actual": actual,
        "message": (
            check.get("pass_message")
            if passed
            else check.get("failure_message")
        ) or ("문서 내용을 확인했습니다." if passed else "문서 내용이 기준과 일치하지 않습니다."),
        "detail": detail,
    }


def _content_check_expected(check, context):
    if "expected" in check:
        return _resolve_rule_value(str(check.get("expected") or ""), context)
    if "texts" in check:
        return ", ".join(
            _resolve_rule_value(str(text), context)
            for text in check.get("texts") or []
            if str(text).strip()
        )
    if "text" in check:
        return _resolve_rule_value(str(check.get("text") or ""), context)
    return str(check.get("type") or "")


def _read_excel_workbook(file_info):
    extension = file_info.extension.lower()
    data = _read_file_bytes(file_info)
    if extension == ".xlsx":
        return _read_xlsx_workbook(data)
    if extension == ".xls":
        return _read_xls_workbook(data)
    raise DownloadReviewInspectionError(f"지원하지 않는 Excel 확장자입니다: {extension or '(없음)'}")


def _read_xlsx_workbook(data):
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=False)
    except Exception as exc:
        raise DownloadReviewInspectionError("xlsx 파일을 읽을 수 없습니다.") from exc

    sheets = []
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows():
            rows.append([_excel_cell_text(cell.value) for cell in row])
        sheets.append(
            ExcelSheet(
                name=worksheet.title,
                rows=_trim_empty_edges(rows),
                header_text=_worksheet_header_text(worksheet),
                footer_text=_worksheet_footer_text(worksheet),
            )
        )
    workbook.close()
    return ExcelWorkbook(sheets=sheets)


def _worksheet_header_text(worksheet):
    parts = []
    for header in (worksheet.oddHeader, worksheet.evenHeader, worksheet.firstHeader):
        for section in (header.left, header.center, header.right):
            text = getattr(section, "text", "") or ""
            if text:
                parts.append(text)
    return _normalize_spaces(" ".join(parts))


def _worksheet_footer_text(worksheet):
    parts = []
    for footer in (worksheet.oddFooter, worksheet.evenFooter, worksheet.firstFooter):
        for section in (footer.left, footer.center, footer.right):
            text = getattr(section, "text", "") or ""
            if text:
                parts.append(text)
    return _normalize_spaces(" ".join(parts))


def _read_xls_workbook(data):
    try:
        import xlrd
    except ImportError as exc:
        raise DownloadReviewInspectionError("xls 파일을 읽으려면 xlrd 패키지가 필요합니다.") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=data)
    except Exception as exc:
        raise DownloadReviewInspectionError("xls 파일을 읽을 수 없습니다.") from exc

    # xlrd는 인쇄 머리글/바닥글(BIFF HEADER/FOOTER 레코드)을 노출하지 않으므로 직접 파싱한다.
    headers_by_sheet = _xls_print_headers(data)
    footers_by_sheet = _xls_print_footers(data)

    sheets = []
    for worksheet in workbook.sheets():
        rows = []
        for row_index in range(worksheet.nrows):
            rows.append([
                _excel_cell_text(worksheet.cell_value(row_index, col_index))
                for col_index in range(worksheet.ncols)
            ])
        sheets.append(
            ExcelSheet(
                name=worksheet.name,
                rows=_trim_empty_edges(rows),
                header_text=headers_by_sheet.get(worksheet.name, ""),
                footer_text=footers_by_sheet.get(worksheet.name, ""),
            )
        )
    return ExcelWorkbook(sheets=sheets)


def _xls_print_headers(data):
    return _xls_print_records(data, opcode=0x0014)


def _xls_print_footers(data):
    return _xls_print_records(data, opcode=0x0015)


def _xls_print_records(data, *, opcode):
    """`.xls` 워크북 스트림을 직접 파싱해 시트별 인쇄 머리글/바닥글을 추출한다.

    xlrd는 BIFF HEADER/FOOTER 레코드를 노출하지 않으므로, OLE2 컨테이너에서
    Workbook 스트림을 꺼낸 뒤 BOUNDSHEET(0x85)로 시트별 substream 위치를 찾고
    각 substream 첫 대상 레코드를 읽는다. 실패하면 빈 매핑을 반환한다.
    """
    try:
        from io import StringIO

        from xlrd.compdoc import CompDoc

        compdoc = CompDoc(data, logfile=StringIO())
        mem, base, size = compdoc.locate_named_stream("Workbook")
        if mem is None or not size:
            return {}
        stream = mem[base:base + size]
    except Exception:
        return {}

    target_opcode = opcode
    try:
        boundsheets = []
        pos = 0
        total = len(stream)
        while pos + 4 <= total:
            record_opcode, length = struct.unpack("<HH", stream[pos:pos + 4])
            body = stream[pos + 4:pos + 4 + length]
            if record_opcode == 0x0085 and len(body) >= 8:  # BOUNDSHEET
                ply_pos = struct.unpack("<I", body[0:4])[0]
                name, _ = _xls_unicode_string(body, 6, 1)
                boundsheets.append((ply_pos, name))
            pos += 4 + length

        records = {}
        for ply_pos, name in boundsheets:
            records[name] = _xls_print_record_at(stream, ply_pos, target_opcode)
        return records
    except Exception:
        return {}


def _xls_print_record_at(stream, start, target_opcode):
    pos = start
    total = len(stream)
    depth = 0
    while pos + 4 <= total:
        opcode, length = struct.unpack("<HH", stream[pos:pos + 4])
        body = stream[pos + 4:pos + 4 + length]
        if opcode == 0x0809:  # BOF
            depth += 1
        elif opcode == 0x000A:  # EOF
            depth -= 1
            if depth <= 0:
                return ""
        elif opcode == target_opcode:
            if length == 0:
                return ""
            text, _ = _xls_unicode_string(body, 0, 2)
            return text
        pos += 4 + length
    return ""


def _xls_unicode_string(body, offset, cch_size):
    if cch_size == 1:
        cch = body[offset]
        cursor = offset + 1
    else:
        cch = struct.unpack("<H", body[offset:offset + 2])[0]
        cursor = offset + 2
    grbit = body[cursor]
    cursor += 1
    if grbit & 0x01:
        text = body[cursor:cursor + cch * 2].decode("utf-16-le", "replace")
    else:
        text = body[cursor:cursor + cch].decode("cp949", "replace")
    return text, cursor


def _excel_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return _normalize_spaces(value)


def _trim_empty_edges(rows):
    trimmed = [list(row) for row in rows]
    while trimmed and not any(cell for cell in trimmed[-1]):
        trimmed.pop()
    max_width = 0
    for row in trimmed:
        for index, cell in enumerate(row):
            if cell:
                max_width = max(max_width, index + 1)
    if not max_width:
        return []
    return [row[:max_width] + [""] * max(0, max_width - len(row)) for row in trimmed]


def _find_cell_containing(rows, text):
    needle = str(text or "")
    if not needle:
        return None
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            if needle in value:
                return {"row": row_index, "column": col_index, "value": value}
    return None


def _find_cell_with_date_range(rows, start_date, end_date):
    """`{시작일} ~ {종료일}` 기간이 적힌 셀을 날짜 정규화 기준으로 찾는다.

    셀 안 날짜 구분자(`.`/`-`/`/`/`년월일`)가 달라도 같은 날짜면 인정한다.
    """
    start_norm = _format_dot_date(start_date)
    end_norm = _format_dot_date(end_date)
    if not (start_norm and end_norm):
        return None
    date_pattern = r"\d{4}\s*[.\-/년]\s*\d{1,2}\s*[.\-/월]\s*\d{1,2}\s*일?\.?"
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            found = [_format_dot_date(token) for token in re.findall(date_pattern, str(value or ""))]
            if start_norm in found and end_norm in found:
                return {"row": row_index, "column": col_index, "value": value}
    return None


def _find_labeled_date_range_cell(rows, label, start_date, end_date):
    """라벨(공백 제거 후 포함)과 {시작일}~{종료일}을 함께 담은 셀을 날짜 정규화로 찾는다.

    셀 텍스트의 공백을 무시하고, 날짜 형식(`.`/`-`/`/`/`년월일`, 2~4자리 연도)이 달라도
    실제 날짜가 같으면 일치로 본다.
    """
    start_norm = _format_dot_date(start_date)
    end_norm = _format_dot_date(end_date)
    if not (start_norm and end_norm):
        return None
    needle = _normalize_no_space(label)
    date_pattern = r"\d{2,4}\s*[.\-/년]\s*\d{1,2}\s*[.\-/월]\s*\d{1,2}\s*일?\.?"
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            text = str(value or "")
            if needle and needle not in _normalize_no_space(text):
                continue
            found = {_format_dot_date(token) for token in re.findall(date_pattern, text)}
            if start_norm in found and end_norm in found:
                return {"row": row_index, "column": col_index, "value": text}
    return None


def _excel_area_from_column_anchor(rows, label, *, column_index):
    for row_index, row in enumerate(rows):
        if column_index >= len(row) or label not in row[column_index]:
            continue

        last_row = row_index
        last_col = column_index
        for scan_row_index in range(row_index, len(rows)):
            scan_row = rows[scan_row_index]
            populated_columns = [
                col_index
                for col_index, value in enumerate(scan_row[column_index:], start=column_index)
                if value
            ]
            if populated_columns:
                last_row = scan_row_index
                last_col = max(last_col, max(populated_columns))
        return {
            "start_row": row_index + 1,
            "start_column": column_index + 1,
            "end_row": last_row + 1,
            "end_column": last_col + 1,
            "range": f"{_excel_column_name(column_index + 1)}{row_index + 1}:{_excel_column_name(last_col + 1)}{last_row + 1}",
        }
    return None


def _excel_column_name(index):
    name = ""
    value = index
    while value:
        value, remainder = divmod(value - 1, 26)
        name = chr(ord("A") + remainder) + name
    return name


def _date_range_start(value):
    date_value = _parse_date(value)
    if not date_value:
        return None
    return datetime.combine(date_value, time.min)


def _date_range_end(value):
    date_value = _parse_date(value)
    if not date_value:
        return None
    return datetime.combine(date_value, time.max)


def _parse_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    normalized = text.replace("년", ".").replace("월", ".").replace("일", "")
    normalized = normalized.replace("/", ".").replace("-", ".")
    match = re.search(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", normalized)
    if not match:
        return None
    try:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()
    except ValueError:
        return None


def _inspection_folder_tree(verify_result):
    folders = set()
    file_folders = set()
    for file_info in _inspection_files(verify_result):
        segments = tuple(_folder_segments(file_info.path))
        file_folders.add(segments)
        for index in range(1, len(segments) + 1):
            folders.add(segments[:index])

    for folder in _local_download_folders(getattr(verify_result, "download_dir", "")):
        folders.add(folder)

    return folders, file_folders


def _folder_tree_from_files(files):
    """주어진 파일 목록만으로 폴더 트리(folders, file_folders)를 구성한다."""
    folders = set()
    file_folders = set()
    for file_info in files:
        segments = tuple(_folder_segments(file_info.path))
        file_folders.add(segments)
        for index in range(1, len(segments) + 1):
            folders.add(segments[:index])
    return folders, file_folders


def _local_download_folders(download_dir):
    base = Path(download_dir or "")
    if not base.is_dir():
        return set()
    folders = set()
    for path in base.rglob("*"):
        if not path.is_dir():
            continue
        parts = tuple(path.relative_to(base).parts)
        for index in range(1, len(parts) + 1):
            folders.add(parts[:index])
    return folders


def _find_folder_by_keyword_chain(folders, keyword_chain):
    chain = [str(item).strip() for item in keyword_chain if str(item).strip()]
    for folder in sorted(folders, key=lambda item: (len(item), "/".join(item))):
        if _folder_keyword_chain_indices(list(folder), chain):
            return folder
    return ()


def _folder_direct_files(folder, files):
    """해당 폴더 바로 아래에 있는 파일 목록(FileInfo)을 반환한다."""
    target = tuple(folder)
    return [file_info for file_info in (files or []) if tuple(_folder_segments(file_info.path)) == target]


def _files_under_folder(folder, files):
    """해당 폴더(및 모든 하위)에 속한 파일 목록(FileInfo)을 반환한다."""
    target = tuple(folder)
    depth = len(target)
    return [
        file_info
        for file_info in (files or [])
        if tuple(_folder_segments(file_info.path))[:depth] == target
    ]


def _any_file_name_contains(files, keyword):
    """파일명에 keyword(문자열 또는 문자열 목록) 중 하나라도 포함되면 True.

    목록으로 주면 OR 매칭(하나라도 포함되면 통과)이다. 0바이트 파일도
    이름만 맞으면 그대로 인정한다(내용까지 확인하지 않음).
    """
    needles = [keyword] if isinstance(keyword, str) else list(keyword or [])
    needles = [str(needle or "").strip() for needle in needles if str(needle or "").strip()]
    if not needles:
        return False
    return any(
        needle in str(file_info.name or "")
        for file_info in (files or [])
        for needle in needles
    )


def _exception_words_text(keyword):
    """_any_file_name_contains 용 예외 단어를 메시지에 쓸 문자열로 만든다."""
    needles = [keyword] if isinstance(keyword, str) else list(keyword or [])
    needles = [str(needle or "").strip() for needle in needles if str(needle or "").strip()]
    return "/".join(needles)


def _run_folder_check(folders, file_folders, folder_check, files=None):
    keyword = str(folder_check.get("keyword") or "").strip()
    failure_message = str(folder_check.get("failure_message") or "폴더 구조 확인 불가")
    # 특정 단어(문자열 또는 목록, OR 매칭)가 든 파일이 있으면 폴더 구조(폴더
    # 존재 여부/하위 폴더 개수)와 무관하게 무조건 통과 처리하는 예외
    # (예: '수행 안 함'/'대상 아님' 안내 파일, 0바이트여도 이름만 맞으면 인정).
    name_exception = folder_check.get("pass_if_file_name_contains") or ""
    name_exception_text = _exception_words_text(name_exception)
    if name_exception and _any_file_name_contains(files, name_exception):
        return {
            "keyword": keyword,
            "folder": "",
            "passed": True,
            "message": f"'{name_exception_text}' 포함 파일 존재하여 예외 통과",
            "actual": f"'{name_exception_text}' 포함 파일 존재",
        }

    # rawdata zip 안 어디에 있든 키워드를 포함한 폴더를 직접 찾는다.
    folder = _find_folder_by_keyword_chain(folders, [keyword]) if keyword else ()
    if not folder:
        return {
            "keyword": keyword,
            "folder": "",
            "passed": False,
            "message": failure_message,
            "actual": "폴더 없음",
        }

    # 쓸데없는 래퍼 폴더(예: 성능시험/주요기능)가 하나 끼어 있을 수 있다.
    # 하위 폴더가 정확히 1개뿐이면 그 안으로 내려가 원래 기대한 구조를 찾는다.
    if folder_check.get("unwrap_single_folder"):
        guard = 0
        while guard < 5:
            children = list(_immediate_child_folders(folder, folders))
            if len(children) == 1:
                folder = children[0]
                guard += 1
            else:
                break

    # 폴더 안에 이미지 파일이 최소 N개 있어야 한다(예: 결함 rawdata 스크린샷).
    min_images = folder_check.get("min_images")
    if min_images is not None:
        image_count = len([
            file_info
            for file_info in _files_under_folder(folder, files)
            if file_info.extension.lower() in IMAGE_EXTENSIONS
        ])
        if image_count < int(min_images):
            return {
                "keyword": keyword,
                "folder": "/".join(folder),
                "passed": False,
                "message": failure_message,
                "actual": f"이미지 {image_count}개",
            }

    exact_child_folders = folder_check.get("exact_child_folders")
    min_child_folders = folder_check.get("min_child_folders")
    child_folders = _immediate_child_folders(folder, folders)
    child_files = _immediate_child_file_folders(folder, file_folders)

    # 예외: 하위 폴더 없이 txt 파일만 있으면 '수행안함' 안내로 보고 통과 처리한다.
    # (예: 보안시험이 설치형 PC 프로그램이라 미수행 안내 txt만 존재하는 경우)
    if folder_check.get("txt_only_pass") and not child_folders:
        direct_files = _folder_direct_files(folder, files)
        if direct_files and all(f.extension.lower() == ".txt" for f in direct_files):
            return {
                "keyword": keyword,
                "folder": "/".join(folder),
                "passed": True,
                "message": "수행안함 안내(txt)만 존재하여 예외 통과",
                "actual": "txt 안내 파일만 존재: " + ", ".join(f.name for f in direct_files[:3]),
            }

    # 예외 단어 매칭은 함수 맨 위에서 이미 확인했다(매칭됐으면 여기 도달하지
    # 않고 통과로 반환됨). 하위 폴더가 없는데도 여기까지 왔다는 건 예외 단어도
    # 없었다는 뜻이므로 실패로 본다.
    if name_exception and not child_folders:
        return {
            "keyword": keyword,
            "folder": "/".join(folder),
            "passed": False,
            "message": failure_message,
            "actual": "하위 폴더 없음",
        }

    if exact_child_folders is not None and len(child_folders) != int(exact_child_folders):
        return {
            "keyword": keyword,
            "folder": "/".join(folder),
            "passed": False,
            "message": failure_message,
            "actual": f"하위 폴더 {len(child_folders)}개",
        }

    if min_child_folders is not None and len(child_folders) < int(min_child_folders):
        return {
            "keyword": keyword,
            "folder": "/".join(folder),
            "passed": False,
            "message": failure_message,
            "actual": f"하위 폴더 {len(child_folders)}개",
        }

    if folder_check.get("each_child_has_entry"):
        empty_children = [
            child
            for child in child_folders
            if not _immediate_child_folders(child, folders)
            and not _immediate_child_file_folders(child, file_folders)
        ]
        if empty_children:
            return {
                "keyword": keyword,
                "folder": "/".join(folder),
                "passed": False,
                "message": failure_message,
                "actual": "빈 하위 폴더: " + ", ".join("/".join(child) for child in empty_children[:5]),
            }

    min_entries = folder_check.get("min_entries")
    if min_entries is not None:
        entry_count = len(child_folders) + len(child_files)
        if entry_count < int(min_entries):
            return {
                "keyword": keyword,
                "folder": "/".join(folder),
                "passed": False,
                "message": failure_message,
                "actual": f"항목 {entry_count}개",
            }

    return {
        "keyword": keyword,
        "folder": "/".join(folder),
        "passed": True,
        "message": "정상",
        "actual": {
            "child_folder_count": len(child_folders),
            "child_file_count": len(child_files),
        },
    }


def _find_immediate_descendant_folder(base, folders, keyword):
    candidates = [
        folder
        for folder in folders
        if len(folder) == len(base) + 1
        and folder[: len(base)] == base
        and keyword in folder[-1]
    ]
    return sorted(candidates, key=lambda item: "/".join(item))[0] if candidates else ()


def _immediate_child_folders(parent, folders):
    return {
        folder
        for folder in folders
        if len(folder) == len(parent) + 1
        and folder[: len(parent)] == parent
    }


def _immediate_child_file_folders(parent, file_folders):
    return {
        folder
        for folder in file_folders
        if folder == parent
    }


def _safe_artifact_id(value):
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "").strip())
    return cleaned.strip("._-") or "artifact"


def _docx_defect_report_round_dates(file_info):
    """시험성적서에서 '결함리포트 송부' 표의 차수별 보고일자를 읽는다.

    최신 시험성적서 서식에는 이 표 자체가 없는 경우가 있다(예: 재인증/간소화
    시험처럼 결함 이력을 별도 요약하지 않는 서식). 표가 없다고 시험성적서 전체
    점검을 에러로 처리하면 안 되므로, 못 찾으면 예외 대신 빈 dict를 반환한다.
    결함차수는 이 표가 아니어도 결함리포트 파일 자체에서 추론할 수 있다
    (_infer_defect_round_count_from_files 참고).
    """
    tables = _docx_tables(file_info)
    target_text = ""
    for table in tables:
        flattened = " ".join(cell for row in table for cell in row if cell)
        if "결함리포트 송부" in flattened:
            target_text = flattened
            break
    if not target_text:
        return {}

    rounds = {}
    for match in re.finditer(r"([1-9]\d*)\s*차\s*[:：]\s*(\d{4})\s*[./-]\s*(\d{1,2})\s*[./-]\s*(\d{1,2})", target_text):
        round_no, year, month, day = match.groups()
        rounds[f"{int(round_no)}차"] = f"{int(year):04d}.{int(month):02d}.{int(day):02d}."
    return rounds


def _docx_first_table_after_text(file_info, marker):
    root = _docx_root(file_info)
    ns = _word_ns()
    marker_text = str(marker or "")
    marker_seen = False
    body = root.find(".//w:body", namespaces=ns)
    if body is None:
        return []

    for child in body:
        tag_name = etree.QName(child).localname
        if tag_name == "tbl":
            rows = _word_table_rows(child, ns)
            flattened = " ".join(cell for row in rows for cell in row if cell)
            if marker_seen:
                return rows
            if marker_text and marker_text in flattened:
                marker_seen = True
            continue

        text = _word_element_text(child, ns)
        if marker_text and marker_text in text:
            marker_seen = True
    return []


def _docx_tables(file_info):
    root = _docx_root(file_info)
    ns = _word_ns()
    return [
        _word_table_rows(table, ns)
        for table in root.xpath(".//w:tbl", namespaces=ns)
    ]


def _docx_table_with_header(file_info, header):
    """헤더 셀 텍스트로 표를 직접 찾는다(선행 마커 텍스트에 의존하지 않음).

    _docx_first_table_after_text는 마커가 목차에도 그대로 나오면(예: 절 제목이
    '4.4 시험일정'처럼 목차 항목과 동일한 텍스트) 목차 바로 뒤의 엉뚱한 표를
    집어오는 문제가 있다. 헤더가 문서 전체에서 그 표에만 유일하게 있다면
    이 방식이 더 안전하다.
    """
    for table in _docx_tables(file_info):
        if _find_cell_containing(table, header):
            return table
    return []


def _word_table_rows(table, ns):
    rows = []
    for row in table.xpath("./w:tr", namespaces=ns):
        cells = []
        for cell in row.xpath("./w:tc", namespaces=ns):
            cells.append(_word_cell_text(cell, ns))
        rows.append(cells)
    return rows


def _word_element_text(element, ns):
    return _normalize_spaces("".join(element.xpath(".//w:t/text()", namespaces=ns)))


def _docx_root(file_info):
    data = _read_file_bytes(file_info)
    with ZipFile(BytesIO(data)) as docx_file:
        return etree.fromstring(docx_file.read("word/document.xml"))


def _docx_paragraphs(file_info):
    root = _docx_root(file_info)
    ns = _word_ns()
    paragraphs = []
    for paragraph in root.xpath(".//w:p", namespaces=ns):
        text = "".join(paragraph.xpath(".//w:t/text()", namespaces=ns)).strip()
        if text:
            paragraphs.append(_normalize_spaces(text))
    return paragraphs


def _docx_all_text(file_info):
    data = _read_file_bytes(file_info)
    ns = _word_ns()
    paragraphs = []
    try:
        with ZipFile(BytesIO(data)) as docx_file:
            target_names = [
                name
                for name in docx_file.namelist()
                if name.startswith("word/")
                and name.endswith(".xml")
                and (
                    name == "word/document.xml"
                    or PurePosixPath(name).name.startswith("header")
                    or PurePosixPath(name).name.startswith("footer")
                )
            ]
            for name in sorted(target_names):
                try:
                    root = etree.fromstring(docx_file.read(name))
                except etree.XMLSyntaxError:
                    continue
                # 한 문단 안의 run(w:t)은 공백 없이 이어붙인다. Word가 숫자/단어를
                # 여러 run으로 쪼개므로 공백 join을 쓰면 "2026"이 "20 2 6"이 된다.
                for paragraph in root.xpath(".//w:p", namespaces=ns):
                    text = "".join(paragraph.xpath(".//w:t/text()", namespaces=ns))
                    if text.strip():
                        paragraphs.append(text)
    except (KeyError, BadZipFile) as exc:
        raise DownloadReviewInspectionError("docx 본문/머리말/바닥글을 읽을 수 없습니다.") from exc
    return _normalize_spaces(" ".join(paragraphs))


def _docx_footer_text(file_info):
    data = _read_file_bytes(file_info)
    ns = _word_ns()
    texts = []
    try:
        with ZipFile(BytesIO(data)) as docx_file:
            footer_names = sorted(
                name
                for name in docx_file.namelist()
                if name.startswith("word/footer") and name.endswith(".xml")
            )
            for footer_name in footer_names:
                root = etree.fromstring(docx_file.read(footer_name))
                # 문단 안 run은 공백 없이 이어붙여 "2026"이 쪼개지지 않게 한다.
                for paragraph in root.xpath(".//w:p", namespaces=ns):
                    text = _normalize_spaces("".join(paragraph.xpath(".//w:t/text()", namespaces=ns)))
                    if text:
                        texts.append(text)
    except (KeyError, BadZipFile, etree.XMLSyntaxError) as exc:
        raise DownloadReviewInspectionError("docx 바닥글을 읽을 수 없습니다.") from exc
    return _normalize_spaces(" ".join(texts))


def _docx_header_text(file_info):
    data = _read_file_bytes(file_info)
    ns = _word_ns()
    texts = []
    try:
        with ZipFile(BytesIO(data)) as docx_file:
            header_names = sorted(
                name
                for name in docx_file.namelist()
                if name.startswith("word/header") and name.endswith(".xml")
            )
            for header_name in header_names:
                root = etree.fromstring(docx_file.read(header_name))
                for paragraph in root.xpath(".//w:p", namespaces=ns):
                    text = _normalize_spaces("".join(paragraph.xpath(".//w:t/text()", namespaces=ns)))
                    if text:
                        texts.append(text)
    except (KeyError, BadZipFile, etree.XMLSyntaxError) as exc:
        raise DownloadReviewInspectionError("docx 머리글을 읽을 수 없습니다.") from exc
    return _normalize_spaces(" ".join(texts))


def _docx_table_rows(file_info):
    root = _docx_root(file_info)
    ns = _word_ns()
    rows = []
    for table in root.xpath(".//w:tbl", namespaces=ns):
        for row in table.xpath("./w:tr", namespaces=ns):
            cells = []
            for cell in row.xpath("./w:tc", namespaces=ns):
                cells.append(_word_cell_text(cell, ns))
            rows.append(cells)
    return rows


def _word_cell_text(cell, ns):
    paragraphs = []
    for paragraph in cell.xpath("./w:p", namespaces=ns):
        text = "".join(paragraph.xpath(".//w:t/text()", namespaces=ns)).strip()
        if text:
            paragraphs.append(text)
    if not paragraphs:
        paragraphs = ["".join(cell.xpath(".//w:t/text()", namespaces=ns)).strip()]
    return _normalize_spaces(" ".join(paragraphs))


def _word_ns():
    return {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def _find_next_cell_by_label(rows, label):
    target = _normalize_label(label)
    for row in rows:
        for index, cell in enumerate(row[:-1]):
            if _normalize_label(cell) == target:
                return row[index + 1].strip()
    return ""


def _pdf_page_text(file_info, *, page_index=0):
    import fitz

    data = _read_file_bytes(file_info)
    with fitz.open(stream=data, filetype="pdf") as document:
        if document.page_count <= page_index:
            return ""
        return document[page_index].get_text("text")


def _find_matching_paragraph(paragraphs, expected_text, check):
    index = _find_matching_paragraph_index(paragraphs, expected_text, check)
    return paragraphs[index] if index is not None else None


def _find_matching_paragraph_with_all(paragraphs, expected_texts, check):
    index = _find_matching_paragraph_index_with_all(paragraphs, expected_texts, check)
    return paragraphs[index] if index is not None else None


def _find_matching_paragraph_index(paragraphs, expected_text, check):
    expected = _normalize_content(expected_text, check)
    for index, paragraph in enumerate(paragraphs):
        actual = _normalize_content(paragraph, check)
        if expected and expected in actual:
            return index
    return None


def _find_matching_paragraph_index_with_all(paragraphs, expected_texts, check):
    expected_values = [
        _normalize_content(text, check)
        for text in expected_texts
        if _normalize_content(text, check)
    ]
    if not expected_values:
        return None
    for index, paragraph in enumerate(paragraphs):
        actual = _normalize_content(paragraph, check)
        if all(expected in actual for expected in expected_values):
            return index
    return None


def _normalize_content(value, check):
    text = str(value or "")
    if check.get("remove_whitespace"):
        return re.sub(r"\s+", "", text)
    if check.get("normalize_whitespace", True):
        return _normalize_spaces(text)
    return text


def _normalize_compare(value, check):
    text = str(value or "")
    if check.get("remove_whitespace"):
        return re.sub(r"\s+", "", text)
    return _normalize_spaces(text) if check.get("normalize_whitespace") else text.strip()


def _normalize_label(value):
    return re.sub(r"[\s:：]+", "", str(value or "")).lower()


def _normalize_spaces(value):
    # None\ub9cc \ube48 \ubb38\uc790\uc5f4\ub85c \ubcf8\ub2e4. `value or ""`\ub97c \uc4f0\uba74 \uc815\uc218 0\u00b7False\uac00 falsy\ub77c
    # \ube48 \ubb38\uc790\uc5f4\uc774 \ub418\uc5b4 "0"\uc774 \uc0ac\ub77c\uc9c0\ub294 \ubc84\uadf8\uac00 \uc788\uc5c8\ub2e4.
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.replace("\u00a0", " ")).strip()


def _first_line(value):
    text = str(value or "")
    for separator in ("\r\n", "\n", "\r"):
        if separator in text:
            text = text.split(separator, 1)[0]
            break
    return text.strip()


def _extension_from_file_type(file_type):
    value = str(file_type or "").strip().lower()
    if not value or value == "any":
        return ""
    if value.startswith("."):
        return value
    return f".{value}"


def _representative_name(files):
    if not files:
        return ""
    if len(files) == 1:
        return files[0].name
    return f"{files[0].name} 외 {len(files) - 1}개"


def _representative_path(files, project_number):
    if not files:
        return ""
    return _display_path(files[0].path, project_number) or files[0].name


def _display_path(path, project_number):
    normalized = str(path or "").replace("\\", "/")
    index = normalized.find(project_number)
    if index >= 0:
        return normalized[index:]
    return Path(normalized).name


# ── 산출물 저장: 어댑터 sink 위임 ───────────────────────────────────────────────
def _store_pdf_first_page_artifact(project, rule, file_info, *, artifact_id, label):
    return _current_sink().store_pdf_first_page(
        project, rule, file_info, artifact_id=artifact_id, label=label
    )


def _store_pdf_download_artifact(project, rule, file_info, *, artifact_id, label):
    return _current_sink().store_pdf_download(
        project, rule, file_info, artifact_id=artifact_id, label=label
    )


def _store_excel_area_artifact(project, rule, sheet, area, *, artifact_id, label, source_file):
    return _current_sink().store_excel_area(
        project, rule, sheet, area, artifact_id=artifact_id, label=label, source_file=source_file
    )


def set_artifact_sink(sink):
    _ARTIFACT_SINK_VAR.set(sink if sink is not None else NoOpArtifactSink())


class _VerifyResult:
    """엔진 파일 모델이 기대하는 verify_result 어댑터 (.files 만 필요)."""

    def __init__(self, files):
        self.files = list(files or [])


class _ProjectStub:
    def __init__(self, project_number, project_id=""):
        self.project_number = project_number
        self.id = project_id


def build_context(*, project_number="", product_name="", company="", pl="", wd="",
                  start_date="", end_date="", request_date="", contract_date="",
                  certification_committee_date="", center=""):
    """프로젝트 메타데이터로 RuleContext 를 만든다(웹/로컬 공용)."""
    product_raw = _first_line(product_name)
    product, version = _split_product_and_version(product_raw)
    return RuleContext(
        project_number=project_number or "",
        product_raw=product_raw,
        product=product,
        version=version,
        company=_first_line(company),
        pl=_first_pl_name(pl),
        wd=wd or "",
        start_date=start_date or "",
        end_date=end_date or "",
        year=_project_year(project_number or ""),
        request_date=request_date or "",
        contract_date=contract_date or "",
        certification_committee_date=certification_committee_date or "",
        derived_variables={},
        center=center or "",
    )


def evaluate_rules(rules, context, files, *, project=None, sink=None):
    """규칙 목록을 평가해 RuleEvaluation 리스트를 반환한다.

    rules: rule_type/code/name/config_json/target_file_type/target_file_pattern
           속성을 가진 객체 목록(웹=DownloadReviewRule, 로컬=RuleSpec).
    context: RuleContext (build_context 로 생성).
    files: 최상위 파일(FileInfo 등 .name/.path/.size/.extension/.modified_at) 목록,
           또는 .files 속성을 가진 verify_result 류 객체. 후자를 넘기면 zip 확장
           캐시·오류 정보가 그 객체에 그대로 누적된다(웹 호환). zip 은 엔진이 확장한다.
    project: .project_number/.id 를 가진 객체. 없으면 context 로 스텁 생성.
    sink: ArtifactSink. 없으면 no-op(산출물 미생성).
    """
    set_artifact_sink(sink)
    _DOC_CONVERT_CACHE.clear()
    if project is None:
        project = _ProjectStub(context.project_number)
    verify_result = files if hasattr(files, "files") else _VerifyResult(files)
    evaluations = []
    for sequence, rule in enumerate(rules, start=1):
        evaluation = _evaluate_rule(rule, sequence, project, context, verify_result, None)
        evaluation = _apply_disabled_sub_checks(rule, evaluation)
        evaluations.append(evaluation)
        _collect_evaluation_variables(context, evaluation)
    return evaluations
