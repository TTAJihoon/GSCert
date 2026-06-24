"""Word/Excel/PDF 파서 (Django 비종속, bytes 입력).

서버 ecm_download_review_inspection.py 의 검증된 파싱 로직을 그대로 이식했다.
파일 입출력(zip 내부/.doc 변환/디스크 읽기)은 어댑터 책임이며, 여기서는
이미 읽은 bytes 만 받는다.
"""

from __future__ import annotations

import re
import struct
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any
from zipfile import BadZipFile, ZipFile

from lxml import etree

WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


class DocumentReadError(RuntimeError):
    """문서를 파싱할 수 없을 때 발생."""


# ── 공통 텍스트 정규화 ────────────────────────────────────────────────────────

def normalize_spaces(value: Any) -> str:
    # None만 빈 문자열로 본다. `value or ""`를 쓰면 정수 0·False가 falsy라
    # "0"이 사라지는 버그가 있었다.
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.replace(" ", " ")).strip()


def excel_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return normalize_spaces(value)


def trim_empty_edges(rows: list[list[str]]) -> list[list[str]]:
    trimmed = [list(row) for row in rows]
    while trimmed and not any(cell for cell in trimmed[-1]):
        trimmed.pop()
    max_width = 0
    for row in trimmed:
        for index, cell in enumerate(row):
            if cell:
                max_width = max(max_width, index + 1)
    if not max_width:
        return []
    return [row[:max_width] for row in trimmed]


# ── Word (.docx) ─────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class WordDocument:
    paragraphs: list[str]
    tables: list[list[list[str]]]
    header_text: str
    footer_text: str
    # 본문 순서(문단/표 혼합) — "특정 문구 다음 첫 표" 같은 순서 의존 검사용.
    # 각 항목은 ("p", text) 또는 ("tbl", rows).
    body: list[tuple[str, Any]] = field(default_factory=list)

    @property
    def full_text(self) -> str:
        table_text = " ".join(
            cell for table in self.tables for row in table for cell in row if cell
        )
        return normalize_spaces(
            " ".join([*self.paragraphs, table_text, self.header_text, self.footer_text])
        )


def _word_cell_text(cell) -> str:
    paragraphs = []
    for paragraph in cell.xpath("./w:p", namespaces=WORD_NS):
        text = "".join(paragraph.xpath(".//w:t/text()", namespaces=WORD_NS)).strip()
        if text:
            paragraphs.append(text)
    if not paragraphs:
        paragraphs = ["".join(cell.xpath(".//w:t/text()", namespaces=WORD_NS)).strip()]
    return normalize_spaces(" ".join(paragraphs))


def _table_rows(table) -> list[list[str]]:
    rows = []
    for row in table.xpath("./w:tr", namespaces=WORD_NS):
        rows.append([_word_cell_text(cell) for cell in row.xpath("./w:tc", namespaces=WORD_NS)])
    return rows


def _part_text(data: bytes, archive: ZipFile, prefix: str) -> str:
    texts = []
    names = sorted(
        name for name in archive.namelist()
        if name.startswith(prefix) and name.endswith(".xml")
    )
    for name in names:
        try:
            root = etree.fromstring(archive.read(name))
        except etree.XMLSyntaxError:
            continue
        for paragraph in root.xpath(".//w:p", namespaces=WORD_NS):
            # 문단 안 run(w:t)은 공백 없이 이어붙인다(숫자 "2026"이 쪼개지지 않게).
            text = normalize_spaces("".join(paragraph.xpath(".//w:t/text()", namespaces=WORD_NS)))
            if text:
                texts.append(text)
    return normalize_spaces(" ".join(texts))


def read_docx(data: bytes) -> WordDocument:
    try:
        with ZipFile(BytesIO(data)) as archive:
            try:
                document_xml = archive.read("word/document.xml")
            except KeyError as exc:
                raise DocumentReadError("docx 본문을 찾을 수 없습니다.") from exc
            header_text = _part_text(data, archive, "word/header")
            footer_text = _part_text(data, archive, "word/footer")
    except BadZipFile as exc:
        raise DocumentReadError("docx 파일을 읽을 수 없습니다.") from exc

    try:
        root = etree.fromstring(document_xml)
    except etree.XMLSyntaxError as exc:
        raise DocumentReadError("docx 본문을 해석할 수 없습니다.") from exc

    body_root = root.find("w:body", namespaces=WORD_NS)
    body: list[tuple[str, Any]] = []
    paragraphs: list[str] = []
    tables: list[list[list[str]]] = []

    if body_root is not None:
        for child in body_root:
            tag = etree.QName(child).localname if child.tag is not etree.Comment else ""
            if tag == "p":
                text = "".join(child.xpath(".//w:t/text()", namespaces=WORD_NS))
                if text.strip():
                    paragraphs.append(text)
                    body.append(("p", text))
            elif tag == "tbl":
                rows = _table_rows(child)
                tables.append(rows)
                body.append(("tbl", rows))

    return WordDocument(
        paragraphs=paragraphs,
        tables=tables,
        header_text=header_text,
        footer_text=footer_text,
        body=body,
    )


# ── Excel (.xlsx / .xls) ─────────────────────────────────────────────────────

@dataclass(frozen=True)
class ExcelSheet:
    name: str
    rows: list[list[str]]
    header_text: str = ""
    footer_text: str = ""


@dataclass(frozen=True)
class ExcelWorkbook:
    sheets: list[ExcelSheet]

    @property
    def sheet_names(self) -> list[str]:
        return [sheet.name for sheet in self.sheets]


def read_excel(data: bytes, extension: str) -> ExcelWorkbook:
    ext = (extension or "").lower()
    if ext == ".xlsx":
        return _read_xlsx(data)
    if ext == ".xls":
        return _read_xls(data)
    raise DocumentReadError(f"지원하지 않는 Excel 확장자입니다: {ext or '(없음)'}")


def _read_xlsx(data: bytes) -> ExcelWorkbook:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=False)
    except Exception as exc:
        raise DocumentReadError("xlsx 파일을 읽을 수 없습니다.") from exc

    sheets = []
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows():
            rows.append([excel_cell_text(cell.value) for cell in row])
        sheets.append(
            ExcelSheet(
                name=worksheet.title,
                rows=trim_empty_edges(rows),
                header_text=_worksheet_header_text(worksheet),
                footer_text=_worksheet_footer_text(worksheet),
            )
        )
    workbook.close()
    return ExcelWorkbook(sheets=sheets)


def _worksheet_header_text(worksheet) -> str:
    parts = []
    for header in (worksheet.oddHeader, worksheet.evenHeader, worksheet.firstHeader):
        for section in (header.left, header.center, header.right):
            text = getattr(section, "text", "") or ""
            if text:
                parts.append(text)
    return normalize_spaces(" ".join(parts))


def _worksheet_footer_text(worksheet) -> str:
    parts = []
    for footer in (worksheet.oddFooter, worksheet.evenFooter, worksheet.firstFooter):
        for section in (footer.left, footer.center, footer.right):
            text = getattr(section, "text", "") or ""
            if text:
                parts.append(text)
    return normalize_spaces(" ".join(parts))


def _read_xls(data: bytes) -> ExcelWorkbook:
    try:
        import xlrd
    except ImportError as exc:
        raise DocumentReadError("xls 파일을 읽으려면 xlrd 패키지가 필요합니다.") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=data)
    except Exception as exc:
        raise DocumentReadError("xls 파일을 읽을 수 없습니다.") from exc

    headers_by_sheet = _xls_print_records(data, opcode=0x0014)
    footers_by_sheet = _xls_print_records(data, opcode=0x0015)

    sheets = []
    for worksheet in workbook.sheets():
        rows = []
        for row_index in range(worksheet.nrows):
            rows.append([
                excel_cell_text(worksheet.cell_value(row_index, col_index))
                for col_index in range(worksheet.ncols)
            ])
        sheets.append(
            ExcelSheet(
                name=worksheet.name,
                rows=trim_empty_edges(rows),
                header_text=headers_by_sheet.get(worksheet.name, ""),
                footer_text=footers_by_sheet.get(worksheet.name, ""),
            )
        )
    return ExcelWorkbook(sheets=sheets)


def _xls_print_records(data: bytes, *, opcode: int) -> dict[str, str]:
    """.xls 워크북 스트림을 직접 파싱해 시트별 인쇄 머리글/바닥글을 추출한다.

    xlrd는 BIFF HEADER/FOOTER 레코드를 노출하지 않으므로 OLE2 컨테이너에서
    Workbook 스트림을 꺼낸 뒤 BOUNDSHEET(0x85)로 시트별 substream 위치를 찾고
    각 substream 첫 대상 레코드를 읽는다. 실패하면 빈 매핑을 반환한다.
    """
    try:
        from io import StringIO

        from xlrd.compdoc import CompDoc

        compdoc = CompDoc(data, logfile=StringIO())
        mem, base, size = compdoc.locate_named_stream("Workbook")
        if mem is None or not size:
            return {}
        stream = mem[base:base + size]
    except Exception:
        return {}

    try:
        boundsheets = []
        pos = 0
        total = len(stream)
        while pos + 4 <= total:
            record_opcode, length = struct.unpack("<HH", stream[pos:pos + 4])
            body = stream[pos + 4:pos + 4 + length]
            if record_opcode == 0x0085 and len(body) >= 8:  # BOUNDSHEET
                ply_pos = struct.unpack("<I", body[0:4])[0]
                name, _ = _xls_unicode_string(body, 6, 1)
                boundsheets.append((ply_pos, name))
            pos += 4 + length

        records = {}
        for ply_pos, name in boundsheets:
            records[name] = _xls_print_record_at(stream, ply_pos, opcode)
        return records
    except Exception:
        return {}


def _xls_print_record_at(stream: bytes, start: int, target_opcode: int) -> str:
    pos = start
    total = len(stream)
    depth = 0
    while pos + 4 <= total:
        opcode, length = struct.unpack("<HH", stream[pos:pos + 4])
        body = stream[pos + 4:pos + 4 + length]
        if opcode == 0x0809:  # BOF
            depth += 1
        elif opcode == 0x000A:  # EOF
            depth -= 1
            if depth <= 0:
                return ""
        elif opcode == target_opcode:
            if length == 0:
                return ""
            text, _ = _xls_unicode_string(body, 0, 2)
            return text
        pos += 4 + length
    return ""


def _xls_unicode_string(body: bytes, offset: int, cch_size: int):
    if cch_size == 1:
        cch = body[offset]
        cursor = offset + 1
    else:
        cch = struct.unpack("<H", body[offset:offset + 2])[0]
        cursor = offset + 2
    grbit = body[cursor]
    cursor += 1
    if grbit & 0x01:
        text = body[cursor:cursor + cch * 2].decode("utf-16-le", "replace")
    else:
        text = body[cursor:cursor + cch].decode("cp949", "replace")
    return text, cursor


# ── PDF ──────────────────────────────────────────────────────────────────────

def read_pdf_text(data: bytes, *, page_limit: int | None = None) -> str:
    try:
        import fitz
    except ImportError as exc:
        raise DocumentReadError("PDF를 읽으려면 PyMuPDF(fitz)가 필요합니다.") from exc

    try:
        texts = []
        with fitz.open(stream=data, filetype="pdf") as document:
            limit = document.page_count if page_limit is None else min(page_limit, document.page_count)
            for page_index in range(limit):
                texts.append(document[page_index].get_text("text"))
        return normalize_spaces(" ".join(texts))
    except DocumentReadError:
        raise
    except Exception as exc:
        raise DocumentReadError("PDF 파일을 읽을 수 없습니다.") from exc


def read_pdf_page_text(data: bytes, *, page_index: int = 0) -> str:
    """특정 페이지 한 장의 원본 텍스트(정규화 없이)를 반환한다."""
    try:
        import fitz
    except ImportError as exc:
        raise DocumentReadError("PDF를 읽으려면 PyMuPDF(fitz)가 필요합니다.") from exc

    data_bytes = data
    with fitz.open(stream=data_bytes, filetype="pdf") as document:
        if document.page_count <= page_index:
            return ""
        return document[page_index].get_text("text")
