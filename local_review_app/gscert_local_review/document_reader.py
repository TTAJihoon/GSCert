from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile


WORD_NAMESPACE = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
TEXT_TAG = f"{{{WORD_NAMESPACE}}}t"
PARAGRAPH_TAG = f"{{{WORD_NAMESPACE}}}p"
TABLE_TAG = f"{{{WORD_NAMESPACE}}}tbl"
ROW_TAG = f"{{{WORD_NAMESPACE}}}tr"
CELL_TAG = f"{{{WORD_NAMESPACE}}}tc"


class DocumentReadError(RuntimeError):
    """Raised when a local document cannot be parsed."""


@dataclass(frozen=True)
class WordDocument:
    paragraphs: list[str]
    tables: list[list[list[str]]]
    header_text: str
    footer_text: str

    @property
    def full_text(self) -> str:
        table_text = " ".join(cell for table in self.tables for row in table for cell in row if cell)
        return _normalize_spaces(" ".join([*self.paragraphs, table_text, self.header_text, self.footer_text]))


@dataclass(frozen=True)
class ExcelSheet:
    name: str
    rows: list[list[str]]


@dataclass(frozen=True)
class ExcelWorkbook:
    sheets: list[ExcelSheet]

    @property
    def sheet_names(self) -> list[str]:
        return [sheet.name for sheet in self.sheets]


def read_word_document(path: Path) -> WordDocument:
    try:
        with ZipFile(path) as archive:
            document_xml = _read_required_xml(archive, "word/document.xml")
            header_text = _read_part_texts(archive, "word/header")
            footer_text = _read_part_texts(archive, "word/footer")
    except (OSError, BadZipFile, KeyError, ElementTree.ParseError) as exc:
        raise DocumentReadError(f"Word 파일을 읽을 수 없습니다: {path.name}") from exc

    root = ElementTree.fromstring(document_xml)
    paragraphs = [
        text
        for paragraph in root.iter(PARAGRAPH_TAG)
        if (text := _element_text(paragraph))
    ]
    tables = [_table_rows(table) for table in root.iter(TABLE_TAG)]
    return WordDocument(
        paragraphs=paragraphs,
        tables=tables,
        header_text=header_text,
        footer_text=footer_text,
    )


def read_pdf_text(path: Path, *, page_limit: int | None = None) -> str:
    try:
        import fitz

        texts: list[str] = []
        with fitz.open(stream=path.read_bytes(), filetype="pdf") as document:
            limit = document.page_count if page_limit is None else min(page_limit, document.page_count)
            for page_index in range(limit):
                texts.append(document[page_index].get_text("text"))
        return _normalize_spaces(" ".join(texts))
    except Exception as exc:
        raise DocumentReadError(f"PDF 파일을 읽을 수 없습니다: {path.name}") from exc


def read_excel_workbook(path: Path) -> ExcelWorkbook:
    if path.suffix.lower() != ".xlsx":
        raise DocumentReadError(f"현재 로컬 앱은 .xlsx Excel 파일만 읽을 수 있습니다: {path.name}")
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets: list[ExcelSheet] = []
        for worksheet in workbook.worksheets:
            rows = [
                [_excel_cell_text(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            sheets.append(ExcelSheet(name=worksheet.title, rows=_trim_empty_edges(rows)))
        workbook.close()
        return ExcelWorkbook(sheets=sheets)
    except Exception as exc:
        raise DocumentReadError(f"Excel 파일을 읽을 수 없습니다: {path.name}") from exc


def _read_required_xml(archive: ZipFile, name: str) -> bytes:
    with archive.open(name) as handle:
        return handle.read()


def _read_part_texts(archive: ZipFile, prefix: str) -> str:
    texts: list[str] = []
    for name in sorted(archive.namelist()):
        if not name.startswith(prefix) or not name.endswith(".xml"):
            continue
        try:
            root = ElementTree.fromstring(archive.read(name))
        except (KeyError, ElementTree.ParseError):
            continue
        text = _element_text(root)
        if text:
            texts.append(text)
    return _normalize_spaces(" ".join(texts))


def _table_rows(table: ElementTree.Element) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in table.findall(ROW_TAG):
        rows.append([_element_text(cell) for cell in row.findall(CELL_TAG)])
    return rows


def _element_text(element: ElementTree.Element) -> str:
    return _normalize_spaces("".join(node.text or "" for node in element.iter(TEXT_TAG)))


def _normalize_spaces(value: str) -> str:
    return " ".join(str(value or "").split())


def _excel_cell_text(value: Any) -> str:
    if value is None:
        return ""
    return _normalize_spaces(str(value))


def _trim_empty_edges(rows: list[list[str]]) -> list[list[str]]:
    trimmed = [list(row) for row in rows]
    while trimmed and not any(cell for cell in trimmed[0]):
        trimmed.pop(0)
    while trimmed and not any(cell for cell in trimmed[-1]):
        trimmed.pop()
    if not trimmed:
        return []
    max_width = max((len(row) for row in trimmed), default=0)
    right = max_width
    while right > 0 and all(right > len(row) or not row[right - 1] for row in trimmed):
        right -= 1
    return [row[:right] for row in trimmed]
